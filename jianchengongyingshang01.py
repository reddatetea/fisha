'''
仓库供应商简称，运行后形成的字典保表到py文件
'''
import openpyxl
import pprint

def jianchenGongyingshang(fname):


    wb = openpyxl.load_workbook(fname)


    ws1= wb['供应商简称0']

    if  '供应商简称' not in wb.sheetnames:
        wb.create_sheet(title = '供应商简称')
    else :
        pass

    ws2 = wb['供应商简称']

    jianchenGongyingshangs = {}

    for row in range(1,ws1.max_row+1):
        if row == 1 :
            ws2.cell(row, 1).value = '仓库供应商'
            ws2.cell(row, 2).value = '仓库供应商简称'
        else :
            if ws1.cell(row, 2).value not in [None,'']:
                ws2.cell(row, 1).value = ws1.cell(row, 1).value
                ws2.cell(row, 2).value = ws1.cell(row, 2).value
                jianchenGongyingshangs[ws2.cell(row, 1).value]=ws1.cell(row, 2).value
            else:
                continue

    wb.save(fname)
    resultFile = open('jianchenGongyingshangs.py', 'w', encoding='utf-8')
    resultFile.write('jianchen = ' + pprint.pformat(jianchenGongyingshangs))
    return jianchenGongyingshangs

def main():
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\仓库供应商.xlsx'
    jianchenGongyingshang(fname)


if __name__=='__main__':
    main()

