'''
将暂估xls升级为xlsx,然后按标准格式导入到当月暂估文件中去
'''
import os
import openpyxl
import excelmessage
import datetime
import easygui

def zanguChuli(fname):
    dirname,filename = os.path.split(fname)
    os.chdir(dirname)
    wb0 = openpyxl.load_workbook(fname)
    ws0 = wb0.active

    print('请输入当月暂估的路径和文件名：')
    fname1 = excelmessage.wenjian()
    fname1 = excelmessage.excelMessage(fname1)
    wb1 = openpyxl.load_workbook(fname1,data_only=True)
    sheetnames = wb1.sheetnames
    msg = '请选择当月暂估工作表：'
    print('请选择当月暂估工作表：')
    sheet0 = easygui.buttonbox(msg = msg,title = '',choices=sheetnames)
    ws1 = wb1['双佳']
    ws2 = wb1['%s'%sheet0]      #ws2 = wb[sheet0]不行

    nian = ws0.cell(1,1).value[:4]

    #双佳
    for index,row in enumerate(ws0.values):
        if index == 0 :
            continue
        else :
            yue = row[0]
            ri = row[1]
            pingzhenhao = row[6]
            zhaiyao = row[7]
            jiefang = row[8]
            daifang =row[9]
            fangxiang = row[10]
            yuer = row[11]
            shuangjia_hang = (yue, ri, pingzhenhao, zhaiyao, jiefang, daifang, fangxiang, yuer)
            ws1.append(shuangjia_hang)

    for index, row in enumerate(ws0.values):
        if index == 0:
            continue
        else:
            if row[0] in ['', None] :
                continue
            else:
                yue = row[0]
                ri = row[1]
                pingzhenhao = row[6]
                zhaiyao = row[7]
                daifang = row[9]
                daifang_hanshui = round(daifang * 1.13, 2)
                if daifang < 0  or '调整' in zhaiyao:     #贷方金额为负数或摘要中有‘调整’两字，则piao='有'
                    piao = '有'
                else :
                    piao = '无'
                riqi = nian + '-' + yue + '-' + ri
                riqi8 = datetime.datetime.strptime(riqi, '%Y-%m-%d')
                dangyue_hang = (yue, ri, pingzhenhao, zhaiyao, '', daifang, daifang_hanshui, piao, riqi8)
                ws2.append(dangyue_hang)

    wb1.save(fname1)
    return fname1

def main():
    path = r'F:\a00nutstore\006\zw\yingfu\2021yingfu'
    os.chdir(path)
    print('请分别输入临时暂估路径和文件名：')
    fname = excelmessage.wenjian()
    fname  = excelmessage.excelMessage(fname)
    fname1 = zanguChuli(fname)
    os.startfile(fname1)

if __name__ == '__main__':
    main()



