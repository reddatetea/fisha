'''
对流水账进行操作,仅取供应商不为空的！
'''
# _*_ coding = uft-8 _*_
from openpyxl import Workbook,load_workbook
import re
import time
import datetime
import os

def liushuiZhang02(fname):

    wb1 = load_workbook(fname)
    sheet1 = wb1.active

    ss_wb2 = Workbook()
    ss_sheet2 = ss_wb2.active
    ss_sheet2.title = '流水账'

    for index, row in enumerate(sheet1.values):
        if index == 0:
            ss_sheet2.append(('日期','单据号','供货单位','品名','单位','入库数量','入库单价','入库金额'))
        else :
            print(row)
            if row[5] not in [None,''] :
                riqi = row[0]
                danjuhao = row[1]
                gonghuoshang = row[5]
                pinming = row[7]
                danwei = row[9]
                shuliang = row[13]
                danjia = row[11]
                jiner = row[12]

                meihang = (riqi,danjuhao,gonghuoshang,pinming,danwei,shuliang,danjia,jiner)
                ss_sheet2.append(meihang)
            else:
                continue

    fname = fname[:-5]+'001.xlsx'

    ss_wb2.save(fname)
    time.sleep(3)

    return fname

def main():
    fname = r'D:\a00nutstore\fishc\5.25流水账.xlsx'
    liushuiZhang02(fname)

if __name__=='__main__':
    main()









