import openpyxl
wb=openpyxl.load_workbook('工资表.xlsx')
sheet=wb['工资表']
#第一种
for cell in sheet['B2:B4']:
    print(cell[0].value)
#第二种
for row in sheet:
    for cell in row:
        print(cell.value,end=',')
    print()
#第三种
for row in sheet.iter_rows(min_row=1,max_row=4,max_col=4):
    for cell in row:
        print(cell.value,end=',')
    print()
#第四种
for col in sheet.columns:
    for cell in col:
        print(cell.value,end=',')
    print()
#第五种
for col in sheet.iter_cols(min_col=1,max_col=4,min_row=1,max_row=4):
    for cell in col:
        print(cell.value,end=',')
    print()

