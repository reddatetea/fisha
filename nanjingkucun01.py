'''
南京库存报表
'''
# _*_ conding:utf-8 _*_
import openpyxl
import os
import canchengpinrename01
import excelmessage


#期初数
def qichukucun():
    path = r'F:\a00nutstore\006\zw\fengognsi\南京\南京2020\南京202004'
    os.chdir(path)
    filename = r'202004月报南京分(新格式）.xlsx'
    fname = os.path.join(path, filename)
    wb = openpyxl.load_workbook(fname)
    ws = wb['01月进销存表']
    max_row = ws.max_row

    #将库存商品命名标准化
    for row in  range(2, max_row + 1):
        oldpinming = ws.cell(row, 1).value
        pinming = canchengpinrename01.canchengpinrename(oldpinming)
        ws.cell(row,18,value=pinming)
    wb.save(fname)


    #pinmings,qichu_ben,qichu_jian,qichu_jiner
    path = r'F:\a00nutstore\006\zw\fengognsi\南京\南京2020\南京202004'
    os.chdir(path)
    filename = r'202004月报南京分(新格式）.xlsx'
    fname = os.path.join(path, filename)
    wb = openpyxl.load_workbook(fname)
    ws = wb['01月进销存表']
    max_row = ws.max_row

    pinmings = set('')
    qichus = {}

    qichu_ben_total = 0
    qichu_jian_total = 0
    qichu_jiner_total = 0
    for row in range(2, max_row + 1):
        pinming = ws.cell(row,18).value
        if pinming not in ['',None]:
            pinmings.add(pinming)
        else :
            continue

        qichus.setdefault(pinming, {'qichu_ben': 0, 'qichu_jian': 0,'qichu_jiner':0})

        qichu_ben = round(ws.cell(row,14).value, 2)
        qichu_ben_total += qichu_ben
        qichu_jian = round(ws.cell(row,15).value, 2)
        qichu_jian_total += qichu_jian
        qichu_jiner = round(ws.cell(row,17).value,2)
        qichu_jiner_total += qichu_jiner

        qichus[pinming]['qichu_ben'] += qichu_ben
        qichus[pinming]['qichu_jian'] += qichu_jian
        qichus[pinming]['qichu_jiner'] += qichu_jiner


    print(qichus)
    print('qichu_ben', qichu_ben_total)
    print('qichu_jian', qichu_jian_total)
    print('qichu_jiner', qichu_jiner_total)

    wb.close()
    return qichus,pinmings

def nanjingruku(pinmings):
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    print(fname)
    wb = openpyxl.load_workbook(fname)
    ws = wb['sheet1']
    max_row = ws.max_row
    # 将库存商品命名标准化
    for row in range(2, max_row + 1):
        oldpinming = ws.cell(row,14).value
        pinming = canchengpinrename01.canchengpinrename(oldpinming)
        ws.cell(row,15, value=pinming)
    wb.save(fname)

    wb = openpyxl.load_workbook(fname)

    rukus = {}

    ruku_ben_total = 0
    ruku_jian_total = 0
    ruku_jiner_total = 0
    for row in range(2, max_row):
        pinming = ws.cell(row,15).value
        rukus.setdefault(pinming, {'ruku_ben': 0, 'ruku_jian': 0, 'ruku_jiner': 0})
        yewuyuan = ws.cell(row,4).value
        if yewuyuan == '唐加权':
            if pinming not in ['', None]:
                pinmings.add(pinming)
                ruku_ben = ws.cell(row, 8).value
                if ruku_ben in ['',None]:
                    ruku_ben = 0
                else :
                    ruku_ben = ruku_ben

                ruku_ben_total += ruku_ben

                ruku_jian = ws.cell(row, 9).value

                if ruku_jian in ['',None]:
                    ruku_jian = 0
                else :
                    ruku_jian = ruku_jian


                ruku_jian_total += ruku_jian

                ruku_jiner = ws.cell(row, 10).value
                if ruku_jiner in ['',None]:
                    ruku_jiner = 0
                else :
                    ruku_jiner = round(ruku_jiner,2)

                ruku_jiner_total += ruku_jiner

                rukus[pinming]['ruku_ben'] += ruku_ben
                rukus[pinming]['ruku_jian'] += ruku_jian
                rukus[pinming]['ruku_jiner'] += ruku_jiner

                print(rukus)
                print('ruku_ben_total', ruku_ben_total)
                print('ruku_jian_total', ruku_jian_total)
                print('ruku_jiner_total',ruku_jiner_total)

            else:
                continue


        else:
            continue

    wb.save(fname)
    print(rukus)


    return rukus,pinmings


#出库汇总
def nanjingchukus(pinmings):
    path = r'F:\a00nutstore\006\zw\fengognsi\南京\南京2020\南京202004'
    os.chdir(path)
    filename = r'202004月报南京分(新格式）.xlsx'
    fname = os.path.join(path, filename)
    wb = openpyxl.load_workbook(fname)
    ws = wb['01月结算清单汇总']
    max_row = ws.max_row

    # 将库存商品命名标准化
    for row in range(2, max_row + 1):
        oldpinming = ws.cell(row,5).value
        pinming = canchengpinrename01.canchengpinrename(oldpinming)
        ws.cell(row,15,value=pinming)
    wb.save(fname)

    wb = openpyxl.load_workbook(fname)

    chukus = {}
    chuku_jian_total = 0
    chuku_ben_total = 0

    for row in range(2, max_row + 1):
        pinming = ws.cell(row,15).value
        if pinming not in ['', None]:
            pinmings.add(pinming)
            chukus.setdefault(pinming, { 'chuku_jian': 0,'chuku_ben': 0,})

            chuku_jian = ws.cell(row, 7).value
            if chuku_jian not in [0,'',None]:

                chuku_jian = chuku_jian
                chuku_jian_total += chuku_jian
                # 含量
                hanliang = ws.cell(row, 8).value
                if hanliang not in [0, '', None]:
                    hanliang = hanliang
                    # 本数=件数*含量
                    chuku_ben = chuku_jian * hanliang
                    chuku_ben_total += chuku_ben

                    chukus[pinming]['chuku_jian'] += chuku_jian
                    chukus[pinming]['chuku_ben'] += chuku_ben

                else:
                    hangliang = 0



            else:
                chuku_jian = 0


            print(chukus)
            print('chuku_jian', chuku_jian_total)
            print('chuku_ben', chuku_ben_total)
            print(pinmings)

        else:
            continue

    wb.save(fname)
    return chukus, pinmings


#def nanjingkucun(pinmings,qichus,rukus,chukus):



def main():
    qichus, pinmings = qichukucun()
    rukus,pinmings = nanjingruku(pinmings)
    chukus,pinmings = nanjingchukus(pinmings)
    #nanjingkucun(pinmings,qichus,rukus,chukus)


if __name__ == '__main__':
    main()












