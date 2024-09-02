from openpyxl.styles import PatternFill,GradientFill
from openpyxl import load_workbook
fname = r'F:\a00nutstore\fishc\test.xlsx'
wb = load_workbook(fname)
ws = wb.active
cell_d2 = ws['D2']
cell_d3 = ws['D3']

pattern_fill = PatternFill(fill_type = 'solid',fgColor='FFFF00')  #纯黄
cell_d2.fill = pattern_fill

gradient_fill = GradientFill(stop=('FFFFFF','FFFF00','000000'))   #白，纯黄，黑
cell_d3.fill = gradient_fill


wb.save(fname)