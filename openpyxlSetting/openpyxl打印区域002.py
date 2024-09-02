import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

fname = r'D:\a00nutstore\fishc\test.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active

max_row = ws.max_row      #最大行数
print(max_row)
max_column = ws.max_column   #最大列数
print(max_column)

max_column_letter = get_column_letter(max_column)    # 根据列的数字返回字母

# 根据字母返回列的数字
#print(column_index_from_string('D')) # 4

max_cell = '$%s$%s'%(max_column_letter,max_row)   #max_row和mar_column对应的单元格
print_area = '$A$1:{}'.format(max_cell)        #打印区域
ws.print_area = print_area                     #设置打印区域

wb.save(fname)