import openpyxl
from openpyxl.styles import Font,PatternFill,GradientFill,Side,Border,Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import xlwings as xw

fname = r'D:\a00nutstore\fishc\test.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
max_row = ws.max_row

ws.row_dimensions[1].height = 20
ws.column_dimensions['B'].width = 20

font = Font(name = '微软雅黑',size=12,bold=True,italic=True,color='000000')
pattern_fill = PatternFill(fill_type = 'solid',fgColor='40E0D0')  #绿宝石
alignment = Alignment(horizontal = 'center',vertical = 'center',text_rotation = 0,wrap_text = False)

side = Side(style = 'thin',color = '000000')        #边框 线型为thin,黑色
border = Border(left = side,right = side,top = side,bottom = side)    #定义边框的左右上下


for first_row in ws.iter_rows(min_row=1,  max_row=1):

    i = 0
    for cell in first_row:
        i = i + 1
        cell.font = font
        cell.fill = pattern_fill
        cell.alignment = alignment
        cell.border = border
        #width = len(str(cell.value))*2 + 1
        cell_column_letter = get_column_letter(i)
        #ws.column_dimensions[cell_column_letter].width = 20        #设置列宽

font1 = Font(name = '微软雅黑',size=10,bold=False,italic=False,color='000000')


for nextrows in ws.iter_rows(min_row=2, max_row= max_row):
    i = 0
    for cell in nextrows:
        i = i + 1
        cell.font = font1
        cell.border = border


wb.save(fname)

app = xw.App(visible = False,add_book = False)
wb = app.books.open(fname)
ws = wb.sheets.active
ws.autofit()
wb.save(fname)
wb.close()
app.quit()