import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import easygui
import excelmessage06

def setPrintArea(fname):

    #fname = r'D:\a00nutstore\fishc\test.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选要设置打印区域的excel工作表'
    if len(sheetnames) == 1 :
        ws = wb.active
    else :
        print(msg)
        choice = easygui.buttonbox(msg,title='设置打印区域',choices=sheetnames)
        ws = wb['%s'%choice]

    max_row = ws.max_row      #最大行数
    print(max_row)
    max_column = ws.max_column   #最大列数
    print(max_column)

    max_column_letter = get_column_letter(max_column)    # 根据列的数字返回字母

    # 根据字母返回列的数字
    #print(column_index_from_string('D')) # 4

    max_cell = '$%s$%s'%(max_column_letter,max_row)   #max_row和mar_column对应的单元格
    print_area = '$A$1:{}'.format(max_cell)        #打印区域
    ws.print_area = print_area                     #设置打印区域

    wb.save(fname)

def main():
    fname = excelmessage06.excelMessage()
    msg = '请点选要进行打印设置的excel文件'
    #fname = easygui.fileopenbox(msg,title = '设置打印区域')
    setPrintArea(fname)

if __name__ == '__main__':
    main()