from openpyxl.styles import Font
from openpyxl import load_workbook
fname = r'F:\a00nutstore\fishc\test.xlsx'
wb = load_workbook(fname)
ws = wb.active
cell = ws['A2']
font = Font(name = '微软雅黑',size=20,bold=True,italic=True,color='87CEEB')   #SkyBlue
cell.font = font
wb.save(fname)
