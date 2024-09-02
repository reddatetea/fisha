import openpyxl

fname = r'D:\a00nutstore\fishc\openpyxl设置\openpyxl数字格式01.xlsx'

wb = openpyxl.Workbook()
ws = wb.active

ws['a1'] = -1000012.12
ws['A1'].number_format = '#,##0.00'

ws['A2'] = 111111.55
ws['A2'].number_format = '"￥"#,##0.00;-"￥"#,##0.00'

ws.move_range('a1:a2',0,2)
ws.delete_cols(1,1)
ws.column_dimensions['B'].width = 30.88
wb.copy_worksheet(ws).title = 'caigou'

wb.save(fname)
