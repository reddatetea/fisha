#获取字体样式
from openpyxl.styles import Font,PatternFill,GradientFill,Side,Border,Alignment
from openpyxl import load_workbook
fname = r'D:\a00nutstore\006\zw\2020\202008\202008双佳材料出库正正.xlsx'
wb = load_workbook(fname)
ws = wb.active
cell = ws['C1']
font = cell.font
pattern_fill = cell.fill
alignment = cell.alignment
border = cell.border
print(font.name,font.size,font.bold,font.italic,font.color)      #获取单元格字体信息：名字、大小、粗体、斜体、颜色
#print(font)
print(pattern_fill)     #获取单元格填充信息：名字、大小、粗体、斜体、颜色
print(alignment)         #对齐
print(border)             #边框

wb.close()
