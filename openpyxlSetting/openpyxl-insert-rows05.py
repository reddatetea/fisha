import os
import openpyxl
import easygui
from openpyxl.utils import get_column_letter, column_index_from_string

fname = r'D:\a00nutstore\006\zw\2020\202008\202008双佳材料出库正正.xlsx'
path,filename = os.path.split(fname)
os.chdir(path)
wb = openpyxl.load_workbook(fname)

sheetnames = wb.sheetnames
print(sheetnames)

# msg = '请选择工作表'
# choice = easygui.buttonbox(msg,title = msg,choices = sheetnames)
# print(choice)
ws = wb['出库参考 Copy']

max_row = ws.max_row
print(max_row)

kemubianma= set()      #创建一个空集合

kemudic = {'1403001':'纸', '1403006':'灰板', '1403003':'辅料', '1403002':'包装物'}

chukushuliang_col = column_index_from_string('AH')
chukudanjia_col = column_index_from_string('AI')
chukujiner_col = column_index_from_string('AJ')

for i in ws.iter_cols(min_row = 2,min_col = 3,max_row = max_row,max_col=3):
    for cell in i :

        kemubianma.add(cell.value[:7])    #集合中添加元素是add，不是append

        if cell.value[:7] == '1403002'  and  (ws.cell(row =cell.row-1,column = 3).value[:7]== '1403001'):
            zhiXiaoji_row = cell.row
            ws.insert_rows(idx = zhiXiaoji_row)
            ws.cell(row=zhiXiaoji_row, column=3).value = '纸小计'
            print(zhiXiaoji_row)
            ws['AJ{}'.format(zhiXiaoji_row)] = '=SUM(AJ2:AJ{})'.format(zhiXiaoji_row-1)


        if cell.value[:7] == '1403003' and  (ws.cell(row =cell.row-1,column = 3).value[:7]== '1403002') :
            huibanXiaoji_row = cell.row
            ws.insert_rows( idx= huibanXiaoji_row)
            ws.cell(row=huibanXiaoji_row, column=3).value = '灰板小计'
            print(huibanXiaoji_row)
            ws['AH{}'.format(huibanXiaoji_row)] = '=SUM(AH{}:AH{})'.format(zhiXiaoji_row+1, huibanXiaoji_row - 1)
            ws['AJ{}'.format(huibanXiaoji_row)] = '=SUM(AJ{}:AJ{})'.format(zhiXiaoji_row+1,huibanXiaoji_row-1)

        if cell.value[:7] == '1403006' and  (ws.cell(row =cell.row-1,column = 3).value[:7]== '1403003') :
            fuliaoXiaoji_row = cell.row
            ws.insert_rows(idx=fuliaoXiaoji_row)
            ws.cell(row=fuliaoXiaoji_row, column=3).value = '辅料小计'
            print(fuliaoXiaoji_row)
            ws['AJ{}'.format(fuliaoXiaoji_row)] = '=SUM(AJ{}:AJ{})'.format(huibanXiaoji_row + 1, fuliaoXiaoji_row - 1)

        if cell.value[:7] == '1403006' and  (ws.cell(row =cell.row+1,column = 3).value in ['',None]) :
            baozhuangwuXiaoji_row = cell.row +1
            ws.cell(row=baozhuangwuXiaoji_row, column=3).value = '包装物小计'
            print(baozhuangwuXiaoji_row)
            ws['AJ{}'.format(baozhuangwuXiaoji_row)] = '=SUM(AJ{}:AJ{})'.format(fuliaoXiaoji_row + 1, baozhuangwuXiaoji_row - 1)

heji_row = baozhuangwuXiaoji_row + 1
ws.cell(row=heji_row, column=3).value = '合计'
print(heji_row)
ws['AJ{}'.format(heji_row)] = '=AJ{}+AJ{}+AJ{}+AJ{}'.format(zhiXiaoji_row,huibanXiaoji_row,fuliaoXiaoji_row,baozhuangwuXiaoji_row)

print(zhiXiaoji_row,huibanXiaoji_row,fuliaoXiaoji_row,baozhuangwuXiaoji_row)
print(baozhuangwuXiaoji_row)
print(kemubianma)

wb.save(fname)
