import openpyxl

fname3 = r'F:\a00nutstore\006\zw\baiyun\白云当月入库.xlsx'
wb3 = openpyxl.load_workbook(fname3)
ws3 = wb3['当月']

def deleKonghang():
    for i in range(2, ws3.max_row + 1):
        if ws3.cell(i,1).value == None:
            delelist.append(i)
        else :
            continue






ws3.delete_rows(delelist)

wb3.save(fname3)