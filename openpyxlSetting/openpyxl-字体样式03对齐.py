#对齐字体样式
from openpyxl.styles import Alignment
from openpyxl import load_workbook
fname = r'F:\a00nutstore\fishc\test.xlsx'
wb = load_workbook(fname)
ws = wb.active
cell = ws['A3']
alignment = Alignment(horizontal = 'center',vertical = 'center',text_rotation = 45,wrap_text = True)
cell.alignment = alignment

wb.save(fname)
