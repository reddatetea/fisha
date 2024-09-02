import os
import openpyxl
import excelmessage

# fname = excelmessage.wenjian()
# fname = excelmessage.excelMessage(fname)
fname = r'F:\a00nutstore\fishc\888.xlsx'
wb = openpyxl.load_workbook(fname)
ws3 = wb.worksheets[1]
max_row = ws3.max_row
print(max_row)
for cell in ws3['A']:
       if cell.value == '(I订制缝线本)小计:':
            chaoben_row = cell.row + 1
            print(chaoben_row)

       if cell.value == '(X装订配件)小计:':
            zhangben_row = cell.row + 1
            print(zhangben_row)

ws3.cell(chaoben_row,1).value = '抄本合计（含锐意）'
ws3.cell(zhangben_row,1).value = '账本合计'
wb.save('888-1.xlsx')










