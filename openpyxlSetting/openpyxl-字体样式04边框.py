from openpyxl.styles import Side,Border
from openpyxl import load_workbook
fname = r'F:\a00nutstore\fishc\test.xlsx'
wb = load_workbook(fname)
ws = wb.active
cell = ws['D6']
side1 = Side(style = 'thin',color = '00008B')
side2 = Side(style = 'thick',color = '00008B')
border = Border(left = side1,right = side2,top = side2,bottom = side1)
cell.border = border

wb.save(fname)