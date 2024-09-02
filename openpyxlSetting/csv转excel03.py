import openpyxl
import csv
 
# 创建工作簿对象
work_book = openpyxl.Workbook()
 
# 创建sheet
work_sheet = work_book.create_sheet(title=u"双佳数量")
 
# 打开csv文件
csvfile = open('双佳数量000.csv')
 
# 获取csv.reader
lines = csv.reader(csvfile)
 
# row
row = 1
 
# csv文件中假如不包含标题可以使用如下代码加入
# title_list 中写入每个标题名称以逗号相隔
# title_list = [u'标题']
#
# # 写入第一行，标题
# for i in range(1, title_list.__len__() + 1):
#     work_sheet.cell(row=row, column=i).value = title_list[i - 1]
 
# 写入从csv读取的内容 如使用了以上代码 这里行数要加一
for line in lines:
    # print(line)
    lin=1
    
    chang = len(line)
    
    for j in range(chang):
        if j==0 :
            continue
            j=j+1
            lin=lin+1
        else :
            work_sheet.cell(row=row, column=lin).value = line[j]
            lin=lin+1
            j = j+1
    row += 1
 
# 关闭文件
csvfile.close()
 
# 保存工作表
work_book.save('双佳数量000.xlsx')
