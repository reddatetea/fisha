import openpyxl

wb = openpyxl.load_workbook('openpyxl创建新工作表.xlsx')
sheetnames = wb.sheetnames
print(sheetnames)
'''
ws2 = wb.create_sheet(title = 'two')
ws3 = wb.create_sheet(title = 'three')
ws4 = wb.create_sheet(title = 'four')
ws5 = wb.create_sheet(title = 'five')

wb.save('openpyxl创建新工作表.xlsx')

'''

delelist = ['two1','three1','five1','four1']

#批量删除工作表
for i in delelist:
    ws = wb['%s'%i]
    wb.remove_sheet(ws)

wb.save('openpyxl创建新工作表.xlsx')



