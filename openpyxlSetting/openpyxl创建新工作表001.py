import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet(title = 'one')
ws['A1']= 'zww'
wb.save('openpyxl创建新工作表.xlsx')