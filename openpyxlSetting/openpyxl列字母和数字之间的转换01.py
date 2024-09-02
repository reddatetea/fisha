import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('openpyxl创建新工作表.xlsx')
sheetnames = wb.sheetnames

