import os
import openpyxl
import easygui

fname = r'D:\a00nutstore\006\zw\2020\202008\202008双佳材料出库正正.xlsx'
path,filename = os.path.split(fname)
os.chdir(path)
wb = openpyxl.load_workbook(fname)

sheetnames = wb.sheetnames
print(sheetnames)

msg = '请选择工作表'
choice = easygui.buttonbox(msg,title = msg,choices = sheetnames)
print(choice)
ws = wb['%s'%choice]

max_row = ws.max_row
print(max_row)

wb.create_sheet('hihi')                   #创建一个名为hihi的工作表
wb.copy_worksheet(ws)                #复制工作表，到工作表最后一个

sheetnames = wb.sheetnames
ws_name = sheetnames[-1]
ws = wb[ws_name]
ws.title = 'tang'
print(sheetnames)

wb.save(fname)

