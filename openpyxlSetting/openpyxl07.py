'''
此版可将盘存表抬头的字典全部打印出来
'''
from openpyxl import Workbook,load_workbook
#from openpyxl.utils import get_column_letter
import re
#import chaiwuName


wb =load_workbook('abc.xlsx')
sheet = wb.active

#print(sheet.max_row)
#print(sheet.max_column)
jishu = 0
first_list = []
for index,row in enumerate(sheet.values):
    print(row[1])
    first_list.append(row[1])


    if row[0] == None:
        jishu = jishu +1
        #print(jishu)



'''
实际行数mrows,当一行中第一个元素的值为None时，表示这一行实际数据是没有的！从这一行开始计数并汇总，将这个汇总数从
max_row中减去后，就是实际行数！
'''

mrows = sheet.max_row -jishu
#print(mrows)

#print(first_list)
#haiwuName.zileicailiao_rename(first_list)



file1 = open('F:\\00nutstore\\fishc\\材料标准命名.txt','w',encoding = 'utf-8')
file2 = open('F:\\00nutstore\\fishc\\异常命名明细.txt','w',encoding = 'utf-8')

cw_names = []
#不能匹配的列表abnormal_list
abnormal_lists = []

for int_string in first_list:

    if int_string =='':
        cw_name = ''
        cw_names.append(cw_name)

    else :
        if '卷筒' in int_string:
            if "上白" in int_string or "中白" in int_string or "下白" in int_string or "上红" in int_string or "中红" in int_string or "上红" in int_string or "上蓝" in int_string or "中蓝" in int_string or "下蓝" in int_string:
                leixing = '无碳复写卷筒纸'

            else:
                leixing = '卷筒纸'

            pattern = r'(?P<ke>\d{2,3})g?(?:.*)(?P<leixing>卷筒).*'
            regexp = re.compile(pattern)
            mat = regexp.search(int_string)

            if mat == None :
                cw_name = int_string
                cw_names.append(cw_name)
                abnormal_lists.append(cw_name)

            else :
                # 克重
                ke = mat.group('ke')

                zddd = ''
                # 该材料的财务软件标准命名
                # cw_name = ke + 'g' + leixing + zddd
                cw_name = leixing + zddd
                # print(cw_name)

                # 该材料的财务软件标准命名
                # cw_name = ke + 'g' + leixing + zddd

                # print(cw_name)

                cw_names.append(cw_name)









        elif '不干胶' in int_string:
            cw_name = '不干胶'
            cw_names.append(cw_name)



        elif '箱' in int_string:
            if '封箱'  in int_string:
                cw_name = '封箱胶'
                cw_names.append(cw_name)
            else :
                cw_name = '纸箱'
                cw_names.append(cw_name)


        elif '卷膜POF' in int_string:
            cw_name = '卷膜POF'
            cw_names.append(cw_name)

        elif '哑膜' in int_string:
            cw_name = '哑膜'
            cw_names.append(cw_name)

        elif '光膜' in int_string:
            cw_name = '光膜'
            cw_names.append(cw_name)

        elif '缠绕卷膜' in int_string:
            cw_name = '缠绕卷膜'
            cw_names.append(cw_name)

        elif '打包带' in int_string:
            cw_name = '打包带'
            cw_names.append(cw_name)

        elif '收缩袋' in int_string:
            cw_name = '收缩袋'
            cw_names.append(cw_name)

        elif '空白袋' in int_string or '吊袋' in int_string:
            cw_name = '空白袋'
            cw_names.append(cw_name)

        elif '手提' in int_string or '提手' in int_string:
            cw_name = '手提袋'
            cw_names.append(cw_name)

        elif '书套' in int_string:
            cw_name = '书套'
            cw_names.append(cw_name)

        elif '小文具胶' in int_string:
            cw_name = '小透明胶'
            cw_names.append(cw_name)

        elif '封面' in int_string or '衬膜' in int_string:
            cw_name = '封面'
            cw_names.append(cw_name)

        elif '胶套' in int_string:
            if '水胶套' in int_string :
                cw_name = int_string
                cw_names.append(cw_name)
            else :
                cw_name = '胶套'
                cw_names.append(cw_name)


        elif '热熔胶' in int_string or '波士胶' in int_string:
            cw_name = '热熔胶'
            cw_names.append(cw_name)

        elif '816胶' in int_string or '8475胶' in int_string:
            cw_name = '816胶'
            cw_names.append(cw_name)

        elif '封箱胶' in int_string:
            cw_name = '封箱胶'
            cw_names.append(cw_name)

        elif '啫喱胶' in int_string:
            cw_name = '啫喱胶'
            cw_names.append(cw_name)

        elif '大双面胶德莎' in int_string or '3mmE1060H双面胶' in int_string:
            cw_name = '进口双面胶'
            cw_names.append(cw_name)

        elif '聚乙烯醇' in int_string:
            cw_name = '聚乙烯醇'
            cw_names.append(cw_name)

        elif '线圈' in int_string or '线环' in int_string:
            cw_name = '线圈'
            cw_names.append(cw_name)

        elif '书边带' in int_string:
            cw_name = '堵头布'
            cw_names.append(cw_name)

        elif '无纺布' in int_string:
            cw_name = '无纺布'
            cw_names.append(cw_name)

        elif '电化铝' in int_string:
            cw_name = '电化铝'
            cw_names.append(cw_name)

        elif '票据夹' in int_string:
            cw_name = '票据夹'
            cw_names.append(cw_name)

        elif '票据夹' in int_string:
            cw_name = '票据夹'
            cw_names.append(cw_name)

        elif '16KPP账夹' in int_string:
            cw_name = '16kpp帐夹'
            cw_names.append(cw_name)

        elif '25KPP账夹' in int_string:
            cw_name = '25kpp帐夹'
            cw_names.append(cw_name)

        elif 'PP铁夹' in int_string:
            cw_name = 'PP铁夹'
            cw_names.append(cw_name)

        elif '袖珍账夹' in int_string:
            cw_name = '袖珍帐夹'
            cw_names.append(cw_name)

        elif '账钉' in int_string or '活页钉' in int_string:
            cw_name = '账钉'
            cw_names.append(cw_name)

        elif '挂锁' in int_string:
            cw_name = '挂锁'
            cw_names.append(cw_name)

        elif '圆珠笔' in int_string:
            cw_name = '圆珠笔'
            cw_names.append(cw_name)

        elif '背条' in int_string:
            cw_name = '背条（米）'
            cw_names.append(cw_name)

        elif 'PP片' in int_string:
            cw_name = 'PP片'
            cw_names.append(cw_name)

        elif '铁丝' in int_string:
            cw_name = '铁丝'
            cw_names.append(cw_name)

        elif '白海绵' in int_string:
            cw_name = '海绵（张）'
            cw_names.append(cw_name)

        elif '3MM海绵' in int_string:
            cw_name = '海绵'
            cw_names.append(cw_name)

        elif '橡皮筋' in int_string:
            cw_name = '橡皮筋'
            cw_names.append(cw_name)

        elif '7MM油双面胶' in int_string:
            cw_name = '7MM双面胶'
            cw_names.append(cw_name)

        elif '丝带' in int_string :
            cw_name = '1分丝带'
            cw_names.append(cw_name)





        elif '灰板' in int_string or '双灰' in int_string:
            # cw_name = '灰板'
            # cw_names.append(cw_name)

            pattern = r'(?P<ke>\d\.?\d?\d?)(MM)?(?:.*?)(?P<leixing>灰板|双灰)(?:.*?)'
            regexp = re.compile(pattern)
            mat = regexp.search(int_string)

            if mat == None :
                cw_name = int_string
                cw_names.append(cw_name)
                abnormal_lists.append(cw_name)

            else :
                # 克重
                ke = mat.group('ke')

                # 类型
                leixing = mat.group('leixing')

                if 'MM' in int_string:
                    ke = str(int(10 * float(ke)))
                    cw_name = ke + 'mm' + leixing

                else:
                    cw_name = ke + 'g' + leixing

                # print(cw_name)
                cw_names.append(cw_name)

        else:
            pattern = r'(?P<ke>\d{2,3})g?(?:.*?)(?P<leixing>双胶|白卡|牛|铜|涂布|无碳|有光纸|黑卡|莱妮纹|镭射金|浅绿皮纹|特种纸)(?:.*?)(?P<changKuan>(\d{3}\*\d{4})|(\d{4}\*\d{3})|(\d{3}\*\d{3}))'
            regexp = re.compile(pattern)
            mat = regexp.search(int_string)

            if mat == None :
                cw_name = int_string
                cw_names.append(cw_name)
                abnormal_lists.append(cw_name)

            else:
                # 克重
                ke = mat.group('ke')

                # 类型
                leixing = mat.group('leixing')

                # 长和宽
                changKuan = mat.group('changKuan')

                if '牛' in leixing or '牛卡' in leixing or '牛皮' in leixing:
                    leixing = '牛卡'

                elif '铜' in leixing or '铜版' in leixing:
                    leixing = '铜版'

                elif '莱妮纹' in leixing or '特种纸' in leixing or '镭射金' in leixing or '浅绿皮纹' in leixing:
                    leixing = '特种纸'

                elif '有光' in leixing:
                    leixing = '有光纸'

                elif '黑卡' in leixing:
                    leixing = '黑卡'

                elif '白卡' in leixing:
                    leixing = '白卡'

                elif '无碳' in leixing or '无碳复写' in leixing or '无碳复写纸' in leixing:
                    leixing = '无碳复写纸'

                else:
                    leixing = leixing

                # 标准命名
                pinming = ke + 'g' + leixing + changKuan
                # print(pinming)

                # 长宽
                chang = re.compile('(?P<chang>\d{3,4})\*(?P<kuan>\d{3,4})').search(changKuan).group('chang')
                kuan = re.compile('(?P<chang>\d{3,4})\*(?P<kuan>\d{3,4})').search(changKuan).group('kuan')
                # print(chang)
                # print(kuan)

                '''
                if danwei = kg :
                   zhongLiang = rukushuliang
                else :
                   zhongLiang = ke/1000*chang*kuang *500 *rukushuliang

                卷筒纸的命名法有点不同

                '''
                # 判断该材料是正度还是大度
                mianji = int(chang) * int(kuan)
                # print(mianji)
                if mianji / 787 / 1092 <= 1.05:
                    zddd = '正度'
                else:
                    zddd = '大度'
                # print(zddd)

                # 该材料的财务软件标准命名
                cw_name = ke + 'g' + leixing + zddd
                cw_names.append(cw_name)

                # print(cw_name)

                # cw_names.append(cw_name)

cw_Newnames = []
for cw_name in cw_names:
    i = 0
    if cw_name == '17mm灰板':
        cw_name = '18mm灰板'
        cw_Newnames.append(cw_name)
    elif cw_name == '7mm灰板':
        cw_name = '75dm灰板'
        cw_Newnames.append(cw_name)
    elif '无碳复写纸' in cw_name:
        cw_name = '无碳复写纸'
        cw_Newnames.append(cw_name)

    else:
        cw_name = cw_name
        cw_Newnames.append(cw_name)



    file1.write('%s\n'%cw_name.strip())



    #print(cw_name)




#打印没有匹配的材料名称
#print('**********************')
#print('下面是没有匹配的材料：')
for abnormal_list in abnormal_lists:
    file2.write('%s\n'%abnormal_list)
    #print(abnormal_list)

file1.close()
file2.close()

#数据去向工作簿 Workbook
ss_wb = Workbook()
ss_sheet = ss_wb.active
ss_sheet.title = '流水账'


for index,row in enumerate(sheet.values):
    #print(index,cw_Newnames[index])
    ss_sheet.append(row+(cw_Newnames[index],))

ss_wb.save('abcd.xlsx')

wb1 =load_workbook('abcd.xlsx')
sheet1 = wb1.active

#创立七个字典，分别是期初数、入库数量、采购入库，半成品入库，出库数量，生产领料，结余数
qichu_dic = {}
ruku_dic = {}
caigou_dic ={}
banchengpin_dic = {}
chuku_dic ={}
lingliao_dic = {}
qimo_dic = {}

qichu = 0
ruku = 0
caigou = 0
banchengpin = 0
chuku = 0
lingliao = 0
qimo  = 0

yuancailiao = list(set(cw_Newnames))

print(yuancailiao)
yuancailiao.remove('品名')


#斯初数
qichu_total = 0
for i in yuancailiao :
    qichu = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue

            # 将上面计算出的实际行数，运行到下面的程序！不这样就会报错！
        if index < mrows:
            if row[10] == i:
                qichu = qichu + row[3]
            else:
               continue


    qichu_dic[i] = qichu
    qichu_total = qichu_total +qichu
#print(qichu_dic)

#入库数
ruku_total = 0
i = 0
for i in yuancailiao :
    ruku = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                ruku = ruku + row[5]
            else:
               continue
    ruku_dic[i] = ruku
    ruku_total = ruku_total + ruku
#print(ruku_dic)

#采购入库
caigou_total = 0
i = 0
for i in yuancailiao :
    caigou = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                caigou = caigou + row[4]
            else:
               continue
    caigou_dic[i] = caigou
    caigou_total = caigou_total + caigou

#半成品入库
banchengpin_total = 0
i = 0
for i in yuancailiao :
    banchengpin = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                banchengpin =banchengpin + row[9]
            else:
               continue
    banchengpin_dic[i] = banchengpin
    banchengpin_total = banchengpin_total + banchengpin

#出库数量
chuku_total = 0
i = 0
for i in yuancailiao :
    chuku = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                chuku =chuku + row[6]
            else:
               continue
    chuku_dic[i] = chuku
    chuku_total = chuku_total + chuku

#生产领料
lingliao_total = 0
i = 0
for i in yuancailiao :
    lingliao = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                lingliao =lingliao + row[8]
            else:
               continue
    lingliao_dic[i] = lingliao
    lingliao_total = lingliao_total + lingliao

#期末
qimo_total = 0
i = 0
for i in yuancailiao :
    qimo = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                qimo =qimo + row[7]
            else:
               continue
    qimo_dic[i] = qimo
    qimo_total = qimo_total + qimo



taitou_lists = ['qichu','ruku','caigou','banchengpin','chuku','lingliao','qimo']

print('******************************')
print('打印期初字典')
print('期初合计为：%s'%qichu_total)
print('******************************')

for key,value in qichu_dic.items():
    print('{key}:{value}'.format(key = key, value = value))

print('******************************')
print('打印入库数量字典')
print('入库数量合计为：%s'%ruku_total)
print('******************************')

for key,value in ruku_dic.items():
    print('{key}:{value}'.format(key = key, value = value))

print('******************************')
print('打印采购入库字典')
print('采购入库合计为：%s'%caigou_total)
print('******************************')

for key,value in caigou_dic.items():
    print('{key}:{value}'.format(key = key, value = value))

print('******************************')
print('打印半成品入库字典')
print('半成品入库合计为：%s'%banchengpin_total)
print('******************************')

for key,value in banchengpin_dic.items():
    print('{key}:{value}'.format(key = key, value = value))

print('******************************')
print('打印出库数量字典')
print('出库数量合计为：%s'%chuku_total)
print('******************************')

for key,value in chuku_dic.items():
    print('{key}:{value}'.format(key = key, value = value))

print('******************************')
print('打印生产领料字典')
print('生产领料合计为：%s'%lingliao_total)
print('******************************')

for key,value in lingliao_dic.items():
    print('{key}:{value}'.format(key = key, value = value))

print('******************************')
print('打印结存字典')
print('结存合计为：%s'%qimo_total)
print('******************************')

for key,value in qimo_dic.items():
    print('{key}:{value}'.format(key = key, value = value))











