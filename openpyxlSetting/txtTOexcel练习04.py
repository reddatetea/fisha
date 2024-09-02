# -*- coding:utf-8 -*-
import re
import datetime
from openpyxl import load_workbook

with open('C:\\Users\\asus\\dal\\fishc\\abc.txt','r',encoding='utf-8') as file :
    dataMat=[]
    for imput_str in file.readlines() :
        #去除字符串中间的空格
        #print(imput_str)
        temp = imput_str.split()
        imput_str = ' '.join(temp)
        dataMat.append(imput_str)
print(dataMat)

gz_dic = {'胡国华': '叉车工', '李进军': '搬运工', '徐露': '原材料库管员', '王诗兵': '成品库管员', '黄志刚': '搬运工', '戴玲': '会计', '刘革红': '叉车工', '吴永乐': '叉车工', '胡红霞': '成品库管员', '葛翠平': '零配件库管员', '周宗华': '叉车工', '肖腊美': '会计', '白亚会': '会计', '彭志刚': '叉车工', '江小珍': '成品库管员', '赵付平': '仓库主管', '周秀芝': '原材料库管员', '周太云': '原材料库管员', '刘俊英': '财务经理', '张文伟': '会计', '周文婷': '会计', '马金蓉': '会计'}

wb = load_workbook('东山头.xlsx')

# 选择默认的工作表
ws = wb.active

# 给工作表重命名
ws.title = '双佳仓库及财务信息表'

# 写入一行数据
#row = ['序号','姓名','年龄','性别','电话','身份证号', '序号','现居住所在地','是否有已满14天的健康检测证明','是否与确诊患者或疑似新冠肺炎患者有过接触','备注']
#ws.append(row)
i = 0 
for imput_str in dataMat:
    imput_re1 = r'^[\u4e00-\u9fa5]{2,3}'
    pattern1 = re.compile(imput_re1)
    res1 = re.search(pattern1,imput_str)
    name = res1.group()
    print(name)

    #读取身份证
    imput_re2 = r'(\d{6})(\d{4})(\d{2})(\d{2})(\d{3})([0-9]|X|x)'
    pattern2 = re.compile(imput_re2)
    res2 = re.search(pattern2,imput_str)
    idcard = res2.group().strip()
    birthYear = int(idcard[6:10])
    tdyear = int(datetime.date.today().strftime('%Y'))
    age= tdyear - birthYear
    print(len(idcard))
    senventeen = int(idcard[-2])
    if senventeen%2==0 :
        gender = '女'
    else :
        gender = '男'
    print(idcard,age,senventeen,gender)

    #读取住址
    imput_re3 = r'[\u4e00-\u9fa5]{4,30}\w*[\u4e00-\u9fa5]{4,30}\w*[-]*\d*'
    pattern3 = re.compile(imput_re3)
    res3 = re.search(pattern3,imput_str)
    address = res3.group()
    print(address)

    #读取电话号码
    imput_re4 = r'1[0-9]{10}$'
    pattern4 = re.compile(imput_re4)
    res4 = re.search(pattern4,imput_str)
    phonenumber = res4.group()
    print(phonenumber)

    #读取工种
    gongzhong = gz_dic[name]

#开始逐行写入数据：
    ws.cell(row = i +4 ,column = 1).value = i + 1
    ws.cell(row = i +4 ,column = 2).value = name
    ws.cell(row = i +4 ,column = 3).value = age
    ws.cell(row = i +4 ,column = 4).value = gender
    ws.cell(row = i +4 ,column = 5).value = phonenumber
    ws.cell(row = i +4 ,column = 6).value = idcard
    ws.cell(row = i +4 ,column = 7).value = gongzhong
    ws.cell(row = i +4 ,column = 8).value = address
    ws.cell(row = i +4 ,column = 9).value = '正在办理'
    ws.cell(row = i +4 ,column = 10).value = '无'

    i = i + 1

ws.cell(row = 2 ,column = 4).value = '湖北双佳纸品有限公司'
wb.save('fangong04.xlsx')
























