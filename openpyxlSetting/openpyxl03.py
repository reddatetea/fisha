'''
复制一个工作表的数据到另一工作表
并在数据表的后面加上公式，进行自动计算
计算实际最高行数，避免清除数据后会导致实际行数与max_row不一样而报错的问题！
'''
from openpyxl import Workbook,load_workbook
#from openpyxl.utils import get_column_letter
import re
#import chaiwuName
cw_names = range(6)


wb =load_workbook('123.xlsx')
sheet = wb.active

print(sheet.max_row)
print(sheet.max_column)
jishu = 0
first_list = []
for index,row in enumerate(sheet.values):
    print(row[1])
    first_list.append(row[1])


    if row[0] == None:
        jishu = jishu +1
        print(jishu)



'''
实际行数mrows,当一行中第一个元素的值为None时，表示这一行实际数据是没有的！从这一行开始计数并汇总，将这个汇总数从
max_row中减去后，就是实际行数！
'''

mrows = sheet.max_row -jishu
print(mrows)


#数据去向工作簿 Workbook
ss_wb = Workbook()
ss_sheet = ss_wb.active
ss_sheet.title = '流水账'

for index,row in enumerate(sheet.values):
    print(index,cw_names[index])
    ss_sheet.append(row+(cw_names[index],))

ss_wb.save('1234.xlsx')








