'''
此版本将去除数量中科目名称加进来
'''
# -*-coding = utf-8 -*-
from openpyxl import Workbook,load_workbook
#from openpyxl.utils import get_column_letter
import re
#import chaiwuName

#此版本的原材料盘存表是去除空行和小计后的
wb =load_workbook('原材料盘存表.xlsx')
sheet = wb.active

#print(sheet.max_row)
#print(sheet.max_column)
jishu = 0
first_list = []
for index,row in enumerate(sheet.values):
    print(row[1])
    first_list.append(row[1])

    if row[0] == None:
        jishu = jishu +1
        #print(jishu)

mrows = sheet.max_row -jishu

file1 = open('F:\\00nutstore\\fishc\\材料标准命名.txt','w',encoding = 'utf-8')
file2 = open('F:\\00nutstore\\fishc\\异常命名明细.txt','w',encoding = 'utf-8')

cw_names = []
#不能匹配的列表abnormal_list
abnormal_lists = []

for in_string in first_list:

    if in_string =='':
        cw_name = ''
        cw_names.append(cw_name)

    else :
        if '卷筒' in in_string:
            if "上白" in in_string or "中白" in in_string or "下白" in in_string or "上红" in in_string or "中红" in in_string or "上红" in in_string or "上蓝" in in_string or "中蓝" in in_string or "下蓝" in in_string:
                leixing = '无碳复写卷筒纸'

            else:
                leixing = '卷筒纸'

            pattern = r'(?P<ke>\d{2,3})g?(?:.*)(?P<leixing>卷筒).*'
            regexp = re.compile(pattern)
            mat = regexp.search(in_string)

            if mat == None :
                cw_name = in_string
                cw_names.append(cw_name)
                abnormal_lists.append(cw_name)

            else :
                # 克重
                ke = mat.group('ke')

                zddd = ''
                # 该材料的财务软件标准命名
                # cw_name = ke + 'g' + leixing + zddd
                cw_name = leixing + zddd
                # print(cw_name)

                # 该材料的财务软件标准命名
                # cw_name = ke + 'g' + leixing + zddd

                # print(cw_name)

                cw_names.append(cw_name)

        elif '不干胶' in in_string:
            cw_name = '不干胶'
            cw_names.append(cw_name)

        elif '箱' in in_string:
            if '封箱'  in in_string:
                cw_name = '封箱胶'
                cw_names.append(cw_name)
            else :
                cw_name = '纸箱'
                cw_names.append(cw_name)


        elif '卷膜POF' in in_string:
            cw_name = '卷膜POF'
            cw_names.append(cw_name)

        elif '哑膜' in in_string:
            cw_name = '哑膜'
            cw_names.append(cw_name)

        elif '光膜' in in_string:
            cw_name = '光膜'
            cw_names.append(cw_name)

        elif '缠绕卷膜' in in_string:
            cw_name = '缠绕卷膜'
            cw_names.append(cw_name)

        elif '打包带' in in_string:
            cw_name = '打包带'
            cw_names.append(cw_name)

        elif '收缩袋' in in_string:
            cw_name = '收缩袋'
            cw_names.append(cw_name)

        elif '空白袋' in in_string or '吊袋' in in_string:
            cw_name = '空白袋'
            cw_names.append(cw_name)

        elif '手提' in in_string or '提手' in in_string:
            cw_name = '手提袋'
            cw_names.append(cw_name)

        elif '书套' in in_string:
            cw_name = '书套'
            cw_names.append(cw_name)

        elif '小文具胶' in in_string:
            cw_name = '小透明胶'
            cw_names.append(cw_name)

        elif '封面' in in_string or '衬膜' in in_string:
            cw_name = '封面'
            cw_names.append(cw_name)

        elif '胶套' in in_string:
            if '水胶套' in in_string :
                cw_name = in_string
                cw_names.append(cw_name)
            else :
                cw_name = '胶套'
                cw_names.append(cw_name)


        elif '热熔胶' in in_string or '波士胶' in in_string:
            cw_name = '热熔胶'
            cw_names.append(cw_name)

        elif '816胶' in in_string or '8475胶' in in_string:
            cw_name = '816胶'
            cw_names.append(cw_name)

        elif '封箱胶' in in_string:
            cw_name = '封箱胶'
            cw_names.append(cw_name)

        elif '啫喱胶' in in_string:
            cw_name = '啫喱胶'
            cw_names.append(cw_name)

        elif '大双面胶德莎' in in_string or '3mmE1060H双面胶' in in_string:
            cw_name = '进口双面胶'
            cw_names.append(cw_name)

        elif '聚乙烯醇' in in_string:
            cw_name = '聚乙烯醇'
            cw_names.append(cw_name)

        elif '线圈' in in_string or '线环' in in_string:
            cw_name = '线圈'
            cw_names.append(cw_name)

        elif '书边带' in in_string:
            cw_name = '堵头布'
            cw_names.append(cw_name)

        elif '无纺布' in in_string:
            cw_name = '无纺布'
            cw_names.append(cw_name)

        elif '电化铝' in in_string:
            cw_name = '电化铝'
            cw_names.append(cw_name)

        elif '票据夹' in in_string:
            cw_name = '票据夹'
            cw_names.append(cw_name)

        elif '票据夹' in in_string:
            cw_name = '票据夹'
            cw_names.append(cw_name)

        elif '16KPP账夹' in in_string:
            cw_name = '16kpp帐夹'
            cw_names.append(cw_name)

        elif '25KPP账夹' in in_string:
            cw_name = '25kpp帐夹'
            cw_names.append(cw_name)

        elif 'PP铁夹' in in_string:
            cw_name = 'PP铁夹'
            cw_names.append(cw_name)

        elif '袖珍账夹' in in_string:
            cw_name = '袖珍帐夹'
            cw_names.append(cw_name)

        elif '账钉' in in_string or '活页钉' in in_string:
            cw_name = '账钉'
            cw_names.append(cw_name)

        elif '挂锁' in in_string:
            cw_name = '挂锁'
            cw_names.append(cw_name)

        elif '圆珠笔' in in_string:
            cw_name = '圆珠笔'
            cw_names.append(cw_name)

        elif '背条' in in_string:
            cw_name = '背条（米）'
            cw_names.append(cw_name)

        elif 'PP片' in in_string:
            cw_name = 'PP片'
            cw_names.append(cw_name)

        elif '铁丝' in in_string:
            cw_name = '铁丝'
            cw_names.append(cw_name)

        elif '白海绵' in in_string:
            cw_name = '海绵（张）'
            cw_names.append(cw_name)

        elif '3MM海绵' in in_string:
            cw_name = '海绵'
            cw_names.append(cw_name)

        elif '橡皮筋' in in_string:
            cw_name = '橡皮筋'
            cw_names.append(cw_name)

        elif '7MM油双面胶' in in_string:
            cw_name = '7MM双面胶'
            cw_names.append(cw_name)

        elif '丝带' in in_string :
            cw_name = '1分丝带'
            cw_names.append(cw_name)

        elif '灰板' in in_string or '双灰' in in_string:
            # cw_name = '灰板'
            # cw_names.append(cw_name)

            pattern = r'(?P<ke>\d\.?\d?\d?)(MM)?(?:.*?)(?P<leixing>灰板|双灰)(?:.*?)'
            regexp = re.compile(pattern)
            mat = regexp.search(in_string)

            if mat == None :
                cw_name = in_string
                cw_names.append(cw_name)
                abnormal_lists.append(cw_name)

            else :
                # 克重
                ke = mat.group('ke')

                # 类型
                leixing = mat.group('leixing')

                if 'MM' in in_string:
                    ke = str(int(10 * float(ke)))
                    cw_name = ke + 'mm' + leixing

                else:
                    cw_name = ke + 'g' + leixing

                # print(cw_name)
                cw_names.append(cw_name)

        else:
            pattern = r'(?P<ke>\d{2,3})g?(?:.*?)(?P<leixing>双胶|白卡|牛|铜|涂布|无碳|有光纸|黑卡|莱妮纹|镭射金|浅绿皮纹|特种纸)(?:.*?)(?P<changKuan>(\d{3}\*\d{4})|(\d{4}\*\d{3})|(\d{3}\*\d{3}))'
            regexp = re.compile(pattern)
            mat = regexp.search(in_string)

            if mat == None :
                cw_name = in_string
                cw_names.append(cw_name)
                abnormal_lists.append(cw_name)

            else:
                # 克重
                ke = mat.group('ke')

                # 类型
                leixing = mat.group('leixing')

                # 长和宽
                changKuan = mat.group('changKuan')

                if '牛' in leixing or '牛卡' in leixing or '牛皮' in leixing:
                    leixing = '牛卡'

                elif '铜' in leixing or '铜版' in leixing:
                    leixing = '铜版'

                elif '莱妮纹' in leixing or '特种纸' in leixing or '镭射金' in leixing or '浅绿皮纹' in leixing:
                    leixing = '特种纸'

                elif '有光' in leixing:
                    leixing = '有光纸'

                elif '黑卡' in leixing:
                    leixing = '黑卡'

                elif '白卡' in leixing:
                    leixing = '白卡'

                elif '无碳' in leixing or '无碳复写' in leixing or '无碳复写纸' in leixing:
                    leixing = '无碳复写纸'

                else:
                    leixing = leixing

                # 标准命名
                pinming = ke + 'g' + leixing + changKuan
                # print(pinming)

                # 长宽
                chang = re.compile('(?P<chang>\d{3,4})\*(?P<kuan>\d{3,4})').search(changKuan).group('chang')
                kuan = re.compile('(?P<chang>\d{3,4})\*(?P<kuan>\d{3,4})').search(changKuan).group('kuan')

                # 判断该材料是正度还是大度
                mianji = int(chang) * int(kuan)
                # print(mianji)
                if mianji / 787 / 1092 <= 1.05:
                    zddd = '正度'
                else:
                    zddd = '大度'
                # print(zddd)

                # 该材料的财务软件标准命名
                cw_name = ke + 'g' + leixing + zddd
                cw_names.append(cw_name)

cw_Newnames = []
for cw_name in cw_names:
    i = 0
    if cw_name == '17mm灰板':
        cw_name = '18mm灰板'
        cw_Newnames.append(cw_name)
    elif cw_name == '7mm灰板':
        cw_name = '75dm灰板'
        cw_Newnames.append(cw_name)
    elif '无碳复写纸' in cw_name:
        cw_name = '无碳复写纸'
        cw_Newnames.append(cw_name)

    else:
        cw_name = cw_name
        cw_Newnames.append(cw_name)
    file1.write('%s\n'%cw_name.strip())

#打印没有匹配的材料名称
#print('**********************')
#print('下面是没有匹配的材料：')
for abnormal_list in abnormal_lists:
    file2.write('%s\n'%abnormal_list)
    #print(abnormal_list)

file1.close()
file2.close()

#数据去向工作簿 Workbook
ss_wb = Workbook()
ss_sheet = ss_wb.active
ss_sheet.title = '流水账'


for index,row in enumerate(sheet.values):
    #print(index,cw_Newnames[index])
    ss_sheet.append(row+(cw_Newnames[index],))

ss_wb.save('abcd.xlsx')

wb1 =load_workbook('abcd.xlsx')
sheet1 = wb1.active

#创立七个字典，分别是期初数、入库数量、采购入库，半成品入库，出库数量，生产领料，结余数
qichu_dic = {}
ruku_dic = {}
caigou_dic ={}
banchengpin_dic = {}
chuku_dic ={}
lingliao_dic = {}
qimo_dic = {}

qichu = 0
ruku = 0
caigou = 0
banchengpin = 0
chuku = 0
lingliao = 0
qimo  = 0

yuancailiao = list(set(cw_Newnames))

print(yuancailiao)
yuancailiao.remove('品名')

#斯初数
qichu_total = 0
for i in yuancailiao :
    qichu = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue

            # 将上面计算出的实际行数，运行到下面的程序！不这样就会报错！
        if index < mrows:
            if row[10] == i:
                qichu = qichu + row[3]
            else:
               continue

    qichu_dic[i] = qichu
    qichu_total = qichu_total +qichu
#print(qichu_dic)

#入库数
ruku_total = 0
i = 0
for i in yuancailiao :
    ruku = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                ruku = ruku + row[5]
            else:
               continue
    ruku_dic[i] = ruku
    ruku_total = ruku_total + ruku
#print(ruku_dic)

#采购入库
caigou_total = 0
i = 0
for i in yuancailiao :
    caigou = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                caigou = caigou + row[4]
            else:
               continue
    caigou_dic[i] = caigou
    caigou_total = caigou_total + caigou

#半成品入库
banchengpin_total = 0
i = 0
for i in yuancailiao :
    banchengpin = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                banchengpin =banchengpin + row[9]
            else:
               continue
    banchengpin_dic[i] = banchengpin
    banchengpin_total = banchengpin_total + banchengpin

#出库数量
chuku_total = 0
i = 0
for i in yuancailiao :
    chuku = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                chuku =chuku + row[6]
            else:
               continue
    chuku_dic[i] = chuku
    chuku_total = chuku_total + chuku

#生产领料
lingliao_total = 0
i = 0
for i in yuancailiao :
    lingliao = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                lingliao =lingliao + row[8]
            else:
               continue
    lingliao_dic[i] = lingliao
    lingliao_total = lingliao_total + lingliao

#期末
qimo_total = 0
i = 0
for i in yuancailiao :
    qimo = 0
    for index, row in enumerate(sheet1.values):
        if index == 0:
            continue
        if index < mrows:
            if row[10] == i:
                qimo =qimo + row[7]
            else:
               continue
    qimo_dic[i] = qimo
    qimo_total = qimo_total + qimo

taitou_lists = ['qichu','ruku','caigou','banchengpin','chuku','lingliao','qimo']

print('******************************')
print('打印期初字典')
print('期初合计为：%s'%qichu_total)
print('******************************')

for key,value in qichu_dic.items():
    print('{key}:{value}'.format(key = key, value = value))
print('******************************')
print('打印入库数量字典')
print('入库数量合计为：%s'%ruku_total)
print('******************************')

for key,value in ruku_dic.items():
    print('{key}:{value}'.format(key = key, value = value))
print('******************************')
print('打印采购入库字典')
print('采购入库合计为：%s'%caigou_total)
print('******************************')

for key,value in caigou_dic.items():
    print('{key}:{value}'.format(key = key, value = value))
print('******************************')
print('打印半成品入库字典')
print('半成品入库合计为：%s'%banchengpin_total)
print('******************************')

for key,value in banchengpin_dic.items():
    print('{key}:{value}'.format(key = key, value = value))
print('******************************')
print('打印出库数量字典')
print('出库数量合计为：%s'%chuku_total)
print('******************************')

for key,value in chuku_dic.items():
    print('{key}:{value}'.format(key = key, value = value))
print('******************************')
print('打印生产领料字典')
print('生产领料合计为：%s'%lingliao_total)
print('******************************')
for key,value in lingliao_dic.items():
    print('{key}:{value}'.format(key = key, value = value))

print('******************************')
print('打印结存字典')
print('结存合计为：%s'%qimo_total)
print('******************************')
for key,value in qimo_dic.items():
    print('{key}:{value}'.format(key = key, value = value))

wb =load_workbook('双佳数量.xlsx')
sheet = wb.active

jishu = 0
for index,row in enumerate(sheet.values):
    #print((row[0]))
    if row[0] == None:
        jishu = jishu + 1
        contiune


mrows = sheet.max_row -jishu
print(mrows)


kemuList0 = []
kemuList = []
danjiaList = []
qichushuliangList = []
qichujinerList = []
qimoshuliangList = []
qimojinerList = []

for index,row in enumerate(sheet.values):
    if index == 0:
        continue



    else :
        if index < mrows:
            #kemu0 = row[2].strip()
            kemuList0.append(row[2])

            try:
                danjia = float(row[13]) / float(row[12])
            except:
                danjia = 0

            danjiaList.append(danjia)


            if row[4] == '借':
                qichushuliang = float(row[5])
            else:
                qichushuliang = -1*float(row[5])
            qichushuliangList.append(qichushuliang)

            if row[4] == '借':
                qichujiner = float(row[6])
            else:
                qichujiner = -1*float(row[6])
            qichujinerList.append(qichujiner)

            if row[11] == '借':
                qimoshuliang = float(row[12])
            else:
                qimoshuliang = -1 * float(row[12])
            qimoshuliangList.append(qimoshuliang)

            if row[11] == '借':
                qimojiner = float(row[13])
            else:
                qimojiner = -1 * float(row[13])
            qimojinerList.append(qimojiner)

for kemu0 in kemuList0 :
    print(kemu0)
    #kemu0.strip()
    #kemuList.append(kemu0)

print(kemuList0)
print(danjiaList)
print(qichushuliangList)
print(qichujinerList)
print(qimoshuliangList)
print(qimojinerList)

temp = []
kemuList = []
for i in kemuList0 :
    if i != None:
        temp.append(i)
    else :
        temp.append('')

for i in temp :
    i = i.strip()
    kemuList.append(i)

print(kemuList)
print(len(kemuList))
print(len(kemuList0))
print(len(danjiaList))



#数据去向工作簿 Workbook
ss_wb = Workbook()
ss_sheet = ss_wb.active
ss_sheet.title = '数量汇总'


#各字典加一个空对空，避免excle科目那一列为空值时，找不到元素的情况
qichu_dic[''] = ''
ruku_dic[''] = ''
caigou_dic[''] = ''
banchengpin_dic['']= ''
chuku_dic[''] = ''
lingliao_dic[''] = ''
qimo_dic[''] = ''

print('字典长度')
print(len(qichu_dic))


#新增抬头
new_taitou = ('科目','单价','期初数量','期初金额','期末数量','期末金额','仓库期初','入库数量','采购入库','半成品入库','出库数量','生产领料','仓库结存')
i =0
row  = 0


for index,row in enumerate(sheet.values):
    if index == 0 :
        ss_sheet.append(row + new_taitou)
    else:
        if index <= mrows:
            i = index -1



            #ss_sheet.append(row + (kemuList[i],danjiaList[i],qichushuliangList[i],qichujinerList[i],qimoshuliangList[i],qimojinerList[i],qichu,ruku,caigou,banchengpin,chuku,lingliao,qimo))
            ss_sheet.append(row + (kemuList[i],danjiaList[i],qichushuliangList[i],qichujinerList[i],qimoshuliangList[i],qimojinerList[i]))
ss_wb.save('数量汇总.xlsx')


wb =load_workbook('数量汇总.xlsx')
sheet = wb.active

jishu = 0
first_list = []
for index,row in enumerate(sheet.values):
    #print(row[1])
    first_list.append(row[1])

    if row[0] == None:
        jishu = jishu +1
mrows = sheet.max_row -jishu

#数据去向工作簿 Workbook
ss_wb = Workbook()
ss_sheet = ss_wb.active
ss_sheet.title = '数量汇总8'

for index,row in enumerate(sheet.values):
    if index == 0 :
        ss_sheet.append(row)
    else:
        if index <= mrows:
            i = index -1
            if row[16] in qichu_dic :
                qichu = qichu_dic[row[16]]
            else :
                qichu = 0

            if row[16] in ruku_dic:
                ruku = ruku_dic[row[16]]
            else:
                ruku = 0

            if row[16] in caigou_dic :
                caigou = caigou_dic[row[16]]
            else :
                caigou = 0

            if row[16] in banchengpin_dic :
                banchengpion = banchengpin_dic[row[16]]
            else :
                banchengpin = 0

            if row[16] in chuku_dic :
                chuku = chuku_dic[row[16]]
            else :
                chuku = 0

            if row[16] in lingliao_dic :
                lingliao = lingliao_dic[row[16]]
            else :
                lingliao = 0

            if row[16] in qimo_dic :
                qimo = qimo_dic[row[16]]
            else :
                qimo = 0

            ss_sheet.append(row + (qichu, ruku, caigou, banchengpin, chuku, lingliao, qimo))

ss_wb.save('数量汇总8.xlsx')
















