import openpyxl

fname = r'F:\a00nutstore\006\zw\yingfu\2021yingfu\应付账款2021-0630.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb['2021-0630 (2)']
font_name =ws['A1'].font.name
# 获取字体名称
font_size = ws['A1'].font.size
# 获取字体大小
bold = ws['A1'].font.bold
# 获取是否加粗，True--加粗，False--未加粗
color = ws['A1'].font.color
# 获取字体颜色
print(font_name,font_size,bold,color)

wb.close()