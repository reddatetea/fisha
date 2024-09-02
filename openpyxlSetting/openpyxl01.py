'''
复制一个工作表的数据到另一工作表
并在数据表的后面加上公式，进行自动计算
计算实际最高行数，避免清除数据后会导致实际行数与max_row不一样而报错的问题！
'''
from openpyxl import Workbook,load_workbook
#from openpyxl.utils import get_column_letter

wb =load_workbook('abc.xlsx')
sheet = wb.active

print(sheet.max_row)
print(sheet.max_column)
jishu = 0
for index,row in enumerate(sheet.values):
    print((row[0]))


    if row[0] == None:
        jishu = jishu +1
        print(jishu)

'''
实际行数mrows,当一行中第一个元素的值为None时，表示这一行实际数据是没有的！从这一行开始计数并汇总，将这个汇总数从
max_row中减去后，就是实际行数！
'''

mrows = sheet.max_row -jishu
print(mrows)

#数据去向工作簿 Workbook
ss_wb = Workbook()
ss_sheet = ss_wb.active
ss_sheet.title = '流水账'
xl = 'aaaa'
for index,row in enumerate(sheet.values):
    if index ==0 :
        ss_sheet.append(row+('销售额',xl))
        continue

#将上面计算出的实际行数，运行到下面的程序！不这样就会报错！
    if  index < mrows :

        #row[2]表示这一行第三个元素！列表 row[0]表示第一个元素
        shuliang = int(row[2])
        danjia = int(row[3])

        #用变量来设置公式比较方便 je = shuliang * danjia
        xiaoshoue = shuliang * danjia

        #列表的经典表达方式，如果只有一个元素的元组用逗号！(+(xiaoshoue,))
        ss_sheet.append(row+(xiaoshoue,xl))

ss_wb.save('abcabc.xlsx')





