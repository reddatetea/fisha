'''
从实时流水账中引入当月的白云入库，公式有，但去重没达到预期
'''
import BaiyunDangyueToBaiyun
import lszTobaiyunDangyue
import os
import quchong
# import pandas08
import openpyxl
import allyueprice
import pricelisttoprice
import baiyunleibie
import pandas as pd

baiyunQijian = input('请输入白云期间：格式为2020-04：\n')
piaojuhao = input('请输入开始的票据号：格式为90033617：')
lszTobaiyunDangyue.lszTobaiyundangyue(baiyunQijian,piaojuhao)
dunjia_dic = BaiyunDangyueToBaiyun.dunjiaDic()
fname2,maxrow2 = BaiyunDangyueToBaiyun.baiyunDangyueToby()
BaiyunDangyueToBaiyun.jiagongsi(fname2,maxrow2,dunjia_dic)

fname = fname2

#数据去重
#in_subject = ['公司', '开票日期', '入库单号','品名', '数量(令)', '计算重量','仓库材料']
in_subject = ['开票日期', '入库单号','品名', '数量(令)', '计算重量','仓库材料']
wb = openpyxl.load_workbook(fname)
sheetnames = wb.sheetnames
print('选择工作表')
sheetname = '2020'
quchong.delchongfu(fname,sheetname,in_subject)
gongying_dic = dunjia_dic
#加公式
def jiagongsi(fname,gongying_dic):
    wb = openpyxl.load_workbook(fname)
    ws = wb['2020']
    maxrow = ws.max_row
    for i in range(2, maxrow + 1):
        gongyingshang = '驻马店白云纸业有限公司'
        pinming = ws.cell(i, 17).value
        if ws.cell(i,14).value in ['',None]:
            ws.cell(i, 11).value = gongying_dic.get((gongyingshang, pinming), 0)  # 取最新价格字典，没有则为0
            # 金额用公式表达
            #pricename = ws.cell(i,17).value
            #danjia = gongying_dic['驻马店白云纸业有限公司'][baiyunQijian][pricename]
            ws.cell(i, 12, value='=round(J' + str(i) + ' * ' + 'K' + str(i) + ', 2)')
            # 不含税金额用公式表达
            ws.cell(i, 13, value='=round(L' + str(i) + '/1.13 , 2)')
        else :
            continue
    wb.save(fname)

#加类别
dicnew = baiyunleibie.baiyunLeibie()
def jialeibie(dicnew):
    wb = openpyxl.load_workbook(fname)
    ws = wb['2020']
    maxrow = ws.max_row
    for i in range(2, maxrow + 1):
        if ws.cell(i, 14).value in ['', None]:
            pricename = ws.cell(i, 17).value
            #写入白云类别
            dicnew.setdefault(pricename,'')
            ws.cell(i, 19, value = dicnew[pricename])
        else:
            continue
    wb.save(fname)

jiagongsi(fname,dunjia_dic)
jialeibie(dicnew)
os.startfile(fname)




