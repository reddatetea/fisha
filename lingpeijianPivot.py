import os
import openpyxl
import excelseting

newpath_name = r'零配件各供应商价格明细'
if os.path.exists(newpath_name):
    pass
else:
    os.makedirs(newpath_name)
path = os.getcwd()
newpath = os.path.join(path,newpath_name)
fname = r'零配件入库20181226-20210827.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
gys = [j.value for j in ws['C']][1:]
gongyingshangs = []
for j in gys:
    if j not in gongyingshangs:
        gongyingshangs.append(j)
gys_pinming_danjia = {key:{} for key in gongyingshangs}   #避fromkeys坑
for row in list(ws.values)[1:]:
    gongyingshang = row[2]
    pinming = row[4]
    danjia = row[9]
    gys_pinming_danjia[gongyingshang][pinming]=danjia
for gongyingshang,pinming_danjia in gys_pinming_danjia.items():
    nwb = openpyxl.Workbook()
    nws = nwb.active
    nws.title = gongyingshang
    filename = '{}.xlsx'.format(gongyingshang)
    nfname = os.path.join(newpath,filename)
    nws.append(['零配件名称','单价'])
    for pinming,danjia in pinming_danjia.items():
        nws.append([pinming,danjia])
    nwb.save(nfname)
    excelseting.fastseting(nfname,gongyingshang,gongyingshang)
wb.close()


