'''
根据白云价格字典和纸张字典，制作所有期间原材料价格明细
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
import yuancailiaoBaiyunPriceAll
import yuancailiaoZhiPriceAll

#建新的当月价格表excel
def jiannewpiao():
    #print('请输入价格表所在路径:')
    #path = input('')
    path = r'F:\a00nutstore\006\zw\price'
    os.chdir(path)
    filename = 'all月原材料价格表.xlsx'
    fname = os.path.join(path, filename)
    wb = openpyxl.Workbook(fname)
    ws = wb.create_sheet(title='price')

    taitou = ('供应商','期间', '品名', '价格', '令数', '吨数', '金额')
    ws.append(taitou)
    wb.save(fname)
    return fname

#新建字典
def jiannewdic():

    byprices = yuancailiaoBaiyunPriceAll.bypriceDic()
    zhiprices = yuancailiaoZhiPriceAll.zhipriceDic()
    return byprices,zhiprices

#写入白云价格
def writebyprice(fname,byprices,gongyingshang =  '驻马店白云纸业有限公司'):
    wb = openpyxl.load_workbook(fname)
    ws = wb['price']
    for key, value in byprices.items():
        ws.append((gongyingshang,)+key + value)
    wb.save(fname)

def writezhiprice(fname,zhiprices):
    wb = openpyxl.load_workbook(fname)
    ws = wb['price']
    for key, value in zhiprices.items():
        ws.append(key + value)
    wb.save(fname)

#删除数据为0的行
def dangyueprice(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.create_sheet(title = 'all月')
    maxrow = ws.max_row

    ws1 = wb['price']
    for index,row in enumerate(ws1.values):
        if (row[4]==0 and row[5]== 0 and row[6]==0) or row[3]==0:
            continue
        else:
            ws.append(row)

    wb.save(fname)

def main():
    fname = jiannewpiao()
    #qijian = input('请输入期间，如2020-04\n')
    byprices,zhiprices = jiannewdic()
    writebyprice(fname,byprices,gongyingshang =  '驻马店白云纸业有限公司')
    writezhiprice(fname, zhiprices)
    dangyueprice(fname)
    os.startfile(fname)

if __name__ == '__main__':
    main()

