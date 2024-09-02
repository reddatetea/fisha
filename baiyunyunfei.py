import openpyxl
import os
import easygui
import datetime
from openpyxl.utils import column_index_from_string
import excelseting

fname = r'F:\a00nutstore\006\zw\baiyun\2020白云入库.xlsx'
start = input('请输入送货单开始日期"2021-06-03"\n')
end=  input('请输入送货单开始日期"2021-07-15"\n')

path,filename = os.path.split(fname)
os.chdir(path)
new_filename = r'白云运费{}.xlsx'.format(end)
new_fname = os.path.join(path,new_filename)


wb = openpyxl.load_workbook(fname)
ws = wb['2020']

nwb = openpyxl.Workbook()
nws = nwb.active
nws.title = 'yf{}'.format(end[-5:])
yf_print_ws_name = 'yfdy{}'.format(end[-5:])
yf_print_ws = nwb.create_sheet(yf_print_ws_name)

pricename_num = column_index_from_string('G')
ling_num = column_index_from_string('H')
dun_num = column_index_from_string('J')
dic = {}

new_values = []
start_riqi = datetime.datetime.strptime(start, '%Y-%m-%d')
end_riqi = datetime.datetime.strptime(end, '%Y-%m-%d')
for index,row in enumerate(ws.values):
    if index == 0:
        nws.append(row)
        new_values.append(row)
    else :
        songhuodan_riqi = row[2]
        if  songhuodan_riqi >= start_riqi and  songhuodan_riqi <= end_riqi :
            pricename = row[pricename_num-1]
            ling = row[ling_num-1]
            dun = row[dun_num-1]
            yunfei = round(dun*115,2)
            buhanshui = round(yunfei/1.09,2)
            nws.append(row)
            new_values.append(row)
            if len(list(filter(lambda x: x in pricename,['价差','返利','冲减'])))==0:                #剔除pricename中含价差、返利、冲减
                if pricename not in dic.keys():
                    dic[pricename] = (ling, dun, yunfei, buhanshui)
                else:
                    ling += dic[pricename][0]
                    dun += dic[pricename][1]
                    yunfei += dic[pricename][2]
                    buhanshui += dic[pricename][3]
                    dic[pricename] = (ling, dun, yunfei, buhanshui)
            else :
                continue
        else  :
            continue

yf_print_ws.append(['品名','数量（令）','数量（吨）','运费','不含税运费'])
for key,value in dic.items():
    yf_print_ws.append((key,)+value)

ling_total = sum([j[0] for j in dic.values()])
dun_total = sum([j[1] for j in dic.values()])
yunfei_total = sum([j[2] for j in dic.values()])
buhanshui_total = sum([j[3] for j in dic.values()])
yf_print_ws.append(['合计',ling_total,dun_total,yunfei_total,buhanshui_total])
wb.close()
nwb.save(new_fname)

taitou = '白云运费{}至{}'.format(start,end)
excelseting.fastseting(new_fname,yf_print_ws_name,taitou)
os.startfile(new_fname)




