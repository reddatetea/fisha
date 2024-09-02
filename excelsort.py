#本模块对excel进行排序
import win32com.client
import os
import openpyxl
import easygui
from openpyxl.utils import get_column_letter,column_index_from_string

def excelSort(fname, ws_name, ziduan_col_letter,max_column_letter, max_row,max_column,sort_choice):
    excel = win32com.client.Dispatch("Excel.Application")

    wb = excel.Workbooks.Open(fname)
    ws = wb.Worksheets(ws_name)
    #ziduan_range = '{}2:{}{}'.format(ziduan_col_letter, ziduan_col_letter,max_row)
    ziduan_range = 'A2:{}{}'.format(max_column_letter, max_row)
    zidun_key = '{}1'.format(ziduan_col_letter)
    #ws.Range('B2:B92').Sort(Key1=ws.Range('B1'), Order1=1, Orientation=1)
    ws.Range(ziduan_range).Sort(Key1=ws.Range(zidun_key), Order1=sort_choice, Orientation=1)

    wb.Save()
    excel.Application.Quit()
    return fname


def choiceSheet():

    msg = '请点选要进行排序的excel文件'
    print(msg)
    fname = easygui.fileopenbox(msg, title='excel')
    path, filename = os.path.split(fname)
    os.chdir(path)

    wb = openpyxl.load_workbook(fname,data_only=True)
    sheetnames = wb.sheetnames

    if len(sheetnames) == 1:
        ws_name = sheetnames[0]
    else:
        msg = '请选择要排序的工作表'
        print(msg)
        ws_name = easygui.buttonbox(msg, title='sheet', choices=sheetnames)

    ws = wb[ws_name]

    max_row = ws.max_row
    max_column = ws.max_column

    firstrow = ws[1]
    print(firstrow)
    ziduans = []
    for i in firstrow:
        ziduans.append(i.value)

    msg = '请选定要排序的字段'
    print(msg)
    ziduan = easygui.buttonbox(msg, title='请选择字段', choices=ziduans)
    place = ziduans.index(ziduan)
    ziduan_col_letter = get_column_letter(idx=place + 1)
    max_column_letter = get_column_letter(idx=max_column)

    msg = '请选择升序or降序'
    print(msg)
    orders = ['升序','降序']
    order_fangsi = easygui.buttonbox(msg, title='请选择字段', choices=orders)
    print(order_fangsi)
    if  order_fangsi == '升序':
       sort_choice = 1
    else:
       sort_choice = 2



    wb.close()
    return fname, ws_name, ziduan_col_letter,max_column_letter, max_row,max_column,sort_choice

def main():
    fname, ws_name, ziduan_col_letter, max_column_letter, max_row, max_column,sort_choice = choiceSheet()
    excelSort(fname, ws_name, ziduan_col_letter, max_column_letter, max_row, max_column,sort_choice)

if __name__ == '__main__':
    main()







