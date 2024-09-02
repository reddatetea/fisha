'''
打开再关闭，以解决dataonly打开后出现公式丢失的问题,
难道加公式后都要打开一下，然后再进行其它操作？
'''
from win32com.client import Dispatch
import easygui
import openpyxl

def just_open(fname):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(fname)
    xlBook.Save()
    xlBook.Close()
    return fname

def main():
    wb = openpyxl.Workbook()
    ws =  wb.active
    ws['a1'].value = 10
    ws['b1'].value = 20
    ws['c1'].value  = '=sum(a1:b1)'
    fname = r'e:\excel_data_only_test.xlsx'
    wb.save(fname)
    fname = just_open(fname)                       #如果没有此句，则整个程序运行后，c1和a2单元格里啥者没有！
    wb = openpyxl.load_workbook(fname,data_only=True)
    ws = wb.active
    ws['a2']  = ws['c1'].value
    wb.save(fname)
    # fname = just_open(fname)

    #
    # msg = '请点选要打开和关闭的excel文件'
    # fname =easygui.fileopenbox(msg)
    # just_open(fname)

if __name__ == "__main__":
    main()

