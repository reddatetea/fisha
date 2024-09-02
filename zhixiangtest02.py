import openpyxl
import os
import zhixiangtodic01

path = r'F:\a00nutstore\006\zw\ZHIXIANG'
os.chdir(path)

filename = '2020纸箱入库 - 副本.xlsx'
fname = os.path.join(path, filename)

wb = openpyxl.load_workbook(fname)
sheetnames = wb.sheetnames
print(sheetnames)
ws1=wb['九安源2001']
zhixiangdic = zhixiangtodic01.zhixiangtoDic()
zhixiangdic.setdefault('BSSSSS',(0,0,0))
s=zhixiangdic['BSSSSS']
print(s)

wb.close()

