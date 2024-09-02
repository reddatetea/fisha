# -*-coding = utf-8 -*-
from openpyxl import Workbook,load_workbook
import cangkutotal
import os

def zhongTotal(qichu_dic, ruku_dic, caigou_dic, banchengpin_dic, chuku_dic, lingliao_dic, qimo_dic, taitou_lists, cangkucailiaolist1,fname2):
    #cw_names, cw_Newnames, fname1 = cailiaorenames.jiacaiwuname(fname1)
    #qichu_dic, ruku_dic, caigou_dic, banchengpin_dic, chuku_dic, lingliao_dic, qimo_dic, taitou_lists,cangkucailiaolist1 = cangkutotal.cangkuTotal(fname1)
    #打开'数量金额汇总.xlsx'
    dirname,filename = os.path.split(fname2)
    wb2 = load_workbook(fname2)
    sheet2 = wb2.active

    jishu = 0
    first_list = []
    for index, row in enumerate(sheet2.values):
        # print(row[1])
        first_list.append(row[1])

        if row[0] == None:
            jishu = jishu + 1
    mrows = sheet2.max_row - jishu

    # 数据去向工作簿 Workbook
    ss_wb1 = Workbook()
    ss_sheet1 = ss_wb1.active
    ss_sheet1.title = '数量金额盘存表'

    new_taitou = ('仓库期初', '入库数量', '采购入库', '半成品入库', '出库数量', '生产领料', '仓库结存')

    for index, row in enumerate(sheet2.values):
        if index == 0:
            ss_sheet1.append(row + new_taitou)
        else:
            if index <= mrows:
                i = index - 1
                if row[17] in qichu_dic:
                    qichu = qichu_dic[row[17]]
                else:
                    qichu = 0

                if row[17] in ruku_dic:
                    ruku = ruku_dic[row[17]]
                else:
                    ruku = 0

                if row[17] in caigou_dic:
                    caigou = caigou_dic[row[17]]
                else:
                    caigou = 0

                if row[17] in banchengpin_dic:
                    banchengpin = banchengpin_dic[row[17]]
                else:
                    banchengpin = 0

                if row[17] in chuku_dic:
                    chuku = chuku_dic[row[17]]
                else:
                    chuku = 0

                if row[17] in lingliao_dic:
                    lingliao = lingliao_dic[row[17]]
                else:
                    lingliao = 0

                if row[17] in qimo_dic:
                    qimo = qimo_dic[row[17]]
                else:
                    qimo = 0

                ss_sheet1.append(row + (qichu, ruku, caigou, banchengpin, chuku, lingliao, qimo))

    #cangkucailiaolist1 = cangkucailiaolist1
    filename = r'数量金额盘存表.xlsx'
    fname = os.path.join(dirname,filename)
    ss_wb1.save(fname)
    return fname


def main():
    zhongTotal(qichu_dic, ruku_dic, caigou_dic, banchengpin_dic, chuku_dic, lingliao_dic, qimo_dic, taitou_lists, cangkucailiaolist1,fname2)

if __name__== '__main__':
    main()



