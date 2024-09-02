#本版本实现多个记账日期查询
import os
import xlwings as xw
import pandas as pd
import openpyxl
import easygui
from openpyxl.utils import  get_column_letter,column_index_from_string
# import excelcellfillby
import excelseting
import excelprint

def choiceSheet():
    fname =r'F:\a00nutstore\006\zw\baiyun\2020白云入库.xlsx'
    path, filename = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname,data_only=True)
    ws_name = '2020'
    ws = wb[ws_name]
    max_row = ws.max_row
    zhiziduan =  ['数量(令)','计算重量', '数量(吨)', '金额', '不含税金额']
    hangziduan ='品名'
    jisuanfangsi = 'sum'
    #saixuanliezhi = input('请输入日期（如9/30/2020)\n')
    jizhang_number = column_index_from_string('N')
    # saixuanliezhi = input('请输入日期（如9/30/2020)\n')
    saixuanliezhis = set()
    for jizhang_row in ws.iter_rows(min_row=2, min_col=jizhang_number, max_row=max_row, max_col=jizhang_number):
        for j in jizhang_row:
            if j.value not in ['', None, 0]:
                saixuanliezhis.add(j.value)
            else:
                continue

    saixuanliezhis = list(saixuanliezhis)
    saixuanliezhis.sort()
    msg = '请选择记账日期'
    print(msg)
    saixuanliezhis = easygui.multchoicebox(msg, title=msg, choices=saixuanliezhis)
    print(saixuanliezhis)
    wb.close()
    return fname, ws_name, zhiziduan,hangziduan, jisuanfangsi,saixuanliezhis

def toushiBiao(fname, ws_name, zhiziduan,hangziduan,jisuanfangsi,saixuanliezhis):
    app = xw.App(visible = False,add_book = False)
    workbook = app.books.open(fname)
    worksheets = workbook.sheets
    ws = worksheets[ws_name]
    values = ws.range('A1').expand('table').options(pd.DataFrame).value
    filter_js =[]
    for  j in range(len(saixuanliezhis)):
        filter_j= values[values['记账'] == '{}'.format(saixuanliezhis[j])]
        filter_js.append(filter_j)
    filtered = pd.concat(filter_js)

    print(filtered)
    print(fname, ws_name, zhiziduan,hangziduan, jisuanfangsi,saixuanliezhis)
    pivottable = pd.pivot_table(filtered, values=zhiziduan, index=hangziduan, columns=None, aggfunc=jisuanfangsi, fill_value =0,margins=True, margins_name='合计')
    order = zhiziduan
    pivottable = pivottable[order]
    nws = worksheets['toushibiao']  # 新表
    nws.clear()
    nws.range('A1').value = pivottable
    nws.range('B1').expand('down').number_format = '#,##0.00'
    nws.range('E1').expand('down').number_format = '#,##0.00'
    nws.range('F1').expand('down').number_format = '#,##0.00'
    nws.range('C1').expand('down').number_format = '#,##0.000'
    nws.range('D1').expand('down').number_format = '#,##0.000'
    workbook.save()
    workbook.close()
    app.quit()

def main():
    fname, ws_name, zhiziduan, hangziduan, jisuanfangsi,saixuanliezhis= choiceSheet()
    toushiBiao(fname, ws_name, zhiziduan,hangziduan,jisuanfangsi,saixuanliezhis)
    wb = openpyxl.load_workbook(fname)
    ws_name = 'toushibiao'
    ws = wb[ws_name]
    max_row = ws.max_row
    max_col = ws.max_column
    min_row = 1
    min_col = 1
    # excelcellfillby.excelcellFill(fname,ws_name,min_row,min_col,max_row,max_col)
    fname, ws_name = excelseting.setPrintArea(fname,ws_name)
    excelseting.firstRowSeting(fname, ws_name)
    excelseting.yemei(fname, ws_name,'白云')
    os.system(fname)
    #excelprint.wsPrint(fname, ws_name)

if __name__ == '__main__':
    main()