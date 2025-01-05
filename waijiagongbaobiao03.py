import easygui
import excelmessage
import openpyxl
from openpyxl.utils import column_index_from_string
import waijiagongDanjia
import waijiagongKaibieJisuan
import os

def  baobiaoTaitou(fname):
    wb = openpyxl.load_workbook(fname)
    ws_name = easygui.enterbox('请输入期间"2024-12"')
    ws0 = wb['外加工样表']
    taitou = [j.value for j in ws0[1]]
    # print(taitou)
    wb.create_sheet(title = ws_name)
    ws = wb[ws_name]
    ws.append(taitou)
    wb.save(fname)
    return fname,ws_name

def biginTwo(a,b):
    if a >= b :
        big = a

    else :
        big = b
    return big


def addDangyue(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    msg = '请点选当月外加工excel文件'
    fname1 = easygui.fileopenbox(msg)
    fname1 = excelmessage.excelMessage(fname1)
    wb1 = openpyxl.load_workbook(fname1)
    sheetnames = wb1.sheetnames
    msg = '请点选工作表'
    if len(sheetnames) != 1:
        ws_name1 = easygui.choicebox(msg, choices=sheetnames)
        ws1 = wb1[ws_name1]
    else :
        ws1 = wb1[wb1.sheetnames[0]]

    max_row1 = ws1.max_row
    for index,row in enumerate(ws1.values):
        if index==0:
            continue
        else:
            pinming = row[1]
            shangri = row[3]
            ruku = row[4]
            chuku = row[5]
            benri = row[6]
            if (pinming not in['',None]) and (shangri != 0 or  ruku != 0 or chuku != 0  or benri != 0) :
                hang = [ws_name,'',pinming,shangri,ruku,chuku,benri]
                ws.append(hang)
    wb1.close()
    wb.save(fname)
    return fname

def baobiao(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    kaibie_num = column_index_from_string('H')
    fumo_num = column_index_from_string('I')
    fumo_danjia_num = column_index_from_string('J')
    fumo_jiner_num = column_index_from_string('K')
    yawen_danjia_num = column_index_from_string('M')
    yawen_jiner_num = column_index_from_string('N')
    caichong_danjia_num = column_index_from_string('P')
    caichong_jiner_num = column_index_from_string('Q')
    tangjin_danjia_num = column_index_from_string('S')
    tangjin_jiner_num = column_index_from_string('T')

    dic = waijiagongDanjia.danjia()
    max_row = ws.max_row
    for j in range(2,max_row+1):
        kaibie = ws.cell(j,kaibie_num).value
        print(kaibie)
        zhang = waijiagongKaibieJisuan.kaibie(kaibie)
        ling = ws.cell(j,6).value
        print('ling',ling)
        zhangshu = ling*zhang
        print('zhangshu',zhangshu)
        ws.cell(j,7).value = zhangshu
        jiagong_fangsi = ws.cell(j,fumo_num).value
        fangsis = ['光膜','亮膜','哑膜','压纹','烫金','彩葱']
        shiji_fangsis = []
        for  fangsi in fangsis:
            if fangsi in jiagong_fangsi:
                shiji_fangsis.append(fangsi)
            else :
                continue
        print(shiji_fangsis)
        fumo_danjia = 0
        yawen_danjia = 0
        tangjin_danjia = 0
        caichong_danjia = 0
        fumo_kaiji = 0
        yawen_kaiji = 0
        tangjin_kaiji = 0
        caichong_kaiji = 0
        for shiji_fangsi in shiji_fangsis:
            # danjia = dic[shiji_fangsi][kaibie]
            # kaiji = dic[shiji_fangsi]['开机费']

            if shiji_fangsi =='光膜' or shiji_fangsi =='亮膜':
                fumo_danjia =  dic[shiji_fangsi][kaibie]
                fumo_kaiji = dic[shiji_fangsi]['开机费']
            elif shiji_fangsi =='哑膜' :
                fumo_danjia =  dic[shiji_fangsi][kaibie]
                fumo_kaiji = dic[shiji_fangsi]['开机费']
            elif shiji_fangsi =='压纹' :
                yawen_danjia =  dic[shiji_fangsi][kaibie]
                yawen_kaiji = dic[shiji_fangsi]['开机费']
            elif shiji_fangsi =='烫金' :
                tangjin_danjia =  dic[shiji_fangsi][kaibie]
                tangjin_kaiji = dic[shiji_fangsi]['开机费']
            else :
                caichong_danjia =  dic[shiji_fangsi][kaibie]
                caichong_kaiji = dic[shiji_fangsi]['开机费']
        fumo_jiner = biginTwo(zhangshu*fumo_danjia,fumo_kaiji)
        yawen_jiner = biginTwo(zhangshu * yawen_danjia,yawen_kaiji)
        tangjin_jiner = biginTwo(zhangshu * tangjin_danjia,tangjin_kaiji)
        caichong_jiner = biginTwo(zhangshu * caichong_danjia,caichong_kaiji)

        ws.cell(j,fumo_danjia_num).value = fumo_danjia
        ws.cell(j, yawen_danjia_num).value = yawen_danjia
        ws.cell(j, tangjin_danjia_num).value = tangjin_danjia
        ws.cell(j, caichong_danjia_num).value = caichong_danjia

        ws.cell(j, fumo_jiner_num).value = fumo_jiner
        ws.cell(j, yawen_jiner_num).value = yawen_jiner
        ws.cell(j, tangjin_jiner_num).value = tangjin_jiner
        ws.cell(j, caichong_jiner_num).value = caichong_jiner
        ws['X{}'.format(j)]='=K{} + N{} + Q{} + T{} + U{}'.format(j,j,j,j,j)
        ws['Y{}'.format(j)]='=(E{} + D{}) * 0.002'.format(j,j)
        ws['Z{}'.format(j)]='=IF((D{} + E{}) * (1 - Y{}) - F{} - AB{} < 0, 0, (D{} + E{}) * (1 - Y{}) - F{} - AB{})'.format(j,j,j,j,j,j,j,j,j,j)
        ws['AA{}'.format(j)] = '=Z{}*1200'.format(j)


    ws['C{}'.format(max_row+1)].value = '合计'
    ws['AA{}'.format(max_row +1)].value = '=SUM(AA2:AA{})'.format(max_row)
    ws['AB{}'.format(max_row + 1)].value = '=SUM(AB2:AB{})'.format(max_row)
    for j in 'DEFGKNQTXWZ':
        ws['{}{}'.format(j,max_row +1)].value = '=SUM({}2:{}{})'.format(j,j,max_row)

    weis = ['添美加工费：','加网版费','减：超耗扣款','减：扣工费','应付加工费']
    for j in range(len(weis)):
        wei = weis[j]
        ws['G{}'.format(max_row+3+j)].value = wei
    ws['X{}'.format(max_row+3)].value = ws['X{}'.format(max_row+1)].value
    ws['X{}'.format(max_row + 5)].value = ws['AA{}'.format(max_row + 1)].value
    ws['X{}'.format(max_row+7)].value='=X{} + X{}  -X{} - X{}'.format(max_row+3,max_row+4,max_row+5,max_row+6)




    # wb.save(fname)
    wb.save(fname)




def main():
    fname = r'F:\a00nutstore\006\zw\jiagong\2019\复膜\鑫添美外加工.xlsx'
    msg = '请点选本月外加工表'
    fname,ws_name = baobiaoTaitou(fname)
    addDangyue(fname,ws_name)
    os.startfile(fname)
    msg = '是否已输入开别和加工方式'
    # print(msg)
    titleYN = easygui.ccbox(msg, title='请选择Yes or No', choices=('yes', 'no'), image=None)
    if titleYN == True:
        baobiao(fname,ws_name)
    else:
        pass
    os.startfile(fname)



if __name__ == '__main__':
    main()
