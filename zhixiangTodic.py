'''
根据纸箱尺寸，制作纸箱尺寸字典
'''

# _*_ conding:utf-8 _*_

import os
import openpyxl
from openpyxl.utils import column_index_from_string
import easygui

def zhixiangdic():
    #fname = r'D:\a00nutstore\006\zw\ZHIXIANG\2020纸箱入库.xlsx'
    fname = r'F:\a00nutstore\006\zw\ZHIXIANG\2020纸箱入库.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb['纸箱尺寸']
    maxrows = ws.max_row
    pinming_number = column_index_from_string('E')
    chang_number = column_index_from_string('F')
    kuan_number = column_index_from_string('G')
    gao_number = column_index_from_string('H')
    zhixiang_dic = {}
    for row in range(2,maxrows+1):
        pinming = ws.cell(row,pinming_number).value
        chang = ws.cell(row, chang_number).value
        kuan = ws.cell(row, kuan_number).value
        gao = ws.cell(row, gao_number).value
        zhixiang_dic[pinming]=(chang,kuan,gao)                  #(长，宽，高)
        wb.close()
    return zhixiang_dic

def main():
    zhixiangdic()

if __name__ == '__main__':
    main()




