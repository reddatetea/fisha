'''
删除工作表01,py
可以在不打开excel文件的情况下，删除指定的若干个工作表
'''

# _*_ conding:utf-8 _*_

import openpyxl
import excelmessage06
import easygui

def delsheets(fname,del_sheets):
    wb = openpyxl.load_workbook(fname)
    for del_sheet in del_sheets:
        ws = wb[del_sheet]
        wb.remove(ws)
    wb.save(fname)

def main():
    fname = excelmessage06.excelMessage()
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选要删除的工作表'
    print(msg)
    title = 'sheetname'
    del_sheets = easygui.multchoicebox(msg=msg, title=title, choices=sheetnames)
    delsheets(fname, del_sheets)

if __name__ == '__main__' :
    main()


