import openpyxl
import pprint

def yuancailiaoprice():
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料价格表.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb['价格表']
    mrows = ws.max_row
    rcols = ws.max_column
    print(mrows)

    gongying_dic = {}
    qijian_dic = {}

    for i in range(2,mrows-1):


        gongyingshang = ws.cell(i,1).value
        qijian = ws.cell(i,2).value
        pinming = ws.cell(i,3).value
        dunjia = ws.cell(i,5).value
        gongying_dic.setdefault(gongyingshang,{})
        gongying_dic[gongyingshang].setdefault(qijian,{})
        gongying_dic[gongyingshang][qijian][pinming] = dunjia


    print(qijian_dic)
    print(gongying_dic)
    resultFile = open('yuancailiaoPriceDic.py', 'w',encoding='utf-8')
    resultFile.write('allDate = ' + pprint.pformat(gongying_dic))
    resultFile.close()

    wb.close()
    return gongying_dic

def main():
    yuancailiaoprice()


if __name__=="__main__":
    main()




