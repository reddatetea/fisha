import openpyxl
import easygui
import os
import excelmessage
import datetime
import formatPainter
from openpyxl.styles import Font, Border, Side, Fill, Alignment


def printseting(fname1, riqi):
    # 设置字体、边框、对齐等常量
    font = Font(name='宋体', size=10)
    font_xiaoji = Font(name='宋体', bold=True, size=10)
    font_firstrow = Font(name='宋体', bold=True, size=20)
    thin_danbian = Side(border_style='thin')
    double_danbian = Side(border_style='double')
    thin_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=thin_danbian)
    double_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=double_danbian)
    wb = openpyxl.load_workbook(fname1)
    for ws in wb.worksheets:
        if ws.title == 'Sheet':
            del wb[ws.title]
        else:
            # 插入第一行
            ws.insert_rows(1)
            ws['A1'].value = '双佳' + ws.title + ' ' + riqi
            ws['A1'].font = font_firstrow
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            max_row = ws.max_row
            max_column = ws.max_column
            if ws.title != '合计':
                # 单元格宽度
                ws.column_dimensions['A'].width = 8.56
                ws.column_dimensions['B'].width = 17.33
                ws.column_dimensions['C'].width = 13.56
                ws.column_dimensions['D'].width = 3.67
                for j in 'EFGHIJK':
                    ws.column_dimensions['{}'.format(j)].width = 7.44
                    ws.merge_cells('A1:K1')
                for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
                    if row == 2:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        for col in range(1, max_column + 1):
                            if col <= 3:
                                ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                ws.freeze_panes = ws['E3']
            else:
                # 单元格宽度
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['G'].width = 12
                ws.merge_cells('A1:H1')
                # ws.freeze_panes = ws['E3']
                for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
                    if row == 2:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        for col in range(1, max_column + 1):
                            if col <= 1:
                                ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                ws.cell(row, col).alignment = Alignment(horizontal='right', vertical='center')
            for cell in ws['A']:
                row = cell.row
                if row == 1:
                    continue
                else:
                    if '小计' in cell.value:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).font = font_xiaoji
                            ws.cell(row, col).border = double_bian
                            ws.merge_cells('A{}:C{}'.format(row, row))
                    else:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).font = font
                            ws.cell(row, col).border = thin_bian
            # ws.freeze_panes = ws['E3']
            ws.print_title_rows = '1:2'  # the first row
            # 页脚设置
            ws.oddFooter.center.text = " &[Page] / &N"  # 1/n
            ws.oddFooter.center.size = 12  # 页脚中字体大小
            ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
            ws.oddFooter.center.color = "000000"  # 页脚中字体颜色
            ws.oddFooter.right.text = "张文伟 &[Date]"  # 页脚右 文字
            ws.oddFooter.right.size = 12  # 页脚右 字体大小
            ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
            ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
            ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.5, left=0.21, right=0.24,
                                                                  footer=0.5,
                                                                  bottom=1)
            for i in ws.iter_rows():      #将零值替换为空
                for j in i:
                    if j.value == 0:
                        j.value = ''
                    else:
                        continue
    wb.save(fname1)


def res_path(relative_path):
    """获取资源绝对路径"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


msg = '请点选"产成品当日和累计"工作表'
fname = excelmessage.wenjian(msg)
fname = excelmessage.excelMessage(fname)
today_date = datetime.date.today()
yesterday_date = (today_date + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")
msg = '报表日期是{}?'.format(yesterday_date)
choice = easygui.ccbox(msg, title='请选择"是"或"否"', choices=('是', '否'))
if choice:
    riqi = yesterday_date
else:
    msg = '请输入报表日期'
    riqi = easygui.enterbox(msg, title=" 昨天日期")
fname_canchengpin = '产成品日报表{}.xlsx'.format(riqi)
path, filename = os.path.split(fname)
os.chdir(path)
wb = openpyxl.load_workbook(fname)
sheetnames = wb.sheetnames
if 'AAA' in sheetnames:
    wsa = wb['AAA']
else:
    msg = '请点选"当日"工作表'
    print(msg)
    choice = easygui.choicebox(msg, choices=sheetnames)
    wsa = wb[choice]

if 'BBB' in sheetnames:
    wsb = wb['BBB']
else:
    msg = '请点选"累计"工作表'
    print(msg)
    choice = easygui.choicebox(msg, choices=sheetnames)
    wsb = wb[choice]

def getNoTotal(ws):
    df = list(ws.values)
    df1 = [i for i in df if ('累计' not in i[0]) and ('合   计' not in i[0])]  # AAA表中删除有累计和合计的行
    return df1

dfa1 = getNoTotal(wsa)
dfb1 = getNoTotal(wsb)
wb.close()
yesterday = [i[5] for i in dfa1]
num_in = [i[6] for i in dfa1]
num_out = [i[7] for i in dfa1]
category = [i[0] for i in dfa1]
name = [i[1] for i in dfa1]
category_name = zip(category, name)
yesterday_nums = zip(yesterday, num_in, num_out)
category_name_yesterday_nums = dict(zip(category_name, yesterday_nums))

dfb2 = []
for index, i in enumerate(dfb1):
    if index == 0:
        newrow = ['类别', '货号', '品名', '含量', '月初', '上日', '本日入库', '入库累计', '本日出库', '出库累计', '结存']
    else:
        category_name = (i[0], i[1])
        yesterday = category_name_yesterday_nums.get(category_name, [0, 0, 0])[0]
        num_in = category_name_yesterday_nums.get(category_name, [0, 0, 0])[1]
        num_out = category_name_yesterday_nums.get(category_name, [0, 0, 0])[2]
        newrow = list(i[:4]) + [i[5]] + [yesterday] + [num_in] + [i[6]] + [num_out] + [i[7]] + [i[8]]



    dfb2.append(newrow)
df_zero = [i for i in dfb2 if (i[-1] == 0) and ('小计' not in i[0])]
dfb_noZero = [i for i in dfb2 if (i[-1] != 0) or (('小计' in i[0]) and (i[-1] == 0))]  # 不删小计

def classify(str):
    return list(str) + ['(' + i for i in list(str)]

cata_dianshang = classify('A')                         #类别
cata_chaoben = classify('BCDEFGHI')
cata_xinrui = classify('JKL')
cata_ruiyi = classify('M')
cata_zhangben = classify('NOPQRSTUVWX')
cata_waimao = ['Y出口订制产品', '(Y出口订制产品']
cata_dinzhicanpin = ['Y订制产品', '(Y订制产品']
cata_dinzhizhangbu = ['Y订制账簿产品', '(Y订制账簿产品']
cata_feicipin = classify('Z')

title = ['类别', '货号', '品名', '含量', '月初', '上日', '本日入库', '入库累计', '本日出库', '出库累计', '结存']
df_dianshang = []
df_chaoben = []
df_xinrui = []
df_ruiyi = []
df_zhangben = []
df_waimao = []
df_dinzhicanpin = []
df_dinzhizhangbu = []
df_feicipin = []

for index, j in enumerate(dfb_noZero):
    if index == 0:
        continue
    else:
        i = j[0]
        one = i[0]
        two = i[:2]
        five = i[:5]
        six = i[:6]
        seven = i[:7]
        eight = i[:8]
        nine = i[:9]
        if (one in cata_dianshang) or (two in cata_dianshang):
            df_dianshang.append(j)
        elif (one in cata_chaoben) or (two in cata_chaoben):
            df_chaoben.append(j)
        elif (one in cata_xinrui) or (two in cata_xinrui):
            df_xinrui.append(j)
        elif (one in cata_ruiyi) or (two in cata_ruiyi):
            df_ruiyi.append(j)
        elif (one in cata_zhangben) or (two in cata_zhangben):
            df_zhangben.append(j)
        elif (seven in cata_waimao) or (eight in cata_waimao):
            df_waimao.append(j)
        elif (five in cata_dinzhicanpin) or (six in cata_dinzhicanpin):
            df_dinzhicanpin.append(j)
        elif (eight in cata_dinzhizhangbu) or (nine in cata_dinzhizhangbu):
            df_dinzhizhangbu.append(j)
        else:
            df_feicipin.append(j)

dfs = [df_chaoben, df_zero, df_xinrui, df_ruiyi, df_zhangben, df_waimao, df_dianshang, df_dinzhicanpin,
       df_dinzhizhangbu, df_feicipin]
dfs_str = ['df_chaoben', 'df_zero', 'df_xinrui', 'df_ruiyi', 'df_zhangben', 'df_waimao', 'df_dianshang',
           'df_dinzhicanpin', 'df_dinzhizhangbu', 'df_feicipin']
sheetnames = ['抄本及作业本', '零结存', '新锐', '锐意', '账本', '外贸', '电商', '订制本', '订制账簿', '废次品']


def addLastrow(df, total, name):
    last_row = df[-1].copy()
    last_row[0] = '{} 合计'.format(name)
    last_row[1:4] = ['', '', '']
    last_row[4:] = total
    df.append(last_row)
    return df


cate_nums = {}
for df, x, sheetname in zip(dfs, dfs_str, sheetnames):
    if sheetname != '零结存':
        begin = sum([i[4] for i in df if '小计' in i[0]])
        yesterday = sum([i[5] for i in df if '小计' in i[0]])
        num_in = sum([i[6] for i in df if '小计' in i[0]])
        total_in = sum([i[7] for i in df if '小计' in i[0]])
        num_out = sum([i[8] for i in df if '小计' in i[0]])
        total_out = sum([i[9] for i in df if '小计' in i[0]])
        end = sum([i[10] for i in df if '小计' in i[0]])

    else:
        begin = sum([i[4] for i in df])
        yesterday = sum([i[5] for i in df])
        num_in = sum([i[6] for i in df])
        total_in = sum([i[7] for i in df])
        num_out = sum([i[8] for i in df])
        total_out = sum([i[9] for i in df])
        end = sum([i[10] for i in df])
    total = [begin, yesterday, num_in, total_in, num_out, total_out, end]
    df = addLastrow(df, total, sheetname)
    cate_nums[sheetname] = total
    exec('{}'.format(x) + '=df')

df_chaoben_dinzhi = df_chaoben + df_dinzhicanpin[:-1] + df_dinzhizhangbu[:-1] + df_feicipin[:-1]
begin = 0
yesterday = 0
num_in = 0
total_in = 0
num_in = 0
total_out = 0
end = 0
for i in ['抄本及作业本', '订制本', '订制账簿', '废次品']:
    begin = cate_nums[i][0] + begin
    yesterday = cate_nums[i][1] + yesterday
    num_in = cate_nums[i][2] + num_in
    total_in = cate_nums[i][3] + total_in
    num_out = cate_nums[i][4] + num_out
    total_out = cate_nums[i][5] + total_out
    end = cate_nums[i][6] + end

total = [begin, yesterday, num_in, total_in, num_out, total_out, end]
df_chaoben_dinzhi = addLastrow(df_chaoben_dinzhi, total, '抄本作业本及订制')

df_heji = []
title2 = ['类别', '月初', '上日数','本日入库','当月入库','本日出库', '当月出库', '结存']
df_heji.append(title2)
for i in ['抄本及作业本', '电商', '订制本', '新锐', '锐意', '账本', '外贸', '订制账簿', '废次品']:
    v = cate_nums[i]
    # value = [v[0], v[3], v[5], v[6]]
    df_heji.append([i] + v)
begin = sum([i[0] for k, i in cate_nums.items() if k != '零结存'])
yesterday = sum([i[1] for k, i in cate_nums.items() if k != '零结存'])
num_in = sum([i[2] for k, i in cate_nums.items() if k != '零结存'])
total_in = sum([i[3] for k, i in cate_nums.items() if k != '零结存'])
num_out = sum([i[4] for k, i in cate_nums.items() if k != '零结存'])
total_out = sum([i[5] for k, i in cate_nums.items() if k != '零结存'])
end = sum([i[6] for k, i in cate_nums.items() if k != '零结存'])
df_heji.append(['合计', begin, yesterday,num_in,total_in,num_out,total_out, end])

wb = openpyxl.Workbook()
ws1 = wb.active
dfs = [df_chaoben_dinzhi, df_dianshang, df_xinrui, df_ruiyi, df_zhangben, df_waimao, df_zero, df_heji]
sheetnames = ['抄本作业及订制', '电商', '新锐', '锐意', '账本', '外贸', '零结存', '合计']
for df, sheetname in zip(dfs, sheetnames):
    if sheetname == '抄本作业及订制':
        ws1.title = sheetname
        df = [title] + df
        for i in df:
            ws1.append(i)
    elif sheetname == '合计':
        ws = wb.create_sheet(sheetname)
        for i in df:
            ws.append(i)fname0 = res_path('img/leiji.xlsx')
wb0 = openpyxl.load_workbook(res_path('img/leiji.xlsx'))
ws0 = wb0.active
area0 = ws0['A1:H12']
wb = openpyxl.load_workbook(fname_canchengpin)
ws = wb['合计']
cell_start1 = ws['A1']
formatPainter.stylesFormat(ws, area0, cell_start1)
wb0.close()
wb.save(fname_canchengpin)
    else:
        ws = wb.create_sheet(sheetname)
        df = [title] + df
        for i in df:
            ws.append(i)
wb.save(fname_canchengpin)
printseting(fname_canchengpin, riqi)
# easygui.msgbox(msg = '程序结束运行')

os.startfile(fname_canchengpin)



