import openpyxl
import os
import sys
import  excelprint

def zanguTiaozhen(gongyingshang,pinming,jiner):

    path = r'F:\a00nutstore\006\zw'
    os.chdir(path)
    filename = r'暂估调整.xlsx'
    fname = os.path.join(path,filename)
    ws_name =  '暂估'
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    ws.cell(1, 1, value=gongyingshang)
    ws.cell(3, 4, value=pinming)
    ws.cell(3, 5, value=jiner)

    wb.save(fname)
    return fname,ws_name


def main():
    print('请输入供应商:')
    gongyingshang = input('')

    print('请输入品名：')
    pinming = input('')

    print('请输入金额：')
    jiner = float(input(''))

    fname,ws_name = zanguTiaozhen(gongyingshang, pinming, jiner)

    excelprint.wsPrint(fname,ws_name)


#subprocess.Popen(['start','暂估调整.xlsx'],shell=True)

if __name__=="__main__" :
    main()











