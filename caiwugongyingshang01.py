'''
财务供应商，运行后形成的列表保表到py文件
'''
import openpyxl
import pprint

def caiwuGongyingshang(fname2):


    wb2 = openpyxl.load_workbook(fname2)

    if  '财务供应商' not in wb2.sheetnames:
        wb2.create_sheet(title = '财务供应商')
    else :
        pass

    ws1 = wb2['Vendor']
    ws2 = wb2['财务供应商']

    chaiwuGongyingshangs = []

    for row in range(1,ws1.max_row+1):
        if row == 1 :
            ws2.cell(row, 1).value = '财务供应商'
        else :
            ws2.cell(row, 1).value = ws1.cell(row, 2).value
            chaiwuGongyingshangs.append(ws2.cell(row, 1).value)


    wb2.save(fname2)
    resultFile = open('chaiwuGongyingshangs.py', 'w', encoding='utf-8')
    resultFile.write('chaiwu = ' + pprint.pformat(chaiwuGongyingshangs))
    return chaiwuGongyingshangs

def main():
    fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\财务供应商.xlsx'
    caiwuGongyingshang(fname2)

if __name__=='__main__':
    main()






