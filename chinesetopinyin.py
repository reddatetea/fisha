#本模块实现中文字符串转拼音
import os
import easygui
from xpinyin import Pinyin
import openpyxl
import afterSortedYuanxuhao

def getpinyin(chinese_str):
    p = Pinyin()
    pinyins = p.get_pinyin(chinese_str, '-')
    pinyin_str = ' '.join([i.capitalize() for i in pinyins.split('-')] )
    return pinyin_str

def main():
    # chinese_str = input('请输入需要转为拼音的中文字符串\n')
    # pinyin_str = getpinyin(chinese_str)
    # print(pinyin_str)
    fname = r'f:\a00nutstore\006\zw\yingfu\2021yingfu\20210531应付.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    # wb.create_sheet('供应商排序')
    # nws = wb['供应商排序']
    lst = list(ws.values)[1:-1]      #去掉首行和尾行
    taitou = [j.value for j in ws[1]]
    mohang = [j.value for j in ws[ws.max_row]]
    gongyings = [j[0].value for j in ws['d2:d{}'.format(ws.max_row-1)]]    #ws['d2:d100']提取出的是元组,元组个数由列数决定
    gonyings_pinyin = list(map(getpinyin,gongyings))
    ws.delete_rows(1, ws.max_row - 1 + 1)
    ws.append(taitou)
    xhs = afterSortedYuanxuhao.sorted_yuanxuhao(gonyings_pinyin)
    for xh in xhs:
        ws.append(lst[xh])
    ws.append(mohang)
    wb.save(fname)

if __name__ == '__main__':
    main()