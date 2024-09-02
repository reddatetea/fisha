'''
出库录入后核对，将数量金额明细账与材料出库EXCEL核对
pandas处理
'''
# _*_ coding:utf-8 _*_
import os
import pandas as pd
import easygui
import openpyxl
from chukuPd import totalShuliangjiner

def benyue():
    msg = '请点选当月出库参考'
    fname = easygui.fileopenbox(msg, title='出库excel')
    path, filename = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选出库参考copy'
    sheet_name = easygui.buttonbox(msg, title='出库参考copy', choices=sheetnames)
    df0 = pd.read_excel(fname,sheet_name = sheet_name)
    wb.close()
    return df0

def benyueshuru():
    df1,qijian = totalShuliangjiner()
    return df1,qijian

def hedui(df0,df1):
    df1 = df1.loc[:,['科目名称','本期贷方数量','期末数量']]
    df1 = df1.rename({'本期贷方数量':'实际出库数量','期末数量':'实际期末数量'},axis=1)
    df8 = pd.merge(df0,df1,on='科目名称',how='left')
    df8['出库差异']=df8['实际出库数量']-df8['出库']
    df8['期末差异'] = df8['实际期末数量'] - df8['本日数量']-df8['未达项']
    return df8

def main():
    df0 = benyue()
    df1,qijian = benyueshuru()
    df8 = hedui(df0,df1)
    fname = '{}出库差异.xlsx'.format(qijian)
    df8.to_excel(fname, sheet_name='出库差异', index=False)
    os.startfile(fname)


if __name__ == '__main__':
    main()











