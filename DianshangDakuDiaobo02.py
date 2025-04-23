'''
1. 计算002库调拨至003聚水潭库
2. 计算003聚水潭调味回002电商库
计算电商002库调拨至003聚水潭库净调拨数（1-2）

'''
import pandas as pd
import numpy as np
import re
import easygui
import os
import openpyxl
import bianmabiaozhunhua06
from bianmabiaozhunhua06 import quShu,chuliFirstHenggang,bianmaBiaozhun 
from bianmabiaozhunhua06 import huizhongXiaoshoubuDiaobojia,chuliXiaoshoubuDiaobojia

dic = {'code': '存货编码',
 'chuku_ben_diaobo': '出库本数',
 'chuku_jian_diaobo': '出库件数',
 'chuku_jiner': '出库金额',
 'ruku_ben_diaobo': '入库本数',
 'ruku_jian_diaobo': '入库件数',
 'ruku_jiner': '入库金额',
 'jiner': '净出库金额',
 'price': '调拨价',
 'bianma': '编码参考'}

#大库调拨单
def getDiaobo(fname_diaobo):
    df_diaobo = pd.read_excel(fname_diaobo,dtype = {'单据日期':'datetime64[ns]','存货编码':str})
    df_diaobo = df_diaobo.iloc[:-1]
    df_diaobo = df_diaobo[['存货编码', '存货名称', '调出仓库', '调入仓库', '数量（本）', '数量2（件）']]
    dic_diaobo = dict(zip(['存货编码', '存货名称', '调出仓库', '调入仓库', '数量（本）', '数量2（件）'],
                          ['code', 'stock', '调出仓库', '调入仓库', 'ben', 'jian']))
    df_diaobo = df_diaobo.rename(columns=dic_diaobo)
    # 调拨入库
    df_diaobo_ruku = df_diaobo.loc[df_diaobo['调入仓库'] == '发货仓库']
    df_diaobo_ruku = df_diaobo_ruku[['code', 'stock', 'ben', 'jian']]
    df_diaobo_ruku.columns = ['code', 'stock', 'ruku_ben_diaobo', 'ruku_jian_diaobo']
    # 调拨出库
    df_diaobo_chuku = df_diaobo.loc[df_diaobo['调出仓库'] == '发货仓库']
    df_diaobo_chuku = df_diaobo_chuku[['code', 'stock', 'ben', 'jian']]
    df_diaobo_chuku.columns = ['code', 'stock', 'chuku_ben_diaobo', 'chuku_jian_diaobo']
    df_diaobo_chuku
    return df_diaobo_ruku, df_diaobo_chuku

#透视表
def pivotDf(df,index,values,aggfunc):
    pivot= pd.pivot_table(df,index = index,values = values,aggfunc = aggfunc)
    # pivot = pivot.reset_index()
    return pivot
def jiaHeji(df,dic):
    total = df.sum().to_list()
    total[0] = '合计'
    total[-1] = ''
    total[-2] = ''
    df.loc[len(df)] = total
    df = df.rename(columns = dic)
    return df

# fname_diaobo = r"F:\a00nutstore\008\zww08\产成品\调拨\调拨单20250226-0325.xlsx"
fname_diaobo = easygui.fileopenbox('请点选调拨单“0126-0225”')
path,file = os.path.split(fname_diaobo)
os.chdir(path)
df_diaobo_ruku, df_diaobo_chuku = getDiaobo(fname_diaobo)

index = 'code'
aggfunc ='sum'
# values = ['chuku_ben_diaobo','chuku_jian_diaobo']
chuku = pivotDf(df_diaobo_chuku,index,values = ['chuku_ben_diaobo','chuku_jian_diaobo'],aggfunc=aggfunc)
ruku = pivotDf(df_diaobo_ruku,index,values = ['ruku_ben_diaobo','ruku_jian_diaobo'],aggfunc=aggfunc)

#连接调拨出库和调拨入库
diaobo = pd.merge(chuku,ruku,left_index = True,right_index = True,how = 'outer')
diaobo = diaobo.fillna(0)

#销售部内部调拨价格汇总
isno = easygui.boolbox('是否对销售部调拨价格进行汇总')
df_xiaoshou1 = bianmabiaozhunhua06.huizhongXiaoshoubuDiaobojia(isno)
#将销售部价格中含字符的全部处理为数值
df_xiaoshou1,df_xiaoshou2,df_xiaoshou,bianma_diaobojia = chuliXiaoshoubuDiaobojia(df_xiaoshou1)
diaobo = diaobo.assign(price = diaobo.index.map(bianma_diaobojia))
diaobo.to_excel('diaobo.xlsx')
#将大库的编码打短
diaobo1= diaobo.copy()
diaobo1 = diaobo1.reset_index()
diaobo1 = diaobo1.assign(bianma = diaobo1.code)
diaobo1.price = diaobo1.price.fillna(0)
#一分为二，处理后再合并
diaobo11 = diaobo1[diaobo1['price']  == 0]
diaobo12 = diaobo1[~(diaobo1['price']  == 0)]

diaobo11 = diaobo11.assign(bianma = diaobo11.code)
diaobo11['bianma'] = diaobo11['bianma'].map(chuliFirstHenggang)
diaobo11['bianma'] = diaobo11['bianma'].map(bianmaBiaozhun)
diaobo11 = diaobo11.assign(price = diaobo11['bianma'].map(bianma_diaobojia))
diaobo3 = pd.concat([diaobo11,diaobo12])
diaobo3.to_excel('diaobo3.xlsx')

diaobo4 = diaobo3.copy()
diaobo4.insert(3,'chuku_jiner',round(diaobo4['chuku_ben_diaobo']*diaobo4['price'],2))
diaobo4.insert(6,'ruku_jiner',round(diaobo4['ruku_ben_diaobo']*diaobo4['price'],2))
diaobo4 = diaobo4.sort_values(['code'])
diaobo4.to_excel('diaobo4.xlsx',index = False)

chuku = diaobo4[['code', 'chuku_ben_diaobo', 'chuku_jian_diaobo', 'chuku_jiner','price', 'bianma']]
ruku = diaobo4[['code','ruku_ben_diaobo', 'ruku_jian_diaobo', 'ruku_jiner', 'price', 'bianma']]
chuku1 = jiaHeji(chuku,dic)
ruku1 = jiaHeji(ruku,dic)
jindiaobo = diaobo4.copy()
jindiaobo.insert(7,'jiner',jindiaobo.chuku_jiner - jindiaobo.ruku_jiner)
jindiaobo1 = jindiaobo.copy()
jindiaobo1 = jiaHeji(jindiaobo1,dic)

chuku2 = chuku1[~((chuku1['出库本数']) == 0 & (chuku1['出库件数'] == 0)) ]
ruku2 = ruku1[~((ruku1['入库本数'] == 0) & (ruku1['入库件数'] == 0)) ]
path = r"F:\a00nutstore\008\zww08\产成品\调拨"
os.chdir(path)
qijian = easygui.enterbox('请输入期间"2025-03"')
fname = f'大库调拨电商库明细{qijian}.xlsx'
wb=openpyxl.Workbook()
wb.save(fname)
with pd.ExcelWriter(fname, engine='openpyxl') as writer:
    chuku2.to_excel(writer, sheet_name='出库',  index=False)
    ruku2.to_excel(writer, sheet_name='入库',  index=False)
    jindiaobo1.to_excel(writer, sheet_name='调拨',  index=False)
wb = openpyxl.load_workbook(fname)
for i in wb.sheetnames:
    if i == 'Sheet':
        ws = wb[i]
        wb.remove_sheet(ws)
    else :
        ws = wb[i]
        ws.freeze_panes = 'B2'
wb.save(fname)
os.startfile(fname)





