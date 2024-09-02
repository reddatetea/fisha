# -*- coding:utf-8 -*-

import openpyxl
import datetime
import os

# 当天日期
dtrq = datetime.date.today().strftime('%Y%m%d')
path = r'F:\a00nutstore\006\zw\ZHIXIANG'
os.chdir(path)

filename = '纸箱当月入库%s.xlsx' % dtrq
fname = os.path.join(path, filename)

wb1 = openpyxl.Workbook()
ws1 = wb1.create_sheet(title='当月')
print(ws1)
wb1.save(fname)


fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
wb2 = openpyxl.load_workbook(fname2)
ws2 = wb2['流水账']

# qijian = '2020-04'
print(ws2.max_row)