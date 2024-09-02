#本版尝试不要网格线
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font,PatternFill,GradientFill,Side,Border,Alignment
import xlwings as xw
import easygui
import excelmessage
import os
import subprocess
import excelprint

def setPrintArea(fname):        #打印区域设置
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选要设置打印区域的excel工作表'
    # if len(sheetnames) == 1 :
    #     ws = wb.active
    # else :
    #     print(msg)
    #     choice = easygui.buttonbox(msg,title='设置打印区域',choices=sheetnames)
    #     ws = wb['%s'%choice]

    ws =wb['财务汇总']

    ws_name = ws.title

    max_row = ws.max_row      #最大行数
    #print(max_row)
    max_column = ws.max_column   #最大列数
    #print(max_column)

    max_column_letter = get_column_letter(max_column)    # 根据列的数字返回字母

    # 根据字母返回列的数字
    #print(column_index_from_string('D')) # 4

    for  index,rows in  enumerate(ws.values):
        if  rows[0] in ['',None] :
            start_row = index +2

            continue
        if rows[0].strip() in  ['总计' ]:
            #print(rows[0])
            max_row = index + 1
            break

    max_cell = '$%s$%s'%(max_column_letter,max_row)   #max_row和mar_column对应的单元格
    #print_area = '$A$1:{}'.format(max_cell)        #打印区域
    print_area = '$A${}:{}'.format(start_row,max_cell)        #打印区域
    ws.print_area = print_area                     #设置打印区域

    wb.save(fname)
    return fname,ws_name,start_row,max_row


def firstRowSeting(fname,ws_name,start_row,max_row):        #首行设置、其它行设置
    wb = openpyxl.load_workbook(fname)
    ws = wb['%s'%ws_name]

    #sheet_view.topLeftCell = 'A1'     #避坑
    ws.print_title_cols = 'A:B'  # the first two cols
    #ws.print_title_rows = '1:1'  # the first row
    ws.print_title_rows = '{}:{}'.format(start_row,start_row)  # the first row
    ws.freeze_panes = 'A{}'.format(start_row+1)
    # ws.row_dimensions[1].height = 20
    # ws.column_dimensions['B'].width = 20

    font = Font(name = '微软雅黑',size=12,bold=True,italic=True,color='000000')
    pattern_fill = PatternFill(fill_type = 'solid',fgColor='40E0D0')  #绿宝石
    pattern_fill1 = PatternFill(fill_type='solid', fgColor='AFEEEE')  #苍白的 绿宝石

    alignment = Alignment(horizontal = 'center',vertical = 'center',text_rotation = 0,wrap_text = False)

    #side = Side(style ="thin'',color = '000000')        #边框 线型为thin,黑色
    side = Side(style=None, color='000000')  # 边框 线型为thin,黑色
    border = Border(left = side,right = side,top = side,bottom = side)    #定义边框的左右上下


    for first_row in ws.iter_rows(min_row=start_row,  max_row=start_row):

        i = 0
        for cell in first_row:
            i = i + 1
            cell.font = font
            cell.fill = pattern_fill
            cell.alignment = alignment
            cell.border = border
            #width = len(str(cell.value))*2 + 1
            cell_column_letter = get_column_letter(i)
            #ws.column_dimensions[cell_column_letter].width = 20        #设置列宽

    font1 = Font(name = '微软雅黑',size=10,bold=False,italic=False,color='000000')

    for nextrows in ws.iter_rows(start_row +1, max_row= max_row-1):

        i = 0
        for cell in nextrows:
            i = i + 1
            cell.font = font1
            cell.border = border
            if  cell.row%2 != 0:
                cell.fill = pattern_fill1
            else :
                continue

    for nextrows in ws.iter_rows(min_row=max_row, max_row=max_row):
        for cell in nextrows:
            cell.font = font
            cell.border = border
            cell.fill = pattern_fill
            cell.alignment = alignment
    wb.save(fname)

    app = xw.App(visible = False,add_book = False)
    wb = app.books.open(fname)

    ws = wb.sheets[ws_name]

    ws.autofit()                #自动适应单元格
    wb.save(fname)
    wb.close()
    app.quit()
    return fname


def yemei(fname,ws_name,start_row):
    wb = openpyxl.load_workbook(fname)
    #print('现在设置页眉页脚')
    ws = wb['%s'%ws_name]
    # ws.print_options.horizontalCentered = True
    # ws.print_options.verticalCentered = True
    # msg = '请选择是否输入标题'
    # print(msg)
    # titleYN = easygui.ccbox(msg,title='请选择Yes or No',choices=('yes','no'),image = None)
    # if titleYN == True :
    #     title = input('请输入打印标题：\n')
    # else :
    #     title = ws_name

    title = "{}{}材料入库".format(ws.cell(row = 1,column = 2).value,' ')

    #冻结单元格
    ws.freeze_panes = 'A{}'.format(start_row+1)

    #打印标题
    ws.print_title_cols = 'A:B' # the first two cols   #左端标题列（即从左侧重复的列数）
    ws.print_title_rows = '{}:{}'.format(start_row,start_row)   # the first row     #顶端标题行（即从上端重复的行数）

    #页眉设置
    ws.oddHeader.center.text =  '%s'%title    #1/n      页眉中文字
    ws.oddHeader.center.size = 18                   # 页眉中字体大小
    # ws.oddHeader.center.font = '宋体,Italic'         # 页眉中字体,宋体、倾斜
    # ws.oddHeader.center.font = '宋体,Bold'         # 页眉中字体,宋体、加粗
    ws.oddHeader.center.font = '宋体,Bold Italic'     # 页眉中字体,宋体、加粗、倾斜

    ws.oddHeader.center.color = "000000"            # 页眉中字体颜色

    #页脚设置
    ws.oddFooter.center.text =  " &[Page] / &N"     #1/n      页脚中文字
    ws.oddFooter.center.size = 12                   # 页脚中字体大小
    ws.oddFooter.center.font = "Tahoma"             # 页脚中字体
    ws.oddFooter.center.color = "000000"            # 页脚中字体颜色

    ws.oddFooter.right.text =  "张文伟 &[Date]"     #页脚右 文字
    ws.oddFooter.right.size = 12                    #页脚右 字体大小
    ws.oddFooter.right.font = "书体坊米芾体"        #页脚右 字体
    ws.oddFooter.right.color = "000000"             #页脚右 字体颜色


    wb.save(fname)

def  rowHeight(fname,ws_name,start_row,max_row):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    for i in range(start_row,max_row+1):
        if  i==start_row  or  i==max_row:
            ws.row_dimensions[i].height = 30
        else :
            ws.row_dimensions[i].height = 27

    wb.save(fname)


def main():
    #fname = excelmessage.excelMessage()
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    path,filename = os.path.split(fname)
    os.chdir(path)
    fname,ws_name,start_row,max_row = setPrintArea(fname)
    fname = firstRowSeting(fname,ws_name,start_row,max_row)
    yemei(fname,ws_name,start_row)
    rowHeight(fname,ws_name,start_row,max_row)
    # wb = openpyxl.load_workbook(fname)
    # ws = wb[ws.title]
    # ws.row_dimensions[star_row].height =30

    #subprocess.Popen(['start', {}].format(fname), shell=True)         #命令行打开excel
    #subprocess.Popen(fname,shell= True)3
    #os.system(fname)              #打开excel文件
    excelprint.wsPrint(fname,ws_name)


if __name__ == '__main__':
    main()