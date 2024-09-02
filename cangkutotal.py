from openpyxl import Workbook,load_workbook
import cailiaorenames

def cangkuTotal(fname):
    cw_names, cw_Newnames, fname = cailiaorenames.jiacaiwuname(fname)
    #打开 '盘存表材料更名.xlsx'
    wb1 =load_workbook(fname)
    sheet1 = wb1.active

    jishu = 0
    first_list=[]
    for index, row in enumerate(sheet1.values):
        print(row[1])
        first_list.append(row[1])

        if row[0] == None:
            jishu = jishu + 1
            # print(jishu)

    mrows = sheet1.max_row - jishu

    #创立七个字典，分别是期初数、入库数量、采购入库，半成品入库，出库数量，生产领料，结余数
    qichu_dic = {}
    ruku_dic = {}
    caigou_dic ={}
    banchengpin_dic = {}
    chuku_dic ={}
    lingliao_dic = {}
    qimo_dic = {}

    qichu = 0
    ruku = 0
    caigou = 0
    banchengpin = 0
    chuku = 0
    lingliao = 0
    qimo  = 0

    #cw_names,cw_Newnames,fname = cailiaorenames.jiacaiwuname(fname)
    yuancailiao = list(set(cw_Newnames))

    print(yuancailiao)
    yuancailiao.remove('品名')

    #斯初数
    qichu_total = 0
    for i in yuancailiao :
        qichu = 0
        for index, row in enumerate(sheet1.values):
            if index == 0:
                continue

                # 将上面计算出的实际行数，运行到下面的程序！不这样就会报错！
            if index < mrows:
                if row[10] == i:
                    qichu = qichu + row[3]
                else:
                   continue

        qichu_dic[i] = qichu
        qichu_total = qichu_total +qichu
    #print(qichu_dic)

    #入库数
    ruku_total = 0
    i = 0
    for i in yuancailiao :
        ruku = 0
        for index, row in enumerate(sheet1.values):
            if index == 0:
                continue
            if index < mrows:
                if row[10] == i:
                    ruku = ruku + row[5]
                else:
                   continue
        ruku_dic[i] = ruku
        ruku_total = ruku_total + ruku
    #print(ruku_dic)

    #采购入库
    caigou_total = 0
    i = 0
    for i in yuancailiao :
        caigou = 0
        for index, row in enumerate(sheet1.values):
            if index == 0:
                continue
            if index < mrows:
                if row[10] == i:
                    caigou = caigou + row[4]
                else:
                   continue
        caigou_dic[i] = caigou
        caigou_total = caigou_total + caigou

    #半成品入库
    banchengpin_total = 0
    i = 0
    for i in yuancailiao :
        banchengpin = 0
        for index, row in enumerate(sheet1.values):
            if index == 0:
                continue
            if index < mrows:
                if row[10] == i:
                    banchengpin =banchengpin + row[9]
                else:
                   continue
        banchengpin_dic[i] = banchengpin
        banchengpin_total = banchengpin_total + banchengpin

    #出库数量
    chuku_total = 0
    i = 0
    for i in yuancailiao :
        chuku = 0
        for index, row in enumerate(sheet1.values):
            if index == 0:
                continue
            if index < mrows:
                if row[10] == i:
                    chuku =chuku + row[6]
                else:
                   continue
        chuku_dic[i] = chuku
        chuku_total = chuku_total + chuku

    #生产领料
    lingliao_total = 0
    i = 0
    for i in yuancailiao :
        lingliao = 0
        for index, row in enumerate(sheet1.values):
            if index == 0:
                continue
            if index < mrows:
                if row[10] == i:
                    lingliao =lingliao + row[8]
                else:
                   continue
        lingliao_dic[i] = lingliao
        lingliao_total = lingliao_total + lingliao

    #期末
    qimo_total = 0
    i = 0
    for i in yuancailiao :
        qimo = 0
        for index, row in enumerate(sheet1.values):
            if index == 0:
                continue
            if index < mrows:
                if row[10] == i:
                    qimo =qimo + row[7]
                else:
                   continue
        qimo_dic[i] = qimo
        qimo_total = qimo_total + qimo

    taitou_lists = ['qichu','ruku','caigou','banchengpin','chuku','lingliao','qimo']

    print('******************************')
    print('打印期初字典')
    print('期初合计为：%s'%qichu_total)
    print('******************************')

    for key,value in qichu_dic.items():
        print('{key}:{value}'.format(key = key, value = value))
    print('******************************')
    print('打印入库数量字典')
    print('入库数量合计为：%s'%ruku_total)
    print('******************************')

    for key,value in ruku_dic.items():
        print('{key}:{value}'.format(key = key, value = value))
    print('******************************')
    print('打印采购入库字典')
    print('采购入库合计为：%s'%caigou_total)
    print('******************************')

    for key,value in caigou_dic.items():
        print('{key}:{value}'.format(key = key, value = value))
    print('******************************')
    print('打印半成品入库字典')
    print('半成品入库合计为：%s'%banchengpin_total)
    print('******************************')

    for key,value in banchengpin_dic.items():
        print('{key}:{value}'.format(key = key, value = value))
    print('******************************')
    print('打印出库数量字典')
    print('出库数量合计为：%s'%chuku_total)
    print('******************************')

    for key,value in chuku_dic.items():
        print('{key}:{value}'.format(key = key, value = value))
    print('******************************')
    print('打印生产领料字典')
    print('生产领料合计为：%s'%lingliao_total)
    print('******************************')
    for key,value in lingliao_dic.items():
        print('{key}:{value}'.format(key = key, value = value))

    print('******************************')
    print('打印结存字典')
    print('结存合计为：%s'%qimo_total)
    print('******************************')
    for key,value in qimo_dic.items():
        print('{key}:{value}'.format(key = key, value = value))

    # 各字典加一个空对空，避免excle科目那一列为空值时，找不到元素的情况
    qichu_dic[''] = ''
    ruku_dic[''] = ''
    caigou_dic[''] = ''
    banchengpin_dic[''] = ''
    chuku_dic[''] = ''
    lingliao_dic[''] = ''
    qimo_dic[''] = ''
    print('字典长度')
    print(len(qichu_dic))

    cangkucailiaolist1 = yuancailiao

    return qichu_dic,ruku_dic,caigou_dic,banchengpin_dic,chuku_dic,lingliao_dic,qimo_dic,taitou_lists,cangkucailiaolist1

def main():
    fname = r'F:\a00nutstore\006\zw\2021\202102\盘存表材料更名.xlsx'
    cangkuTotal(fname)

if __name__== '__main__':
    main()



