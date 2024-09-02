'''
本模块是从流水账中将纸箱入库数据引入，使用类
'''
# _*_ coding:utf-8 _*_
import openpyxl
import os
import datetime
import zhixiangTodic
from openpyxl.utils import column_index_from_string
import zhixiang

def zhixiangDangyue(qijian):
    # 当天日期
    dtrq = datetime.date.today().strftime('%Y%m%d')
    path = r'F:\a00nutstore\006\zw\ZHIXIANG'
    #path = r'D:\a00nutstore\006\zw\ZHIXIANG'
    os.chdir(path)

    filename = '纸箱当月入库%s.xlsx'%dtrq
    fname = os.path.join(path, filename)

    wb1 = openpyxl.Workbook()
    ws1 = wb1.create_sheet(title='当月')

    fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    #fname2 = r'D:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['流水账']

    #qijian = '2020-04'
    print(ws2.max_row)
    jianchen = {'武汉市九安源纸业有限公司': '武汉九安源纸业有限公司',
                '恒龙包装': '武汉恒龙包装有限公司',
                '孝感鑫荣环保包装': '孝感鑫荣环保包装科技有限公司'
                }
    for i in range(1, ws2.max_row + 1):
        if i == 1:
            taitou = (
                '单位','日期', '单号', '箱型', '品名', '长', '宽', '高', '长*宽*高', '适装产品', '数量', '单个面积', '面积',
                '制版费', '合同单价', '合同金额', '对账单单价', '对账单金额', '多计', '多计金额', '备注','pricename')
            ws1.append(taitou)
        else:
            if (ws2.cell(i, 3).value in jianchen) and (ws2.cell(i, 11).value == qijian) :

                ws1.cell(i, 1, value=ws2.cell(i, 3).value)
                ws1.cell(i, 2, value=ws2.cell(i, 1).value)
                ws1.cell(i,3, value=ws2.cell(i, 2).value)
                ws1.cell(i, 4, value=ws2.cell(i, 4).value)
                ws1.cell(i, 5, value=ws2.cell(i, 4).value)

                zhixiangdic = zhixiangTodic.zhixiangdic()
                zhixiangdic.setdefault(ws2.cell(i, 4).value,(0,0,0))
                chang = zhixiangdic[ws2.cell(i, 4).value][0]
                kuan = zhixiangdic[ws2.cell(i, 4).value][1]
                gao = zhixiangdic[ws2.cell(i, 4).value][2]

                ws1.cell(i, 6, value=chang)
                ws1.cell(i, 7, value=kuan)
                ws1.cell(i, 8, value=gao)
                ws1.cell(i, 9, value=str(chang)+'*'+str(kuan)+'*'+str(gao))
                ws1.cell(i, 11, value=ws2.cell(i,6).value)
                ws1.cell(i, 17, value=round(ws2.cell(i,7).value, 2))
                ws1.cell(i, 18, value=round(ws2.cell(i,8).value, 2))
                ws1.cell(i,24,value=ws2.cell(i,10).value)


            else:
                continue
    wb1.save(fname)
    wb2.close()

    # 当月正
    fname3 = r'F:\a00nutstore\006\zw\ZHIXIANG\纸箱当月入库%s.xlsx'%dtrq
    #fname3 = r'D:\a00nutstore\006\zw\ZHIXIANG\纸箱当月入库%s.xlsx' % dtrq
    wb3 = openpyxl.load_workbook(fname3)
    ws3 = wb3['当月']

    ws4 = wb3.create_sheet(title='当月正')

    for index, row in enumerate(ws3.values):
        if index == 0:
            ws4.append(row)
        else:
            if row[0] != None:
                ws4.append(row)
            else:
                continue

    wb3.save(fname3)

    #添加合同价格，计算合同金额，多计单计，多计金额
    wb3 = openpyxl.load_workbook(fname3)
    ws3 = wb3['当月正']
    max_rows = ws3.max_row
    pinming_num = column_index_from_string('E')

    chang_num = column_index_from_string('F')
    kuan_num = column_index_from_string('G')
    gao_num = column_index_from_string('H')
    shuliang_num = column_index_from_string('K')

    dange_mianji_num = column_index_from_string('L')
    mianji_num = column_index_from_string('M')
    hetong_danjia_num = column_index_from_string('O')
    hetong_jiner_num = column_index_from_string('P')
    duoji_num = column_index_from_string('S')
    duojijiner_num = column_index_from_string('T')
    songhuo_danjia_num = column_index_from_string('Q')
    songhuo_jiner_num = column_index_from_string('R')

    for row in range(2,max_rows+1):
        gongyingshang = ws3.cell(row, 1).value
        pinming = ws3.cell(row,pinming_num).value
        chang = ws3.cell(row, chang_num).value
        kuan = ws3.cell(row, kuan_num).value
        gao = ws3.cell(row, gao_num).value
        shuliang = ws3.cell(row, shuliang_num).value
        songhuo_danjia = ws3.cell(row, songhuo_danjia_num).value
        songhuo_jiner = ws3.cell(row, songhuo_jiner_num).value


        leibie = zhixiang.zhixiangLeibian(pinming)
        pinmin, chang, kuan, gao, dange_mianji = leibie.zhixiangmessage()
        singlePrice = leibie.singleprice(gongyingshang)

        #合同单价
        hetong_danjia = round(dange_mianji * singlePrice, 2)

        #合同金额
        hetong_jiner = round(hetong_danjia*shuliang,2)

        #写入单个面积
        ws3.cell(row,dange_mianji_num,value = dange_mianji)

        # 写入面积
        ws3.cell(row, mianji_num, value=round(dange_mianji*shuliang,2))

        #写入合同单价
        ws3.cell(row,hetong_danjia_num,value = hetong_danjia)

        # 写入合同金额
        ws3.cell(row, hetong_jiner_num, value=round(hetong_danjia*shuliang,2))

        # 写入多计
        ws3.cell(row, duoji_num, value=songhuo_danjia-hetong_danjia)

        # 写入多计金额
        ws3.cell(row, duojijiner_num, value=songhuo_jiner - hetong_jiner)

    wb3.save(fname3)

    #九安源
    fname3 = r'F:\a00nutstore\006\zw\ZHIXIANG\纸箱当月入库%s.xlsx' % dtrq
    #fname3 = r'd:\a00nutstore\006\zw\ZHIXIANG\纸箱当月入库%s.xlsx' % dtrq
    wb3 = openpyxl.load_workbook(fname3)
    ws3 = wb3['当月正']

    qishu = qijian[2:4]+qijian[-2:]

    ws4 = wb3.create_sheet(title='九安源%s'%qishu)
    ws5 = wb3.create_sheet(title='恒龙%s'%qishu)
    ws6 = wb3.create_sheet(title='孝感鑫荣%s' % qishu)



    for index, row in enumerate(ws3.values):
        if index == 0:
            ws4.append(row)
            ws5.append(row)
            ws6.append(row)
        else:
            if row[0] == '武汉市九安源纸业有限公司':
                ws4.append(row)

            elif row[0] == '恒龙包装':
                ws5.append(row)

            elif row[0] == '孝感鑫荣环保包装':
                ws6.append(row)
            else:
                continue

    wb3.save(fname3)



def main():
    qijian = input('请输入期间：格式为2020-04：\n')
    #piaojuhao = input('请输入票据号：')
    zhixiangDangyue(qijian)

if __name__ == '__main__':
    main()



















