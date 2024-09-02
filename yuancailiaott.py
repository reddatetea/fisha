'''
yuancailiaott.py
本模块将仓库的盘存表和财务的数量金额账完全合并
本版增加导入当月出库，作为出库参考
耗时5分钟
'''
# -*-coding : utf-8 -*-
from openpyxl import Workbook,load_workbook
#from openpyxl.utils import get_column_letter
import re
import csv
import openpyxl
import win32com.client as win32
import os
import excelmessage
import time
import cailiaorenames
import cangkutotal
import chaiwutotal
import zhongtotal
import addqita
import addqitaheji
import addlists
import chukuchuli
import dangyueweida

#删除盘存表小计和合计行
def shanchukonghang(fname):
    dirname, filename = os.path.split(fname)
    wb0 =load_workbook(fname)
    sheetnames =wb0.sheetnames

    # 第一个工作表的名字
    for sheetname in sheetnames:
        if 'AAA' in sheetname:
            sheetname0 = sheetname
        else :
            continue

    sheet0 = wb0[sheetname0]
    jishu0 = 0
    first_list0 = []
    for index,row in enumerate(sheet0.values):
        print(row[1])
        first_list0.append(row[1])

        if row[0] == None:
            jishu0 = jishu0 +1

    mrows0 = sheet0.max_row -jishu0

    #删除空行
    sk_wb = Workbook()
    sk_sheet = sk_wb.active
    sk_sheet.title = '盘存表'

    for index,row in enumerate(sheet0.values):
        #print(index,cw_Newnames[index])
        if row[1]=='' or row[1]==None :
            continue
        sk_sheet.append(row)

    filename = '盘存表.xlsx'
    fname = os.path.join(dirname,filename)
    sk_wb.save(fname)
    time.sleep(3)
    return fname

def addqitas(fname1):

    cw_names, cw_Newnames, fname1 = cailiaorenames.jiacaiwuname(fname1)

    qichu_dic, ruku_dic, caigou_dic, banchengpin_dic, chuku_dic, lingliao_dic, qimo_dic, taitou_lists, cangkucailiaolist1 = cangkutotal.cangkuTotal(fname1)

    list1 = cangkucailiaolist1

    print(list1)
    print('请分别输入财务数量金额明细账的路径和文件名：')
    list2,kemuname_dic,fname2 = chaiwutotal.chaiwuTotal()
    print(list2)

    addList1s, addList2s = addlists.addLists(list1, list2)

    fname = zhongtotal.zhongTotal(qichu_dic, ruku_dic, caigou_dic, banchengpin_dic, chuku_dic, lingliao_dic, qimo_dic, taitou_lists, cangkucailiaolist1,fname2)
    dirfile, filename = os.path.split(fname)
    os.chdir(dirfile)


    #打开'数量金额盘存表.xlsx'

    wb6 = load_workbook(fname)
    ws6 = wb6.active

    jishu = 0
    first_list = []
    for index, row in enumerate(ws6.values):
        # print(row[1])
        first_list.append(row[1])

        if row[0] == None:
            jishu = jishu + 1
    mrows = ws6.max_row - jishu
    print(mrows)

    # 插入N行，n等于列表个数,在第mrows-1行上面行开始处
    ws6.insert_rows(mrows - 1, len(addList2s))

    # 批量写入数据, 写入科目addlist1
    for i in range(len(addList2s)):
        ws6.cell(mrows - 1 + i, 18, addList2s[i])

    # 批量写入数据, 按列写入，第18列，第mrows-1行开始,
    # 写入期初数量
    for i in range(len(addList2s)):
        try:
            a = qichu_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0

        ws6.cell(mrows - 1 + i, 24, a)

    # 写入入库数量
    for i in range(len(addList2s)):
        try:
            a = ruku_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0
        ws6.cell(mrows - 1 + i, 25, a)

    # 写入采购数量
    for i in range(len(addList2s)):
        try:
            a = caigou_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0

        ws6.cell(mrows - 1 + i, 26, a)

    # 写入半成品数量
    for i in range(len(addList2s)):
        try:
            a = banchengpin_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0

        ws6.cell(mrows - 1 + i, 27, a)

    # 写入出库数量
    for i in range(len(addList2s)):
        try:
            a = chuku_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0
        ws6.cell(mrows - 1 + i, 28, a)

    # 写入生产领料数量
    for i in range(len(addList2s)):
        try:
            a = lingliao_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0
        ws6.cell(mrows - 1 + i, 29, a)

    # 写入期末数量
    for i in range(len(addList2s)):
        try:
            a = qimo_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0

        ws6.cell(mrows - 1 + i, 30, a)

    # 删除一行，删除第morws行
    # ws6.delete_rows(mrows)

    newname = filename[:-5] + '8.xlsx'
    fname = os.path.join(dirfile,newname)
    wb6.save(fname)
    return kemuname_dic,fname



def main():
    print('请点选原材料盘存表文件')
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    fname1 = shanchukonghang(fname)

    kemuname_dic, fname = addqitas(fname1)

    newname = addqitaheji.addqitaHeji(kemuname_dic, fname)
    print('最终文件:数量金额盘存表18.xlsx')
    #导入当月出库，并形成出库参考
    print('请点选本月材料出库文件')
    fname2 = excelmessage.wenjian()
    fname2 = excelmessage.excelMessage(fname2)
    path,filename = os.path.split(fname)
    os.chdir(path)

    print('计算当月未达，即财务做了入库记账，而仓库算作下个期间')
    qijian = input('请输入当前期间(格式如2020-04)：')
    xia_qijian = qijian[:-2] + '%02d' % (int(qijian[-2:]) + 1)
    fname = dangyueweida.baiyunweida(xia_qijian)
    dangyueweida.zhiweida(xia_qijian, fname)
    dangyueweida.fuliaoweida(xia_qijian,fname)
    dangyueweida.weida(fname)
    weida_dic = dangyueweida.weidaDic(fname)
    chukuchuli.chukuChuli(fname=fname2,newname=newname,weida_dic=weida_dic)

if __name__=='__main__':
    main()