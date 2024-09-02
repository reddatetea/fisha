import  os
import  openpyxl
import  excelmessage
import easygui

def  getrenshu(ws):

    max_row = ws.max_row
    for i in range(1,max_row+1):
        if ws.cell(i,1).value =='营业区: 鸣鹤中餐厅':
            minhe_hang = i
            #print(minhe_hang)
        elif ws.cell(i,1).value =='营业区: 江畔咖啡厅':
            jiangpan_coffee_hang = i
            #print(jiangpan_coffee_hang)
        elif ws.cell(i,1).value =='营业区: 汇吧':
            huiba_hang = i
            #print(huiba_hang)
        elif ws.cell(i,1).value =='营业区: 江畔露台':
            jiangpan_lutai_hang = i
            #print(jiangpan_lutai_hang)
        elif ws.cell(i,1).value =='营业区: 汇吧':
            huiba_hang = i
            #print(huiba_hang)
        elif ws.cell(i,1).value =='营业区: 客房送餐':
            kefangsongcan_hang = i
            #print(kefangsongcan_hang)
        elif ws.cell(i,1).value =='营业区: 宴会厅':
            yanhuiting_hang = i
            #print(yanhuiting_hang)
        elif ws.cell(i,1).value =='营业区: 季节售卖':
            huizhong_hang = i
            #print(huizhong_hang)

        elif ws.cell(i, 1).value == '付款':
            fukuan_hang = i
            break
        else :
            continue


    return minhe_hang,jiangpan_coffee_hang,huiba_hang,jiangpan_lutai_hang,kefangsongcan_hang,yanhuiting_hang,huizhong_hang,fukuan_hang


def  countrenshu(ws,min_row,max_row):
    renshu_Breakfast, renshu_Lunch, renshu_Dinner,renshu_Supper,renshu_CoffeeBreak = 0,0,0,0,0
    for row in ws.iter_rows(min_row, max_row, max_col=12):

        if row[0].value =='Breakfast':
            renshu_Breakfast = row[-1].value
            #print(renshu_Breakfast)
        elif row[0].value =='Lunch':
            renshu_Lunch = row[-1].value
            #print(renshu_Lunch)
        elif row[0].value =='Dinner':
            renshu_Dinner = row[-1].value
            #print(renshu_Dinner)
        elif row[0].value =='Supper':
            renshu_Supper = row[-1].value
            #print(renshu_Supper)
        elif row[0].value =='Coffee Break':
            renshu_CoffeeBreak = row[-1].value
            #print(renshu_CoffeeBreak)
        else :
            continue

    return renshu_Breakfast,renshu_Lunch,renshu_Dinner,renshu_Supper,renshu_CoffeeBreak

def quZhongrenshu(ws,min_row,max_row):
    for row in ws.iter_rows(min_row, max_row, max_col=12):
        if row[0].value == '小计':
            zhongrenshu = row[-1].value
        else:
           continue
    return zhongrenshu

def  daorurenshu(fname1,minhe,jiangpan_coffee,huiba,jiangpan_lutai,kefangsongcan,yanhuiting):
    wb1 = openpyxl.load_workbook(fname1)
    ws1 = wb1.active
    ws1.cell(10,13).value = jiangpan_coffee[0]
    ws1.cell(11, 13).value = jiangpan_coffee[1]
    ws1.cell(12, 13).value = jiangpan_coffee[2]
    ws1.cell(13, 13).value = jiangpan_coffee[3]
    ws1.cell(14, 13).value = minhe[1]
    ws1.cell(15, 13).value = minhe[2]
    ws1.cell(24, 13).value = huiba[1]
    ws1.cell(25, 13).value = huiba[2]
    ws1.cell(30, 13).value = kefangsongcan[0]
    ws1.cell(31, 13).value = kefangsongcan[1]
    ws1.cell(32, 13).value = kefangsongcan[2]
    ws1.cell(33, 13).value = kefangsongcan[3]
    ws1.cell(34, 13).value = yanhuiting[0]
    ws1.cell(35, 13).value = yanhuiting[1]
    ws1.cell(36, 13).value = yanhuiting[2]
    ws1.cell(37, 13).value = yanhuiting[4]
    wb1.save(fname1)
    return fname1

def main():
    msg = '请点选cross_outlet_revenue.xls'
    print(msg)
    fname = easygui.fileopenbox(msg,title =msg)
    fname  = excelmessage.excelMessage(fname)
    #fname =r'C:\Users\asus\Desktop\晴川跨地区营业收入表cross_outlet_revenue.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    getrenshu(ws)
    minhe_hang,jiangpan_coffee_hang,huiba_hang,jiangpan_lutai_hang,kefangsongcan_hang,yanhuiting_hang,huizhong_hang,fukuan_hang = getrenshu(ws)
    minhe = countrenshu(ws,minhe_hang, jiangpan_coffee_hang)
    jiangpan_coffee = countrenshu(ws,jiangpan_coffee_hang,huiba_hang)
    huiba = countrenshu(ws,huiba_hang,jiangpan_lutai_hang)
    jiangpan_lutai = countrenshu(ws,jiangpan_lutai_hang,kefangsongcan_hang)
    kefangsongcan = countrenshu(ws,kefangsongcan_hang,yanhuiting_hang)
    yanhuiting = countrenshu(ws,yanhuiting_hang,huizhong_hang)
    #print(minhe,jiangpan_coffee,huiba,jiangpan_lutai,kefangsongcan,yanhuiting)
    renshuheji = sum(minhe)+sum(jiangpan_coffee)+sum(huiba)+sum(jiangpan_lutai)+sum(kefangsongcan)+sum(yanhuiting)
    #print(renshuheji)
    zhongrenshu = quZhongrenshu(ws,huizhong_hang,fukuan_hang)
    wb.close()
    if renshuheji == zhongrenshu:
        msg = '请点选"导入SUN的总人数.xlsx"'
        print(msg)
        fname1 = easygui.fileopenbox(msg)
        fname1 = daorurenshu(fname1,minhe,jiangpan_coffee,huiba,jiangpan_lutai,kefangsongcan,yanhuiting)
    else :
        print('人数有误，请手工计算')

    os.system(fname1)



if __name__ =='__main__':
    main()


