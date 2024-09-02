'''
#本模块删除excel中的重复行,本版直接数据清洗,用inplace，失败不知原因
'''
# 导入pandas包并重命名为pd
import pandas as pd
import os
import xlsxlsx
import excelmessage
import openpyxl


def delchongfu(fname, excelname, sheetname):
    # dirname, excelname = os.path.split(fname)
    # wb = openpyxl.load_workbook(fname)
    # ws = wb.active
    # sheetname = ws.title
    # wb.close()
    # excelname = fname

    # 读取Excel中Sheet1中的数据
    data = pd.DataFrame(pd.read_excel(excelname, sheetname))

    # 查看读取数据内容
    print(data)

    # 查看是否有重复行
    re_row = data.duplicated()
    print(re_row)

    # 查看去除重复行的数据,
    no_re_row = data.drop_duplicates(subset=['单位', '供应商', '时间','单号', '材料', '入库','入库（kg）'], keep='first')
    #no_re_row = data.drop_duplicates(keep='first')
    print(no_re_row)

    # 查看基于[科目编码]列去除重复行的数据
    wp = data.drop_duplicates()
    # wp = data.drop_duplicates(['科目编码'])
    print(wp)

    # 将去除重复行的数据输出到excel表中
    duplicated_name = excelname[:-5] + '过滤重复行' + '.xlsx'
    no_re_row.to_excel("%s" % duplicated_name)

    return duplicated_name


def main():
    print('请输入要删除重复数据的excel名字:')
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)

    dirname, excelname = os.path.split(fname)
    os.chdir(dirname)

    wb = openpyxl.load_workbook(fname)
    ws = wb.active

    # sheetname = wb.sheetnames[0]
    sheetname = ws.title
    print(sheetname)
    wb.close()

    delchongfu(fname, excelname, sheetname)


if __name__ == '__main__':
    main()










