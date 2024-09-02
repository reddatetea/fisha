# _*_ coding:utf-8 _*_
import os
import easygui
import shishiliushuizhang
import lingpeijianliushuizhang
import re
import excelmessage
import quchong
import openpyxl
import datetime


def pipeifile(guanjianzi, filename):
    # regax = r'.*%s.*\.[x|X][l|L][s|S][x|X]?'%guanjianzi
    regax = r'.*(?P<guanjian>%s).*\.[x|X][l|L][s|S][x|X]?' % guanjianzi
    pattern = re.compile(regax)
    mat = pattern.search(filename)
    return mat

def main():
    # 公司
    desktopPath = r'D:\ribaobiao'
    os.chdir(desktopPath)
    filenames = os.listdir(desktopPath)
    in_subject=['入货单号', '入库日期', '相关单位', '单据附注', '货品编码', '货品名称', '规格', '所属类别', '单位', '入库数量', '单价', '金额', '备注', '期间', '标准日期']
    guanjianzi = '统计'

    for filename in filenames:
        mat = pipeifile(guanjianzi, filename)

        if mat:
            if mat.group('guanjian') == '统计':
                fname = r'F:\a00nutstore\006\zw\lingpeijian\零配件实时入库2020.xlsx'

                fname2 = os.path.join(desktopPath, filename)
                try :

                    fname2 = excelmessage.excelMessage(fname2)

                    fname2 = lingpeijianliushuizhang.addruku(fname, fname2)
                    #lingpeijianliushuizhang.ruku(fname)
                except :
                    print('"零配件入库统计"文件可能有问题！')

        else:
            pass

        #下面对零配件数据去重
        fname = r'F:\a00nutstore\006\zw\lingpeijian\零配件实时入库2020.xlsx'
        path, excelname = os.path.split(fname)
        os.chdir(path)
        sheetname = 'ssrk'
        fname = quchong.delchongfu(fname, sheetname,in_subject)

        wb = openpyxl.load_workbook(fname)
        ws = wb['ssrk']
        max_row = ws.max_row
        for cell in ws['B']:
            riqi = cell.value
            if riqi == '入库日期':
                continue
            else:
                cell.value = datetime.datetime.strftime(riqi, '%Y-%m-%d')
        wb.save(fname)



    guanjianzi = '流水'

    for filename in filenames:
        mat = pipeifile(guanjianzi, filename)

        if mat:
            if mat.group('guanjian') == '流水':
                oldname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'

                fname = os.path.join(desktopPath, filename)
                try :
                    fname = excelmessage.excelMessage(fname)
                    shishiliushuizhang.liushuizhangMeitian(oldname, fname)
                except:
                    print('"原材料流水账"文件可能有问题！')
        else:
            continue


if __name__ == '__main__':
    main()







