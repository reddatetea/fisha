'''
本模块将仓库的盘存表和财务的数量金额账合并，合并进来的盘存表以数量金额账的科目为限
'''
# -*-coding = utf-8 -*-
from openpyxl import Workbook,load_workbook
#from openpyxl.utils import get_column_letter
import re
import csv
import openpyxl
import win32com.client as win32
import os
import excelmessage
import time
import cailiaorenames
import cangkutotal
import chaiwutotal
import zhongtotal
import addqita
import addqitaheji
import addlists

#删除盘存表小计和合计行
def shanchukonghang(fname):
    wb0 =load_workbook(fname)
    sheetname0s = wb0.sheetnames

    # 第一个工作表的名字
    sheetname0 = wb0.sheetnames[0]
    sheet0 = wb0[sheetname0]
    jishu0 = 0
    first_list0 = []
    for index,row in enumerate(sheet0.values):
        print(row[1])
        first_list0.append(row[1])

        if row[0] == None:
            jishu0 = jishu0 +1

    mrows0 = sheet0.max_row -jishu0

    #删除空行
    sk_wb = Workbook()
    sk_sheet = sk_wb.active
    sk_sheet.title = '盘存表'

    for index,row in enumerate(sheet0.values):
        #print(index,cw_Newnames[index])
        if row[1]=='' or row[1]==None :
            continue
        sk_sheet.append(row)


    sk_wb.save('盘存表.xlsx')
    time.sleep(3)


def main():
    fname = excelmessage.wenjian()
    excelmessage.excelMessage(fname)
    print(fname)

    fname = fname + "x"
    print(fname)

    shanchukonghang(fname)
    cailiaorenames.jiacaiwuname()

    time.sleep(2)

    #cangkutotal.cangkuTotal()

    #chaiwutotal.chaiwuTotal()
    #list1 = cangkutotal.cangkuTotal()


    list2 = chaiwutotal.chaiwuTotal()
    print(list2)

    #zhongtotal.zhongTotal()
    list1 = zhongtotal.zhongTotal()
    #qichu_dic, ruku_dic, caigou_dic, banchengpin_dic, chuku_dic, lingliao_dic, qimo_dic, taitou_lists, cangkucailiaolist1 = cangkutotal.cangkuTotal()
    #list1 = cangkucailiaolist1
    print(list1)

    addList1s, addList2s = addlists.addLists(list1, list2)
    #addqita.addQita(list1,list2)

    #addqitaheji.addqitaHeji()

if __name__=='__main__':
    main()