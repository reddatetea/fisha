#本版本实现点选 两个日期对应的暂估单元格区域，复制到粘贴版
import os
import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter
import excelcellfill
# import excelfastsetingzg
# import  excelfastseting
import  excelseting
import pyperclip
import os


def clearsheet(fname1):
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open(fname1)
    worksheets = workbook.sheets
    ws = worksheets['zg']
    ws.clear()
    ws1 = worksheets['暂估调整']
    ws1.clear()
    ws2 = worksheets['折扣分配']
    ws2.clear()
    workbook.save()
    workbook.close()
    app.quit()

def zanguRange():
    cell_range = pyperclip.paste()
    #ws1_title = ('品名','入库(令)','入库(吨)','金额','不含税金额')
    fname1 = r'F:\a00nutstore\fishc\纸暂估调整.xlsx'
    wb1 = openpyxl.load_workbook(fname1)
    ws1 = wb1[ 'zg']
    with open('twozangu.txt', 'w') as f:
        f.write(cell_range)
    shuju = []
    with open('twozangu.txt', 'r') as f:
        for eachline in f:
            print(eachline)
            eachline = eachline.strip('\n')
            print(eachline)
            newhang = list(eachline.split('\t'))
            print(newhang)
            if newhang == ['']:
                continue
            else:
                shuju.append(newhang)
                ws1.append(newhang)
    wb1.save(fname1)
    return fname1

def  makeTwodics(fname1):
    wb = openpyxl.load_workbook(fname1)
    ws =  wb['zg']
    max_row = ws.max_row
    max_col = ws.max_column
    #gongyingshang = ws['A1'].value
    dic1 = {}
    dic2 ={}
    #获得品名列表
    pinmings = []
    for index, row in enumerate(ws.values):
        if row[0]in [None,'']:
            konghang = index+ 1

    for  cell in ws['A']:
        if  (cell.value not in ['品名','总计','',None] )  and  (cell.value not in  pinmings):
            pinmings.append(cell.value)

    print('pinmings:',pinmings)
    fields = [i.value for i in ws[1]]
    ws_title = fields.copy()
    print('fields;',fields)
    print('konghang:',konghang)
    for row in range(2,konghang):
        print(row)
        pinming = ws.cell(row,1).value

        if pinming not in ['品名','总计','',None] :
            print(pinming)
            print(row)
            hang_shuju = []
            for col in range(2,max_col+1):
                fields[col-2] = float(ws.cell(row,col).value)
                hang_shuju.append(fields[col-2])

            dic1[pinming] = hang_shuju
        else :
            break

        #调整后暂估字典

        for row in range(konghang+1, max_row + 1):
            print(row)
            pinming2 = ws.cell(row, 1).value
            if pinming2 not in ['品名','总计','',None] :

                hang_shuju = []
                for col in range(2, max_col + 1):
                    fields[col - 2] = float(ws.cell(row, col).value)
                    hang_shuju.append(fields[col - 2])

                dic2[pinming2] = hang_shuju
    print('dic1:',dic1)
    print('dic2:',dic2)
    wb.close()
    return pinmings,ws_title,dic1,dic2

def  zangutiaozheng(fname1,pinmings,ws_title,dic1,dic2):
    wb =  openpyxl.load_workbook(fname1)
    ws = wb['暂估调整']
    ws.append(ws_title)
    for row in range(2, len(pinmings) + 2):
        ws.cell(row, 1).value = pinmings[row - 2]
    wb.save(fname1)

    wb = openpyxl.load_workbook(fname1)
    ws = wb['暂估调整']

    max_row = len(pinmings)+1
    max_col = len(ws_title)
    for row in range(2,max_row+1):
        pinming =ws.cell(row,1).value
        kongzhi = [i*0 for i in range(max_col)]
        # 字典取值时，如果没有键值，则对应的值为[0，0，0，0]
        shuju2 = dic2.get(pinming,kongzhi)
        print(shuju2)
        shuju1 = dic1.get(pinming,kongzhi)
        print(shuju1)
        for col in range(2,max_col+1):
            ws.cell(row,col).value = shuju2[col-2]-shuju1[col-2]

    #加最后一行，并加上求和公式
    ws.cell(max_row + 1, 1).value = '合计'
    for col in range(2, max_col + 1):
        string = get_column_letter(col)
        ws.cell(max_row + 1, col).value = '=sum({}2:{}{})'.format(string,string,max_row)

    max_row = max_row +1

    wb.save(fname1)
    return max_row,max_col

def  zekoufenpei(fname,ws_title,dic2):
    wb = openpyxl.load_workbook(fname)
    ws2= wb['折扣分配']
    ws1 =wb['暂估调整']
    max_row = ws1.max_row
    shuju = ws_title
    for  i in range(len(shuju)):
        if '吨' in shuju[i]:
            dun_biaoji = i
            break
    for i in range(len(shuju)):
        if '金额' in shuju[i]:
            jiner_biaoji = i
            break

    zekoujiner = 0
    for index,row in enumerate(ws1.values):
        if index ==0:
            ws2.append(ws_title)
        elif index == max_row -1 :
            continue
        else :
            #print(row[jiner_biaoji])
            if   ('折扣'   in  row[0] ) or  ('价差'  in row[0]) or ('返利'   in row[0]) or ('返利'  in row[0])  :
                zekoujiner = zekoujiner +row[jiner_biaoji] # 将折扣金额累计
                continue
            else :
                ws2.append(row)
            print(zekoujiner)
        wb.save(fname)

    wb = openpyxl.load_workbook(fname)
    ws = wb['折扣分配']
    max_row = ws.max_row
    max_col = len(ws_title)
    dun_total = 0
    kongzhi = [i*0 for i in range(len(ws_title)-1)]
    for index,row in enumerate(ws.values):
        if index == 0:
            continue
        else :
            pinming = row[0]
            shuju2 = dic2.get(pinming, kongzhi)
            for  col in range(2,max_col+1):
                shu2 = shuju2[col-2]
                ws.cell(row=index + 1, column=col).value = shu2
            dun2 = shuju2[dun_biaoji - 1]
            dun_total = dun_total + dun2
            ws.cell(row=index + 1, column=dun_biaoji + 1).value = dun2

    #加最后一行，并加上求和公式

    for col in range(2,max_col+1):
        string = get_column_letter(col)
        ws.cell(max_row+1, col).value = '=sum({}2:{}{})'.format(string,string,max_row)

    ws.cell(max_row+1,dun_biaoji+1).value = dun_total
    ws.cell(max_row+1,jiner_biaoji+1).value = zekoujiner
    ws.cell(max_row + 1, 1).value = '合计'

    wb.save(fname)
    wb = openpyxl.load_workbook(fname)
    sheetname = '折扣分配'
    ws = wb[sheetname]
    max_row = ws.max_row
    max_col = ws.max_column
    heji = 0
    for index,row in enumerate(ws.values):
        if index==0 :
            continue
        elif index == max_row -1 :
            continue
        else :
            bili = row[dun_biaoji]/dun_total
            zekou = round(bili*zekoujiner,2)
            ws.cell(index+1,jiner_biaoji+1).value = zekou
            ws.cell(index + 1, jiner_biaoji+2).value = round(zekou/1.13,2)
            heji = zekou + heji

    #ws.cell(max_row-1,jiner_biaoji).value = ws.cell(max_row-1,jiner_biaoji).value-(heji-zekoujiner)

    wb.save(fname)
    return fname,sheetname,max_row,max_col


def  main():
    fname1 = r'F:\a00nutstore\fishc\纸暂估调整.xlsx'
    clearsheet(fname1)
    zanguRange()
    #fname1 =  r'D:\a00nutstore\fishc\纸暂估调整.xlsx'
    pinmings,ws_title,dic1,dic2 = makeTwodics(fname1)
    max_row, max_col = zangutiaozheng(fname1,pinmings, ws_title, dic1, dic2)
    sheetname = '暂估调整'
    title = sheetname
    fname = fname1
    excelcellfill.excelcellFill(fname,sheetname,min_row=1,min_col=1,max_row=max_row,max_col=max_col)

    tmp = input('请输入供应商名称\n')
    title = '{} {}'.format(tmp,title)
    # fname, ws_name = excelseting.setPrintArea(fname1, sheetname)
    # fname = excelseting.firstRowSeting(fname1, ws_name)
    # excelseting.yemei(fname1, ws_name,title)
    excelseting.fastseting(fname, sheetname, title)


    fname = fname1
    fname, sheetname, max_row, max_col = zekoufenpei(fname, ws_title, dic2)
    excelcellfill.excelcellFill(fname, sheetname, min_row=1, min_col=1, max_row=max_row, max_col=max_col)
    print('请点选折扣分配')
    fname, ws_name = excelseting.setPrintArea(fname)
    fname = excelseting.firstRowSeting(fname, ws_name)
    excelseting.yemei(fname, ws_name,title)

    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open(fname)
    worksheets = workbook.sheets
    ws1 = worksheets[sheetname]
    ws2 = worksheets['暂估调整']
    for col in range(2,max_col+1):
        string = get_column_letter(col)
        ws1.range('{}1'.format(string)).expand('down').number_format = '#,##0.00'
        ws2.range('{}1'.format(string)).expand('down').number_format = '#,##0.00'

    
    workbook.save()
    workbook.close()
    app.quit()
    os.startfile(fname1)


if  __name__ == '__main__':
    main()



