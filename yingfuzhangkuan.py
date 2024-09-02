'''
对应付账款计算
本版本改进参数传递，先运行zanguchuli.py，以对当月暂估进行处理！并核对无误后再运行此程序！
'''
# _*_ conding:utf-8 _*_
import openpyxl
import os
import excelmessage
import easygui
import yingfuprintseting
from openpyxl.utils import get_column_letter,column_index_from_string
import gongyingshangshouxin

#暂估汇总，暂估字典，有票金额汇总，无票金额汇总
def zangutotal(fname):
    dirname,filename = os.path.split(fname)
    os.chdir(dirname)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请选择暂估工作表'
    title = '当月暂估'
    print(msg)
    sheetname0 = easygui.buttonbox(msg=msg,title = title,choices=sheetnames)
    ws = wb['%s'%sheetname0]
    max_row = ws.max_row
    zangus = {}
    jiner_you_total = 0
    jiner_wu_total = 0
    for row in range(2, max_row + 1):
        gongyingshang = ws.cell(row, 5).value
        zangus.setdefault(gongyingshang, {'jiner_you': 0, 'jiner_wu': 0})
        if ws.cell(row, 8).value == '有':
            jiner_you = ws.cell(row, 6).value
            jiner_you_total += jiner_you
            zangus[gongyingshang]['jiner_you'] += jiner_you
        else:
            jiner_wu = ws.cell(row, 6).value
            #jiner_wu = int(jiner_wu)
            jiner_wu_total += jiner_wu
            zangus[gongyingshang]['jiner_wu'] += jiner_wu
    print(zangus)
    print('jiner_you', jiner_you_total)
    print('jiner_wu', jiner_wu_total)
    wb.close()
    return zangus,jiner_you_total,jiner_wu_total,fname

#打开当月应付账款，当月应付账款字典{gongyingshang:(kemudaima,qichu,fukuan,gouru)},供应商集合gongyingshangs{}
def gongyingShangs(fname2):
    yingfuzhangkuandic={}
    dirname,filename = os.path.split(fname2)
    os.chdir(dirname)
    wb = openpyxl.load_workbook(fname2)
    ws = wb.active
    max_row = ws.max_row
    gongyingshangs=set()      #供应商集合初始化
    for row in range(2,max_row):
        if ws.cell(row,4).value not in ['',None]:
            gongyingshangs.add(ws.cell(row,4).value)   #供应商集合添加当月供应商
            gongyingshang = ws.cell(row,4).value
            kemudaima = ws.cell(row,3).value
            if ws.cell(row,5).value == '贷':
                qichu = float(ws.cell(row,6).value)      #期初数，贷方取正值
            else :
                qichu = float(ws.cell(row,6).value)*(-1)   #期初数，借方方取负值
            fukuan = float(ws.cell(row,7).value)         #当月付款
            gouru = float(ws.cell(row,8).value)          #当月购货
            yingfuzhangkuandic[gongyingshang]=(kemudaima,qichu,fukuan,gouru)    #当月应付账款字典
        else :
            continue
    wb.close()
    return gongyingshangs,yingfuzhangkuandic,fname2

def zanguDic(zangus,gongyingshangs):
    zangudic = zangus
    for jianchen in list(zangudic.keys()):
        for gongyingshang in gongyingshangs:
            if jianchen in gongyingshang :
               zangudic[gongyingshang]=zangudic.pop('%s'%jianchen)    #替换字典中的键值的标准写法
            else :
                continue
    for jianchen in zangudic.keys():
        gongyingshangs.add(jianchen)
    for gongyingshang in gongyingshangs:
        print(gongyingshang)
    return zangudic,gongyingshangs

def zangudicPrint(zangudic,gongyingshangs):
    jiner_you_total = 0
    jiner_wu_total = 0
    for gongyingshang,jiner in zangudic.items():
        jiner_you_total+=jiner['jiner_you']
        jiner_wu_total+=jiner['jiner_wu']

    # 将扣款明细的供应商也加入
    koukuan = {}
    #打开'供应商扣款明细.xlsx'
    msg = '请点选供应商扣款明细账文件'
    print(msg)
    fname4 = excelmessage.wenjian()
    fname4 = excelmessage.excelMessage(fname4)
    dirname,filename = os.path.split(fname4)
    wb2 = openpyxl.load_workbook(fname4)
    ws2 = wb2['扣款']
    max_row2 = ws2.max_row
    print(max_row2)
    for row in range(2, max_row2):
        if ws2.cell(row, 2).value not in ['', None]:
            gongyingshangs.add(ws2.cell(row, 2).value)
            key = ws2.cell(row,2).value
            value =ws2.cell(row,3).value
            koukuan[key] =round(value,2)
        else:
            continue
    wb2.close()

    print('jiner_wu', jiner_you_total)
    print('jiner_wu', jiner_wu_total)

    return koukuan,gongyingshangs

#写入应付账款数据
def yingfuZhangkuan(fname2,gongyingshangs,yingfuzhangkuandic,koukuan,zangudic,shouxine_dic):
    dirname,filename = os.path.split(fname2)
    os.chdir(dirname)
    wb = openpyxl.load_workbook(fname2)
    ws = wb.create_sheet(title = '应付账款参考')

#写入应付账款抬头
    taitou = ('科目代码', '供应商', '期初余额', '付款', '购入', '期末余额', '期初0', '购入0', '累计未认证发票', '本月新增未认证', '扣款',  '暂估有票','暂估无票','授信额')
    ws.append(taitou)
#写入应付账款科目名称，供应商，期初余额0，付款，购货0
    for gongyingshang in gongyingshangs:
        yingfuzhangkuandic.setdefault(gongyingshang,(0, 0, 0, 0))
        #写入供应商
        gongyingshang = gongyingshang
        print(gongyingshang)
        #写入科目代码
        kemudaima = yingfuzhangkuandic[gongyingshang][0]
        #写入期初余额
        qichu = yingfuzhangkuandic[gongyingshang][1]
        #写入付款金额
        fukuan = yingfuzhangkuandic[gongyingshang][2]
        #写入购货金额
        gouhuo =yingfuzhangkuandic[gongyingshang][3]
        meihang = (kemudaima,gongyingshang,0,fukuan,0,0,qichu,gouhuo)
        ws.append(meihang)
    filename = '应付账款参考.xlsx'
    fname3 = os.path.join(dirname,filename)
    wb.save(fname3)
#写入扣款
    wb = openpyxl.load_workbook(fname3)
    ws = wb['应付账款参考']
    #写入扣款 ,写入暂估有票和暂估无票
    maxrow = ws.max_row
    for row in range(2,maxrow+1):
        gongyingshang = ws.cell(row,2).value
        koukuan.setdefault(gongyingshang,0)
        koukuanjiner=koukuan[gongyingshang]*(-1)
        ws.cell(row,11,value=koukuanjiner)
        #写入暂估有票和暂估无票
        zangudic.setdefault(gongyingshang,{'jiner_you':0,'jiner_wu':0})
        ws.cell(row,12, value=zangudic[gongyingshang]['jiner_you'])
        ws.cell(row,13, value=zangudic[gongyingshang]['jiner_wu'])
        if round(ws.cell(row, 12).value * 1.13, 2) - ws.cell(row, 11).value < 0:
            kouqian = round(ws.cell(row, 12).value * 1.13, 2)
        else:
            kouqian = ws.cell(row, 11).value

        # 计算调整后期初，购入，及期末余额
        ws.cell(row, 3, value=ws.cell(row, 7).value + kouqian)
        ws.cell(row, 5, value=ws.cell(row, 8).value)
        ws.cell(row, 6, value=ws.cell(row, 3).value + ws.cell(row, 5).value - ws.cell(row, 4).value)

        # 写入供应商授信额
        # msg = '请点选供应商授信额文件'
        # print(msg)
        # fname_shouxin = easygui.fileopenbox(msg)
        # shouxine_dic = gongyingshangshouxin.shouxine(fname_shouxin)
        shouxine_dic.setdefault(gongyingshang, 0)
        ws.cell(row, 14, value=shouxine_dic[gongyingshang])

    wb.save(fname3)

def main():

    print('请输入暂估路径和文件名')
    fname = excelmessage.wenjian()
    fname  = excelmessage.excelMessage(fname)
    zangus, jiner_you_total, jiner_wu_total, fname = zangutotal(fname)

    print('请输入临时应付账款路径和文件名')
    fname2 = excelmessage.wenjian()
    fname2 = excelmessage.excelMessage(fname2)
    gongyingshangs,yingfuzhangkuandic,fname2 = gongyingShangs(fname2)

    zangudic,gongyingshangs = zanguDic(zangus,gongyingshangs)
    koukuan, gongyingshangs = zangudicPrint(zangudic,gongyingshangs)
    msg = '请点选供应商授信额文件'
    print(msg)
    fname_shouxin = easygui.fileopenbox(msg)
    shouxine_dic = gongyingshangshouxin.shouxine(fname_shouxin)
    yingfuZhangkuan(fname,gongyingshangs,yingfuzhangkuandic,koukuan,zangudic,shouxine_dic)
    print(gongyingshangs)
    print(yingfuzhangkuandic)
    print('形成"应付账款参考.xlsx"')
    yingfuprintseting.main(jiner_wu_total)


if __name__ == '__main__':
    main()










