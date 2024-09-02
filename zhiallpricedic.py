'''
根据纸入库制作纸价格字典，制作纸所有期间价格字典
注意金额不能为空！
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
from openpyxl.utils import column_index_from_string

def zhipriceDic():
    #print('请输入纸入库文件所在路径：')
    #path = input('')

    path = r'F:\a00nutstore\006\zw\else'
    os.chdir(path)
    filename = r'2020入库.xlsx'
    fname = os.path.join(path,filename)
    wb = openpyxl.load_workbook(fname)
    nws = wb.copy_worksheet(wb['入库'])
    nws.title = '入库copy'
    wb.save(fname)

    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb['入库copy']

    maxrow = ws.max_row
    for i in 'GKLMNOPQ':
        for j in range(2,maxrow+1):
            if ws['{}{}'.format(i,j)].value in ['',None]:
                ws['{}{}'.format(i,j)].value = 0
            else :
                continue
    wb.save(fname)

def main():
    #qijian = input('请输入期间，如2020-04\n')
    zhipriceDic()


if __name__ == '__main__':
    main()







