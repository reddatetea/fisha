import openpyxl
from openpyxl.styles import Font,PatternFill,GradientFill,Side,Border,Alignment
import easygui
import os

def excelcellFill(fname,ws_name,min_row,min_col,max_row,max_col):


    font = Font(name='微软雅黑', size=12, bold=True, italic=True, color='000000')
    font1 = Font(name='微软雅黑', size=10, bold=False, italic=False, color='000000')
    pattern_fill = PatternFill(fill_type='solid', fgColor='40E0D0')  # 绿宝石
    pattern_fill1 = PatternFill(fill_type='solid', fgColor='AFEEEE')  # 苍白的绿宝石
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False)

    wb = openpyxl.load_workbook(fname)

    ws = wb[ws_name]
    ws.print_title_cols = 'A:B'  # the first two cols
    ws.print_title_rows = '1:1'  # the first row
    ws.freeze_panes = 'B2'

    if min_row%2 == 0 :      #min_row为偶数

        for i in ws.iter_cols(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in i:
                cell.font = font1

                if cell.row == min_row:
                    cell.font = font
                    cell.fill = pattern_fill
                    cell.alignment = alignment
                elif cell.row == max_row:
                    cell.font = font
                    cell.alignment = alignment
                    cell.fill = pattern_fill
                elif cell.row % 2 == 0 :
                    cell.fill = pattern_fill1
                else :
                    continue
    else :
        for i in ws.iter_cols(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in i:
                cell.font = font1
                if cell.row == min_row:
                    cell.fill = pattern_fill
                    cell.font = font
                    cell.alignment = alignment

                elif cell.row == max_row:
                    cell.font = font
                    cell.fill = pattern_fill
                    cell.alignment = alignment
                elif (cell.row+1) % 2 == 0 :
                    cell.fill = pattern_fill1
                else:
                    continue



    wb.save(fname)
    return fname

def main():
    msg = '请点选要设置单元格填充的excel文件'
    fname = easygui.fileopenbox(msg,title=msg)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选要设置单元格填充的excel工作表'
    ws_name = easygui.buttonbox(msg, title='工作表', choices=sheetnames)
    ws = wb[ws_name]

    max_col = ws.max_column

    for cell in ws['B']:
        if cell.value in ['',None]:
            max_row = cell.row-1              #合计行的行数
            print(max_row)
            break

    wb.close()
    min_col = 1
    min_row = 1
    excelcellFill(fname,ws_name,min_row, min_col, max_row=max_row, max_col=max_col)

if __name__== '__main__':
    main()








