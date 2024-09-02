import openpyxl
import pprint

def yuancailiaoprice():
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料价格表.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb['价格表']
    mrows = ws.max_row
    rcols = ws.max_column
    print(mrows)

    pinming_dic={}
    qijian_dic={}
    gongying_dic = {}
    for i in range(mrows-1):
        pinming_dic[ws.cell(i+2,3).value] = ws.cell(i+2,5).value
        qijian_dic[ws.cell(i + 2, 2).value] = pinming_dic
        gongying_dic[ws.cell(i + 2, 1).value] = qijian_dic
    print(pinming_dic)
    print(qijian_dic)
    print(gongying_dic)
    resultFile = open('yuancailiaoPriceDic.py', 'w',encoding='utf-8')
    resultFile.write('allDate = ' + pprint.pformat(gongying_dic))
    resultFile.close()

    wb.close()
    return gongying_dic,qijian_dic,pinming_dic

def main():
    yuancailiaoprice()


if __name__=="__main__":
    main()




