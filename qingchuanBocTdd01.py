'''
银行对账表BOC与Tdd对账差异明细 '
'''

import pdfplumber
import re
import openpyxl
from openpyxl.styles import Alignment
import easygui
import os
from  xlsxlsx import xlsXlsx
import numpy as np
import pandas as pd

def getBoc():
    fname = easygui.fileopenbox('请点选Bocvc*.pdf文件')
    path, file = os.path.split(fname)
    os.chdir(path)
    pdf = pdfplumber.open(fname)
    filename = 'boc.txt'
    with open(filename, 'w') as f:
        f.write('')
    with open(filename, 'a') as f:
        for page in pdf.pages:
            content = page.extract_text()
            f.write(content)
    pdf.close()
    data0 = []
    with open(filename) as f:
        for readline in f.readlines():
            data0.append(readline.strip())
    data1 = []
    pattern = r'^\|\s{1,2}(\d{1,2})\s{1,2}\|\.*'
    regex = re.compile(pattern)
    for j in data0:
        mat = regex.search(j)
        if mat and int(mat.group(1)) >= 1 and int(mat.group(1)) <= 99:
            i = j.split('|')
            boc_date = i[2].strip()
            boc_details = i[6].strip()
            boc_notes = i[-2].strip()
            boc_debit = i[-6].strip()
            boc_credit = i[-5].strip()
            if boc_debit == None or boc_debit == '':
                boc_debit = 0
            else:
                boc_debit = float(boc_debit.replace(',', ''))
                boc_credit = i[-5].strip()
            if boc_credit == None or boc_credit == '':
                boc_credit = 0
            else:
                boc_credit = float(boc_credit.strip().replace(',', ''))
            data1.append([boc_date, boc_details, boc_notes, boc_debit, boc_credit])
        else:
            continue
    # data1.sort(key = lambda x:x[1])
    lst1 = [i[3] for i in data1]
    lst2 = [i[4] for i in data1]
    dates0 = [i[0] for i in data1]
    dates = ['20' + str(x)[:2] + '/' + str(x)[2:4] + '/' + str(x)[-2:] for x in dates0]
    # dates = [pd.Timestamp(i) for i in dates1]
    detals = [i[1] for i in data1]
    notes = [i[2] for i in data1]
    data1.sort(key=lambda x: x[0])
    boc_debit_num, boc_debit_detal_note = getDic(lst1, dates, detals, notes)
    boc_credit_num, boc_credit_detal_note = getDic(lst2, dates, detals, notes)
    return data1,boc_debit_num,boc_debit_detal_note,boc_credit_num, boc_credit_detal_note

def getTdd():
    easygui.msgbox(msg='请复制"tt+-20**.xls"中的相关数据区域后,按"OK"按钮', title='请从11500那一行开始 ', ok_button='OK')
    df = pd.read_clipboard()
    df.columns = ['Transaction Reference', 'Transaction Date', 'Journal No.', 'Debits/Credits', 'Base Amount',
                  'Description']
    df['Transaction Date'] = df['Transaction Date'].astype('datetime64[ns]')
    df['Transaction Date'] = df['Transaction Date'].map(lambda x: x.strftime('%Y/%m/%d'))
    df['Base Amount'] = df['Base Amount'].str.replace('(', '').str.replace(')', '')
    df['debit'] = np.where(df['Debits/Credits'] == 'D', df['Base Amount'], 0)
    df['credit'] = np.where(df['Debits/Credits'] == 'C', df['Base Amount'], 0)
    df['debit'] = df['debit'].str.replace(',', '')
    df['debit'] = df['debit'].astype('float')
    df['credit'] = df['credit'].str.replace(',', '')
    df['credit'] = df['credit'].astype('float')
    df['credit'] = df['credit'].fillna(0)
    df['debit'] = df['debit'].fillna(0)
    df.sort_values('Transaction Date')
    lst1 = df['debit'].to_list()
    lst2 = df['credit'].to_list()
    dates = df['Transaction Date'].to_list()
    detals = df['Description'].to_list()
    data2 = list(zip(dates,detals,lst1,lst2))
    data2.sort(key = lambda x:x[0])
    tdd_debit_num,tdd_debit_detal = getDic2(lst1, dates, detals)
    tdd_credit_num,tdd_credit_detal  = getDic2(lst2, dates, detals)
    return data2,tdd_debit_num,tdd_debit_detal,tdd_credit_num,tdd_credit_detal

def getDic(lst, dates, detals, notes):
    dic_num = {}
    num_detal_note = {}
    for i in range(len(lst)):
        num = lst[i]
        date = dates[i]
        detal = detals[i]
        note = notes[i]
        num_total = lst.count(num)
        dic_num[num] = num_total
        if num_total == 1:
            num_detal_note[num] = [date, detal, note]
    return dic_num, num_detal_note

def getDic2(lst,dates,detals):
    dic_num = {}
    num_detal = {}
    for i in range(len(lst)):
        num = lst[i]
        date = dates[i]
        detal = detals[i]
        num_total = lst.count(num)
        dic_num[num] = num_total
        if num_total == 1:
            num_detal[num] = [date,detal]
    return dic_num,num_detal

def getDicDif(dic1, dic2):
    dif = []
    dic_dif = {}
    for k1, v1 in dic1.items():
        if k1 == 0:
            continue
        else:
            for k2, v2 in dic2.items():
                if k1 == k2:
                    if v1 > v2:
                        dic_dif[k1] = [v1, v2]
                    else:
                        continue
                else:
                    continue
    for k1 in dic1.keys():
        if k1 not in dic2.keys():
            if dic1.get(k1, 0) == 1:
                dif.append(k1)
            elif dic1.get(k1, 0) > 1:
                dic_dif[k1] = [v1, 0]
            else:
                continue
        else:
            continue
    dif.sort()
    return dic_dif, dif

#表12
def getBiao12(data1,data2,boc_dc_detal_note,boc_dc_dif_dic,boc_dc_dif,wb,ws,a,b,c,d):

    for i in range(3, len(boc_dc_dif) + 3):
        amount = boc_dc_dif[i - 3]
        date = boc_dc_detal_note.get(amount, (0, 0, 0))[0]
        detal = boc_dc_detal_note.get(amount, (0, 0, 0))[1]
        notes = boc_dc_detal_note.get(amount, (0, 0, 0))[2]

        ws[f'{a}{i}'].value = date
        ws[f'{b}{i}'].value = detal
        ws[f'{c}{i}'].value = notes
        ws[f'{d}{i}'].value = amount

        dif_rows = len(boc_dc_dif_dic)
        if dif_rows == 0:
            continue
        else:
            row_start = 3 + len(boc_dc_dif)

            data1.sort(key = lambda x:x[0])
            for j in data1:
                amount1 = j[3]
                date1 = j[0]
                detal1 = j[1]
                notes1 = j[2]
                for k, v in boc_dc_dif_dic.items():
                    if amount1 == k:
                        row_start = row_start + 1
                        ws[f'{a}{row_start}'].value = date1
                        ws[f'{b}{row_start}'].value = detal1
                        ws[f'{c}{row_start}'].value = notes1
                        ws[f'{d}{row_start}'].value = k
                    else:
                        continue
            data2.sort(key = lambda x:x[0])
            for j in data2:
                amount2 = j[3]
                date2 = j[0]
                detal2 = j[1]
                for k, v in boc_dc_dif_dic.items():
                    if amount2 == k:
                        row_start = row_start + 1
                        ws[f'{a}{row_start}'].value = date2
                        ws[f'{b}{row_start}'].value = detal2
                        ws[f'{d}{row_start}'].value = k * (-1)
                    else:
                        continue

    return wb,ws

def getBiao34(data1,data2,tdd_dc_detal,tdd_dc_dif_dic,tdd_dc_dif,wb,ws,a,b,c,d):
    for i in range(3, len(tdd_dc_dif) + 3):
        amount = tdd_dc_dif[i - 3]
        date = tdd_dc_detal.get(amount, (0, 0, 0))[0]
        notes = tdd_dc_detal.get(amount, (0, 0, 0))[1]
        ws[f'{a}{i}'].value = date
        ws[f'{b}{i}'].value = notes
        ws[f'{d}{i}'].value = amount
        dif_rows = len(tdd_dc_dif_dic)
        if dif_rows == 0:
            continue
        else:
            row_start = 3 + len(tdd_dc_dif)
            data2.sort(key = lambda x:x[0])
            for j in data2:
                amount1 = j[2]
                date1 = j[0]
                detal1 = j[1]
                for k, v in tdd_dc_dif_dic.items():
                    if amount1 == k:
                        row_start = row_start + 1
                        ws[f'{a}{row_start}'].value = date1
                        ws[f'{b}{row_start}'].value = detal1
                        ws[f'{d}{row_start}'].value = k
                    else:
                        continue
            data1.sort(key = lambda x:x[0])
            for j in data1:
                amount2 = j[4]
                date2 = j[0]
                detal2 = j[1]
                notes2 = j[2]
                for k, v in tdd_dc_dif_dic.items():
                    if amount2 == k:
                        row_start = row_start + 1
                        ws[f'{a}{row_start}'].value = date2
                        ws[f'{b}{row_start}'].value = detal2
                        ws[f'{c}{row_start}'].value = notes2
                        ws[f'{d}{row_start}'].value = k * (-1)
                    else:
                        continue
    return wb,ws


def getExcel():
    wb = openpyxl.Workbook()
    ws = wb.active
    align = Alignment(horizontal='center', vertical='center')
    ws['A1'].value = 'BOC借方有，TTC方没有'
    ws['E1'].value = 'BOC贷方有，TTD方没有'
    ws['I1'].value = 'TTD方有，BOC借方没有'
    ws['M1'].value = 'TTC方有，BOC贷方没有'
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = align
    ws.merge_cells('E1:H1')
    ws['E1'].alignment = align
    ws.merge_cells('I1:L1')
    ws['I1'].alignment = align
    ws.merge_cells('M1:P1')
    ws['M1'].alignment = align
    a = [*'ABCDEFGHIJKLMNOP']
    b = ['日期', '摘要', '备注', '金额'] * 4
    c = list(zip(a, b))
    for index, i in enumerate(c):
        ws[f'{i[0]}2'].value = i[1]
        ws[f'{i[0]}2'].alignment = align
    return wb,ws


def main():
    data1,boc_debit_num,boc_debit_detal_note,boc_credit_num, boc_credit_detal_note = getBoc()
    data2,tdd_debit_num,tdd_debit_detal,tdd_credit_num,tdd_credit_detal = getTdd()
    boc_debit_dif_dic, boc_debit_dif = getDicDif(boc_debit_num, tdd_credit_num)  #boc
    boc_credit_dif_dic, boc_credit_dif = getDicDif(boc_credit_num, tdd_debit_num)  #boc
    tdd_debit_dif_dic, tdd_debit_dif = getDicDif(tdd_debit_num, boc_credit_num)   #tdd
    tdd_credit_dif_dic, tdd_credit_dif = getDicDif(tdd_credit_num, boc_debit_num)  #tdd
    wb,ws = getExcel()
    wb,ws = getBiao12(data1,data2,boc_debit_detal_note,boc_debit_dif_dic,boc_debit_dif,wb,ws,'A','B','C','D')
    wb, ws = getBiao12(data1, data2, boc_credit_detal_note, boc_credit_dif_dic, boc_credit_dif, wb, ws, 'E', 'F', 'G', 'H')

    wb,ws = getBiao34(data1, data2,tdd_debit_detal, tdd_debit_dif_dic, tdd_debit_dif, wb, ws, 'I', 'J', 'K', 'L')
    wb, ws = getBiao34(data1, data2, tdd_credit_detal, tdd_credit_dif_dic, tdd_credit_dif, wb, ws, 'M', 'N', 'O', 'P')

    fname1 = r'boc-tdd对比图.xlsx'
    wb.save(fname1)
    os.startfile(fname1)

if __name__ == '__main__':
    main()