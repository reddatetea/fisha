import openpyxl
import os

def danjia():
    fname = r'F:\a00nutstore\006\zw\jiagong\2019\鑫添美加工价格表.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    dic = {}

    jiagong_fangsis = [j.value for j in ws['A'][1:7]]
    guiges = [j.value for j in ws[1][2:9]]
    # print(jiagong_fangsis)
    # print(guiges)

    for row in ws.iter_rows(min_row=2, min_col=1, max_row=7, max_col=9):
        guige_jiage = {}
        fangsi = row[0].value
        jiages = [j.value for j in row[2:]]
        guige_jiage = dict(zip(guiges,jiages))
        dic[fangsi] = guige_jiage

    wb.close()
    print(dic)
    return dic

def main():
    danjia()

if __name__ == '__main__':
    main()