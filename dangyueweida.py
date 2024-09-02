'''
dangyueweida03.py
当月白云入库和2020原材料入库中，录入的本应属于下个期间的材料的数量，对这些进行汇总，以便于仓库原材料核对
形成的‘当月未达'存放于白云入库路径下，且自动覆盖原来的‘当月未达’文件
'''
# _*_ conding:utf-8 _*_
import openpyxl
from openpyxl.utils import column_index_from_string
import os
import datetime
import time


def baiyunweida(xia_qijian):
    # 打开白云入库
    path1 = r'F:\a00nutstore\006\zw\baiyun'
    #path1 = r'D:\a00nutstore\006\zw\baiyun'
    filename1 = r'2020白云入库.xlsx'
    fname1 = os.path.join(path1, filename1)
    print(fname1)

    wb1 = openpyxl.load_workbook(fname1,data_only = True)
    ws1 = wb1['2020']
    maxrow1 = ws1.max_row
    # print(maxrow1)
    # print(ws1.cell(maxrow1,4).value)
    # print(ws1.cell(maxrow1, 4).value=='2020-05')
    # print(ws1.cell(maxrow1,14 ).value)

    # 新建excel文件
    #path = r'D:\a00nutstore\006\zw\baiyun'
    path = r'F:\a00nutstore\006\zw\baiyun'
    os.chdir(path)
    filename = '当月未达.xlsx'
    fname = os.path.join(path,filename)

    wb = openpyxl.Workbook(fname)
    ws = wb.create_sheet('白云未达')

    # 写入白云入库列号
    kaipiaoriqi_num = column_index_from_string('B')
    danjuhao_num = column_index_from_string('F')
    cailiaomingchen_num = column_index_from_string('O')
    chaiwumingchen_num = column_index_from_string('G')
    lingshu_num = column_index_from_string('H')
    dunshu_num = column_index_from_string('J')
    jizhang_num = column_index_from_string('N')

    for index,row in enumerate(ws1.values):
        if index == 0 :
            taitou = ('供应商','开票日期','单据号','材料名称','财务名称','令数','吨数','记账日期')
            ws.append(taitou)
        else :
            continue
    wb.save(fname)

    time.sleep(3)

    wb = openpyxl.load_workbook(fname,data_only=True)
    ws = wb['白云未达']

    for row in range(1, maxrow1 + 1):
        if row == 1:
            continue
        else:
            #取期间为下一个期间且已记账的数据
            #if (ws1.cell(row,4).value.strip()==qijian) and (ws1.cell(row,jizhang_num).value not in ['',None]):
            if (ws1.cell(row, 4).value.strip() == xia_qijian) and (ws1.cell(row,jizhang_num).value not in ['',None]):
                ws.cell(row, 1, value='白云')
                ws.cell(row, 2, value=ws1.cell(row, kaipiaoriqi_num).value)
                ws.cell(row, 3, value=ws1.cell(row, danjuhao_num).value)
                ws.cell(row, 4, value=ws1.cell(row, cailiaomingchen_num).value)
                ws.cell(row, 5, value=ws1.cell(row, chaiwumingchen_num).value)
                ws.cell(row, 6, value=ws1.cell(row, lingshu_num).value)
                ws.cell(row, 7, value=ws1.cell(row, dunshu_num).value)
                ws.cell(row, 8, value=ws1.cell(row, jizhang_num).value)
            else:
                continue

    wb.save(fname)
    return fname

def zhiweida(xia_qijian,fname):
    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb.create_sheet('纸未达')

    # 打开原材料纸类入库
    path2 = r'F:\a00nutstore\006\zw\else'
    #path2 = r'D:\a00nutstore\006\zw\else'
    filename2 = r'2020入库.xlsx'
    fname2 = os.path.join(path2, filename2)
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['入库']
    maxrow2 = ws2.max_row

    kaipiaoriqi_num = column_index_from_string('D')
    danjuhao_num = column_index_from_string('E')
    cailiaomingchen_num = column_index_from_string('F')
    chaiwumingchen_num = column_index_from_string('H')
    lingshu_num = column_index_from_string('G')
    dunshu_num = column_index_from_string('M')
    jizhang_num = column_index_from_string('V')

    for row in range(1, maxrow2 + 1):
        if row == 1:
            taitou = ('供应商', '开票日期', '单据号', '材料名称', '财务名称', '令数', '吨数', '记账日期')
            ws.append(taitou)
        else:

            # 取期间为下一个期间且已记账的数据
            if (ws2.cell(row, 3).value.strip() == xia_qijian) and (ws2.cell(row, jizhang_num).value not in ['', None]):
                ws.cell(row, 1, value=ws2.cell(row, 2).value)
                ws.cell(row, 2, value=ws2.cell(row, kaipiaoriqi_num).value)
                ws.cell(row, 3, value=ws2.cell(row, danjuhao_num).value)
                ws.cell(row, 4, value=ws2.cell(row, cailiaomingchen_num).value)
                ws.cell(row, 5, value=ws2.cell(row, chaiwumingchen_num).value)
                ws.cell(row, 6, value=ws2.cell(row, lingshu_num).value)
                ws.cell(row, 7, value=ws2.cell(row, dunshu_num).value)
                ws.cell(row, 8, value=ws2.cell(row, jizhang_num).value)
            else:
                continue

    wb.save(fname)
    time.sleep(3)
    return fname

#从原材料实时流水账，形成输料未达
def fuliaoweida(xia_qijian,fname):
    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb.create_sheet('辅料未达')

    # 打开原材料实时流水账.xlsx
    path3 = r'F:\a00nutstore\006\zw\原材料实时流水账'
    #path3 = r'D:\a00nutstore\006\zw\原材料实时流水账'
    filename3 = r'原材料实时流水账.xlsx'
    fname3 = os.path.join(path3, filename3)
    #fname1 = dataonly02.dataOnly(fname1)
    print(fname3)

    wb3 = openpyxl.load_workbook(fname3,data_only = True)
    ws3 = wb3['流水账']
    maxrow3 = ws3.max_row
    print(maxrow3)

    # 写入辅料库列号
    kaipiaoriqi_num = column_index_from_string('A')
    danjuhao_num = column_index_from_string('B')
    cailiaomingchen_num = column_index_from_string('D')
    chaiwumingchen_num = column_index_from_string('I')
    lingshu_num = column_index_from_string('F')
    dunshu_num = column_index_from_string('O')
    jizhang_num = column_index_from_string('S')

    for index,row in enumerate(ws3.values):
        if index == 0 :
            taitou = ('供应商','开票日期','单据号','材料名称','财务名称','令数','吨数','记账日期')
            ws.append(taitou)
        else :
            continue
    wb.save(fname)

    wb = openpyxl.load_workbook(fname,data_only=True)
    ws = wb['辅料未达']

    for row in range(1, maxrow3 + 1):
        if row == 1:
            continue
        else:
            #取期间为下一个期间且已记账的数据
            #if (ws1.cell(row,4).value.strip()==qijian) and (ws1.cell(row,jizhang_num).value not in ['',None]):
            if (ws3.cell(row,11).value.strip() == xia_qijian) and (ws3.cell(row,jizhang_num).value not in ['',None]):
                ws.cell(row, 1, value=ws3.cell(row, 3).value)
                ws.cell(row, 2, value=ws3.cell(row, kaipiaoriqi_num).value)
                ws.cell(row, 3, value=ws3.cell(row, danjuhao_num).value)
                ws.cell(row, 4, value=ws3.cell(row, cailiaomingchen_num).value)
                ws.cell(row, 5, value=ws3.cell(row, chaiwumingchen_num).value)
                ws.cell(row, 6, value=ws3.cell(row, lingshu_num).value)
                ws.cell(row, 7, value=ws3.cell(row, dunshu_num).value)
                ws.cell(row, 8, value=ws3.cell(row, jizhang_num).value)
            else:
                continue

    wb.save(fname)


def weida(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.create_sheet('未达')
    ws1 = wb['白云未达']
    ws2 = wb['纸未达']
    ws3 = wb['辅料未达']
    for index,row in enumerate(ws1.values):
        if index == 0 :
            ws.append(row)
        else :
            if row[0] != None:
                ws.append(row)
            else:
                continue

    for index,row in enumerate(ws2.values):
        if index == 0 :
            continue
        else :
            if row[0] != None:
                ws.append(row)
            else:
                continue

    for index,row in enumerate(ws3.values):
        if index == 0 :
            continue
        else :
            if row[0] != None:
                ws.append(row)
            else:
                continue

    wb.save(fname)

def weidaDic(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb['未达']
    maxrow= ws.max_row
    weida_dic = {}

    for row in range(2,maxrow+1):
        #当吨数不为0或空时，有时白云入库有价差，那么吨数为空
        if ws.cell(row,7).value in ['',None]:
            continue
        else :
            chaiwumingchen = ws.cell(row, 5).value
            lingshu = float(ws.cell(row, 6).value)
            dunshu = float(ws.cell(row, 7).value) * 1000
            weida_dic.setdefault(chaiwumingchen, {'lingshu': 0, 'dunshu': 0})
            weida_dic[chaiwumingchen]['lingshu'] += lingshu
            weida_dic[chaiwumingchen]['dunshu'] += dunshu



    print(weida_dic)

    wb.close()
    return weida_dic



def main():
    print('计算当月未达，即财务做了入库记账，而仓库算作下个期间')
    qijian = input('请输入当前期间(格式如2020-04)：\n')
    if  int(qijian[-2:]) !=12 :
        xia_qijian = qijian[:-2] + '%02d' % (int(qijian[-2:]) + 1)
    else :
        xia_qijian = str(int(qijian[:-3]) +1)+ '-01'                  #年底处理期间
    fname = baiyunweida(xia_qijian)
    zhiweida(xia_qijian,fname)
    fuliaoweida(xia_qijian, fname)
    weida(fname)
    weidaDic(fname)

if __name__ == '__main__' :
    main()













