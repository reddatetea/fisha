#本版实现供应商多选
import os
import xlwings as xw
import pandas as pd
import openpyxl
import easygui
from openpyxl.utils import  get_column_letter,column_index_from_string
# import excelcellfillruku
import excelseting
import excelprint

def choiceSheet():
    fname =r'F:\a00nutstore\006\zw\else\2020入库.xlsx'
    path, filename = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname,data_only=True)
    ws_name = '入库'
    ws = wb[ws_name]
    max_row = ws.max_row
    zhiziduan =  ['入库', '入库(吨)', '送货单金额', '不含税价']
    hangziduan ='jd'
    jisuanfangsi = 'sum'
    jizhang_number = column_index_from_string('V')
    #saixuanliezhi = input('请输入日期（如9/30/2020)\n')
    saixuanliezhis = set()
    for jizhang_row in ws.iter_rows(min_row=2,min_col=jizhang_number, max_row = max_row,max_col=jizhang_number):
        for j in  jizhang_row:
            if j.value not in ['', None, 0]:
                saixuanliezhis.add(j.value)
            else:
                continue

    saixuanliezhis = list(saixuanliezhis)
    print(saixuanliezhis)
    saixuanliezhis.sort()
    msg = '请选择记账日期'
    print(msg)
    saixuanliezhis = easygui.multchoicebox(msg, title=msg, choices=saixuanliezhis)
    print(saixuanliezhis)
    wb.close()
    return fname, ws_name, zhiziduan,hangziduan, jisuanfangsi,saixuanliezhis

def toushiBiao(fname, ws_name, zhiziduan,hangziduan,jisuanfangsi,saixuanliezhis):
    app = xw.App(visible = False,add_book = False)
    workbook = app.books.open(fname)
    worksheets = workbook.sheets
    ws = worksheets[ws_name]

    gongyingshangs = set()
    for i in ws.range('B2').expand('down').value:
        if i not in ['', None,0]:
            gongyingshangs.add(i)
        else:
            continue
    gongyingshangs = list(gongyingshangs)
    gongyingshangs.sort()
    msg = '请选择供应商'
    gongyingshang = easygui.multchoicebox(msg, title=msg, choices=gongyingshangs)
    print(gongyingshang)
    values = ws.range('A1').expand('table').options(pd.DataFrame).value
    filter_js = []
    for j in range(len(saixuanliezhis)):
        for m in range(len(gongyingshang)):
            filter_jm =values[(values[u'记账']=='{}'.format(saixuanliezhis[j])) & (values[u'供应商']=='{}'.format(gongyingshang[m]))]
            filter_js.append(filter_jm)
    filtered = pd.concat(filter_js)

    print(fname, ws_name, zhiziduan,hangziduan, jisuanfangsi,saixuanliezhis)
    pivottable = pd.pivot_table(filtered, values=zhiziduan, index=hangziduan, columns=None, aggfunc=jisuanfangsi, fill_value =0,margins=True, margins_name='合计')
    order = zhiziduan
    pivottable = pivottable[order]
    nws = worksheets['toushibiao']  # 新表
    nws.clear()
    nws.range('A1').value = pivottable
    nws.range('A1').value = '品名'
    nws.range('B1').value = '入库(令)'
    nws.range('B1').expand('down').number_format = '#,##0.00'
    nws.range('C1').expand('down').number_format = '#,##0.000'
    nws.range('D1').expand('down').number_format = '#,##0.00'
    nws.range('E1').expand('down').number_format = '#,##0.00'

    if len(gongyingshang)  ==1:
        gongyingshang = gongyingshang[0]


    workbook.save()
    workbook.close()
    app.quit()
    return gongyingshang                                #选定的要查询的供应商

def main():
    fname, ws_name, zhiziduan, hangziduan, jisuanfangsi,saixuanliezhis= choiceSheet()
    gongyingshang =toushiBiao(fname, ws_name, zhiziduan,hangziduan,jisuanfangsi,saixuanliezhis)
    wb = openpyxl.load_workbook(fname)
    ws_name = 'toushibiao'
    ws = wb[ws_name]
    max_row = ws.max_row
    max_col = ws.max_column
    min_row = 1
    min_col = 1
    # excelcellfillruku.excelcellFill(fname,ws_name,min_row,min_col,max_row,max_col)
    fname, ws_name = excelseting.setPrintArea(fname,ws_name)
    excelseting.firstRowSeting(fname, ws_name)
    excelseting.yemei(fname, ws_name,gongyingshang)
    os.system(fname)
    #excelprint.wsPrint(fname,ws_name)

if __name__ == '__main__':
    main()