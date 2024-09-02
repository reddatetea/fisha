'''
根据纸入库制作纸价格字典
本版增加传入期间参数处理,取当期
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
from openpyxl.utils import column_index_from_string

def zhipriceDic(qijian):
    #print('请输入纸入库文件所在路径：')
    #path = input('')

    path = r'F:\a00nutstore\006\zw\else'
    os.chdir(path)
    filename = r'2020入库.xlsx'
    fname = os.path.join(path,filename)
    wb = openpyxl.load_workbook(fname,data_only=True,read_only=True)
    ws = wb['入库']
    maxrow = ws.max_row

    zhiprices = {}
    dunshu_total = 0

    #gongyingshang = '白云'

    #获取列号
    gongyingshang_num = column_index_from_string('B')
    qijian_num = column_index_from_string('C')
    pricename_num  = column_index_from_string('X')
    price_num = column_index_from_string('O')
    lingshu_num = column_index_from_string('G')
    dunshu_num = column_index_from_string('M')
    jiner_num = column_index_from_string('R')
    jizhang_num = column_index_from_string('V')



    for row in range(2,maxrow+1):
        dangqian_qijian = ws.cell(row,qijian_num).value
        gongyingshang = ws.cell(row,gongyingshang_num).value
        pricename = ws.cell(row,pricename_num).value
        price = ws.cell(row, price_num).value
        lingshu = ws.cell(row, lingshu_num).value
        dunshu = ws.cell(row, dunshu_num).value
        jiner = ws.cell(row, jiner_num).value
        jizhang = ws.cell(row, jizhang_num).value


        #如果要实现获取全部期间的价格字典则，则将下面注释的四句取消注释即可！

        #zhiprices.setdefault(dangqian_qijian,{})
        #zhiprices[dangqian_qijian].setdefault(gongyingshang,{})
        #zhiprices[dangqian_qijian][gongyingshang].setdefault(pricename,{})
        #zhiprices[dangqian_qijian][gongyingshang][pricename].setdefault(price, {'lingshu':0,'dunshu':0,'jiner':0})


        #期间、令数不为None、吨数不为None、pricename不为None
        if (dangqian_qijian == qijian) and (dunshu not in ['',None]) and (jizhang not in ['',None]):
            zhiprices.setdefault(gongyingshang, {})
            zhiprices[gongyingshang].setdefault(pricename, {})
            zhiprices[gongyingshang][pricename].setdefault(price, {'lingshu': 0, 'dunshu': 0, 'jiner': 0})


            if '卷筒纸' in pricename :
                lingshu = 0
                #zhiprices.setdefault(dangqian_qijian, {})


                zhiprices[gongyingshang][pricename][price]['lingshu'] += lingshu
                zhiprices[gongyingshang][pricename][price]['dunshu'] += dunshu
                zhiprices[gongyingshang][pricename][price]['jiner'] += jiner
                dunshu_total+=dunshu

            else :
                zhiprices[gongyingshang][pricename][price]['lingshu'] += lingshu

                zhiprices[gongyingshang][pricename][price]['dunshu'] += dunshu
                zhiprices[gongyingshang][pricename][price]['jiner'] += jiner
                dunshu_total += dunshu


        else :
            continue

    print(zhiprices)

    #zhiprices = zhiprices['%s' % qijian]
    print(zhiprices)
    print('总吨数',dunshu_total)


    wb.close()
    return zhiprices

def main():
    qijian = input('请输入期间，如2020-04\n')
    zhiprices = zhipriceDic(qijian)
    print(zhiprices)

if __name__ == '__main__':
    main()







