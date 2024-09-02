'''
南京分公司库存报表应收账款与代账公司资损表核对
'''

import os
import openpyxl
from openpyxl.utils import column_index_from_string
import excelmessage


# path = input('请输入月报的路径：\n')
# filename = input('请输入月报的文件名：\n')
path = r'F:\a00nutstore\006\zw\fengognsi\南京\南京2020\南京202004'
filename = r'202004月报南京分(新格式）.xlsx'
fname = os.path.join(path, filename)
os.chdir(path)

wb = openpyxl.load_workbook(fname, data_only=True, read_only=True)

sheetnames = wb.sheetnames
print(sheetnames)
temp = int(input('请输入库存报表的文件名，如第4个请输入3：\n'))
sheetname = sheetnames[temp]
ws = wb['%s' % sheetname]

kucun_maxrow = ws.max_row
chukujiner_num = column_index_from_string('M')
rukujiner_num = column_index_from_string('I')
excel_qimokucun_num = column_index_from_string('Q')

# 出库金额
chukujiner = ws.cell(kucun_maxrow, chukujiner_num).value
print('chukujiner', chukujiner)
# 本期购入金额
rukujiner = ws.cell(kucun_maxrow, rukujiner_num).value
print('rukujiner', rukujiner)
#excel库存报表上期末库存金额
excel_qimokucun = ws.cell(kucun_maxrow,excel_qimokucun_num).value

# 应收账款
temp = int(input('请输入应收账款的文件名，如第4个请输入3：\n'))
sheetname = sheetnames[temp]
ws = wb['%s' % sheetname]
yingshou_maxrow = ws.max_row
yingshou_num = column_index_from_string('I')
excel_yingshouyuer_num = column_index_from_string('K')
for row in range(1, yingshou_maxrow + 1):
    if ws.cell(row, 2).value in ['总  计']:
        yingshou_jiefang = ws.cell(row, yingshou_num).value
        excel_yingshouyuer = ws.cell(row, excel_yingshouyuer_num).value
        break
    else:
        continue
print('yingshou_jiefang', yingshou_jiefang)

# 银行对账单余额
print(sheetnames)
temp = int(input('请输入银行存款的文件名，如第4个请输入3：\n'))
sheetname = sheetnames[temp]
ws = wb['%s' % sheetname]
yinghang_duizhagndan_maxrow = ws.max_row
yinghang_duizhagndan_yuer = ws.cell(yinghang_duizhagndan_maxrow, 7).value

print('yinghang_duizhagndan_yuer', yinghang_duizhagndan_yuer)

# 入库清单金额
sheetnames = wb.sheetnames
print(sheetnames)
temp = int(input('请输入入库清单的文件名，如第4个请输入3：\n'))
sheetname = sheetnames[temp]
ws = wb['%s' % sheetname]
qingdan_maxrow = ws.max_row
print(qingdan_maxrow)
qingdan_num = column_index_from_string('G')
qingdan_jiner = ws.cell(qingdan_maxrow, qingdan_num).value
for row in range(1, qingdan_maxrow + 1):
    if ws.cell(row, 1).value in ['莱特合计']:
        qingdan_jiner = ws.cell(row, qingdan_num).value
        break
    else:
        continue
print('qingdan_jiner', qingdan_jiner)
wb.close()

# 打开代账公司资损表
print('请输入南京代账公司资损表路径和文件名')
fname = excelmessage.wenjian()
fname = excelmessage.excelMessage(fname)
#wb = openpyxl.load_workbook(fname)
#path = input('请输入南京代账公司资损表路径：\n')
#path = r'F:\a00nutstore\006\zw\fengognsi\南京\南京2020\南京202004\LAITE'
#filename = input('请输入南京代账公司资损表文件名：\n')
#filename = r'武汉莱特.xlsx'
#fname = os.path.join(path, filename)
wb = openpyxl.load_workbook(fname)

sheetnames = wb.sheetnames
print(sheetnames)
temp = int(input('请输入资损表的文件名，如第4个请输入3：\n'))
sheetname = sheetnames[temp]
ws = wb['%s' % sheetname]
baobiao_maxrow = ws.max_row

for row in range(1, baobiao_maxrow):

    if ws.cell(row, 1).value == '银行存款':
        qimoyinghang = ws.cell(row, 2).value
    if ws.cell(row, 1).value == '库存商品':
        qimokucun = ws.cell(row, 2).value
    if ws.cell(row, 1).value == '应收账款':
        yingshou_qimo_jiner = ws.cell(row, 2).value
    if ws.cell(row, 3).value == '应付账款（武汉莱特总部）':
        yingfu_kaipiao = ws.cell(row, 4).value
    if ws.cell(row, 3).value == '应付账款（估价武汉莱特）':
        yingfu_zangu = ws.cell(row, 4).value
    if ws.cell(row, 2).value == '商品销售收入':
        shouru = ws.cell(row, 3).value
    if ws.cell(row, 2).value == '商品销售成本':
        chengben = ws.cell(row, 3).value

        


print('银行存款', qimoyinghang)
print('库存商品', qimokucun)
print('应收账款', yingshou_qimo_jiner)
print('应付账款（武汉莱特总部）', yingfu_kaipiao)
print('应付账款（估价武汉莱特）', yingfu_zangu)

wb.close()


print('打开南京分公司对账文件')

path = r'F:\a00nutstore\006\zw\fengognsi\对账\关系\2020'
filename = r'南京勾稽.xlsx'
fname = os.path.join(path,filename)
wb = openpyxl.load_workbook(fname)
sheetnames = wb.sheetnames
print(sheetnames)
temp = int(input('请输入想要复制的工作表名，如第4个请输入3：\n'))
oldname = sheetnames[temp]
ws1 = wb[oldname]
print(oldname)
newname = input('请输入对账期间:\n')
print(newname)

ws2 = wb.copy_worksheet(ws1)
ws2.title = newname
wb.save(fname)

wb = openpyxl.load_workbook(fname)
ws = wb[newname]
maxrow = ws.max_row

#写入库存商品出库金额
ws.cell(3,2,value =chukujiner)
#写入销售成本
ws.cell(4,2,value =chengben)

#应收账款借方发生额
ws.cell(7,2,value = yingshou_jiefang)

#商品销售收入
ws.cell(8,2,value = shouru)

#商品销售收入(含税)
ws.cell(9,2,value = round(shouru*1.13,2))

#库存商品期初余额(不含税)
qichu = ws.cell(13,2).value
ws.cell(12,2,value = qichu )

#库存商品期末余额(不含税)
ws.cell(13,2,value = qimokucun )

#商品销售成本
ws.cell(14,2,value = chengben)

#库存商品本期入库(不含税)
ws.cell(17,2,value = rukujiner)

#库存商品本期入库(不含税)
ws.cell(21,2,value = qingdan_jiner)

#库存商品本期入库(不含税)
ws.cell(23,2,value = rukujiner)

#差额5 银行存款
ws.cell(26,2,value = qimoyinghang)
ws.cell(27,2,value =yinghang_duizhagndan_yuer)

#往来
#总部应收账款余额和分公司应付账款
yingshouzhangkuan = float(input('请输入总部应收账款分公司余额:\n'))
ws.cell(30,2,value = yingshouzhangkuan)
ws.cell(31,2,value = yingfu_kaipiao+yingfu_zangu)

#应收账款余额（资损表和应收账款报表）
ws.cell(39,2,value = yingshou_qimo_jiner)
ws.cell(40,2,value = excel_yingshouyuer)


#库存商品（资损表和库存商品报表）
ws.cell(35,2,value = excel_qimokucun)
ws.cell(36,2,value = qimokucun)


wb.save(fname)





















