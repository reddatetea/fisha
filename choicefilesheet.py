from xlsxlsx import xlsXlsx
import openpyxl
import os
import re
import easygui

def dianxuan(msg,title):
    print(msg)
    fname = easygui.fileopenbox(msg=msg,title=title)
    path, filename = os.path.split(fname)
    os.chdir(path)
    # 用正则表达式查看是低版本还是高版本
    pattern = '\w+(\.[X|x][l|L][Ss]$)'
    rep = re.compile(pattern)
    mat = rep.search(fname)
    if mat != None:
        xlsXlsx(fname)
        fname = fname + 'x'
    else:
        fname = fname

    choice = easygui.ccbox(msg='是否显示excel文件"%s"基本信息 ' % filename, title='excelmessage ', choices=('yes  ', ' no '),
                           image=None)

    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    if choice == True:

        print('下面将显示excel文件%s基本信息 ' % filename)
        print('文件名为:%s' % fname)


        # 第一个工作表的名字
        sheetname = wb.sheetnames[0]
        sheet = wb[sheetname]
        jishu = 0
        first_list = []

        for index, row in enumerate(sheet.values):
            first_list.append(row[1])
            if row[0] == None:
                jishu = jishu + 1

        mrows = sheet.max_row - jishu
        print('工作表名称：', sheetnames)
        print('实际操作的工作表名称：', sheetname)
        print('工作表的最大行数是：', sheet.max_row)
        print('工作表的最大列数是：', sheet.max_column)
        print('工作表计数是：', jishu)
        print('工作表可操作的最大行数：', mrows)


        wb.close()

    else:
        pass

    sheetnames.append('NOCHOICE')
    choice = easygui.buttonbox(msg='请选择工作表',title = '',choices = sheetnames)

    if choice == 'NOCHOICE':
        print('本次未选工作表')
        return fname

    else :
        sheetname = choice
        print('您选择的工作表是：',sheetname)
        return sheetname

def main():
    msg = '请点选excel文件'
    title = ''
    dianxuan(msg,title)

if __name__ == '__main__':
    main()

