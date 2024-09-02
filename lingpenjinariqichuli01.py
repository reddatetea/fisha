import openpyxl
import datetime
fname = r'F:\a00nutstore\006\zw\lingpeijian\零配件实时入库20201.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb['ssrk']
# riqi_string0 = ws['b2'].value
for row in ws['B2:B2523']:
    riqi_string0 = row[0].value
    if type(riqi_string0) == 'datetime.datetime':
        riqi_string= datetime.datetime.strftime(riqi_string0, '%Y-%m-%d')
        row[0].value = riqi_string
    else :
        continue

wb.save(fname)
