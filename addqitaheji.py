from openpyxl import load_workbook, Workbook
from xlsxlsx import xlsXlsx
import os


def addqitaHeji(kemuname_dic,fname):

    #打开'数量金额盘存表8.xlsx'

    #filename = r'数量金额盘存表8.xlsx'

    wb7 = load_workbook(fname)
    ws7 = wb7.active

    jishu = 0
    first_list=[]
    for index, row in enumerate(ws7.values):
        first_list.append(row[1])

        if row[0] == None:
            jishu = jishu + 1


    mrows1 = ws7.max_row - jishu


    mrows2 =ws7.max_row

    #删除一行，删除第morws行
    #ws7.delete_rows(mrows2)

    print(mrows2)

    qichutotal=0
    rukutotal=0
    caigoutotal=0
    banchengpintotal=0
    chukutotal=0
    lingliaototal=0
    qimototal=0

    #期初数量求和
    for i in range(2,mrows2-1):
        qichutotal = qichutotal + float(ws7.cell(row=i,column=24).value)
    ws7.cell(mrows2-1,24,qichutotal)

    #入库数量求和
    for i in range(2,mrows2-2+1):
        rukutotal = rukutotal + float(ws7.cell(row=i,column=25).value)
    ws7.cell(mrows2-1,25,rukutotal)

    #采购数量求和
    for i in range(2,mrows2-2+1):
        caigoutotal = caigoutotal + float(ws7.cell(row=i,column=26).value)
    ws7.cell(mrows2-1,26,caigoutotal)

    #半成品数量求和
    for i in range(2,mrows2-2+1):
        banchengpintotal = banchengpintotal + float(ws7.cell(row=i,column=27).value)
    ws7.cell(mrows2-1,27,banchengpintotal)

    #出库数量求和
    for i in range(2,mrows2-2+1):
        if ws7.cell(row=i,column=28).value ==None:
            ws7.cell(row=i, column=28).value =0
            chukutotal = chukutotal
        else :
            chukutotal = chukutotal + float(ws7.cell(row=i,column=28).value)
    ws7.cell(mrows2-1,28,chukutotal)

    #生产领料数量求和
    for i in range(2,mrows2-2+1):
        if ws7.cell(row=i,column=29).value ==None:
            ws7.cell(row=i, column=29).value =0
            lingliaototal = lingliaototal
        else :
            lingliaototal = lingliaototal + float(ws7.cell(row=i,column=29).value)
    ws7.cell(mrows2-1,29,lingliaototal)

    #期末数量求和
    for i in range(2,mrows2-2+1):
        qimototal = qimototal + float(ws7.cell(row=i,column=30).value)
    ws7.cell(mrows2-1,30,qimototal)

    for i in range(mrows1+1,mrows2-2+1):
        if ws7.cell(row=i,column=18).value in kemuname_dic:
            ws7.cell(i,3,kemuname_dic[ws7.cell(row=i,column=18).value])
        else:
            ws7.cell(i, 3,'')

    #删除一行，删除第morws行
    ws7.delete_rows(mrows2)

    newname = fname[:-6] + '18.xlsx'
    wb7.save(newname)
    return newname

def main():
    addqitaHeji(kemuname_dic,fname)

if __name__=='__main__':
    main()



