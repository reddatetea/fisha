import openpyxl
import  pingpangaoyunhui
import  pingpangshijinsai
import  pingpangshijiebei

fname = r'乒乓球男单历届奥运会战绩.xlsx'
ws_name = '男单'
ay_yundongyuans,ay_guanjun_dic,ay_yajun_dic,ay_jijun_dic = pingpangaoyunhui.aoyunhui(fname,ws_name)

fname = r'乒乓球男单历届世锦赛战绩.xlsx'
ws_name = '男单'
sjs_yundongyuans,sjs_guanjun_dic,sjs_yajun_dic,sjs_jijun_dic = pingpangshijinsai.shijinsai(fname,ws_name)

fname = r'乒乓球男单历届世界杯战绩.xlsx'
ws_name = '男单'
sjb_yundongyuans,sjb_guanjun_dic,sjb_yajun_dic,sjb_jijun_dic = pingpangshijiebei.shijiebei(fname,ws_name)

yundongyuans = set()
for j in list(ay_yundongyuans):
    yundongyuans.add(j)
for j in list(sjs_yundongyuans):
    yundongyuans.add(j)
for j in list(sjb_yundongyuans):
    yundongyuans.add(j)

yundongyuans = list(yundongyuans)
print(yundongyuans)

fname =r'乒乓球男单三大赛战绩.xlsx'
ws_name = r'三大赛'
wb = openpyxl.load_workbook(fname)
ws  = wb[ws_name]
title =[ '运动员','奥运会金牌','奥运会银牌','奥运会铜牌','世锦赛金牌','世锦赛银牌','世锦赛铜牌','世界杯金牌','世界杯银牌','世界杯铜牌','奖牌总数']
ws.append(title)
for  j in range(2,len(yundongyuans)+2):
    ws.cell(j,1).value = yundongyuans[j-2]
wb.save(fname)
wb = openpyxl.load_workbook(fname)
ws = wb[ws_name]
for index,row in enumerate(ws.values):
    if  index==0:
        continue
    else :
        yundongyuan = row[0]
        ay_guanjun_dic.setdefault(yundongyuan, 0)
        ay_yajun_dic.setdefault(yundongyuan, 0)
        ay_jijun_dic.setdefault(yundongyuan, 0)
        sjs_guanjun_dic.setdefault(yundongyuan, 0)
        sjs_yajun_dic.setdefault(yundongyuan, 0)
        sjs_jijun_dic.setdefault(yundongyuan, 0)
        sjb_guanjun_dic.setdefault(yundongyuan, 0)
        sjb_yajun_dic.setdefault(yundongyuan, 0)
        sjb_jijun_dic.setdefault(yundongyuan, 0)

        ayjp = ay_guanjun_dic[yundongyuan]
        ayyp = ay_yajun_dic[yundongyuan]
        aytp = ay_jijun_dic[yundongyuan]

        sjsjp = sjs_guanjun_dic[yundongyuan]
        sjsyp = sjs_yajun_dic[yundongyuan]
        sjstp = sjs_jijun_dic[yundongyuan]

        sjbjp = sjb_guanjun_dic[yundongyuan]
        sjbyp = sjb_yajun_dic[yundongyuan]
        sjbtp = sjb_jijun_dic[yundongyuan]

        heji = ayjp+ayyp+aytp+sjsjp+sjsyp+sjstp+sjbjp+sjbyp+sjbtp
        # hang = [yundongyuan,ayjp,ayyp,aytp,sjsjp,sjsyp,sjstp,sjbjp,sjbyp,sjbtp,heji]
        # ws.append(hang)
        #ws.cell(index, 1).value = yundongyuan
        ws.cell(index+1,2).value = ayjp
        ws.cell(index+1, 3).value = ayyp
        ws.cell(index+1, 4).value = aytp
        ws.cell(index+1, 5).value = sjsjp
        ws.cell(index+1, 6).value = sjsyp
        ws.cell(index+1, 7).value = sjstp
        ws.cell(index+1, 8).value = sjbjp
        ws.cell(index+1, 9).value = sjbyp
        ws.cell(index+1, 10).value = sjbtp
        ws.cell(index+1, 11).value = heji


wb.save(fname)










