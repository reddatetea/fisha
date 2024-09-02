'''
20211011只需分别点选本月应付账款临时文件和暂估文件，自动出应付账款报表！格式也全部设置好，不过要花一分钟微调下！
'''
# _*_ conding:utf-8 _*_
import pandas as pd
import openpyxl
import os
import pypinyin
import easygui
import yingfuprintseting
from openpyxl.utils import get_column_letter, column_index_from_string
import excelseting
import time
import xlwings as xw
import excelautofit
import excelcellfill
import excelseting
import re


# 打开当月应付账款，当月应付账款字典{gongyingshang:(kemudaima,qichu,fukuan,gouru)},供应商集合gongyingshangs{}
def gongyingShangs(fname1):
    # df1 = pd.read_excel(fname1, dtype={'供应商编号': 'string'})  #高版本pandas适用
    df1 = pd.read_excel(fname1, dtype={'供应商编号': str})  # 低版本pandas适用
    df1.rename(columns={'供应商名称': '供应商',
                        '供应商编号': '科目代码',
                        '方向': '方向4',
                        '期初余额金额': '期初余额',
                        '借方金额': '付款',
                        '贷方金额': '购入'},
               inplace=True)
    df8 = df1[['供应商', '科目代码']]
    df1.pop('科目代码')

    df1['期初余额'] = df1['期初余额'] * (df1['方向4'].agg(lambda x: -1 if x == '借' else 1))
    df1.pop('科目编码')
    df1.pop('科目名称')
    df1.pop('方向4')
    df1.pop('方向20')
    df1.pop('期末余额金额')

    df1 = df1.iloc[0:-1, :]
    gongyingshangs = df1.供应商.to_list()
    print(df1.head(20))
    return df1, df8, gongyingshangs


# 暂估汇总，暂估字典，有票金额汇总，无票金额汇总
def zangutotal(gongyingshangs, fname_zangu, ws_name):
    df_zangu = pd.read_excel(fname_zangu, sheet_name=ws_name, usecols=['供应商', '暂估金额', '应付金额'])
    df_zangu = df_zangu.groupby('供应商', as_index=False).sum()
    df_zangu = df_zangu[df_zangu.暂估金额 != 0]
    print('df_zangu', df_zangu)
    # df_zangu = df_zangu['供应商', '应付金额']
    dic = dict(zip(df_zangu.供应商, df_zangu.应付金额))  # 计算暂估多冲金额，形成字典
    for jian in list(dic.keys()):
        for gongyingshang in gongyingshangs:
            if jian in gongyingshang:
                dic[gongyingshang] = dic.pop(jian)  # 将字典中的供应商简称替换为全称
            else:
                continue
    data = []
    for key, value in dic.items():
        if value < 0:
            data.append([key, value])
        else:
            continue
    df_zangu = pd.DataFrame(data, columns=['供应商', 'koukuan2'])
    print(df_zangu)
    return df_zangu


def gongyingshangKoukuan(fname2):
    # df2 = pd.read_excel(fname2, dtype={'科目代码': 'string'})
    df2 = pd.read_excel(fname2, dtype={'科目代码': str})
    df2.rename(columns={'扣款单位名称': '供应商'}, inplace=True)
    df2['减扣款'] = df2['减扣款'] * (-1)
    df2.pop('科目代码')
    df2 = df2.iloc[0:-1, :]
    return df2


def gongyingshangShouxiner(fname3):
    # df5 = pd.read_excel(fname3, dtype={'供应商编号': 'string'})
    df5 = pd.read_excel(fname3, dtype={'供应商编号': str})
    df5.rename(columns={'供应商名称': '供应商'}, inplace=True)
    df5.pop('供应商编号')
    return df5


def shujuLianjieToushi(df1, df2, df_zangu):
    df3 = pd.concat([df1, df2, df_zangu])
    df3['付款'].fillna(0, inplace=True)
    df3['购入'].fillna(0, inplace=True)
    df3['期初余额'].fillna(0, inplace=True)
    df3['减扣款'].fillna(0, inplace=True)
    df3['koukuan2'].fillna(0, inplace=True)
    toushibiao = pd.pivot_table(df3, index='供应商',
                                values=['期初余额', '付款', '购入', '减扣款', 'koukuan2'],
                                aggfunc='sum',
                                fill_value=0,
                                margins=True,
                                margins_name='总计'

                                )
    order = ['期初余额', '付款', '购入', '减扣款', 'koukuan2']
    df4 = toushibiao[order]
    gys = list(df4.index)
    df4['期初余额'] = df4['期初余额'] + df4['减扣款'] + df4['koukuan2']
    order = ['期初余额', '付款', '购入']
    df4 = df4[order]
    df4 = df4.loc[~(df4 == 0).all(axis=1)]
    return gys, df4


def addShouxinBianma(df4, df5, df8):
    df6 = pd.merge(df4, df5, on='供应商', how='left')
    # df6.to_excel('应付账款授信额.xlsx', index=False)
    df7 = pd.merge(df6, df8, on='供应商', how='left')  # df8是df1 只取供应商和编号
    df7['期末余额'] = df7.期初余额 - df7.付款 + df7.购入
    order = ['科目代码', '供应商', '期初余额', '付款', '购入', '期末余额', '授信额']
    df7 = df7[order]
    return df7


def sortGongyingshang(df):
    def getpinyin(gongyingshang):
        gys = pypinyin.pinyin(gongyingshang, style=pypinyin.NORMAL)
        gys_pinyin = ''.join(i[0] for i in gys)
        return gys_pinyin

    # df.sort_values(by='供应商')
    df['gys'] = list(map(getpinyin, df.供应商))
    df.sort_values(by='gys', inplace=True)
    # # gongyingshangs.sort(key = lambda gongs : gys_dic[gongs])
    df.pop('gys')
    df.tail(20)
    return df


def addShuziStyle(fname, ws_name, shuju_row, shuju_col):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    for row in ws.iter_rows(min_row=2, min_col=3, max_row=shuju_row, max_col=shuju_col):
        for cell in row:
            cell.number_format = '#,##0.00'
    wb.save(fname)


def addBeizhu(fname, ws_name):
    wb = openpyxl.load_workbook(fname)
    ws0 = wb[ws_name]
    shuju_row = ws0.max_row
    msg = '选择应付账款备注文件'
    print(msg)
    fname2 = r'F:\a00nutstore\006\zw\yingfu\应付账款重要文档\应付账款备注.xlsx'
    df = pd.read_excel(fname2, sheet_name='备注')
    with pd.ExcelWriter(fname, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        df.to_excel(writer, sheet_name=ws_name, startrow=shuju_row + 2, index=False, header=None)
    # ws0.cell(shuju_row+2,1).value = '备注'
    # writer.save()
    wb = openpyxl.load_workbook(fname)
    ws0 = wb[ws_name]
    ws0.cell(shuju_row + 2, 1).value = '备注'
    wb.save(fname)
    return fname


def copyTwoBiao(fname, ws_name, shuju_row):
    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb[ws_name]
    caigou = ws_name + '采购'
    chaiwu = ws_name + '财务'
    wb.copy_worksheet(ws).title = caigou
    ws_caigou = wb[caigou]
    ws_caigou.move_range('a{}:a{}'.format(shuju_row + 2, ws_caigou.max_row), 0, 1)
    ws_caigou.delete_cols(1, 1)
    ws_caigou.column_dimensions['A'].width = 25  # 列宽
    ws_caigou.column_dimensions['B'].width = 19
    ws_caigou.column_dimensions['C'].width = 16
    ws_caigou.column_dimensions['D'].width = 16
    ws_caigou.column_dimensions['E'].width = 19
    # ws_caigou.column_dimensions['F'].hidden = True

    wb.copy_worksheet(ws_caigou).title = chaiwu
    ws_chaiwu = wb[chaiwu]
    ws_chaiwu.column_dimensions['A'].width = 28  # 列宽
    ws_chaiwu.column_dimensions['B'].width = 19
    ws_chaiwu.column_dimensions['C'].width = 16
    ws_chaiwu.column_dimensions['D'].width = 16
    ws_chaiwu.column_dimensions['E'].width = 19
    ws_chaiwu.column_dimensions['F'].width = 16
    ws_caigou.delete_cols(6, 1)  # 删除F列，‘授信额’
    ws_caigou.page_margins = openpyxl.worksheet.page.PageMargins(top=1, header=0.5, left=1, right=1, footer=0.5,
                                                                 bottom=1)
    ws_chaiwu.page_margins = openpyxl.worksheet.page.PageMargins(top=1, header=0.5, left=1, right=1, footer=0.5,
                                                                 bottom=1)

    wb.save(fname)
    return fname, ws_name, caigou, chaiwu


def main():
    path = r'F:\a00nutstore\006\zw\yingfu\2022yingfu'
    os.chdir(path)
    msg = '请点选本月临时应付账款路径和文件名'
    print(msg)
    fname1 = easygui.fileopenbox(msg)
    df1, df8, gongyingshangs = gongyingShangs(fname1)
    print(gongyingshangs)

    msg = '请点选本月点选暂估文件'
    print(msg)
    fname_zangu = easygui.fileopenbox(msg)
    wb = openpyxl.load_workbook(fname_zangu)
    sheetnames = wb.sheetnames
    wb.close()
    msg = '自动选择暂估excel工作表'
    print(msg)
    ws_zangu = '0112'  # 默认暂估工作表
    regax = r'\d{4}'
    pattern = re.compile(regax)
    for sheetname in sheetnames:
        mat = pattern.search(sheetname)
        if mat != None:
            ws_zangu = mat.group()
            break
        else:
            continue

    df_zangu = zangutotal(gongyingshangs, fname_zangu, ws_zangu)

    # msg = '请点选供应商扣款明细文件名'
    # print(msg)
    # fname2 = easygui.fileopenbox(msg)
    fname2 = r'F:\a00nutstore\006\zw\yingfu\应付账款重要文档\供应商扣款明细.xlsx'
    df2 = gongyingshangKoukuan(fname2)
    gys, df4 = shujuLianjieToushi(df1, df2, df_zangu)

    # msg = '请点选供应商授信额文件'
    # print(msg)
    # fname3 = easygui.fileopenbox(msg)
    fname3 = r'F:\a00nutstore\006\zw\yingfu\应付账款重要文档\供应商授信额.xlsx'
    df5 = gongyingshangShouxiner(fname3)
    df7 = addShouxinBianma(df4, df5, df8)

    df7 = sortGongyingshang(df7)  # 对供应商排序
    shuju_row = df7.shape[0] + 1
    shuju_col = df7.shape[1]

    # ws_name = input('请输入本月应付账款名字"0730"\n')
    print('选择本月暂估文件的日期作为作为应付账款日期')
    ws_name = ws_zangu
    newfilname = '应付账款{}.xlsx'.format(ws_name)
    fname = os.path.join(path, newfilname)
    wb = openpyxl.Workbook()
    wb.save(fname)
    time.sleep(3)
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    ws.title = ws_name

    df7.to_excel(fname, sheet_name=ws_name, index=False)
    # excelautofit.excelAutofit(fname, ws_name)

    addShuziStyle(fname, ws_name, shuju_row, shuju_col)  # 数字格式为千分符
    addBeizhu(fname, ws_name)
    excelseting.setPrintArea(fname, ws_name)
    excelcellfill.excelcellFill(fname, ws_name, shuju_row, shuju_col, min_row=1, min_col=1)
    fname, *ws_names = copyTwoBiao(fname, ws_name, shuju_row)
    yingfu_name = '应付账款' + ws_name
    for name in ws_names:
        excelseting.yemei(fname, name, yingfu_name)

    os.startfile(fname)


if __name__ == '__main__':
    main()










