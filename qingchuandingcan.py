import csv
import easygui
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side
import os

def duquCsv(csv_file):
    # data = []
    with open(csv_file,'r',encoding='UTF-8') as file :   #20210709加encoding='utf-8'
        data = [line.strip().split(',') for line in file.readlines()]
        print(data)
        return data

def  saveToexcel(fname,data):
    jiesuan_riqi = data[1][5]
    wb = openpyxl.load_workbook(fname)
    ws_name = jiesuan_riqi.replace("`",'').replace('-','.')
    wb.create_sheet(title = ws_name)
    ws = wb[ws_name]
    for index,row in enumerate(data):
        if index == 0:
            ws.append(row[:-1])
        else :
            for k in range(7,18):
                if k==12:
                    continue
                else :
                    row[k] = float(row[k])

            ws.append(row[:-1])
    wb.save(fname)
    return fname,ws_name

def setStyles(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    max_row = ws.max_row
    font = Font(name = '等线',size = '11')
    bd = Border(left=Side(border_style="thin"),

                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
                )
    alignment = Alignment(vertical='center', horizontal='center')
    for row in ws.rows:
        for cell in row :
            cell.font = font
            cell.border = bd
            cell.alignment = alignment
    for  n in 'ABCDEFGHIJKLMNOPQR':
        ws.column_dimensions['{}'.format(n)].width = '11.88'
    for  p in range(1,max_row+1):
        ws.row_dimensions[p].height = '38.25'

    wb.save(fname)

def main():
    msg = '请点选订餐对账单csv文件'
    print(msg)
    csv_file = easygui.fileopenbox(msg)
    #csv_file = r'F:\a00nutstore\fishc\qingchuan\订餐对账单_2020_12_30_7c7438f1eae7434f8e5ea9b97db5671a(1).csv'
    data = duquCsv(csv_file)
    msg = '请点选直客通送餐记录excel文件'
    print(msg)
    fname = easygui.fileopenbox(msg)
    #fname = r'F:\a00nutstore\fishc\qingchuan\2020.1.1起直客通送餐记录.xlsx'
    fname,ws_name = saveToexcel(fname,data)
    setStyles(fname, ws_name)
    os.system(fname)


if __name__ == '__main__':
    main()