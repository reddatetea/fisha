# _*_ conding:utf-8 _*_

import os
from datetime import datetime
from easygui import fileopenbox
from openpyxl import  load_workbook
from  xlsxlsx import xlsXlsx

dic01={'Rack Rate(R)  /Corporate Regular': 'Rack Rate     (R)', 'Corp Rate(P) /Corporate Preferred': 'Corporate Preferred (P)', 'Comm Rate(L) /Corporate Negociated': 'Corporate Negociated (L)', 'Transient Dic.(D)': 'Transient Dic.(D)', 'Corporate-Regular(G)': 'Corpoarte Regular(G)', 'Benchmark Rate(B)': 'Benchmark Rate(B)', 'Day Use Rate                               ': 'Day Use Rate                               ', 'Long Staying (Y) /Transient Long Term': 'Long Staying (Y) /Transient Long Term', 'AirCrew(O) /Airline Crews': 'Airline Crews (O)', 'Meetings/Convt. /Corporate Meeting(C)': 'Meetings/Convt. /Corporate Meeting(C)', 'Incentives (I) /Incentives': 'Incentives (I) /Incentives', 'Exhibitions (F) /Trade Fair Exhibition G.': 'Exhibitions (F) /Trade Fair Exhibition .', 'Associations (A) /Association Meeings': 'Associations (A) /Association Meeting', 'Govt/Dip/Military (Q)  /Govt/Dip/Mili.': 'Govt/Dip/Military (Q)  /Govt/Dip/Mili.', 'Sport Events(U1A) /Leisure Group          ': 'Sport Events(U1A) /Leisure Group      (19)', 'Leisure Package-Seasonal(S)': 'Leisure Package-Seasonal(S)        ', 'P&P Discount(V)': 'P&P Discount (V)', 'Leisure Package-Wkend(J)': 'Leisure Pkg- Wkend (J)', 'Contracted Leisure Rate(W1A) / Wholesales': 'Wholesales (W1A)', 'Leisure Package-Other(K)': 'Leisure Package-Other (K) ', 'Group Series (T) /Tour Series': 'Group Series (T) /Tour Series', 'Ad Hoc (U1B) /Leisure Groups': 'Leisure Groups(UIB)', 'Airline Delayed (V1A) ': 'Airline Delayed ', 'Complimentary Rooms (Z)/Compliment': 'Complimentary Rooms (Z)', 'Others (V1B) /P&P Discount & Airline         ': 'House Use'}

def shengji(fname):
    path, filename = os.path.split(fname)
    if os.path.splitext(filename)[1] == '.xls':
        fname = xlsXlsx(fname)
    else:
        fname = fname
    return fname,path

def getData(fname,ws_name):
    wb = load_workbook(fname)
    ws = wb[ws_name]
    riqi = ws['AE49'].value
    month = int(datetime.strftime(riqi,'%Y-%m-%d').split('-')[1])
    xiangmu_jiner ={}
    j = 11
    for j in range(11,46):
        if j in [20,21,22,30,31,32,38,39]:
            continue
        else :
            xiangmu = ws['A{}'.format(j)].value
            t = ws['t{}'.format(j)].value
            u = ws['u{}'.format(j)].value
            w = ws['w{}'.format(j)].value
            x = ws['x{}'.format(j)].value
            ac = ws['ac{}'.format(j)].value
            ad = ws['ad{}'.format(j)].value
            xiangmu_jiner[xiangmu] = (t,u,w,x,ac,ad)
        j = j + 1
    wb.close()
    return month,xiangmu_jiner

def fuzhi(fname, month,xiangmu_jiner,dic01):
    month_dic = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct',
                 11: 'Nov', 12: 'Dec'}
    wb = load_workbook(fname)
    ws_name = month_dic[month]
    ws = wb[ws_name]
    for j in range(6,37):
        if j in [14,15,23,24,30,31]:
            continue
        else :
            xiangmu0 =ws['A{}'.format(j)].value
            xiangmu = dic01[xiangmu0]
            b = xiangmu_jiner[xiangmu][0]
            d = xiangmu_jiner[xiangmu][1]
            f = xiangmu_jiner[xiangmu][2]
            h = xiangmu_jiner[xiangmu][3]
            n = xiangmu_jiner[xiangmu][4]
            p = xiangmu_jiner[xiangmu][5]
            ws['b{}'.format(j)].value = b
            ws['d{}'.format(j)].value = d
            ws['f{}'.format(j)].value = f
            ws['h{}'.format(j)].value = h
            ws['n{}'.format(j)].value = n
            ws['p{}'.format(j)].value = p
        j = j +1
    wb.save(fname)
    os.startfile(fname)

def main():
    msg = '请点选晴川Daily Revenue Report 文件'
    fname = fileopenbox(msg)
    ws_name = r'Segment'
    fname,path = shengji(fname)
    os.chdir(path)
    month, xiangmu_jiner = getData(fname,ws_name)
    msg = '请点选晴川Room Statistic Report 文件'
    fname = fileopenbox(msg)
    fname,path = shengji(fname)
    fuzhi(fname, month,xiangmu_jiner,dic01)

if __name__ == '__main__':
    main()




