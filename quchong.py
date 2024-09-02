# coding: utf-8

import pandas as pd
import os
import xlsxlsx
import excelmessage
import openpyxl
import easygui
import xlwings as xw


def delchongfu(fname,sheetname,in_subject,keep_way ='first' ,index_col=['日期','单据号', '供货单位']):
    app = xw.App(visible = False,add_book = False)
    wb = app.books.open(fname)
    ws = wb.sheets['%s'%sheetname]
    data = pd.DataFrame(pd.read_excel(fname, sheetname))
    col_names = data.columns.to_list()
    data.drop_duplicates(subset=in_subject, keep = keep_way, inplace=True)
    data = data.set_index(index_col)
    data = data.sort_values(by=index_col)
    data = data.reset_index()  # 取消索引
    data = data[col_names]  # 按原来列顺序
    ws.clear()
    ws.range('A1').options(pd.DataFrame, index=False).value = data
    wb.save()
    wb.close()
    app.quit()
    return fname

def main():
    print('请选择要删除重复数据的excel文件:')
    fname =excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    #fname = r'F:\a00nutstore\fishc\原材料实时流水账.xlsx'
    path, excelname = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    print('请选择工作表')
    sheetname = easygui.buttonbox(msg='请选择工作表', title='数据去重工作表', choices=sheetnames)
    ws = wb[sheetname]
    subjects = []
    for first_row in ws.iter_rows(min_row=1, max_row=1):
        for j in first_row:
            subjects.append(j.value)
    print(subjects)

    msg = '请选择判断重复数据所依据的字段'
    print(msg)
    in_subject = easygui.multchoicebox(msg,title='字段',choices=subjects)
    print(in_subject)
    #in_subject = ['日期', '单据号', '供货单位', '品名', '单位', '入库数量', '入库单价', '入库金额', 'cwName', 'priceName', '期间', '送货日期', '白云期间', '令数', '吨数', '令价', '吨价']

    wb.close()
    delchongfu(fname,sheetname,in_subject)


if __name__ == '__main__':
    main()









