import openpyxl
import os
import easygui

def delkonghang(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws =wb[ws_name]
    max_row = ws.max_row

    deleterows = []

    for index,rows in enumerate(ws.values):
        if index == 0:
            continue
        else :
            if ws.cell(index,1).value in [0,'',None]:
                deleterows.append(index)

    deleterows.sort(reverse=True)
    for i in deleterows:
       ws.delete_rows(idx = i)
    wb.save(fname)

def main():
    msg = '请输入要删除空行的文件excel'
    fname = easygui.fileopenbox(msg,title='excel')
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    if len(sheetnames)==1:
        ws = wb.active
        ws_name = ws.title
    else:
        msg = '请选择要删除空行的工作表'
        print(msg)
        ws_name = easygui.buttonbox(msg,title='sheet',choices=sheetnames)

    delkonghang(fname,ws_name)

if __name__ == '__main__':
    main()
