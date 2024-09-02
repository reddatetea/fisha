import openpyxl
import os
import easygui
import tianganDizhiRenzhi

def shufajiaDic(fname = r'书法家档案.xlsx'):
    wb = openpyxl.load_workbook(fname)
    ws = wb['shufajia']
    shufajias = [j.value for j in ws['a'][1:]]
    shufajia_dic = dict.fromkeys(shufajias,[0,0])

    for row in ws.iter_rows(min_row = 2):
        shufajia_dic[row[0].value] = [row[1].value,row[2].value]
    print(shufajia_dic)
    wb.close()
    return shufajia_dic,shufajias

def main():

    tgdz_xh, xh_tgdz = tianganDizhiRenzhi.tianganDizhiSort()
    shufajia_dic,shufajias = shufajiaDic()
    msg = '请点选要查询的书法家名字'
    shufajia = easygui.choicebox(msg,choices=shufajias)
    born_year = shufajia_dic[shufajia][0]
    print('born_year',born_year)
    die_year = shufajia_dic[shufajia][1]
    print('die_year', die_year)
    nianling = die_year - born_year
    print('nianling', nianling)
    msg = '请分别点选书法家作品的天干、地支'

    zuoping_tgdz = tianganDizhiRenzhi.choice_tgdz()
    print('zuoping_tgdz',zuoping_tgdz)
    zuoping_tg_xh = tgdz_xh[zuoping_tgdz]
    print('zuoping_tg_xh',zuoping_tg_xh)

    shufajia_tgdz,dz_sx = tianganDizhiRenzhi.jisuanYearTgdz(born_year, xh_tgdz)
    print('shufajia_tgdz', shufajia_tgdz)
    shufajia_tg_xh = tgdz_xh[shufajia_tgdz]
    print('shufajia_tg_xh', shufajia_tg_xh)
    nian_cha = zuoping_tg_xh - shufajia_tg_xh
    if nian_cha >= 0 :
        nian_cha = nian_cha
    else :
        nian_cha = 60 + nian_cha

    zuoping_year = [str(j ) for j in range(born_year + nian_cha,die_year, 60)]
    print(zuoping_year)
    nianlings =  list(map(lambda x:str(int(x)-born_year),zuoping_year))
    print(nianlings)
    zuoping_year_str = '年、'.join(zuoping_year)+'年'
    nianlings_str = '岁、'.join(nianlings)+'岁'

    print('{}创作这幅作品的年份是{}'.format(shufajia,zuoping_year_str))
    print('{}创作这幅作品的年龄是{}'.format(shufajia, nianlings_str))



if __name__ == '__main__':
    main()