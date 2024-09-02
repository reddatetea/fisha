'''
根据当月白云价格list和纸张list，制作当月原材料价格明细
本版调整列字段位置，将供应商放在第一列，期间放在第二列（与原来相比，互换）
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
import bypricelist02
import zhipricelist02

#建新的当月价格表excel
def jiannewpiao():
    path = r'd:\a00nutstore\006\zw\price'
    os.chdir(path)
    filename = '当月原材料价格表list.xlsx'
    fname = os.path.join(path, filename)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'price'
    taitou = ('供应商', '期间', '品名', '价格', '令数', '吨数', '金额')
    ws.append(taitou)
    wb.save(fname)
    return fname

def writeprice(qijian,fname,prices):
    wb = openpyxl.load_workbook(fname)
    ws = wb['price']
    for i in prices :
        row = (i[0][0],qijian,i[0][1],i[0][2],i[1][0],i[1][1],i[1][2])
        ws.append(row)
    wb.save(fname)

#删除数据为0的行
def dangyueprice(fname,qijian):
    wb = openpyxl.load_workbook(fname)
    ws = wb.create_sheet(title = '%s'%qijian)
    maxrow = ws.max_row
    ws1 = wb['price']
    for index,row in enumerate(ws1.values):
        if (row[4]==0 and row[5]== 0 and row[6]==0) or row[3]==0 :
            continue
        else:
            ws.append(row)
    wb.save(fname)

def main():
    qijian = input('请输入期间，如2020-04\n')
    fname = jiannewpiao()
    byprices = bypricelist02.bypriceDic(qijian)
    zhiprices = zhipricelist02.zhipriceDic(qijian)
    writeprice(qijian,fname,byprices)
    writeprice(qijian,fname,zhiprices)
    dangyueprice(fname,qijian)

if __name__ == '__main__':
    main()

