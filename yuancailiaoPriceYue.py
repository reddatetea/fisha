'''
根据白云价格字典和纸张字典，制作选定期间原材料价格明细
双重排序，供应商、品名
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
import yuancailiaoBaiyunPriceYue
import yuancailiaoZhiPriceYue
import afterSortedYuanxuhao
import chinesetopinyin
import excelseting
import easygui

#建新的当月价格表excel
def jiannewpiao(qijian):
    #print('请输入价格表所在路径:')
    #path = input('')
    path = r'F:\a00nutstore\006\zw\price'
    os.chdir(path)
    filename = '原材料价格表{}.xlsx'.format(qijian)
    fname = os.path.join(path, filename)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title='price')

    taitou = ('供应商','期间', '品名', '价格', '令数', '吨数', '金额')
    ws.append(taitou)
    wb.save(fname)
    return fname

#新建字典
def jiannewdic(qijian):

    byprices = yuancailiaoBaiyunPriceYue.bypriceDic(qijian)
    zhiprices = yuancailiaoZhiPriceYue.zhipriceDic(qijian)
    return byprices,zhiprices

#写入白云价格
def writebyprice(fname,byprices,gongyingshang =  '驻马店白云纸业有限公司'):
    wb = openpyxl.load_workbook(fname)
    ws = wb['price']
    for key, value in byprices.items():
        ws.append((gongyingshang,)+key + value)
    wb.save(fname)

def writezhiprice(fname,zhiprices):
    wb = openpyxl.load_workbook(fname)
    ws = wb['price']
    for key, value in zhiprices.items():
        ws.append(key + value)
    wb.save(fname)

#删除数据为0的行
def dangyueprice(fname,qijian):
    wb = openpyxl.load_workbook(fname)
    ws = wb.create_sheet(title = qijian)
    maxrow = ws.max_row

    ws1 = wb['price']
    for index,row in enumerate(ws1.values):
        if (row[4]==0 and row[5]== 0 and row[6]==0) or row[3]==0:
            continue
        else:
            ws.append(row)

    wb.save(fname)

def sortByGongyingshang(fname,qijian):
    wb = openpyxl.load_workbook(fname)
    ws = wb[qijian]
    values = list(ws.values)[1:]
    gongyingshangs = [j.value for j in ws['a'][1:]]             #供应商列表
    gyss_pinyins = list(map(chinesetopinyin.getpinyin,gongyingshangs))     #供应商拼音列表
    new_values = []
    for j in range(len(values)):
        d = list(values[j])
        d.insert(0, gyss_pinyins[j])                    #只能就地插入元素，不能赋值给另外变量
        new_values.append(d)
    new_values_first_sort = sorted(new_values,key = lambda x :x[2])                           #第一次排序
    new_values_second_sort = sorted(new_values_first_sort ,key = lambda x : x[0])     #第二次排序
    ws.delete_rows(2, ws.max_row - 1)
    for j in new_values_second_sort:
        ws.append(j[1:])                                                  #辅助列就不添加上去了
    wb.save(fname)

def main():
    qijian = easygui.enterbox('请输入期间，如2020-04')
    fname = jiannewpiao(qijian)
    byprices,zhiprices = jiannewdic(qijian)
    writebyprice(fname,byprices,gongyingshang =  '白云')
    writezhiprice(fname, zhiprices)
    dangyueprice(fname,qijian)
    sortByGongyingshang(fname, qijian)
    excelseting.setPrintArea(fname,qijian)
    excelseting.firstRowSeting(fname,qijian)
    taitou = '原材料价格表{}'.format(qijian)
    excelseting.yemei(fname, qijian,taitou )
    os.startfile(fname)






if __name__ == '__main__':
    main()

