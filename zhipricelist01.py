'''
尝试根据纸入库制作纸价格列表，看是不是快些！
本版增加传入期间参数处理,取当期
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
from openpyxl.utils import column_index_from_string

def zhipriceDic(qijian):
    #print('请输入纸入库文件所在路径：')
    #path = input('')

    path = r'd:\a00nutstore\006\zw\else'
    os.chdir(path)
    filename = r'2020入库.xlsx'
    fname = os.path.join(path,filename)
    wb = openpyxl.load_workbook(fname,data_only=True,read_only=True)
    ws = wb['入库']
    maxrow = ws.max_row

    zhiprices = []
    gongying_price = []
    ling_dun_jiner = []

    #获取列号
    gongyingshang_num = column_index_from_string('B')
    qijian_num = column_index_from_string('C')
    pricename_num  = column_index_from_string('X')
    price_num = column_index_from_string('O')
    lingshu_num = column_index_from_string('G')
    dunshu_num = column_index_from_string('M')
    jiner_num = column_index_from_string('R')
    jizhang_num = column_index_from_string('V')

    for row in range(2,maxrow+1):
        dangqian_qijian = ws.cell(row,qijian_num).value
        jizhang = ws.cell(row, jizhang_num).value

        #期间、令数不为None、吨数不为None、pricename不为None
        if (dangqian_qijian == qijian) and (jizhang not in ['',None]):
            gongyingshang = ws.cell(row, gongyingshang_num).value
            pricename = ws.cell(row, pricename_num).value
            price = ws.cell(row, price_num).value
            lingshu = ws.cell(row, lingshu_num).value
            dunshu = ws.cell(row, dunshu_num).value
            jiner = ws.cell(row, jiner_num).value

            if lingshu in ['', None]:
                lingshu = 0
            if dunshu in ['', None]:
                dunshu = 0
            if jiner in ['', None]:
                jiner = 0

            if [gongyingshang,pricename,price] not in gongying_price :
                gongying_price.append([gongyingshang,pricename,price])
                ling_dun_jiner.append([lingshu,dunshu,jiner])
                c = [gongyingshang,pricename,price],[lingshu,dunshu,jiner]
                zhiprices.append(c)

            else :
                k = [zhiprice[0]  for zhiprice in zhiprices]
                print(k)
                j = k.index([gongyingshang,pricename,price])
                zhiprices[j][1][0]=zhiprices[j][1][0]+lingshu
                zhiprices[j][1][1]=zhiprices[j][1][1]+dunshu
                zhiprices[j][1][2]=zhiprices[j][1][2]+jiner

    print(zhiprices)
    wb.close()
    return zhiprices

def main():
    qijian = input('请输入期间，如2020-04\n')
    zhiprices = zhipriceDic(qijian)


if __name__ == '__main__':
    main()







