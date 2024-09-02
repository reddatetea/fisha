'''
bypricedic.py
根据白云入库制作白云指定期间的价格字典
本版增加传入期间参数处理
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
from openpyxl.utils import column_index_from_string

def bypriceDic(qijian):
    #print('请输入白云入库文件所在路径：')
    #path = input('')
    path = r'd:\a00nutstore\006\zw\baiyun'
    os.chdir(path)
    filename = r'2020白云入库.xlsx'
    fname = os.path.join(path,filename)
    wb = openpyxl.load_workbook(fname,data_only=True,read_only=True)
    ws = wb['2020']
    maxrow = ws.max_row

    byprices = []
    gongying_price = []
    ling_dun_jiner = []

    gongyingshang = '驻马店白云纸业有限公司'

    #获取列号
    qijian_num = column_index_from_string('E')
    pricename_num  = column_index_from_string('Q')
    price_num = column_index_from_string('K')
    lingshu_num = column_index_from_string('H')
    dunshu_num = column_index_from_string('J')
    jiner_num = column_index_from_string('L')
    jizhang_num = column_index_from_string('N')

    for row in range(2,maxrow+1):
        danqian_qijian = ws.cell(row,qijian_num).value
        pricename = ws.cell(row,pricename_num).value
        price = ws.cell(row, price_num).value
        lingshu = ws.cell(row, lingshu_num).value
        dunshu = ws.cell(row, dunshu_num).value
        jiner = ws.cell(row, jiner_num).value
        jizhang = ws.cell(row, jizhang_num).value

        for row in range(2, maxrow + 1):
            dangqian_qijian = ws.cell(row, qijian_num).value
            jizhang = ws.cell(row, jizhang_num).value

            # 期间、令数不为None、吨数不为None、pricename不为None
            if (dangqian_qijian == qijian) and (jizhang not in ['', None]):
                pricename = ws.cell(row, pricename_num).value
                price = ws.cell(row, price_num).value
                lingshu = ws.cell(row, lingshu_num).value
                dunshu = ws.cell(row, dunshu_num).value
                jiner = ws.cell(row, jiner_num).value

                if lingshu in ['', None]:
                    lingshu = 0
                if dunshu in ['', None]:
                    dunshu = 0
                if jiner in ['', None]:
                    jiner = 0

                if [gongyingshang, pricename, price] not in gongying_price:
                    gongying_price.append([gongyingshang, pricename, price])
                    ling_dun_jiner.append([lingshu, dunshu, jiner])
                    c = [gongyingshang, pricename, price], [lingshu, dunshu, jiner]
                    byprices.append(c)

                else:
                    k = [byprice[0] for byprice in byprices]
                    print(k)
                    j = k.index([gongyingshang, pricename, price])
                    byprices[j][1][0] = byprices[j][1][0] + lingshu
                    byprices[j][1][1] = byprices[j][1][1] + dunshu
                    byprices[j][1][2] = byprices[j][1][2] + jiner

    print(byprices)
    wb.close()
    return byprices

def main():
    qijian = input('请输入期间，如2020-04\n')
    byprices = bypriceDic(qijian)


if __name__ == '__main__':
    main()







