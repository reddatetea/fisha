#点击‘管理亚马逊的内容’，复制需要添加的内容，拷贝至书单中一个空工作表中，保存后运行，分别将书名、作者、购书日期保存至书单！

import openpyxl
import os
import easygui
import  datetime
import pandas as  pd
import xlwings as xw
import re
import fayinxiang
import deletesheet

def choiceFnameSheet():
    dqrq = datetime.date.today().strftime('%Y%m%d')
    msg = '请输入要删除空行的文件excel'
    fname = easygui.fileopenbox(msg, title='excel')
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    if len(sheetnames) == 1:
        ws = wb.active
        ws_name = ws.title
    else:
        msg = '请选择要删除空行的工作表'
        print(msg)
        ws_name = easygui.buttonbox(msg, title='sheet', choices=sheetnames)

    return dqrq,fname,ws_name


def delkonghang(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws =wb[ws_name]
    max_row = ws.max_row
    deleterows = [cell.row for cell in ws['A'] if cell.value in ['', '...', '已读', None, '0', 0]]
    deleterows.sort(reverse=True)
    print(deleterows)
    for i in deleterows:
        print(i)
        ws.delete_rows(idx = i)
    wb.save(fname)

def delyangzhang(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws =wb[ws_name]
    max_row = ws.max_row
    deleterows = []
    for index,rows in enumerate(ws.values):
        if index == 0:
            continue
        else :
            j= index
            if ws.cell(index+1,1).value =='样章':
                deleterows.append(j-1)
                deleterows.append(j)
                deleterows.append(j + 1)
                deleterows.append(j + 2)

    deleterows.sort(reverse=True)
    for i in deleterows:
       ws.delete_rows(idx = i)
    wb.save(fname)

def borrowedbooks():
    msg = '请输入借阅的书名'
    print(msg)
    target0 = easygui.multenterbox(msg, title='文件后缀',
                                   fields=['book1', 'book2', 'book3', 'book4', 'book5', 'book6', 'book7', 'book8',
                                           'book9', 'book10'])
    jiebooks = []
    for j in target0:
        if j != '':
            jiebooks.append(j)

    print(jiebooks)

    return jiebooks


def  delBorrowedbooks(fname,ws_name,jiebooks):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    max_row = ws.max_row
    deleterows = []
    for book in jiebooks:
        print(book)
        for index, rows in enumerate(ws.values):
            if index == 0:
                continue
            else:
                print(ws.cell(index, 1).value)
                if book in str(ws.cell(index, 1).value):
                    deleterows.append(index)
                    deleterows.append(index + 1)
                    deleterows.append(index + 2)

    deleterows.sort(reverse=True)
    for i in deleterows:
        ws.delete_rows(idx=i)
    wb.save(fname)

def  getbooknames(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws  = wb[ws_name]
    max_row = ws.max_row
    booknames = []
    authors = []
    purchase_dates = []
    for  index,row in enumerate(ws.values):
        row_xuhao = index +1
        if row_xuhao%3==1:
            booknames.append(row[0])
        elif  row_xuhao%3==2:
            authors.append(row[0])
        else:
            purchase_dates.append(row[0])
    wb.close()
    print(booknames)
    print(authors)
    print(purchase_dates)
    print(len(booknames),len(authors),len(purchase_dates))
    return booknames,authors,purchase_dates

def goushuriqi(purchase_dates):
    goushuriqis = []
    regax = r'(.*)年(.*)月(.*)日'
    pattern = re.compile(regax)
    for purchase_date in purchase_dates :
        mat = pattern.search(purchase_date)
        if mat:
            riqi_string0 = mat.group(1) + '-' + mat.group(2) + '-' + mat.group(3)
            biaozun_riqi = datetime.datetime.strptime(riqi_string0, '%Y-%m-%d')
            goushuriqis.append(biaozun_riqi)
        else :
            goushuriqis.append(purchase_date)
    return goushuriqis

def  creatnewsheet(dqrq,fname,ws_name,booknames, authors, goushuriqis):
    wb = openpyxl.load_workbook(fname)
    ws1 = wb[ws_name]
    ws2 = wb.create_sheet(title = dqrq)
    title = ['序号','书名','作者','购书日期','类型']
    ws2.append(title)
    wb.save(fname)
    wb = openpyxl.load_workbook(fname)
    ws  = wb[dqrq]
    for j in  range(2,len(booknames)+2):
        ws.cell(j,2).value = booknames[j-2]
        ws.cell(j,3).value = authors[j-2]
        ws.cell(j,4).value = goushuriqis[j-2]
    wb.save(fname)
    return fname,ws_name

def excelsort(fname,ws_name):
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(fname)
    ws = wb.sheets[ws_name]
    data = pd.DataFrame(pd.read_excel(fname, ws_name))
    data = data.sort_values(by=['购书日期'], ascending=True)                     #对特定列进行排序
    data=data.loc[:,['书名','作者','购书日期']]                 #选取所有行，特定列
    print(data)
    data = data.set_index('书名')
    ws.clear()
    ws.range('A1').value = '序号'
    ws.range('E1').value = '类型'
    ws.range('B1').value = data
    wb.save()
    wb.close()
    app.quit()
    return ws_name

def  addnewbooks(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws1 = wb[ws_name]                                     #新书
    max_row1 = ws1.max_row
    sheetnames = wb.sheetnames
    msg = '请选择要添加新书的工作表'
    print(msg)
    ws_name2 = easygui.buttonbox(msg, title='sheet', choices=sheetnames)
    ws2 = wb[ws_name2]                                       #老书
    max_row2 =  ws2.max_row-1
    for  index,row in enumerate(ws1.values):
        if index==0:
            continue
        else :
            ws2.append(row)
    for  j in range(max_row2+1,max_row2+1+max_row1):
        ws2.cell(j,1).value = j-1
    wb.save(fname)
    return ws_name2

def makeAmazonBookstxt(fname,ws_name2,dqrq):
    wb = openpyxl.load_workbook(fname)
    ws  = wb[ws_name2]
    bookname = r'亚书逊书单{}.txt'.format(dqrq)
    with open(bookname,'w',encoding='utf-8') as f :
        for index,row in enumerate(ws.values):
            for j in row:
                j =str(j)+'\\'
                f.write(j)
            f.write('\n')
    wb.close()
    return  bookname

def  autorename(fname,dqrq):
    wb =openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    regax = r'书单(\d+$)'
    pattern = re.compile(regax)
    for sheetname in sheetnames:
        mat = pattern.search(sheetname)
        if mat:
            wb[sheetname].title = '书单{}'.format(dqrq)
        else :
            continue
    newfilename = r'亚马逊书单{}.xlsx'.format(dqrq)
    wb.save(newfilename)
    return newfilename

def main():
    dqrq, fname, ws_name = choiceFnameSheet()
    delkonghang(fname,ws_name)
    delyangzhang(fname, ws_name)
    jiebooks = borrowedbooks()
    delBorrowedbooks(fname, ws_name, jiebooks)
    booknames, authors, purchase_dates = getbooknames(fname,ws_name)
    #goushuriqis = goushuriqi(purchase_dates)              #在家里用此代码，不用下面四行代码
    #公司用下面这四行代码，在家里
    booknames.reverse()
    authors.reverse()
    purchase_dates.reverse()
    goushuriqis = purchase_dates
    name,ws_name = creatnewsheet(dqrq, fname, ws_name, booknames, authors, goushuriqis)

    ws_name = dqrq
    ws_name = excelsort(fname, ws_name)
    ws_name2 = addnewbooks(fname, ws_name)
    borrowedbooksname = r'amazonBorrowedBooks{}.txt'.format(dqrq)
    with open(borrowedbooksname,'w') as f :
        for jiebook in jiebooks:
            j = jiebook+'\n'
            f.write(j)

    bookname  =makeAmazonBookstxt(fname, ws_name2, dqrq)
    fayinxiang.faYinxiang(bookname)
    fname = autorename(fname, dqrq)
    deletesheet.deleteSheet(fname)


if __name__ == '__main__':
    main()







