'''
20211014用pandas进行暂估处理
'''
import os
import openpyxl
import easygui
import pandas as pd
import numpy as np
import re

def tojianchen(zaiyao,jianchens):
    for j in jianchens:
        if j in zaiyao:
            return j
            break

def zanguChuli(fname,jianchens):
    df0 = pd.read_excel(fname,usecols=[0,1,6,7,8,9,10,11],dtype={0:str,1:str})
    df0 = df0.loc[df0.isnull().sum(1) < 3]
    msg  = '请选择当月暂估文件：'
    fname1 = easygui.fileopenbox(msg)
    wb1 = openpyxl.load_workbook(fname1,data_only=True)
    sheet0 = wb1.sheetnames[1]  # 默认暂估工作表
    regax = r'\d{4}'
    pattern = re.compile(regax)
    for sheetname in wb1.sheetnames:
        mat = pattern.search(sheetname)
        if mat != None:
            sheet0 = mat.group()
            break
        else:
            continue
    ws1 = wb1['双佳']
    ws0= wb1[sheet0]

    max_row = ws1.max_row
    max_row2 = ws0.max_row
    df1 = df0[~df0['摘要'].str.contains('小计|合计')]
    #df1['借方本币']=df1['摘要'].agg(lambda x:x[:8])
    df1['借方本币'] = df1['摘要'].apply(lambda x: tojianchen(x, jianchens))
    df1[ '余额本币'] = df1['余额本币'].agg(lambda x: "")
    df1['方向'] = round(df1['贷方本币']*1.13,2)
    df1  = df1.assign(余额本币=lambda  s:np.where(df1.摘要.str.contains('冲|调整'),'有','无'))
    year = df1.columns[0][:4]
    a = list(df1.columns)
    a[0] = 'month'
    a[1]= 'day'
    df1.columns = a
    df1 = df1.assign(shijian=year+df1.month+df1.day)
    df1.shijian = df1.shijian.transform(lambda x:pd.Timestamp(x))
    with pd.ExcelWriter(fname1,engine='openpyxl') as writer:
        writer.book = wb1
        writer.sheets = dict((ws1.title,ws1) for ws1 in wb1.worksheets)
        df0.to_excel(writer,sheet_name = '双佳',index = False,header=None,startrow=max_row)
        df1.to_excel(writer, sheet_name=sheet0, index=False, header=None, startrow=max_row2)
    #writer.save()
    return fname1

def main():
    path = r'F:\a00nutstore\006\zw\yingfu\2022yingfu'
    os.chdir(path)
    msg = '请选择临时暂估文件：'
    fname = easygui.fileopenbox(msg)
    fname_jianchen = r'F:\a00nutstore\006\zw\yingfu\暂估供应商简称.xlsx'
    sheet_name0 = r'简称'
    df_jianchen = pd.read_excel(fname_jianchen, sheet_name0)
    jianchens = df_jianchen['简称'].to_list()  # 简称列表
    fname1=zanguChuli(fname,jianchens)
    os.startfile(fname1)

if __name__ == '__main__':
    main()



