import  os
import  openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font,PatternFill,GradientFill,Side,Border,Alignment

#fname = r'F:\a00nutstore\fishc\abc统计表.xlsx'
fname = r'D:\a00nutstore\fishc\abc统计表.xlsx'
path,filename = os.path.split(fname)
os.chdir(path)

wb = openpyxl.load_workbook(fname)
ws = wb['abc统计表']

max_row = ws.max_row

ws.print_area = ws.dimensions         #ws.dimensions工作表最大区域 设为 打印区域

side = Side(style = 'thick',color = '000000')        #边框 线型为thin,黑色
side1 = Side(style='thin', color= '40E0D0')  # 边框 线型为thin,白色

border1 = Border(left=side, right=side1, top=side, bottom=side1)  # 定义边框的左右上下
border2 = Border(left=side1, right=side, top=side, bottom=side1)  # 定义边框的左右上下
border3 = Border(left=side1, right=side1, top=side, bottom=side1)  # 定义边框的左右上下

border4 = Border(left=side, right=side1, top=side1, bottom=side1)  # 定义边框的左右上下
border5 = Border(left=side1, right=side, top=side1, bottom=side1)  # 定义边框的左右上下
border6 = Border(left=side1, right=side1, top=side1, bottom=side1)  # 定义边框的左右上下

# for  first_rows in ws.iter_cols(min_row=1, min_col=1, max_row=max_row,max_col=1):       #iter_rows按行遍历，ws.iter_cols按列遍历
#     for cell in first_rows:
#         print(cell.row,cell.column)
#         if cell.column == 1 :
#             cell.border = border1

for  row in ws.rows :
    for cell in row :
        print(cell.coordinate)       #cell.coordinate单元格坐标


wb.save(fname)