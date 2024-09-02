'''
吨价0全部替换,20200902疑问：难道将所有价格都重新更新？
'''
import openpyxl
import pricelisttoprice

def daoruPrice(fname2):

    gongying_dic= pricelisttoprice.yuancailiaoprice()
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2.active
    mrows2 = ws2.max_row
    print(mrows2)

    for i in range(2, mrows2 + 1):
        # 如果吨价0为0，或空，替换为新或0，否则不变

        try:
            # 获取期间字典
            qijian = gongying_dic[ws2.cell(i, 3).value]
            # 获取品名字典,#白云纸业的期间和其他供应商不一样
            if ws2.cell(i, 3).value == '驻马店白云纸业有限公司':
                pinming = qijian[ws2.cell(i, 13).value]
            else:
                pinming = qijian[ws2.cell(i, 11).value]

            # 获取吨价
            dun0 = pinming[ws2.cell(i, 10).value]
            ws2.cell(i, 18).value = dun0
        except:
            ws2.cell(i, 18).value = 0

    wb2.save(fname2)

def main():
    fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'


    daoruPrice(fname2)

if __name__=='__main__':
    main()



