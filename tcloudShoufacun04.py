'''
对T+CLOUD上的收发存汇总表进行二次开发，与公司以前的格式一致
此版本加零结存，将抄本中的小学生按原顺序排列
'''
import os
import datetime
import easygui
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Fill, Alignment
import formatPainter

# store_dic = {'001库':"df.num == '001'",'002电商库':"df.num == '002'",'总库':"(df.num == '001') | (df.num == '002')"}
lst = ['仓库编码',
       '仓库',
       '存货分类 (1级)',
       '存货分类 (2级)',
       '存货分类 (3级)',
       '存货分类 (4级)',
       '存货分类 (5级)',
       '存货编码',
       '存货',
       '存货代码',
       '数量(主单位)',
       '平均单价',
       '金额',
       '数量(辅单位)',
       '计量单位组合',
       '数量(主单位).1',
       '平均单价.1',
       '金额.1',
       '数量(辅单位).1',
       '计量单位组合.1',
       '数量(主单位).2',
       '平均单价.2',
       '金额.2',
       '数量(辅单位).2',
       '计量单位组合.2',
       '数量(主单位).3',
       '平均单价.3',
       '金额.3',
       '辅数量',
       '计量单位组合.3']

lst2 = [
    'store',
    'num',
    'class01',
    'class02',
    'class03',
    'class04',
    'class05',
    'code',
    'stock',
    'content',
    'begin_ben',
    '平均单价',
    '金额',
    'begin_jian',
    '计量单位组合',
    'ruku_ben',
    '平均单价.1',
    '金额.1',
    'ruku_jian',
    '计量单位组合.1',
    'chuku_ben',
    '平均单价.2',
    '金额.2',
    'chuku_jian',
    '计量单位组合.2',
    'end_ben',
    '平均单价.3',
    '金额.3',
    'end_jian',
    '计量单位组合.3']

lst3 = [
    'class02',
    'class05',
    'code',
    'stock',
    'content',
    'begin_jian',
    'ruku_jian',
    'chuku_jian',
    'end_jian',
]

lst01_yesterday = [
    'class02',
    'class05',
    'code',
    'stock',
    'begin_jian',
    'ruku_jian',
    'chuku_jian',
]
lst02_yesterday = [
    'class02',
    'class05',
    'code',
    'stock',
    'yesterday_jian',
    'yesterday_ruku',
    'yesterday_chuku',

]
lst_merge = ['class02', 'class05', 'code', 'stock',
             'content',
             'begin_jian', 'yesterday_jian',
             'yesterday_ruku', 'ruku_jian',
             'yesterday_chuku', 'chuku_jian',
             'end_jian']

lst_result0 = ['class05', 'code', 'stock',
               'content',
               'begin_jian', 'yesterday_jian',
               'yesterday_ruku', 'ruku_jian',
               'yesterday_chuku', 'chuku_jian',
               'end_jian']

lst_result = ['类别', '货号', '品名',
              '含量',
              '月初', '上日',
              '本日入库', '入库累计',
              '本日出库', '出库累计',
              '结余']


def chuli(fname, store_num):
    df = pd.read_excel(fname, skiprows=7)
    df = df.iloc[:, 1:]
    df.columns = lst
    df.columns = lst2
    if store_num == '001库':
        df = df.loc[df.store == '001']
    elif store_num == '002电商库':
        df = df.loc[df.store == '002']
    else:
        df = df.loc[(df.store == '001') | (df.store == '002')]

    df['content'] = df['end_ben'] / df['end_jian']
    df = df[df['store'] != '制表人:']
    df = df[df['store'] != '合计：']
    df = df[df['store'].notnull()]
    df = df.iloc[:, 2:]
    df = df[lst3]
    return df


def printseting(fname1, riqi):
    # 设置字体、边框、对齐等常量
    font = Font(name='宋体', size=10)
    font_xiaoji = Font(name='宋体', bold=True, size=10)
    font_firstrow = Font(name='宋体', bold=True, size=20)
    thin_danbian = Side(border_style='thin')
    double_danbian = Side(border_style='double')
    thin_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=thin_danbian)
    double_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=double_danbian)
    wb = openpyxl.load_workbook(fname1)
    for ws in wb.worksheets:
        if ws.title == 'Sheet':
            del wb[ws.title]
        else:
            # 插入第一行
            ws.insert_rows(1)
            ws['A1'].value = '双佳' + ws.title + ' ' + riqi
            ws['A1'].font = font_firstrow
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            max_row = ws.max_row
            max_column = ws.max_column
            if ws.title != '合计':
                # 单元格宽度
                ws.column_dimensions['A'].width = 8.56
                ws.column_dimensions['B'].width = 17.33
                ws.column_dimensions['C'].width = 13.56
                ws.column_dimensions['D'].width = 3.67
                for j in 'EFGHIJK':
                    ws.column_dimensions['{}'.format(j)].width = 7.44
                    ws.merge_cells('A1:K1')
                for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
                    if row == 2:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        for col in range(1, max_column + 1):
                            if col <= 3:
                                ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                ws.freeze_panes = ws['E3']
            else:
                # 单元格宽度
                ws.column_dimensions['A'].width = 16
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 14
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 12
                ws.merge_cells('A1:H1')
                # ws.freeze_panes = ws['E3']
                for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
                    if row == 2:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        for col in range(1, max_column + 1):
                            if col <= 1:
                                ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                ws.cell(row, col).alignment = Alignment(horizontal='right', vertical='center')
            for cell in ws['A']:
                row = cell.row
                if row == 1:
                    continue
                else:
                    if '小计' in cell.value:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).font = font_xiaoji
                            ws.cell(row, col).border = double_bian
                            ws.merge_cells('A{}:C{}'.format(row, row))
                    else:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).font = font
                            ws.cell(row, col).border = thin_bian
            # ws.freeze_panes = ws['E3']
            ws.print_title_rows = '1:2'  # the first row
            # 页脚设置
            ws.oddFooter.center.text = " &[Page] / &N"  # 1/n
            ws.oddFooter.center.size = 12  # 页脚中字体大小
            ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
            ws.oddFooter.center.color = "000000"  # 页脚中字体颜色
            ws.oddFooter.right.text = "张文伟 &[Date]"  # 页脚右 文字
            ws.oddFooter.right.size = 12  # 页脚右 字体大小
            ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
            ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
            ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.5, left=0.21, right=0.24,
                                                                  footer=0.5,
                                                                  bottom=1)
            for i in ws.iter_rows():  # 将零值替换为空
                for j in i:
                    if j.value == 0:
                        j.value = ''
                    else:
                        continue
    wb.save(fname1)


def res_path(relative_path):
    """获取资源绝对路径"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# first,choice仓库编号
msg = '请点选仓库'
nums = ['001库', '002电商库', '总库']
num = easygui.buttonbox(msg, choices=nums)
msg = '请点选产成品"当日"工作表'
fname = easygui.fileopenbox(msg, title='AAA')
path, filename = os.path.split(fname)
os.chdir(path)
df1 = chuli(fname, num)
msg = '请点选产成品"累计"工作表'
fname2 = easygui.fileopenbox(msg, title='BBB')
df2 = chuli(fname2, num)
df1 = df1[lst01_yesterday]
df1.columns = lst02_yesterday

today_date = datetime.date.today()
yesterday_date = (today_date + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")
msg = '报表日期是{}?'.format(yesterday_date)
choice = easygui.ccbox(msg, title='请选择"是"或"否"', choices=('是', '否'))
if choice:
    riqi = yesterday_date
else:
    msg = '请输入报表日期'
    riqi = easygui.enterbox(msg, title=" 昨天日期")
fname_canchengpin = '产成品日报表{}.xlsx'.format(riqi)
wb = openpyxl.Workbook()
wb.save(fname_canchengpin)

merge = pd.merge(df2, df1, how='left', on=['class02', 'class05', 'code', 'stock'], sort=False)
gp = merge.groupby('class02', sort=False)
data_heji = []


def xiaojiClass05(gp2, key, d):
    d = d[lst_result0]
    d.loc[key] = d.sum()
    d.iloc[-1, :4] = [key + '小计', '', '', '']
    return d


for name, gp002 in gp:
    table = pd.pivot_table(gp002, index=['class05', 'code'], aggfunc=sum, sort=False)
    table = table.reset_index()
    gp2 = table.groupby('class05', sort=False)

    heji = pd.DataFrame(gp002).sum()
    data = []
    for key, gp005 in gp2:
        df05 = xiaojiClass05(gp2, key, gp005)
        data.append(df05)
    df005 = pd.concat(data)
    df005 = df005[lst_result0]

    df005.loc[name] = heji
    df005.iloc[-1, :4] = [name + '合计', '', '', '']
    heji_row = df005.iloc[-1]
    data_heji.append(heji_row)

    with pd.ExcelWriter(fname_canchengpin, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df005.columns = lst_result
        df005.to_excel(writer, sheet_name=name, index=False)
df_heji = pd.DataFrame(data_heji)
df_heji.columns = lst_result
df_heji.loc['合计'] = df_heji.sum(0)

# 零结存
merge1 = merge[lst_merge]
merge1 = merge1.fillna(0)
df_zero = merge1[merge1.end_jian == 0]
df_zero = df_zero.iloc[:, 1:]
df_zero.columns = lst_result
df_zero.loc['合计'] = df_zero.sum(0)
df_zero.iloc[-1, :4] = ['零结存合计', '', '', '']
with pd.ExcelWriter(fname_canchengpin, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_zero.to_excel(writer, sheet_name='零结存', index=False)

with pd.ExcelWriter(fname_canchengpin, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_heji = df_heji[
        ['月初', '上日',
         '本日入库', '入库累计',
         '本日出库', '出库累计',
         '结余']]
    df_heji.to_excel(writer, sheet_name='合计', index_label='类别')

printseting(fname_canchengpin, riqi)
fname0 = res_path('img\leiji.xlsx')
wb0 = openpyxl.load_workbook(res_path('img\leiji.xlsx'))
ws0 = wb0.active
area0 = ws0['A1:H10']
wb = openpyxl.load_workbook(fname_canchengpin)
ws = wb['合计']
cell_start1 = ws['A1']
formatPainter.stylesFormat(ws, area0, cell_start1)
wb0.close()
wb.save(fname_canchengpin)
os.startfile(fname_canchengpin)



