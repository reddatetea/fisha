from openpyxl import load_workbook, Workbook
import addlists
import cangkutotal
import chaiwutotal
import zhongtotal

def addQita(addList1s,addList2s):

    filename = r'数量金额盘存表.xlsx'

    wb6 = load_workbook(filename)
    ws6 = wb6.active

    jishu = 0
    first_list = []
    for index, row in enumerate(ws6.values):
        # print(row[1])
        first_list.append(row[1])

        if row[0] == None:
            jishu = jishu + 1
    mrows = ws6.max_row - jishu
    print(mrows)

    # 插入N行，n等于列表个数,在第mrows-1行上面行开始处
    ws6.insert_rows(mrows - 1, len(addList2s))

    # 批量写入数据, 写入科目addlist1
    for i in range(len(addList2s)):
        ws6.cell(mrows - 1 + i, 18, addList2s[i])

    # 批量写入数据, 按列写入，第18列，第mrows-1行开始,
    # 写入期初数量
    for i in range(len(addList2s)):
        try:
            a = qichu_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0

        ws6.cell(mrows - 1 + i, 24, a)

    # 写入入库数量
    for i in range(len(addList2s)):
        try:
            a = ruku_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0
        ws6.cell(mrows - 1 + i, 25, a)

    # 写入采购数量
    for i in range(len(addList2s)):
        try:
            a = caigou_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0

        ws6.cell(mrows - 1 + i, 26, a)

    # 写入半成品数量
    for i in range(len(addList2s)):
        try:
            a = banchengpin_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0

        ws6.cell(mrows - 1 + i, 27, a)

    # 写入出库数量
    for i in range(len(addList1s)):
        try:
            a = lingliao_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0
        ws6.cell(mrows - 1 + i, 28, a)

    # 写入生产领料数量
    for i in range(len(addList1s)):
        try:
            a = lingliao_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0
        ws6.cell(mrows - 1 + i, 29, a)

    # 写入期末数量
    for i in range(len(addList2s)):
        try:
            a = qimo_dic[ws6.cell(mrows - 1 + i, 18).value]
        except:
            a = 0

        ws6.cell(mrows - 1 + i, 30, a)

    # 删除一行，删除第morws行
    # ws6.delete_rows(mrows)
    newname = filename[:-5] + '8.xlsx'
    wb6.save(newname)


def main():
    fname = r'F:\a00nutstore\006\zw\2021\202102\盘存表材料更名.xlsx'
    qichu_dic, ruku_dic, caigou_dic, banchengpin_dic, chuku_dic, lingliao_dic, qimo_dic, taitou_lists, cangkucailiaolist1 = cangkutotal.cangkuTotal(fname)

    list1 = cangkucailiaolist1
    # list1 = zhongtotal.zhongTotal()
    print(list1)
    list2,kemuname_dic,fname2 = chaiwutotal.chaiwuTotal()
    print(list2)

    addList1s, addList2s = addlists.addLists(list1, list2)
    addQita(addList1s,addList2s)

if __name__=='__main__' :
    main()

