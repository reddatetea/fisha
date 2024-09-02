'''
本版尝试导入纸类材料ke,chang,kuan
计算财务名称cw_name和价格名称price_name
计算期间
计算令数ling和吨数dun

'''
# _*_ coding = uft-8 _*_
from openpyxl import Workbook,load_workbook
import re
import time
import datetime
import qijianchuli
import baiyunqijian
import singleCailiaoRename
import os

def liushuiZhang03(fname):

    wb1 = load_workbook(fname)
    sheet1 = wb1.active

    ss_wb2 = Workbook()
    ss_sheet2 = ss_wb2.active
    ss_sheet2.title = '流水账'

    for index, row in enumerate(sheet1.values):
        if index == 0:
            ss_sheet2.append(('日期','单据号','供货单位','品名','单位','入库数量','入库单价','入库金额','cwName','priceName','期间','送货日期','白云期间','令数','吨数','令价','吨价','吨价0','记账'))
        else :
            #row[14],row[15],row[16] = 0,0,0
            riqi = row[0]
            danjuhao = int(row[1])       #转为整形20200630
            gonghuoshang = row[2]
            pinming = row[3]
            danwei = row[4]
            shuliang = row[5]
            danjia = row[6]
            jiner = row[7]

            #计算财务名称cw_name和价格名称price_name
            cw_name,price_name,ke,chang,kuan = singleCailiaoRename.singlecailiaoName(pinming)

            #计算期间
            qijian = qijianchuli.qijian(riqi)
            baiyun_qijian,songhuoriqi = baiyunqijian.baiyunQijian(riqi)

            #计算令数ling和吨数dun

            if row[4].strip()=='令':
                ling = row[5]
                lingjia = row[6]
                dun = int(ke)/1000/1000*int(chang)*int(kuan)/1000/1000*500*row[5]
                dunjia = 0
            else:
                dun = 0
                ling = 0
                lingjia = 0
                if row[4].strip()== '公斤' or row[4].strip()=='kg':
                    dun = row[5]/1000
                    dunjia = row[6]*1000

                else:
                    ling = 0
                    dunjia = 0


            # 每行数据
            meihang = (riqi, danjuhao, gonghuoshang, pinming, danwei, shuliang, danjia, jiner, cw_name, price_name, qijian, songhuoriqi,baiyun_qijian,ling,dun,lingjia,dunjia)
            ss_sheet2.append(meihang)

    fname = fname[:-8]+ '002.xlsx'
    ss_wb2.save('%s'%fname)
    time.sleep(3)
    return fname

def main():
    fname = r'D:\a00nutstore\fishc\01月25流水账001.xlsx'
    liushuiZhang03(fname)

if __name__ == '__main__':
    main()










