'''
bypricedic.py
根据白云入库制作白云指定期间的价格字典
本版增加传入期间参数处理
金额不能为0
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
from openpyxl.utils import column_index_from_string

def bypriceDic(qijian):
    #print('请输入白云入库文件所在路径：')
    #path = input('')
    path = r'F:\a00nutstore\006\zw\baiyun'
    os.chdir(path)
    filename = r'2020白云入库.xlsx'
    fname = os.path.join(path,filename)
    wb = openpyxl.load_workbook(fname,data_only=True,read_only=True)
    ws = wb['2020']
    maxrow = ws.max_row

    byprices = {}
    dunshu_total = 0

    gongyingshang = '驻马店白云纸业有限公司'

    #获取列号
    qijian_num = column_index_from_string('E')
    pricename_num  = column_index_from_string('Q')
    price_num = column_index_from_string('K')
    lingshu_num = column_index_from_string('H')
    dunshu_num = column_index_from_string('J')
    jiner_num = column_index_from_string('L')
    jizhang_num = column_index_from_string('N')

    for row in range(2,maxrow+1):
        danqian_qijian = ws.cell(row,qijian_num).value
        pricename = ws.cell(row,pricename_num).value
        price = ws.cell(row, price_num).value
        lingshu = ws.cell(row, lingshu_num).value
        dunshu = ws.cell(row, dunshu_num).value
        jiner = ws.cell(row, jiner_num).value
        jizhang = ws.cell(row, jizhang_num).value

        byprices.setdefault(pricename, {})
        byprices[pricename].setdefault(price, {'lingshu':0, 'dunshu': 0, 'jiner':0})

        #期间、令数不为None、吨数不为None、pricename不为None
        if (danqian_qijian == qijian) and dunshu not in ['',None] and jizhang not in  ['',None]:
           if '卷筒纸' in pricename :
                lingshu = 0
                byprices[pricename][price]['lingshu'] += lingshu
                byprices[pricename][price]['dunshu'] += dunshu
                byprices[pricename][price]['jiner'] += jiner
                dunshu_total+=dunshu

           else :
                byprices[pricename][price]['lingshu'] += lingshu
                byprices[pricename][price]['dunshu'] += dunshu
                byprices[pricename][price]['jiner'] += jiner
                dunshu_total += dunshu

        else :
            continue

    wb.close()
    return byprices

def main():
    qijian = input('请输入期间，如2020-04\n')
    byprices = bypriceDic(qijian)
    print(byprices)

if __name__ == '__main__':
    main()







