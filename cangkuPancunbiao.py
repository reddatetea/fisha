import excelmessage
import os
import easygui
import openpyxl
from openpyxl.styles import Font,Border,Side,Fill,Alignment

fname = easygui.fileopenbox(msg='请点选引出的原材料盘存表')
path,filename = os.path.split(fname)
os.chdir(path)
fname = excelmessage.excelMessage(fname)
wb = openpyxl.load_workbook(fname)
ws = wb.active
ws.title = 'AAA'
name = easygui.enterbox(msg='请输入盘存表期间"2022年1月"')
nws_name = '{}盘存表'.format(name)
nws = wb.create_sheet(title=nws_name)
names=['类别', '品名', '单位', '月初','入库数量', '采购入库','半成品入库','出库数量','生产领料','结存']
for index,row in enumerate(ws.values):
    if index==0:
        nws.append(names)
    else :
        if row[3]==0 and row[4]==0 and row[5]==0 and row[6]==0 and row[7]==0 and row[8]==0 and row[9]==0 :
            continue
        else :
            hang = [row[0],row[1],row[2],row[3],row[8],row[4],row[9],row[5],row[7],row[6]]
            nws.append(hang)
filename1 = '{}原材料盘存表.xlsx'.format(name)
fname1 = os.path.join(path,filename1)
wb.save(fname)

def printseting(fname,nws_name):
    #设置字体、边框、对齐等常量
    font = Font(name='宋体', size=10)
    font_xiaoji = Font(name='宋体', bold=True, size=10)
    font_firstrow = Font(name='宋体', bold=True, size=20)
    thin_danbian = Side(border_style='thin')
    double_danbian = Side(border_style='double')
    thin_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=thin_danbian)
    double_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=double_danbian)
    wb = openpyxl.load_workbook(fname)
    # for ws in wb.worksheets :
    ws = wb[nws_name]
    # 单元格宽度
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 4
    for j in 'DEFGHI':
        ws.column_dimensions['{}'.format(j)].width = 7.44
    ws.column_dimensions['J'].width = 9.44
    # 插入第一行
    ws.insert_rows(1)
    ws['A1'].value = nws_name
    ws['A1'].font = font_firstrow
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:J1')
    max_row = ws.max_row
    max_column = ws.max_column
    for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
        if row == 2:
            for col in range(1, max_column + 1):
                ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        else:
            for col in range(1, max_column + 1):
                if col <= 3:
                    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                else:
                    ws.cell(row, col).alignment = Alignment(horizontal='right', vertical='center')

    for cell in ws['A']:
        row = cell.row
        if row == 1:
            continue
        else:
            if '小计' in cell.value:
                for col in range(1, max_column + 1):
                    ws.cell(row, col).font = font_xiaoji
                    ws.cell(row, col).border = double_bian
                    ws.merge_cells('A{}:B{}'.format(row, row))
            else:
                for col in range(1, max_column + 1):
                    ws.cell(row, col).font = font
                    ws.cell(row, col).border = thin_bian
    ws.freeze_panes = ws['D3']
    ws.print_title_rows = '1:2'  # the first row
    # 页脚设置
    ws.oddFooter.center.text = " &[Page] / &N"  # 1/n
    ws.oddFooter.center.size = 12  # 页脚中字体大小
    ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
    ws.oddFooter.center.color = "000000"  # 页脚中字体颜色
    ws.oddFooter.right.text = "张文伟 &[Date]"  # 页脚右 文字
    ws.oddFooter.right.size = 12  # 页脚右 字体大小
    ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
    ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
    ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.1, left=0.51, right=0.1, footer=0.5,
                                                       bottom=1)
    wb.save(fname)
printseting(fname,nws_name)
os.rename(fname,fname1)
os.startfile(fname1)


