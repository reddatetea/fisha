import os
import openpyxl
import xlwings as xw
import excelmessage06
import easygui

def     wsPrint(fname,ws_name):
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(fname)
    ws = wb.sheets[ws_name]
    ws.api.PrintOut()
    wb.close()
    app.quit()

def     main():
    fname = excelmessage06.excelMessage()
    path,filename = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选要打印的excel工作表'
    if len(sheetnames) == 1 :
        ws = wb.active
    else :
        print(msg)
        choice = easygui.buttonbox(msg,title='打印excel工作表',choices=sheetnames)
        ws = wb['%s'%choice]

    ws_name = ws.title
    wb.close()
    wsPrint(fname, ws_name)


if __name__ == '__main__':
    main()

