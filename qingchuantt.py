'''
从tt工作表相关数据区域复制数据到'Reconciliation.xlsx中 '
适用22111,21020,12090
'''

import pandas as pd
import easygui
import os
import excelmessage
import openpyxl
from openpyxl.styles import Alignment
import re

def getMingxiDic():
    easygui.msgbox(msg='请复制"tt+-2023.xls"中的相关数据区域(注意:"上面多选一行")后,按"OK"按钮', title=' 请在数据区域上面多选一行', ok_button='OK')
    df = pd.read_clipboard()
    df.columns = ['Transaction Reference', 'Transaction Date', 'Journal No.', 'Debits/Credits', 'Base Amount',
                  'Description', 'mingxi']
    df['Transaction Date'] = df['Transaction Date'].astype('datetime64[ns]')
    gp = df.groupby('mingxi', sort=['Transaction Date', 'Journal No.'])
    mingxi_numTotal = {}
    mingxi_dic = {}
    for index, row in gp:
        num_total = len(row['Transaction Date'])
        mingxi_numTotal[index] = num_total
        a = row['Transaction Reference'].to_list()
        b = row['Transaction Date'].to_list()
        c = row['Journal No.'].to_list()
        d = row['Debits/Credits'].to_list()
        e = row['Base Amount'].to_list()
        e8 = []
        for i in e:
            j = float(i.replace(',', '').replace('(', '').replace(')', ''))
            e8.append(j)
        f = row['Description'].to_list()
        lst = list(zip(a, b, c, d, e8, f))
        mingxi_dic[index] = lst
    return mingxi_dic, mingxi_numTotal

def getJianQuanDic():
    fname = easygui.fileopenbox('请点选"22111对照表.xlsx"(或"21020对照表.xlsx"或"12090对照表.xlsx"等)')
    path, filename = os.path.split(fname)
    os.chdir(path)
    regax = r'(\d+)\w+'
    pattern = re.compile(regax)
    mat = pattern.search(filename)
    kemu = str(mat.group(1))
    df_duzhaobiao = pd.read_excel(fname)
    jian_quan_dic = dict(zip(df_duzhaobiao['简称'], df_duzhaobiao['全称']))
    quan_jian_dic = dict(zip(df_duzhaobiao['全称'], df_duzhaobiao['简称']))
    return jian_quan_dic, quan_jian_dic, kemu

def getlst(jian_quan_dic, kemu):
    msg = '请点选"Reconciliation Mar,2023 .xls"'
    file = easygui.fileopenbox(msg)
    path, fname0 = os.path.split(file)
    os.chdir(path)
    fname1 = excelmessage.excelMessage(file)
    wb2 = openpyxl.load_workbook(fname1)
    sheetnames = [i for i in wb2.sheetnames if i.startswith(kemu)]
    if len(sheetnames) == 1:
        sheetname = sheetnames[0]
    else:
        sheetname = easygui.choicebox(f'请点选"{kemu}"工作表', choices=sheetnames)
    ws2 = wb2[sheetname]
    Description_lst = [i.value for i in ws2['D']]
    quanchen_rows = {}
    for row, i in enumerate(Description_lst):
        if i in [None, '']:
            continue
        else:
            for j in jian_quan_dic.values():
                if str(i).startswith(j) == True:
                    quanchen_rows[j] = row
    quanchen_start = []
    for quanchen, start in quanchen_rows.items():
        quanchen_start.append([quanchen, start])
    quanchen_start.sort(key=lambda x: x[1])
    quanchen_ends = {}
    for i in quanchen_start:
        quanchen, start = i
        for j in range(start, len(Description_lst)):
            if Description_lst[j] in [None, '']:
                break
            else:
                quanchen_ends[quanchen] = j
    start_quanchen = quanchen_start.copy()
    start_quanchen.sort(key=lambda x: x[1], reverse=True)
    return fname1, wb2, ws2, start_quanchen, quanchen_ends,sheetname

def main():
    mingxi_dic, mingxi_numTotal = getMingxiDic()
    jian_quan_dic, quan_jian_dic, kemu = getJianQuanDic()
    fname1, wb2, ws2, start_quanchen, quanchen_ends, sheetname = getlst(jian_quan_dic, kemu)
    alignment_center = Alignment(horizontal='center', vertical='center')
    for i in start_quanchen:
        quanchen, _ = i
        jianchen = quan_jian_dic[quanchen]
        insert_rows = mingxi_numTotal.get(jianchen, 0)
        if insert_rows == 0:
            continue
        else:
            end_row = quanchen_ends.get(quanchen, 0)
            data0 = mingxi_dic[jianchen]
            data = []
            for m in data0:
                row = []
                for x in m:
                    row.append(x)
                data.append(row)
            data.sort(key=lambda x: [x[1], x[0]], reverse=True)
            j = 0
            for i in data:
                data_len = len(data)
                t = i[1].replace(day=1)
                end_day = t + pd.offsets.MonthEnd()
                date = end_day.strftime('%m/%d/%y')
                Journal_Ref = i[0]
                descrition = i[-1]
                if i[3] == 'C':
                    debit = ''
                    credit = i[4]
                else:
                    credit = ''
                    debit = i[4]
                ws2.insert_rows(end_row + 2, 1)
                if j == data_len - 1:
                    ws2['B%s' % (end_row + 2)].value = date
                    ws2['B%s' % (end_row + 2)].alignment = alignment_center
                    ws2['C%s' % (end_row + 2)].value = Journal_Ref
                    ws2['C%s' % (end_row + 2)].alignment = alignment_center
                    if kemu != '12090':
                        ws2['D%s' % (end_row + 2)].value = descrition
                        ws2['E%s' % (end_row + 2)].value = debit
                        ws2['F%s' % (end_row + 2)].value = credit

                    else:
                        ws2['F%s' % (end_row + 2)].value = debit
                        ws2['G%s' % (end_row + 2)].value = credit
                        ws2['I%s' % (end_row + 2)].value = descrition
                else:
                    ws2['C%s' % (end_row + 2)].value = Journal_Ref
                    ws2['C%s' % (end_row + 2)].alignment = alignment_center
                    ws2['D%s' % (end_row + 2)].value = descrition
                    if kemu != '12090':
                        ws2['D%s' % (end_row + 2)].value = descrition
                        ws2['E%s' % (end_row + 2)].value = debit
                        ws2['F%s' % (end_row + 2)].value = credit
                    else:
                        ws2['F%s' % (end_row + 2)].value = debit
                        ws2['G%s' % (end_row + 2)].value = credit
                        ws2['I%s' % (end_row + 2)].value = descrition
                j = j + 1
    wb2.save(fname1)
    wb2 = openpyxl.load_workbook(fname1)
    ws2 = wb2[sheetname]
    totalRow = [i.value for i in ws2['B']].index('TOTAL')
    if kemu != '12090':
        ws2[f'E{totalRow + 1}'] = f'=SUM(E10:E{totalRow})'
        ws2[f'F{totalRow + 1}'] = f'=SUM(F10:F{totalRow})'
        ws2[f'G{totalRow + 1}'] = f'=SUM(F10:F{totalRow})-SUM(E10:E{totalRow})'
    else :
        ws2[f'F{totalRow + 1}'] = f'=SUM(F10:F{totalRow})'
        ws2[f'G{totalRow + 1}'] = f'=SUM(G10:G{totalRow})'
        ws2[f'H{totalRow + 1}'] = f'=SUM(F10:F{totalRow})-SUM(G10:G{totalRow})'
    wb2.save(fname1)
    os.startfile(fname1)


if __name__ == '__main__':
    main()

