import openpyxl
import os
import easygui

def chaxunAdd(gongyingshang_dic,dic):
    new_gongyingshang = easygui.enterbox('请输入新供应商名字:')
    if new_gongyingshang not in gongyingshang_dic:
        easygui.msgbox('该供应商不在原供应商列表中')
        msg = '请输入"{}"的简称'.format(new_gongyingshang)
        new_jianchen = easygui.enterbox(msg)
        dic[new_gongyingshang] = new_jianchen
    else:
        easygui.msgbox('该供应商已在原供应商列表中,不需要添加')
    return dic

def Gongyingshang(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    maxrows = ws.max_row
    gongyingshang_dic = {}
    for row in range(2, maxrows + 1):
        quchen = ws.cell(row, 1).value
        jianchen = ws.cell(row, 2).value
        gongyingshang_dic[quchen] = jianchen
    wb.close()
    return gongyingshang_dic

def main():
    path = r'F:\a00nutstore\006\zw\price'
    os.chdir(path)
    fname = easygui.fileopenbox(msg='请点选该类材料的供应商名称文件')
    # fname = r'F:\a00nutstore\006\zw\price\纸箱供应商.xlsx'
    gongyingshang_dic = Gongyingshang(fname)
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    jieshuchaxun = True
    dic = {}
    while jieshuchaxun:
        choice = easygui.ccbox('查询和新增纸箱供应商', title='继续查询和添加or退出', choices=('yes', 'no'))
        if choice == True:
            dic = chaxunAdd(gongyingshang_dic,dic)
        else:
            jieshuchaxun = False
    for key,value in dic.items():
        ws.append((key,value))
    wb.save(fname)
    os.startfile(fname)

if __name__=='__main__':
    main()




