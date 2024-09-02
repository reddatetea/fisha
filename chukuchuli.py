'''
chukuchuli.py
将原材料和数量金额汇总后的结果导入，在当月出库中增加一个出库参考工作表，并设公式，免去复制粘贴
本版增加铁丝的处理
本版增加未达的处理(白云、纸、辅料）
'''

# _*_ conding:utf-8 _*_
import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import yuancailiaot
import dangyueweida

def chukuChuli(fname,newname,weida_dic):

    wb1 = openpyxl.load_workbook(fname)
    #ws1 = wb1.create_sheet('出库参考')
    ws1 = wb1['出库参考']
    #取得列号
    chayi_num = column_index_from_string('AE')
    weida_num = column_index_from_string('AF')
    shijichayi_num = column_index_from_string('AG')
    chuku_num = column_index_from_string('AH')
    danjia_num = column_index_from_string('AI')
    jiner_num = column_index_from_string('AJ')

    #科目所在列
    kemumingchen_num = column_index_from_string('R')


    wb2 = openpyxl.load_workbook(newname)
    ws2 = wb2.active
    maxrow2 = ws2.max_row


    for index,row in  enumerate(ws2.values):
        if index == 0:
            ws1.append(row+('差异','未达项','实际差异','出库数量','单价','金额'))
        else :
            ws1.append(row)

    for row in range(2,maxrow2-1):
        kemumingchen = ws1.cell(row,kemumingchen_num).value
        weida_dic.setdefault(kemumingchen,{'lingshu':0,'dunshu':0})

        if ws1.cell(row,kemumingchen_num).value == '铁丝':
            ws1.cell(row, chayi_num, value='=X' + str(row)  + '*2'+ '+Y' + str(row)+'*2' + '-T' + str(row)+'-I' + str(row))
            ws1.cell(row, chuku_num, value='=AB' + str(row)+'*2')

        else :
            ws1.cell(row, chayi_num, value='=X' + str(row) + '+Y' + str(row) + '-T' + str(row) + '-I' + str(row))
            ws1.cell(row, chuku_num, value='=AB' + str(row))

        if '卷筒纸' in kemumingchen :
            ws1.cell(row,weida_num,value =weida_dic[kemumingchen]['dunshu'])
        else :
            ws1.cell(row, weida_num, value=weida_dic[kemumingchen]['lingshu'])

        ws1.cell(row,shijichayi_num,value ='=AE'+str(row)+'+AF'+str(row))

        ws1.cell(row,danjia_num,value ='=S'+str(row))
        ws1.cell(row,jiner_num,value='=IF(T'+str(row)+'+I'+str(row)+'-AH'+str(row)+'=0,U'+str(row)+'+J'+str(row)+',ROUND(AH'+str(row)+'*AI'+str(row)+',2))')

    wb1.save(fname)
    wb2.close()

def main():
    #newname = addqitaheji.addqitaHeji(kemuname_dic,fname)
    newname = r'F:\a00nutstore\006\zw\2020\202004\数量金额盘存表18.xlsx'
    path = input('请输入本月出库路径：\n')
    os.chdir(path)
    filename = input('请输入本月出库文件名：\n')
    fname2 = os.path.join(path, filename)

    print('计算当月未达，即财务做了入库记账，而仓库算作下个期间')
    qijian = input('请输入当前期间(格式如2020-04)：')
    xia_qijian = qijian[:-2] + '%02d' % (int(qijian[-2:]) + 1)
    fname = dangyueweida.baiyunweida(xia_qijian)
    dangyueweida.zhiweida(xia_qijian, fname)
    dangyueweida.weida(fname)
    weida_dic = dangyueweida.weidaDic(fname)



    chukuChuli(fname=fname2,newname=newname,weida_dic=weida_dic)



if __name__ == '__main__':
    main()





