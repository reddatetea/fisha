'''
尝试根据纸入库制作纸价格列表，看是不是快些！
本版增加传入期间参数处理,取当期
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl

def zhipriceDic(qijian):
    path = r'd:\a00nutstore\006\zw\else'
    os.chdir(path)
    filename = r'2020入库.xlsx'
    fname = os.path.join(path,filename)
    wb = openpyxl.load_workbook(fname,data_only=True,read_only=True)
    ws = wb['入库']
    maxrow = ws.max_row
    zhiprices = []
    gongying_price = []
    ling_dun_jiner = []
    for row in range(2,maxrow+1):
        dangqian_qijian = ws['C{}'.format(row)].value
        jizhang = ws['V{}'.format(row)].value
        if (dangqian_qijian == qijian) and (jizhang not in ['',None]):
            gongyingshang = ws['B{}'.format(row)].value
            pricename = ws['X{}'.format(row)].value
            price = ws['O{}'.format(row)].value
            lingshu = ws['G{}'.format(row)].value
            dunshu = ws['M{}'.format(row)].value
            jiner = ws['R{}'.format(row)].value
            if lingshu in ['', None]:
                lingshu = 0
            if dunshu in ['', None]:
                dunshu = 0
            if jiner in ['', None]:
                jiner = 0
            if [gongyingshang,pricename,price] not in gongying_price :
                gongying_price.append([gongyingshang,pricename,price])
                ling_dun_jiner.append([lingshu,dunshu,jiner])
                c = [gongyingshang,pricename,price],[lingshu,dunshu,jiner]
                zhiprices.append(c)
            else :
                k = [zhiprice[0]  for zhiprice in zhiprices]
                print(k)
                j = k.index([gongyingshang,pricename,price])
                zhiprices[j][1][0]=zhiprices[j][1][0]+lingshu
                zhiprices[j][1][1]=zhiprices[j][1][1]+dunshu
                zhiprices[j][1][2]=zhiprices[j][1][2]+jiner
    print(zhiprices)
    wb.close()
    return zhiprices

def main():
    qijian = input('请输入期间，如2020-04\n')
    zhiprices = zhipriceDic(qijian)

if __name__ == '__main__':
    main()







