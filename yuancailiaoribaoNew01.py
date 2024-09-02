import os
import openpyxl
from openpyxl.styles import Font, Border, Side, Fill, Alignment
import excelmessage
import easygui
import datetime


def getNoTotal(ws):
    df = list(ws.values)
    df1 = [i for i in df if '合   计' not in i[0]]  # AAA表中删除有累计和合计的行
    return df1


def quchu(fname, riqi, ku):
    path, filename = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    # "当日"工作表 ws1
    if 'AAA' in sheetnames:
        ws1 = wb['AAA']
    else:
        msg = '请点选"当日"工作表'
        choice = easygui.choicebox(msg, choices=sheetnames)
        ws1 = wb[choice]
    # "累计"工作表 ws2
    if 'BBB' in sheetnames:
        ws2 = wb['BBB']
    else:
        msg = '请点选"累计"工作表'
        choice = easygui.choicebox(msg, choices=sheetnames)
        ws2 = wb[choice]
    dfa1 = getNoTotal(ws1)
    dfb1 = getNoTotal(ws2)
    if ku == '1库':
        wb.close()
        return dfa1, dfb1
    else:

        # 长期库存"当日"工作表 ws1
        if 'CCC' in sheetnames:
            ws11 = wb['CCC']
        else:
            msg = '请点选长期库存"当日"工作表'
            choice = easygui.choicebox(msg, choices=sheetnames)
            ws11 = wb[choice]
        # "累计"工作表 ws2
        if 'DDD' in sheetnames:
            ws22 = wb['DDD']
        else:
            msg = '请点选长期库存"累计"工作表'
            choice = easygui.choicebox(msg, choices=sheetnames)
            ws22 = wb[choice]
        dfa11 = getNoTotal(ws11)
        dfb22 = getNoTotal(ws22)
        wb.close()
        return dfa1, dfb1, dfa11, dfb22


def leibiePinmingDic(dfa):
    # 生成品名的当日字典
    yesterday = [i[3] for i in dfa]
    num_in = [i[4] for i in dfa]
    num_out = [i[5] for i in dfa]
    category = [i[0] for i in dfa]
    name = [i[1] for i in dfa]
    category_name = zip(category, name)
    yesterday_nums = zip(yesterday, num_in, num_out)
    category_name_yesterday_nums = dict(zip(category_name, yesterday_nums))
    return category_name_yesterday_nums


def baobiao(category_name_yesterday_nums, dfb):
    dfb2 = []
    for i in dfb:
        category_name = (i[0], i[1])
        yesterday = category_name_yesterday_nums.get(category_name, [0, 0, 0])[0]
        num_in = category_name_yesterday_nums.get(category_name, [0, 0, 0])[1]
        num_out = category_name_yesterday_nums.get(category_name, [0, 0, 0])[2]
        newrow = list(i[:4]) + [yesterday, num_in, i[4], num_out, i[5], i[6]]
        dfb2.append(newrow)
        df_zero = [i for i in dfb2 if (i[-1] == 0) and ('小计' not in i[0])]
        dfb_noZero = [i for i in dfb2 if (i[-1] != 0) or (('小计' in i[0]) and (i[-1] == 0))]  # 不删小计
    return df_zero, dfb_noZero


def getTotal(df):
    begin = sum([i[3] for i in df if '小计' in i[0]])
    yesterday = sum([i[4] for i in df if '小计' in i[0]])
    num_in = sum([i[5] for i in df if '小计' in i[0]])
    total_in = sum([i[6] for i in df if '小计' in i[0]])
    num_out = sum([i[7] for i in df if '小计' in i[0]])
    total_out = sum([i[8] for i in df if '小计' in i[0]])
    end = sum([i[9] for i in df if '小计' in i[0]])
    total = [begin, yesterday, num_in, total_in, num_out, total_out, end]
    return total


def addLastrow(df, total):
    last_row = df[-1].copy()
    last_row[0] = '合 计'
    last_row[1:3] = ['', '']
    last_row[3:] = total
    df.append(last_row)
    return df


def lingjeicun(df_zero, wb3):
    wb3.create_sheet(title='零结存')
    ws4 = wb3['零结存']
    df_zero
    for index, i in enumerate(df_zero):
        if index == 0:
            ws4.append(['存货大类名称', '品名', '单位', '月初', '上日', '本日入库', '入库累计', '本日出库', '出库累计', '结余'])
        else:
            if i[3] == 0 and i[4] == 0 and i[5] == 0 and i[6] == 0 and i[7] == 0 and i[8] == 0 and i[9] == 0:
                continue
            else:
                ws4.append(i)

    return ws4


def ribao(df, ws):
    for index, i in enumerate(df):
        if index == 0:
            ws.append(['存货大类名称', '品名', '单位', '月初', '上日', '本日入库', '入库累计', '本日出库', '出库累计', '结余'])
        else:
            ws.append(i)


def getRibao(dfa, dfb):
    category_name_yesterday_nums = leibiePinmingDic(dfa)
    df_zero, dfb_noZero = baobiao(category_name_yesterday_nums, dfb)
    total = getTotal(dfb_noZero)
    dfb_noZero = addLastrow(dfb_noZero, total)
    return df_zero, dfb_noZero


def printseting(fname1, riqi):
    # 设置字体、边框、对齐等常量
    font = Font(name='宋体', size=10)
    font_xiaoji = Font(name='宋体', bold=True, size=10)
    font_firstrow = Font(name='宋体', bold=True, size=20)
    thin_danbian = Side(border_style='thin')
    double_danbian = Side(border_style='double')
    thin_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=thin_danbian)
    double_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=double_danbian)
    wb = openpyxl.load_workbook(fname1)
    for ws in wb.worksheets:
        try :
            # 单元格宽度
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 28
            ws.column_dimensions['C'].width = 4
            for j in 'DFGHI':
                ws.column_dimensions['{}'.format(j)].width = 7.44
            ws.column_dimensions['J'].width = 9.44
            # 插入第一行
            ws.insert_rows(1)
            ws['A1'].value = '双佳' + ws.title + ' ' + riqi
            ws['A1'].font = font_firstrow
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A1:J1')
            max_row = ws.max_row
            max_column = ws.max_column
            for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
                if row == 2:
                    for col in range(1, max_column + 1):
                        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                else:
                    for col in range(1, max_column + 1):
                        if col <= 3:
                            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            ws.cell(row, col).alignment = Alignment(horizontal='right', vertical='center')

            for cell in ws['A']:
                row = cell.row
                if row == 1:
                    continue
                else:
                    if '小计' in cell.value:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).font = font_xiaoji
                            ws.cell(row, col).border = double_bian
                            ws.merge_cells('A{}:B{}'.format(row, row))
                    else:
                        for col in range(1, max_column + 1):
                            ws.cell(row, col).font = font
                            ws.cell(row, col).border = thin_bian
            ws.freeze_panes = ws['D3']
            ws.print_title_rows = '1:2'  # the first row
            # 页脚设置
            ws.oddFooter.center.text = " &[Page] / &N"  # 1/n
            ws.oddFooter.center.size = 12  # 页脚中字体大小
            ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
            ws.oddFooter.center.color = "000000"  # 页脚中字体颜色
            ws.oddFooter.right.text = "张文伟 &[Date]"  # 页脚右 文字
            ws.oddFooter.right.size = 12  # 页脚右 字体大小
            ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
            ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
            ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.1, left=0.51, right=0.1,
                                                                  footer=0.5,
                                                                  bottom=1)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 0:
                        cell.value = ''
                    else:
                        continue
        except Exception:
            pass


    wb.save(fname1)


def main():
    msg = '请点选"原材料当日和累计"工作表'
    easygui.msgbox(msg=msg)
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    today_date = datetime.date.today()
    yesterday_date = (today_date + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")
    msg = '报表日期是{}?'.format(yesterday_date)
    choice = easygui.ccbox(msg, title='请选择"是"或"否"', choices=('是', '否'))
    if choice:
        riqi = yesterday_date
    else:
        msg = '请输入报表日期'
        riqi = easygui.enterbox(msg, title=" 昨天日期")

    # 新建原材料日报表工作簿
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = '原材料日报表'
    fname1 = '原材料日报表{}.xlsx'.format(riqi)

    ku = easygui.buttonbox("请点选'1库'或'2库'", "如果需要单独计算长期库存，请选择'2库'", ["1库", "2库"])
    if ku == "2库":
        dfa1, dfb1, dfa11, dfb22 = quchu(fname, riqi, ku)
        df_zero, dfb_noZero = getRibao(dfa1, dfb1)
        ribao(dfb_noZero, ws3)
        ws4 = lingjeicun(df_zero, wb3)

        ws5 = wb3.create_sheet('长期库存表')
        df_zero, dfb_noZero = getRibao(dfa11, dfb22)
        ribao(dfb_noZero, ws5)
    else:
        dfa1, dfb1 = quchu(fname, riqi, ku)
        df_zero, dfb_noZero = getRibao(dfa1, dfb1)
        ribao(dfb_noZero, ws3)
        ws4 = lingjeicun(df_zero, wb3)

    wb3.save(fname1)
    printseting(fname1, riqi)
    easygui.msgbox(msg='程序结束运行')
    os.startfile(fname1)


if __name__ == '__main__':
    main()
