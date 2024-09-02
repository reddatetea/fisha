'''
从粘贴板将内容导入，快速设置，并直接打印
'''
import openpyxl
import  excelseting
import pyperclip
import excelseting
import addshuziqianfenfu
import os
import excelprint
import sys

def getMessage():
    if len(sys.argv) > 1:
        msg = ''.join(sys.argv[1:])
    else:
        msg = pyperclip.paste()

    return msg


def rukuRange(paste_content,fname,ws_name):
    cell_range =paste_content
    content = cell_range.split("\n")
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    ws.delete_rows(1,ws.max_row)
    # wb.save(fname)
    shuju0 = []
    for i in content:
        j = i.split('\t')
        k = [x.strip() for x in j]
        shuju0.append(k)


    # wb = openpyxl.load_workbook(fname)
    # ws = wb['入库']
    for index,row in enumerate(shuju0):
        ws.append(row)
    wb.save(fname)


def main():
    fname = r'普通打印.xlsx'
    ws_name = '打印'
    paste_content =  getMessage()
    rukuRange(paste_content,fname,ws_name)
    os.startfile(fname, 'print')                         #打开excel并打印

if  __name__ == '__main__':
    main()
