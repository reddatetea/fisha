from openpyxl import Workbook,load_workbook
import excelmessage
import win32com.client as win32
import cangkutotal
import os

def chaiwuTotal():

    # 打开双佳数量表，先将xls转化为xlsx
    print('找开双佳数量xls')
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)

    dirname,filename = os.path.split(fname)

    wb = load_workbook('%s' % fname)
    sheet = wb.active
    sheet.title = '双佳数量'
    # sheet = wb.get_sheet_by_name('双佳数量')

    #科目字典，放在这里可以少一个模块
    workbook1 = load_workbook(fname)
    worksheet1 = workbook1.active
    maxrows = worksheet1.max_row
    kemuname = []
    kemubianma = []
    # kemuname_dic ={}
    for i in range(2, maxrows - 2 + 1):
        kemuname.append(worksheet1.cell(row=i, column=4).value.strip())
    print(kemuname)

    for j in range(2, maxrows - 2 + 1):
        kemubianma.append(worksheet1.cell(row=j, column=3).value)
    print(kemubianma)

    kemuname_dic = dict(zip(kemuname, kemubianma))
    print(kemuname_dic)

    jishu = 0
    for index,row in enumerate(sheet.values):
        #print((row[0]))
        if row[0] == None:
            jishu = jishu + 1
        else:
            continue


    mrows = sheet.max_row -jishu
    print(mrows)



    kemuList0 = []
    kemuList = []
    danjiaList = []
    qichushuliangList = []
    qichujinerList = []
    qimoshuliangList = []
    qimojinerList = []

    for index,row in enumerate(sheet.values):
        if index == 0:
            continue



        else :
            if index < mrows:
                #kemu0 = row[2].strip()
                kemuList0.append(row[3])

                #数量为0时，计算单价总是报错，用try except搞定
                try:
                    danjia = float(row[14]) / float(row[13])
                except:
                    danjia = 0

                danjiaList.append(danjia)


                if row[5] == '借':
                    qichushuliang = float(row[6])
                else:
                    qichushuliang = -1*float(row[6])
                qichushuliangList.append(qichushuliang)

                if row[5] == '借':
                    qichujiner = float(row[7])
                else:
                    qichujiner = -1*float(row[7])
                qichujinerList.append(qichujiner)

                if row[12] == '借':
                    qimoshuliang = float(row[13])
                else:
                    qimoshuliang = -1 * float(row[13])
                qimoshuliangList.append(qimoshuliang)

                if row[12] == '借':
                    qimojiner = float(row[14])
                else:
                    qimojiner = -1 * float(row[14])
                qimojinerList.append(qimojiner)

    for kemu0 in kemuList0 :
        print(kemu0)
        #kemu0.strip()
        #kemuList.append(kemu0)

    print(kemuList0)
    print(danjiaList)
    print(qichushuliangList)
    print(qichujinerList)
    print(qimoshuliangList)
    print(qimojinerList)

    temp = []
    kemuList = []
    for i in kemuList0 :
        if i != None:
            temp.append(i)
        else :
            temp.append('')

    for i in temp :
        i = i.strip()
        kemuList.append(i)

    print(kemuList)
    print(len(kemuList))
    print(len(kemuList0))
    print(len(danjiaList))



    #数据去向工作簿 Workbook
    ss_wb = Workbook()
    ss_sheet = ss_wb.active
    ss_sheet.title = '数量金额汇总'

    #新增抬头
    new_taitou = ('科目','单价','期初数量','期初金额','期末数量','期末金额')
    i =0
    row  = 0


    for index,row in enumerate(sheet.values):
        if index == 0 :
            ss_sheet.append(row + new_taitou)
        else:
            if index <= mrows:
                i = index -1



                #ss_sheet.append(row + (kemuList[i],danjiaList[i],qichushuliangList[i],qichujinerList[i],qimoshuliangList[i],qimojinerList[i],qichu,ruku,caigou,banchengpin,chuku,lingliao,qimo))
                ss_sheet.append(row + (kemuList[i],danjiaList[i],qichushuliangList[i],qichujinerList[i],qimoshuliangList[i],qimojinerList[i]))

    filename =  '数量金额汇总.xlsx'
    fname2 = os.path.join(dirname,filename)

    ss_wb.save(fname2)
    return kemuList,kemuname_dic,fname2

def main():
    chaiwuTotal()

if __name__== '__main__':
    main()

