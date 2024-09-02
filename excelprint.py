import os
import openpyxl
import xlwings as xw
import excelmessage
import easygui


def  wsPrint(fname,ws_name):
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(fname)
    ws = wb.sheets[ws_name]
    ws.api.PrintOut()
    wb.close()
    app.quit()

def wsPrints(fname,sheetnames):
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(fname)
    for sheetname in sheetnames:
        ws = wb.sheets[sheetname]
        ws.api.PrintOut()
    wb.close()
    app.quit()



def  main():
    # fname = excelmessage.wenjian()
    # fname = excelmessage.excelMessage(fname)
    # path,filename = os.path.split(fname)
    # os.chdir(path)
    # wb = openpyxl.load_workbook(fname)
    # sheetnames = wb.sheetnames
    # msg = '请点选要打印的excel工作表'
    # if len(sheetnames) == 1 :
    #     ws = wb.active
    # else :
    #     print(msg)
    #     choice = easygui.buttonbox(msg,title='打印excel工作表',choices=sheetnames)
    #     ws = wb['%s'%choice]
    #
    # ws_name = ws.title
    # wb.close()
    # wsPrint(fname, ws_name)
    fname = r'D:\a00nutstore\fishc\材料入库单.xlsx'
    gongyingshangs = ['彩皇','长利五金厂']
    wsPrints(fname,gongyingshangs)


if __name__ == '__main__':
    main()

