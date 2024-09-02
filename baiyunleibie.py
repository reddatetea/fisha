# _*_ conding = utf-8 _*_
'''
制作白云类别字典
'''

import  openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import os
import easygui

def baiyunLeibie():

    fname = r'F:\a00nutstore\006\zw\baiyun\2020白云入库.xlsx'
    path,filename = os.path.split(fname)
    os.chdir(path)

    wb = openpyxl.load_workbook(fname)
    ws = wb['2020']

    maxrow =ws.max_row
    maxcolumn = ws.max_column

    key_choice_col = 'Q'
    key_choice_liehao = column_index_from_string(key_choice_col)

    value_choice_col = 'S'
    value_choice_liehao = column_index_from_string(value_choice_col)

    #下面制作字典
    dicnew = {}
    for i in range(1,maxrow+1):
        dicnew_key = ws.cell(row=i,column = key_choice_liehao).value
        dicnew_value = ws.cell(row=i,column = value_choice_liehao).value

        if (dicnew_key is not None) and  (dicnew_value is  not None and dicnew_value!=0) :
            dicnew[dicnew_key] = dicnew_value
        else :
            continue

    wb.close()

    filename1 = key_choice_col +'_'+ value_choice_col+'_dic.xlsx'
    fname1 = os.path.join(path,filename1)
    wb1 = openpyxl.Workbook(fname1)

    ws1 = wb1.create_sheet(title = key_choice_col +'_'+ value_choice_col)
    taitou =(key_choice_col,value_choice_col)
    #ws1.append(taitou)
    for key,value in dicnew.items():
        newhang =(key,value)
        ws1.append(newhang)

    wb1.save(fname1)
    return dicnew


def main():
    dicnew =baiyunLeibie()
    print(dicnew)


if __name__=='__main__':
    main()






