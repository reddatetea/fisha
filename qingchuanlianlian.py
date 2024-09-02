# _*_ conding:utf-8 _*_

import os
import easygui
import openpyxl
import xlsxlsx

def keynames():
    msg = '请输入若干查询关键字'
    print(msg)
    keynames0 = easygui.multenterbox(msg, title='查询关键字',
                                   fields=['关键字1', '关键字2', '关键字3', '关键字4', '关键字5', '关键字6', '关键字7', '关键字8',
                                           '关键字9', '关键字10','关键字11', '关键字12', '关键字13', '关键字14', '关键字15','关键字16', '关键字17', '关键字18',
                                           '关键字19', '关键字20','关键字21', '关键字22', '关键字23', '关键字24', '关键字25'])
    keynames = []
    for j in keynames0:
        j.strip()
        if j != '' and j not in keynames:
            keynames.append(j)
    return keynames

def openXlsXlsx(file,keyname,mingxi):
    if os.path.splitext(file)[1] == '.xls':
        filename = xlsxlsx.xlsXlsx(file)
    else:
        filename = file
    wb = openpyxl.load_workbook(filename, data_only=True)
    for ws in wb.worksheets:
        for index, row in enumerate(ws.values):
            i = list(row)
            if keyname in str(row[1]):
                i.insert(0, ws['A4'].value.strip('()'))
                print(file)
                print(str(row[1]))
                if i not in mingxi:
                    mingxi.append(i)
                else:
                    break
            else:
                continue
    wb.close()
    return mingxi

def search_file(start_dir, target,keyname,mingxi):
    os.chdir(start_dir)
    for each_file in os.listdir(os.curdir):      #获取当前工作目录os.curdir下的文件列表！
        if  os.path.isfile(each_file):
            file_list = []
            ext = os.path.splitext(each_file)[1]     #文件列表每个文件，os.path.splitext分割为路径和文件名,后缀
            if ext in target :
                file_list.append(os.getcwd() + os.sep + each_file + os.linesep)  # 使用os.sep是程序更标准
                file = os.getcwd() + os.sep + each_file
                mingxi = openXlsXlsx(file,keyname,mingxi)
            else :
                continue
        else:
            mingxi = search_file(each_file, target, keyname, mingxi)  # 递归调用
            try :
                os.chdir(os.pardir)  # 递归调用后切记返回上一层目录
            except:
                print('已搜索到电脑的根目录')
    return mingxi
keynames = keynames()
searchfile_path = r'd:\qingchuan'
if not os.path.exists(searchfile_path):
    os.makedirs(searchfile_path)

print(keynames)
fuhao = ':\/?*[]<>|"'
sheetname = keynames[0]
for j in fuhao:
    if j in keynames[0]:
        sheetname = keynames[0].replace(j, '')
    else:
        continue
fname = searchfile_path + os.sep + '%s合计.xlsx' % sheetname
if os.path.exists(fname):          #删除以前的搜索文件
    os.remove(fname)
msg = '请点选要要查找文件所在文件夹'
start_dir = easygui.diropenbox(msg)
target = ['.xls', '.xlsx']
mingxi = []
for keyname in keynames:
    mingxi = search_file(start_dir, target, keyname, mingxi)

if len(mingxi)>=1:
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = sheetname
    taitou = ['日期', '编号', '名称', '库存代码', '部门', '上菜', '价格', '应用改码后价钱', '数量', '总菜式销售额', '折扣', '内含服务费', '净菜式销售额', '销售额百分比', '成本', '总成本', '成本百分比', '菜式百分比']
    ws1.append(taitou)
    for row in mingxi:
        ws1.append(row)
    wb1.save(fname)

os.startfile(fname)






