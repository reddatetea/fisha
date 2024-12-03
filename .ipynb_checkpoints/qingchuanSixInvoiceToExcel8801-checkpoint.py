'''
读取指定文件夹下所有"CREDIT INVOICE"PDF文件，
并将每张INVOICE中的各项银行支付记录导入
指定的"Remittance"excel文件中相应位置

'''

import pdfplumber
import re
import openpyxl
import easygui
import os

# 先删除原始表的内容，只留表头
msg = '请点选"Remittance.xlsx"'
fname = easygui.fileopenbox(msg)
path, file = os.path.split(fname)
os.chdir(path)
wb = openpyxl.load_workbook(fname)
ws = wb.active
start_row = [i.value for i in ws['A']].index('Customer Number') + 1
max_row = ws.max_row
for i in range(max_row + 1, start_row, -1):
    ws.delete_rows(idx=i)

regex_invoice_number = re.compile(r'Tax Invoice #: (?P<invoice_number>P\d{9})')
regex_invoice_date = re.compile(r'Tax Invoice Date: (?P<invoice_date>\d{2}-\w+-\d{4})')
regex_free_type = re.compile(r'(?P<free_type>.*)\bCNY\b .* CNY')
pdf_files = [i for i in os.listdir(path) if i[-3:] == 'pdf']
item_num = 0
for pdf_file in pdf_files:
    pdf = pdfplumber.open(pdf_file)
    for page in pdf.pages:
        invoice = page.extract_text().split('\n')
        item_line = 0
        for line in invoice:
            if line.startswith('Customer'):
                mat = regex_invoice_number.search(line)
                invoice_number = mat.group('invoice_number')

            elif line.startswith('Holidex'):
                mat = regex_invoice_date.search(line)
                invoice_date = mat.group('invoice_date')
                invoice_date = "'".join(invoice_date.split('-')[::-1][1:])
            else:
                continue
        for line in invoice:
            if line.endswith('CNY'):
                if not line.startswith('Amount Due'):
                    if line.split(' ')[-1] == 'CNY':
                        item_line += 1
                        item_num += 1
                        amounts = line.split(' ')
                        amount = float(amounts[-2].replace('(', '-').replace(')', '').replace(',', ''))
                        mat = regex_free_type.search(line)
                        free_type = mat.group('free_type')
                        row = ['684100', 'P6066', invoice_number, str(item_line), invoice_date, free_type, amount, 'CNY',
                               amount, 'CNY', amount, 'CNY', '', '', 0, 0, '', '', '', '', "", '', amount, '']
                        ws.append(row)
ws[f'w{item_num + start_row + 1}'].value = f'=sum(W{start_row + 1}:W{item_num + start_row})'
wb.save(fname)
os.startfile(fname)












