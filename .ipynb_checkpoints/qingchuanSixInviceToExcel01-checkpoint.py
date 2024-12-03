import pdfplumber
import re
import easygui
import shutil
import os
import openpyxl


def writeDangtian(fname):
    pdf = pdfplumber.open(fname)
    # print('开始读取数据')
    filename = 'cash888.txt'
    with open(filename, 'w') as f:
        f.write('')
    with open(filename, 'a') as f:
        for page in pdf.pages:
            content = page.extract_text()
            f.write(content)
    pdf.close()
    return filename

def matchNumber(filename):
    regex = re.compile(r'\d+(?P<name>\D+)\d+:\d+')
    cash_in = {}
    cash_out = {}
    cashier_names = []
    with open(filename, 'r') as f:
        content = f.readlines()
        j = 0
        for j in range(len(content)):
            hang = content[j]
            if 'Cash CNY' in hang or 'Cash Paid Out CNY' in hang:
                mat = regex.search(hang)
                if mat :
                    name = mat.group('name')
                    cashier_names.append(name)
                    print(name)
                else:
                    continue
                shuju=hang.split()
                jishu = float(shuju[-1].replace(',',''))
                print(shuju)
                print(jishu)
                if  'Cash Paid Out CNY' in hang:
                   cash_out[name] = -1*jishu
                   newhang =content[j+2]
                   jishu = float(newhang.split()[-1].replace(',',''))
                   cash_in[name] = jishu

                else :
                    cash_in[name] = jishu
                j = j + 1

        names = [key for key, value in cash_in.items()] + [key for key, value in cash_out.items()]
        jishus = {name: round(cash_in.get(name, 0) + cash_out.get(name, 0), 2) for name in names}
        print(jishus)

        for hang in content:
            if   'POS - Cash CNY' in hang:
                #print(hang)
                pos_cash = float(hang.split()[-1].replace(',', ''))
                print('pos_cash', pos_cash)
            elif   'Check CNY' in hang:
                #print(hang)
                check_cash = float(hang.split()[-1].replace(',', ''))
                print('check_cash', check_cash)
            else :
                continue
        return cashier_names,jishus,pos_cash,check_cash

def writeNumber(cashier_names,jishus,pos_cash,check_cash):
    msg = '请点选"总出纳报告GCR" excel文件'
    print(msg)
    fname = easygui.fileopenbox(msg)
    path, file = os.path.split(fname)
    # 备份以防万一
    backup_path = r'd:\qingchuangbackup'
    backup_fname = os.path.join(backup_path, file)
    if not os.path.exists(backup_path):
        os.makedirs(backup_path)
    if not os.path.exists(backup_fname):
        shutil.copyfile(fname, backup_fname)
    #fname = r'F:\a00nutstore\fishc\qingchuan\总出纳报告GCR-202103.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    print(sheetnames)
    if len(sheetnames)==1:
        ws = wb.active
    else :
        msg = '请点选当日出纳报告工作表'
        print(msg)
        sheetname = easygui.choicebox(msg,choices=sheetnames)
        ws = wb[sheetname]

    first_col = [i.value for i in ws['A']]
    print(first_col)
    sub_hang1 = first_col.index('SUB-TOTAL  小计') + 1                         #22
    print(sub_hang1)
    yuan_names = [j.value  for row in ws['B9:B22']  for j in row if j.value not in [None,'']]   # if  j.value not in [None,'']
    print('yuan_names',yuan_names)
    add_names = list(set(jishus.keys())-set(yuan_names))
    add_names.sort(key=cashier_names.index)
    print('add_names',add_names)

    k = len(yuan_names)
    print('yuan_names',k)
    l = len(add_names)
    print('add_names',l)
    if k == 0:
        for j in range(9, l + 9):
            name = add_names[j - 9]
            ws.cell(j, 2).value = name
            jishu = jishus.get(name, 0)
            ws.cell(j, 4).value = jishu
    else:
        for j in range(9, k + 9):
            name = yuan_names[j - 9]
            ws.cell(j, 2).value = name
            jishu = jishus.get(name, 0)
            ws.cell(j, 4).value = jishu
        for j in range(k + 9, k + l + 9):
            name = add_names[j - 9 - k]
            ws.cell(j, 2).value = name
            jishu = jishus.get(name, 0)
            ws.cell(j, 4).value = jishu

    ws['I9'].value = pos_cash
    ws.cell(sub_hang1+6,4).value = check_cash
    #wb.close()
    wb.save(fname)
    os.startfile(fname)


def main():
    msg = '请点选Cashier Summary.pdf文件'
    print(msg)
    fname0 = easygui.fileopenbox(msg)
    path, file = os.path.split(fname0)
    os.chdir(path)
    #fname0 = r'F:\a00nutstore\fishc\qingchuan\Cashier Summary004_270793.pdf'
    filename = writeDangtian(fname0)
    cashier_names,jushus,pos_cash,check_cash = matchNumber(filename)
    writeNumber(cashier_names,jushus, pos_cash, check_cash)

if __name__ == '__main__':
    main()


