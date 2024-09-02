import pandas as pd
import excelmessage
import os
import easygui
import openpyxl
from openpyxl.styles import Font,Border,Side,Fill,Alignment

#fname = easygui.fileopenbox(msg='请点选引出的原材料盘存表')
fname = r'F:\a00nutstore\fishc\原材料盘存表20220124.xls'
path,filename = os.path.split(fname)
os.chdir(path)
df0 = pd.read_excel(fname)
dic = {'存货大类名称':'类别','上日数量':'月初','其他入库_入库数量':'入库数量 ','入库数量':'采购入库','半成品入库_入库数量':'半成品入库','生产领料_出库数量':'生产领料','本日数量':'结存'}
df0 = df0.rename(dic,axis=1)
df0 = df0.rename({'入库数量 ':'入库数量'},axis=1)
names=['类别', '品名', '单位', '月初','入库数量', '采购入库','半成品入库','出库数量','生产领料','结存']
df0 = df0[names]
df0 = df0[~((df0['月初']==0) & (df0['入库数量']==0) & (df0['采购入库']==0) & (df0['半成品入库']==0) &  (df0['生产领料']==0) & (df0['结存']==0))]
fname = excelmessage.wenjian()
fname = excelmessage.excelMessage(fname)
wb = openpyxl.load_workbook(fname)
name = easygui.enterbox(msg='请输入盘存表期间"2022年1月"')
nws_name = '{}盘存表'.format(name)
filename1 = '{}盘存表.xlsx'.format(name)
fname1 = os.path.join(path,filename1)
nnws = wb.create_sheet(title=nws_name)
wb.save(fname)
with pd.ExcelWriter(fname, engine='openpyxl')  as writer:
    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    df0.to_excel(writer,nws_name,index=False)
writer.save()

def printseting(fname,nws_name):
    #设置字体、边框、对齐等常量
    font = Font(name='宋体', size=10)
    font_xiaoji = Font(name='宋体', bold=True, size=10)
    font_firstrow = Font(name='宋体', bold=True, size=20)
    thin_danbian = Side(border_style='thin')
    double_danbian = Side(border_style='double')
    thin_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=thin_danbian)
    double_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=double_danbian)
    wb = openpyxl.load_workbook(fname)
    # for ws in wb.worksheets :
    ws = wb[nws_name]
    # 单元格宽度
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 4
    for j in 'DEFGHI':
        ws.column_dimensions['{}'.format(j)].width = 7.44
    ws.column_dimensions['J'].width = 9.44
    # 插入第一行
    ws.insert_rows(1)
    ws['A1'].value = nws_name
    ws['A1'].font = font_firstrow
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:J1')
    max_row = ws.max_row
    max_column = ws.max_column
    for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
        if row == 2:
            for col in range(1, max_column + 1):
                ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        else:
            for col in range(1, max_column + 1):
                if col <= 3:
                    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                else:
                    ws.cell(row, col).alignment = Alignment(horizontal='right', vertical='center')

    for cell in ws['A']:
        row = cell.row
        if row == 1:
            continue
        else:
            if '小计' in cell.value:
                for col in range(1, max_column + 1):
                    ws.cell(row, col).font = font_xiaoji
                    ws.cell(row, col).border = double_bian
                    ws.merge_cells('A{}:B{}'.format(row, row))
            else:
                for col in range(1, max_column + 1):
                    ws.cell(row, col).font = font
                    ws.cell(row, col).border = thin_bian
    ws.freeze_panes = ws['D3']
    ws.print_title_rows = '1:2'  # the first row
    # 页脚设置
    ws.oddFooter.center.text = " &[Page] / &N"  # 1/n
    ws.oddFooter.center.size = 12  # 页脚中字体大小
    ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
    ws.oddFooter.center.color = "000000"  # 页脚中字体颜色
    ws.oddFooter.right.text = "张文伟 &[Date]"  # 页脚右 文字
    ws.oddFooter.right.size = 12  # 页脚右 字体大小
    ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
    ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
    ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.1, left=0.51, right=0.1, footer=0.5,
                                                       bottom=1)
    wb.save(fname)
printseting(fname,nws_name)
os.rename(fname,fname1)
os.startfile(fname1)


