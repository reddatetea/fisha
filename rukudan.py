'''
从粘贴板将材料入库单导入，快速设置，并直接打印
'''
import openpyxl
import pyperclip
import excelseting
import addshuziqianfenfu
import os
import excelprint
import sys

def getMessage():
    if len(sys.argv) > 1:
        msg = ''.join(sys.argv[1:])
    else:
        msg = pyperclip.paste()

    return msg


def rukuRange(paste_content,fname,ws_name):
    cell_range =paste_content
    content = cell_range.split("\n")
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    ws.delete_rows(1,ws.max_row)
    # wb.save(fname)
    shuju0 = []
    for i in content:
        j = i.split('\t')
        k = [x.strip() for x in j]
        shuju0.append(k)
    # gongyinshang = shuju0[0][0] .replace('"','')
    gongyinshang = shuju0[0][1].replace('双佳','白云0')
    for index,row in  enumerate(shuju0):
        if  '品名' in row or  '材料名称'  in row  or  '货品名称' in row:
            start_row = index
        elif  '总计' in row:
            end_row = index
        else :
            continue
    shuju = shuju0[start_row:end_row+1]

    # wb = openpyxl.load_workbook(fname)
    # ws = wb['入库']
    for index,row in enumerate(shuju):
        if index==0:
            ws.append(row)
        else :
            new_row = [row[0]]+[0 if  j=='-'  else  float(j.replace(',','')) for j in row[1:]  ]   #字符串转数字
            ws.append(new_row)
    wb.save(fname)
    return gongyinshang

def main():
    fname = r'F:\a00nutstore\fishc\材料入库单.xlsx'
    ws_name = '入库'
    paste_content =  getMessage()
    gongyinshang = rukuRange(paste_content,fname,ws_name)
    title = '{} 材料入库'.format(gongyinshang)
    #addshuziqianfenfu.addShuziStyle(fname,ws_name,2,3)     #两个参数，则默认2，2
    addshuziqianfenfu.addShuziStyle(fname, ws_name)  # 两个参数，则默认2，2
    excelseting.fastseting(fname,ws_name,title)
    # os.startfile(fname, 'print')                         #打开excel并打印
    excelprint.wsPrint(fname,ws_name)

if  __name__ == '__main__':
    main()
