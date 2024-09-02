
'''
根据零配件入库统计和退货统计，形成零配件实时流水账
'''

# _*_ conding:_*_
import os
import openpyxl
import datetime
from openpyxl.utils import column_index_from_string
import qijianchuli
import quchong
import excelmessage
import easygui
from beifenFile import beifen


def ruku(fname):
    wb = openpyxl.load_workbook(fname,data_only=True)
    ws = wb['ssrk']

    #ws = wb ['2019']
    maxrow = ws.max_row
    print(maxrow)
    #获取列号
    biaozun_riqi_num = column_index_from_string('O')

    qijian_num = column_index_from_string('N')

    for row in range(2,maxrow+1):

        riqi_string0 = ws.cell(row,2).value

        #字符串变标准日期
        #biaozun_riqi = datetime.datetime.strptime(riqi_string0,'%Y-%m-%d')
        #标准日期变字符串
        riqi_string= datetime.datetime.strftime(riqi_string0, '%Y-%m-%d')


        #计算期间
        qijian0 = qijianchuli.qijian(riqi_string)

        #写入数据
        #ws.cell(row, biaozun_riqi_num, value=biaozun_riqi)
        ws.cell(row,qijian_num,value = qijian0)

    wb.save(fname)


def addruku(fname,fname2):
    beifen(fname)   #操作前备份零配件流水账
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2.active
    maxrow2 = ws2.max_row
    rukushuliang_num = column_index_from_string('J')
    rukujiner_num = column_index_from_string('L')

    #入库标记是入库还是退货
    ruku_biaoji = ws2.cell(1,1).value
    if '退货' in ruku_biaoji:
        for row in range(2,maxrow2):
            rukushuliang = ws2.cell(row,rukushuliang_num).value*-1
            rukujiner = ws2.cell(row, rukujiner_num).value * -1
            ws2.cell(row,rukushuliang_num,value = rukushuliang)
            ws2.cell(row, rukujiner_num, value=rukujiner)
            wb2.save(fname2)

    else :
        wb2.close()

    wb1 = openpyxl.load_workbook(fname)
    ws1 = wb1['ssrk']
    #ws1 = wb1['2019']
    maxrow1 = ws1.max_row


    for index,row in enumerate(ws2.values):
        if index == 0 :
            continue

        else :
            if row[0]  in ['',None]:
                continue
            else :
                row1 =list(row)
                riqi_string0 = row[1]
                #print(riqi_string0)
                # 字符串变标准日期
                biaozun_riqi = datetime.datetime.strptime(riqi_string0,'%Y-%m-%d')
                # 标准日期变字符串
                riqi_string = datetime.datetime.strftime(biaozun_riqi, '%Y-%m-%d')

                # 计算期间
                qijian0 = qijianchuli.qijian(riqi_string)
                row1.append(qijian0)
                row1.append(biaozun_riqi)


                ws1.append(row1)
    wb1.save(fname)

    wb2.close()
    return fname2

def main():
    msg = '请选择“零配件实时入库2020.xlsx”文件'
    print(msg)
    fname = r'F:\a00nutstore\006\zw\lingpeijian\零配件实时入库2020.xlsx'

    #fname  = r'F:\a00nutstore\006\zw\lingpeijian\零配件实时入库2020.xlsx'
    #fname = easygui.fileopenbox(msg=msg,title="")
    path,filename = os.path.split(fname)
    os.chdir(path)

    print('请点选 当月零配件入库文件')
    fname2 = excelmessage.wenjian()
    fname2 = excelmessage.excelMessage(fname2)
    path,filename2 = os.path.split(fname2)
    fname2 = addruku(fname,fname2)
    #添加数据后再加上标准日期和期间
    #ruku(fname)
    sheetname = 'ssrk'
    in_subject = ['入货单号', '入库日期', '相关单位', '单据附注', '货品编码', '货品名称', '规格', '所属类别', '单位', '入库数量', '单价', '金额', '备注', '期间',
                  '标准日期']
    quchong.delchongfu(fname, sheetname, in_subject)

if __name__ == '__main__':
    main()









