'''
根据白云价格字典和纸张字典，制作所有期间原材料价格明细
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
import byallpricedic
import zhiallpricedic

#建新的当月价格表excel
def jiannewpiao():
    #print('请输入价格表所在路径:')
    #path = input('')
    path = r'F:\a00nutstore\006\zw\price'
    os.chdir(path)
    filename = 'all月原材料价格表.xlsx'
    fname = os.path.join(path, filename)
    wb = openpyxl.Workbook(fname)
    ws = wb.create_sheet(title='price')

    taitou = ('供应商','期间', '品名', '价格', '令数', '吨数', '金额')
    ws.append(taitou)
    wb.save(fname)
    return fname

#新建字典
def jiannewdic():

    byprices = byallpricedic.bypriceDic()
    zhiprices = zhiallpricedic.zhipriceDic()
    return byprices,zhiprices

#白云pricenames列表
def bypriceNames(byprices):
    # 录入白云pricename列表
    byqijians = []
    bypricenames = []
    bypricelists = []
    for key1,value1 in byprices.items():
        for key2, value2 in value1.items():
            for key3, value3 in value2.items():
                byqijians.append(key1)
                bypricenames.append(key2)
                bypricelists.append(key3)
    return byqijians,bypricenames,bypricelists

#写入白云价格
def writebyprice(fname,byqijians,bypricenames,bypricelists,byprices):
    wb = openpyxl.load_workbook(fname)
    ws = wb['price']
    for row in range(2,len(bypricenames)+2):
        byqijian = byqijians[row - 2]

        bypricename = bypricenames[row-2]
        byprice = bypricelists[row-2]
        byprices[byqijian][bypricename].setdefault(byprice, {'lingshu': 0, 'dunshu': 0, 'jiner': 0})

        lingshu = byprices[byqijian][bypricename][byprice]['lingshu']
        dunshu = byprices[byqijian][bypricename][byprice]['dunshu']
        jiner = byprices[byqijian][bypricename][byprice]['jiner']

        ws.cell(row,1,value = '驻马店白云纸业有限公司')
        ws.cell(row,2,value = byqijian)
        ws.cell(row, 3, value= bypricename)
        ws.cell(row, 4, value=byprice)

        ws.cell(row,5, value= lingshu)
        ws.cell(row, 6, value= dunshu)
        ws.cell(row,7, value= jiner)

    wb.save(fname)

#纸张pricenames列表
def zhipriceNames(zhiprices):
    # 录入纸张zhipricename列表
    zhiqijians = []
    zhigongyingshangs = []
    zhipricenames = []
    zhipricelists = []

    for key1,value1 in zhiprices.items():    #此处key1是供应商,value1是品名
        for key2,value2 in value1.items():              #此处key2是品名,value是价格
            for key3, value3 in value2.items():
                for key4, value4 in value3.items() :  #此处key3是价格,value是(令数吨数金额)
                    zhiqijians.append(key1)
                    zhigongyingshangs.append(key2)
                    zhipricenames.append(key3)
                    zhipricelists.append(key4)


    return zhiqijians,zhigongyingshangs,zhipricenames, zhipricelists


#写入纸张价格
def writezhiprice(fname, zhiqijians,zhigongyingshangs,zhipricenames,zhipricelists,zhiprices):
    wb = openpyxl.load_workbook(fname)
    ws = wb['price']
    maxrow = ws.max_row

    #先追加白云价格并保存后,行数为maxrow0,从maxrow1起写入纸类价格
    maxrow1 = maxrow+1
    for row in range(maxrow+1, len(zhipricelists) + maxrow1):
        zhiqijian = zhiqijians[row - maxrow1]
        zhigongyingshang = zhigongyingshangs[row - maxrow1]
        zhipricename = zhipricenames[row - maxrow1]
        zhiprice = zhipricelists[row - maxrow1]
        zhiprices[zhiqijian][zhigongyingshang][zhipricename].setdefault(zhiprice, {'lingshu': 0, 'dunshu': 0, 'jiner': 0})

        lingshu =zhiprices[zhiqijian][zhigongyingshang][zhipricename][zhiprice]['lingshu']
        dunshu = zhiprices[zhiqijian][zhigongyingshang][zhipricename][zhiprice]['dunshu']
        jiner = zhiprices[zhiqijian][zhigongyingshang][zhipricename][zhiprice]['jiner']

        ws.cell(row, 1, value=zhigongyingshang)
        ws.cell(row, 2, value=zhiqijian)
        ws.cell(row, 3, value=zhipricename)
        ws.cell(row, 4, value=zhiprice)

        ws.cell(row, 5, value=lingshu)
        ws.cell(row, 6, value=dunshu)
        ws.cell(row, 7, value=jiner)

    wb.save(fname)

#删除数据为0的行
def dangyueprice(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.create_sheet(title = 'all月')
    maxrow = ws.max_row

    ws1 = wb['price']
    for index,row in enumerate(ws1.values):
        if (row[4]==0 and row[5]== 0 and row[6]==0) or row[3]==0:
            continue
        else:
            ws.append(row)

    wb.save(fname)

def main():
    #qijian = input('请输入期间，如2020-04\n')
    byprices,zhiprices = jiannewdic()

    byqijians,bypricenames,bypricelists = bypriceNames(byprices)
    zhiqijians,zhigongyingshangs,zhipricenames,zhipricelists = zhipriceNames(zhiprices)
    fname = jiannewpiao()
    writebyprice(fname,byqijians, bypricenames,bypricelists, byprices)
    writezhiprice(fname, zhiqijians,zhigongyingshangs,zhipricenames,zhipricelists,zhiprices)
    dangyueprice(fname)

if __name__ == '__main__':
    main()

