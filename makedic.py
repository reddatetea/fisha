# _*_ conding = utf-8 _*_
'''
利用easygui，通过excel中指定两个字段制作字典
'''

import  openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import os
import easygui

def choiceSheet():
    msg = '请点选要制作字典的excel文件'
    print(msg)
    fname = easygui.fileopenbox(msg=msg)
    path,filename = os.path.split(fname)
    os.chdir(path)

    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选要制作字典的excel工作表'
    choice = easygui.buttonbox(msg=msg, title='工作表', choices=sheetnames)
    ws = wb['%s' % choice]

    maxrow =ws.max_row
    maxcolumn = ws.max_column
    print('最大列号是：',maxcolumn)

    fields = []
    for i in range(1,maxcolumn+1):
        cellvalue = ws.cell(row=1, column=i).value
        fields.append(cellvalue)


    print('你选择的excel文件是：',fname)
    print('你选择的工作表是：',choice)
    print('第一行字段名列表：',fields)

    msg = '请点选要制作字典key值的字段'
    key_choice_col = easygui.buttonbox(msg=msg, title='key', choices=fields)
    print('key_choice_col',key_choice_col)

    field_len = len(fields)
    for i in range(field_len):
        if key_choice_col == fields[i]:
            key_choice_liehao = i +1
            break
        else :
            continue


    #key_choice_liehao = column_index_from_string(key_choice_col)

    msg = '请点选要制作字典value值的字段'
    value_choice_col = easygui.buttonbox(msg=msg, title='value', choices=fields)
    for i in range(field_len):
        if value_choice_col == fields[i]:
            value_choice_liehao = i +1
            break
        else :
            continue
    print('value_choice_col', value_choice_col)

    #value_choice_liehao =  column_index_from_string(value_choice_col)

    #下面制作字典
    dicnew = {}
    for i in range(2,maxrow+1):
        dicnew_key = ws.cell(row=i,column = key_choice_liehao).value
        dicnew_value = ws.cell(row=i,column = value_choice_liehao).value
        dicnew[dicnew_key] = dicnew_value

    print('打印新制作的字典：',dicnew)




    return fname,choice


def main():
    fname =choiceSheet()


if __name__=='__main__':
    main()






