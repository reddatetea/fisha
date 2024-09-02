'''
从实时流水账中引入当月的纸入库
本版采用迭代期间取价格,本版调用本月价格生成实时字典,价格有点小问题
'''

import lszTozhiDangyue
import zhiDangyueToZhi
import os
import openpyxl
import pricelisttoprice
import allyueprice
import qijianchuli04
import quchong
import pandas as pd
import re

def tichuNum(temp):
    pattern = r'卷筒(\d+)$'
    regexp = re.compile(pattern)
    pipei = regexp.search(temp)
    if pipei == None:
        string = temp
    else:
        string = re.sub(pattern, '卷筒', temp)
    return string


qijian = input('请输入入库期间：格式为2020-04：\n')
qijian0 = qijian
piaojuhao = input('请输入票据号：\n')
lszTozhiDangyue.lszTozhidangyue(qijian,piaojuhao)
dunjia_dic = zhiDangyueToZhi.dunjiaDic()
print(dunjia_dic)
fname2,maxrow2 = zhiDangyueToZhi.zhidangyueTozhiruku()
# zhiDangyueToZhi.jiagongsi(fname2,maxrow2,dunjia_dic)
fname = fname2

#数据去重
in_subject = ['单位', '供应商', '时间','单号', '材料', '入库','入库（kg）']
wb = openpyxl.load_workbook(fname)
sheetnames = wb.sheetnames
print('选择工作表')
sheetname = '入库'
quchong.delchongfu(fname,sheetname,in_subject)

# gongying_dic = pricelisttoprice.yuancailiaoprice()
gongying_dic = dunjia_dic
#价格字典，20210923更新

#将去重后的数据重新引入
#pandas08.shujuqingli(duplicated_name, fname)

def jiagongsi(fname,gongying_dic,qijian):
    wb = openpyxl.load_workbook(fname)
    ws = wb['入库']
    maxrow = ws.max_row

    #未记账的加公式
    for i in range(maxrow2,maxrow+1):
        if ws.cell(i,22).value in ['',None]:
            pricename = ws.cell(i,24).value
            cailiao0 = ws.cell(i,6).value
            cailiao =  tichuNum(cailiao0)

            gongyingshang = ws.cell(i,2).value
            # dunjia0=送货单金额/入库吨数
            try:
                danjia0 = round(ws.cell(i, 18).value / ws.cell(i, 13).value, 0)  # 吨价0=送货单金额/吨数
            except:
                danjia0 = 0
            if gongyingshang=='河南省江河纸业有限公司':
                dunjia = gongying_dic.get((gongyingshang,cailiao),0)
            else :
                dunjia = gongying_dic.get((gongyingshang, pricename), 0)

            #金额用公式表达
            ws.cell(i, 14, value=danjia0)
            ws.cell(i, 15, value=dunjia)
            ws.cell(i, 17, value='=round(round(M' + str(i) + ', 3) * ' + 'O' + str(i) + ', 2)')
            #不含税金额用公式表达
            ws.cell(i, 19, value='=round(R' + str(i) + '/1.13 , 2)')
            # 多计用公式表达R2-Q2
            ws.cell(i, 20, value='=R' + str(i) + '-Q'+str(i))
        else :
            continue

    wb.save(fname)

jiagongsi(fname,gongying_dic,qijian)
os.startfile(fname)








