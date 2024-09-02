import os
import openpyxl
import strtofloat01
import datetime


path = r'F:\a00nutstore\006\zw\fengognsi\南京\南京2020\南京202004'
os.chdir(path)

print('请输入银行存款文件名：')
filename = input('')

fname = os.path.join(path, filename)
wb = openpyxl.load_workbook(fname)
ws = wb.active
max_row = ws.max_row


print('请输入南京报表名称：')
filename2 = input('')
fname2 = os.path.join(path, filename2)

wb2 = openpyxl.load_workbook(fname2)
ws2 = wb2['银行参考']

taitou = ('交易时间','借方','贷方','摘要','用途','对方单位名称','余额','个性化信息')

for index,row in enumerate(ws.values):
    if index == 0:
        ws2.append(taitou)
    elif index == 1 :
        pass
    else :
        jiaoyishijian0 = row[3]
        jiaoyishijian = datetime.datetime.strptime(jiaoyishijian0[:10], '%Y-%m-%d')
        jiefang0 = row[5].strip()
        jiefang = strtofloat01.strtoFloat(jiefang0)

        daifang0 =row[6].strip()
        daifang = strtofloat01.strtoFloat(daifang0)

        zhaiyao = row[8]
        yongtu = row[9]
        duifangdanwei = row[10]

        yuer0 = row[11].strip()
        yuer = strtofloat01.strtoFloat(yuer0)

        qita = row[12]

        hang = (jiaoyishijian,jiefang,daifang,zhaiyao,yongtu,duifangdanwei,yuer,qita)
        ws2.append(hang)

wb.close()
wb2.save(fname2)
