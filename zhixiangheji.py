#纸箱求和公式，指定工作簿、工作表和指定求和字段
import  os
import excelmessage
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import easygui

def  sumheji(fname,sheetnames,letters):
    wb = openpyxl.load_workbook(fname)
    for sheetname in sheetnames:
        ws = wb[sheetname]
        max_row = ws.max_row
        for letter in letters:
            ws['A{}'.format(max_row + 1)].value = '合计'
            ws['{}{}'.format(letter, max_row + 1)].value = '=sum({}2:{}{})'.format(letter, letter, max_row)

    wb.save(fname)

def main():
    msg = '请点选要加计合计数的excel文件'
    print(msg)
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    wb = openpyxl.load_workbook(fname)
    strs = wb.sheetnames
    msg = '请点选要求和的工作表'
    sheetnames = easygui.multchoicebox(msg, choices=sts)
    wb.close()
    letters = 'KPRT'
    sumheji(fname,sheetnames,letters)

if __name__ == '__main__':
    main()
