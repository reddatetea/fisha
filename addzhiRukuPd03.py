'''
从实时流水账中引入当月的纸入库
'''
import os
import pandas as pd
import numpy as np
import re
import openpyxl
# import zhigongyingshang
import zhiNewGongyingshang
import xlwings as xw
from beifenFile import beifen
import easygui

#实时吨价
def dunjiaDic():
    fname2 = r'F:\a00nutstore\006\zw\else\2020入库.xlsx'
    df = pd.read_excel(fname2, sheet_name='入库')
    df1 = df[df['pricename'].str.contains('返利|价差|冲减|多计|折扣') == False]  # 品名中包含返利、价差、冲减等不计入字典
    gongyingshangs = list(df1['供应商'])
    pinmings = list(df1['pricename'])
    cailiaos = list(df1['材料'])                                       #list(map(tichuNum,df1['材料']))
    dunjias = list(df1['吨价'])
    jizhangs = list(df1['记账'])
    dunjia_dic = {}
    for j in range(len(dunjias)):
        gongyingshang = gongyingshangs[j]
        if jizhangs[j] not in ["", None]:
            if gongyingshang != '河南省江河纸业有限公司':
                dunjia_dic[(gongyingshangs[j], pinmings[j])] = dunjias[j]
            else:
                cailiao = tichuNum(cailiaos[j])
                dunjia_dic[(gongyingshangs[j], cailiao)] = dunjias[j]
        else:
            continue

    print(dunjia_dic)
    return dunjia_dic

#将材料名称卷筒后面的数字剔除
def tichuNum(temp):
    pattern = r'卷筒(\d+)$'
    regexp = re.compile(pattern)
    pipei = regexp.search(temp)
    if pipei == None:
        string = temp
    else:
        string = re.sub(pattern, '卷筒', temp)
    return string

def liushuizhang():
    df = pd.read_excel(r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx', '流水账')
    df = df.loc[:,
         ['供货单位', '期间', '日期', '单据号', '品名', '令数', 'cwName', '单位', '吨数', '吨价0', '吨价', '令价', '入库金额', 'priceName','送货日期','白云期间']]
    df.insert(0, '单位0', '双佳')
    df.insert(9, '批次', None)
    for j in ['入库(吨)', '入库（kg）']:
        df.insert(11, j, 0)
    for j in ['送货单金额', '不含税金额', '多计', '令价0'][::-1]:
        df.insert(17, j, 0)
    for j in ['记账', '备注'][::-1]:
        df.insert(21, j,None)
    return df

def query():
    start_riqi = pd.Timestamp(easygui.enterbox('请输入入库起始日期期间：格式为2021-11-4：'))
    end_riqi = pd.Timestamp(easygui.enterbox('请输入入库结束日期期间：格式为2021-11-4：'))
    piaojuhao = int(easygui.enterbox('请输入票据号90031065：'))
    return start_riqi,end_riqi,piaojuhao

def qushu(df,start_riqi,end_riqi,piaojuhao,jianchen):
    df.set_index('日期', inplace=True)  # 索引
    df.sort_index(inplace=True)  # 对索引排序
    df = df.truncate(before=start_riqi, after=end_riqi)
    df = df.loc[df['单据号'] >= piaojuhao]
    # jianchen = zhigongyingshang.jianchen
    df = df.loc[df['供货单位'].isin(jianchen)]
    return df

def getruku():
    fname_ruku = r'F:\a00nutstore\006\zw\else\2020入库.xlsx'
    ws_name_ruku = '入库'
    beifen(fname_ruku)         #重要文件备份
    wb = openpyxl.load_workbook(fname_ruku, data_only=True)  # 将有公式的全部转为值，避免以后excel中有公式出现空值
    wb.save(fname_ruku)
    df2= pd.read_excel(fname_ruku,ws_name_ruku)
    max_row = df2.shape[0]
    return fname_ruku,ws_name_ruku,max_row

def chuli(df1):
    df1.reset_index(inplace=True)
    riqi = df1.pop("日期")
    df1.insert(3, "日期", riqi)
    df1['吨数'] = round(df1['吨数'],3)
    df1['入库（kg）'] = df1['吨数']
    df1['入库(吨)'] = df1['吨数']
    df1['送货单金额'] = df1['入库金额']
    dunjia_dic = dunjiaDic()
    df1['priceName'] = np.where(df1['供货单位'] == '河南省江河纸业有限公司', df1['品名'].agg(tichuNum),
                                df1['priceName'])  # 江河的pricename名称按cailiao
    df1 = df1.assign(吨价0=df1.apply(lambda x: dunjia_dic.get((x['供货单位'], x['priceName']),0), axis=1))
    df1['令价'] = round(df1['令价'],2)
    fname_ruku, ws_name_ruku, max_row = getruku()
    wb = openpyxl.load_workbook(fname_ruku)
    with pd.ExcelWriter(fname_ruku, engine='openpyxl')  as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        df1.to_excel(writer, ws_name_ruku, header=None, index=False, startrow=max_row + 1)
    # writer.save()
    return fname_ruku

def delchongfu(fname,sheetname,in_subject,sort_cols):
    app = xw.App(visible = False,add_book = False)
    wb = app.books.open(fname)
    ws = wb.sheets['%s'%sheetname]
    data = pd.DataFrame(pd.read_excel(fname, sheetname))
    data.drop_duplicates(subset=in_subject, keep='first', inplace=True)
    data = data.set_index(in_subject[0])
    data = data.sort_values(by=sort_cols)
    ws.clear()
    ws.range('A1').value = data
    wb.save()
    wb.close()
    app.quit()
    return fname

def jiagongsi(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb['入库']
    maxrow = ws.max_row
    #未记账的加公式
    for i in range(2,maxrow+1):
        if ws.cell(i,22).value in ['',None]:
            #金额用公式表达
            ws.cell(i, 17, value='=round(round(M' + str(i) + ', 3) * ' + 'O' + str(i) + ', 2)')
            #不含税金额用公式表达
            ws.cell(i, 19, value='=round(R' + str(i) + '/1.13 , 2)')
            # 多计用公式表达R2-Q2
            ws.cell(i, 20, value='=R' + str(i) + '-Q'+str(i))
        else :
            continue
    wb.save(fname)

def main():
    start_riqi, end_riqi, piaojuhao = query()
    df = liushuizhang()    #流水账df
    jianchen = zhiNewGongyingshang.zhiGongyingshang()
    df1 = qushu(df,start_riqi,end_riqi,piaojuhao,jianchen)
    fname_ruku = chuli(df1)
    df = pd.read_excel(fname_ruku,'入库')
    in_subject = ['单位', '供应商', '时间', '单号', '材料', '入库', '入库（kg）']
    sort_cols = ['期间','供应商','时间', '单号','材料']
    delchongfu(fname_ruku,'入库', in_subject,sort_cols)
    jiagongsi(fname_ruku)
    os.startfile(fname_ruku)

if __name__ == '__main__':
    main()


