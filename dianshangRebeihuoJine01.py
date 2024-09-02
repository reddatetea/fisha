'''
电商出库将规格型号中的本数提取出来,较多正则处理
2024080加金额
20240824加按品名“汇总”
'''
import re
import os
import pandas as pd
import numpy as np
import openpyxl
import easygui
from openpyxl.styles import Font, Border, Side, Fill, Alignment

fname = easygui.fileopenbox('请点选电商当日出库文件')
# fname = r"F:\repos\fish\8-23.xlsx"
qian, hou = os.path.splitext(fname)
fname_new = qian + "加工" + hou
sheet_name = 0
skiprows = 0
df = pd.read_excel(fname, sheet_name=sheet_name, skiprows=skiprows, dtype={'商品ID': 'str', '运单号': 'str', })
df.dropna(subset=['订单号'], inplace=True)
df['总价'] = df['总价'].str.replace('￥', '')
df['实收'] = df['实收'].str.replace('￥', '')
df = df.astype({'实收': 'float64'})
df = df.astype({'总价': 'float64'})

# df['实收'] = df['实收'] .astype(int)
df1 = df.copy()
lst = ['订单号',
       '发货时间',
       '规格名称',
       '规格编码',
       '数量',
       '总价',
       '实收'
       ]
dic = {'五': 5,
       '六': 6,
       '七': 7,
       '八': 8,
       '九': 9,
       '十': 10,
       '四': 4,
       '三': 3,
       '二': 2,
       '一': 1}


def addBen(string):
    # string = r'(81本)'
    pattern1 = r'(?P<ben>\d+)本'
    pattern2 = r'(?P<ben>[十九八七六五四三二一])本'
    regexp1 = re.compile(pattern1)
    mat1 = regexp1.search(string)
    regexp2 = re.compile(pattern2)
    mat2 = regexp2.search(string)
    if mat1:
        ben = int(mat1.group('ben'))
    else:
        if mat2:
            ben = dic.get(mat2.group('ben'))
        else:
            ben = 1
    return ben
def getGroupby(df):
    gp = df.groupby('规格编码')
    bao = gp['包数'].sum()
    ben = gp['本数'].sum()
    zhongjia = gp['总价'].sum()
    shishou = gp['实收'].sum()
    beihuo = pd.concat([bao, ben,zhongjia,shishou], axis=1)
    beihuo.columns = ['包数', '本数', '总价','实收']
    beihuo.loc['小计'] = [beihuo['包数'].sum(), beihuo['本数'].sum(),beihuo['总价'].sum(),beihuo['实收'].sum()]
    beihuo = beihuo.reset_index()
    return beihuo

df1 = df1[lst]
df1['含量'] = df1['规格名称'].map(addBen)
df1 = df1.assign(本数=df1['数量'] * df1['含量'])
df1 = df1.sort_values(by=['规格编码'])
df1.index = range(1, len(df1) + 1)
df1.index_name = '序号'
df1.insert(0, '序号', range(1, len(df1) + 1))
max_xuhao = len(df1) + 1
df1.loc[max_xuhao] = {'序号': max_xuhao, '数量': df1['数量'].sum(), '本数': df1['本数'].sum(),'总价':df1['总价'].sum(),'实收':df1['实收'].sum()}
df1['序号'] = df1['序号'].replace(max_xuhao, '小计')
df1 = df1[['序号',
       '订单号',
       '发货时间',
       '规格名称',
       '规格编码',
           '本数',
        '含量',
       '数量',
           '总价',
           '实收'
       ]]
df1.columns = ['序号',
       '订单号',
       '发货时间',
       '规格名称',
       '规格编码',
           '本数',
        '含量',
       '包数',
      '总价',
       '实收'
       ]

beihuo = getGroupby(df1)
df2 = df1.copy()
df2 = df2.iloc[:-1,]
df2['规格编码1'] = df2['规格编码'].str.split('+')
df2['len'] = df2['规格编码1'].str.len()
df2 = df2.assign(本数 = np.where(df2.len == 1,df2['本数'] ,df2['本数']/df2['len']))
df2 = df2.assign(包数 = np.where(df2.len == 1,df2['包数'] ,df2['包数']/df2['len']))
df2 = df2.explode('规格编码1')
df2['规格编码'] = df2['规格编码1']
df3 = df2[['序号',
       '订单号',
       '发货时间',
       '规格名称',
       '规格编码',
           '本数',
        '含量',
       '包数',
      '总价',
       '实收'
       ]]
huizhong =getGroupby(df3)
huizhong = huizhong [['规格编码','本数','包数',  '总价','实收']]


with pd.ExcelWriter(fname, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df.to_excel(writer, sheet_name='加工', index=False)
    beihuo.to_excel(writer, sheet_name='备货', index=False)
    df1.to_excel(writer, sheet_name='本数', index=False)
    huizhong.to_excel(writer, sheet_name='汇总', index=False)

wb = openpyxl.load_workbook(fname)


def sheetSetting(sheetname):
    ws = wb[sheetname]
    max_row = ws.max_row
    max_column = ws.max_column

    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 4.5
    ws.column_dimensions['G'].width = 6.13
    ws.freeze_panes = ws['B2']
    ws.print_title_rows = '1:1'  # the first row
    # 页脚设置
    ws.oddFooter.center.text = " &[Page] / &N"  # 1/n
    ws.oddFooter.center.size = 12  # 页脚中字体大小
    ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
    ws.oddFooter.center.color = "000000"  # 页脚中字体颜色
    ws.oddFooter.right.text = "张文伟 &[Date]"  # 页脚右 文字
    ws.oddFooter.right.size = 12  # 页脚右 字体大小
    ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
    ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
    if sheetname == '备货' or sheetname == '汇总':
        ws.column_dimensions['A'].width = 20
        ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.5, left=2, right=0.24,
                                                              footer=0.5,
                                                              bottom=1)

    else:
        ws.column_dimensions['A'].width = 4
        ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.5, left=0.21, right=0.24,
                                                              footer=0.5,
                                                              bottom=1)

    thin_danbian = Side(border_style='thin')
    thin_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=thin_danbian)
    for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
        for col in range(1, max_column + 1):
            ws.cell(row, col).border = thin_bian
    return ws


for sheetname in ['备货', '本数','汇总']:
    ws = sheetSetting(sheetname)

wb.active = ws
wb.save(fname_new)
os.startfile(fname_new)

