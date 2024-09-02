'''
本模块将白云当月入库添加到白云2020入库
本版本模块化
'''
# _*_ conding:utf-8 _*_
import openpyxl
import os
import datetime
import easygui
import pandas as pd


def dunjiaDic():
    fname2 = r'F:\a00nutstore\006\zw\baiyun\2020白云入库.xlsx'
    df = pd.read_excel(fname2, sheet_name='2020')
    df1 = df[df['pricename'].str.contains('返利|价差|冲减|多计') == False]  # 品名中包含返利、价差、冲减等不计入字典
    pinmings = list(df1['pricename'])             #不加list跑不动
    dunjias = list(df1['单价'])
    jizhangs = list(df1['记账'])
    dunjia_dic = {}
    for j in range(len(pinmings)):
        if jizhangs[j] not in ["", None]:
            dunjia_dic[('驻马店白云纸业有限公司', pinmings[j])] = dunjias[j]
        else:
            continue
    print(dunjia_dic)
    return dunjia_dic

def baiyunDangyueToby():
    dtrq = datetime.date.today().strftime('%Y%m%d')
    print('选择白云当月入库路径')
    #path = easygui.diropenbox(msg='请选择白云当月入库路径', title='')
    path = r'F:\a00nutstore\006\zw\baiyun'
    os.chdir(path)
    filename = '白云当月入库%s.xlsx' % dtrq
    fname = os.path.join(path,filename)
    wb1 = openpyxl.load_workbook(fname)
    ws1 = wb1['当月正']

    print('选择白云年入库文件')
    #fname2 = easygui.fileopenbox(msg = '请选择白云年入库文件',title = '2020白云入库')
    fname2 = r'F:\a00nutstore\006\zw\baiyun\2020白云入库.xlsx'
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['2020']
    maxrow2 =ws2.max_row

    dic = {}
    danjuhaos = [j.value for j in ws2['F'][1:]]
    jizhangriqis = [j.value for j in ws2['N'][1:]]
    for j in range(len(danjuhaos)):
        if jizhangriqis[j] not in ['', None]:
            dic[danjuhaos[j]] = jizhangriqis[j]                             #单据号-记账日期字典


    for index, row in enumerate(ws1.values):
        if index == 0:
            pass
        else:
            if row[5] not in dic:                                  #如果未记账（单据号不在单据号-记账日期字典中），则添加！
                ws2.append(row)
            else:
                continue

    wb1.close()
    wb2.save(fname2)
    return fname2,maxrow2

def jiagongsi(fname2,maxrow2,dunjia_dic):
    wb = openpyxl.load_workbook(fname2)
    ws = wb['2020']
    # dun_dic = dunjiaDic(fname2)
    maxrow = ws.max_row
    for i in range(maxrow2+1,maxrow+1):
        gongyingshang = '驻马店白云纸业有限公司'
        pinming = ws.cell(i, 17).value
        ws.cell(i, 11).value = dunjia_dic.get((gongyingshang, pinming), 0)  # 取最新价格字典，没有则为0
        #金额用公式表达
        ws.cell(i,12, value='=round(J' + str(i) + ' * ' + 'K' + str(i) + ', 2)')
        #不含税金额用公式表达
        ws.cell(i,13, value='=round(L' + str(i) + '/1.13 , 2)')
    wb.save(fname2)

def main():
    dunjia_dic = dunjiaDic()
    fname2,maxrow2 = baiyunDangyueToby()
    jiagongsi(fname2,maxrow2,dunjia_dic)



if __name__ == '__main__':
    main()









