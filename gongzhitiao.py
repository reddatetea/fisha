import openpyxl

fname = '员工工资表.xlsx'
wb = openpyxl.load_workbook(fname,data_only=True)    #公取值，不要公式
ws = wb['工资表']
wb.copy_worksheet(ws).title = '工资表打印'
nws = wb['工资表打印']
title = [j.value for j in  ws[2]]

max_col = ws.max_column
for cell in nws['A'][3:]:
    print(cell.row)
    nws.insert_rows(cell.row)
    for j in range(1,max_col+1):
        shuju = nws.cell(cell.row,j).value
        nws.cell(cell.row-1,j).value = title[j-1]


wb.save(fname)