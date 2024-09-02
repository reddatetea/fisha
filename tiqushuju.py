'''
根据指定的供应商和期间，提取对应的原材料流水账数据
本版增加多期间的查询
本版增加os.startfile和easygui
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
from openpyxl.utils import column_index_from_string
import qijiansplit
import easygui
import time


def tiquShuju(fname,im_gongyingshang,im_qijians):
    dirname,filename =os.path.split(fname)
    os.chdir(dirname)

    qj = im_qijians[0] + '~' + im_qijians[-1]
    filename1 = '%s查询%s.xlsx'%(im_gongyingshang,qj)
    fname1 = os.path.join(dirname,filename1)

    wb = openpyxl.load_workbook(fname,data_only=True,read_only=True)
    ws = wb['流水账']

    #获取列号
    qijian_num = column_index_from_string('K')
    gongyingshang_num = column_index_from_string('C')

    wb1 = openpyxl.Workbook()
    #所属期间

    # ws1 = wb1.create_sheet(title='%s%s' % (im_gongyingshang,qj))
    ws1 = wb1.active
    ws_name =  '%s%s' % (im_gongyingshang,qj)
    ws1.title = ws_name

    for index,row in enumerate(ws.values):
        if index == 0 :
            ws1.append(row)
        else :
            qijian = row[qijian_num-1]
            gongyingshang = row[gongyingshang_num-1]

            if qijian in im_qijians and  im_gongyingshang in gongyingshang :
                ws1.append(row)
            else :
                continue
    wb1.save(fname1)
    wb.close()
    return fname1,ws_name

def main():
    #fname = easygui.fileopenbox(msg =  '请选择原材料流水账',title = '实时流水账')
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    path,filename = os.path.split(fname)

    msg = '请输入要查询的供应商和期间'
    title = '供应商流水账查询'
    fields = ['供应商','期间']
    choice = easygui.multenterbox(msg,title,fields)
    im_gongyingshang = choice[0]
    im_qijian0 = choice[1]
    im_qijians = qijiansplit.shijiqijian(im_qijian0)
    fname1 = tiquShuju(fname,im_gongyingshang,im_qijians)
    path,filename = os.path.split(fname1)
    choice = easygui.ccbox(msg='是否打开要查询的文件"%s"' % filename, title='供应商流水账查询', choices=('yes', 'no'))
    if choice == True:
        os.startfile(fname1)
    else:
        pass

    return fname1

if __name__ == '__main__' :
    main()

