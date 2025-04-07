'''
处理聚水潭编码与大库差异
2025-1-27程序的通用性，其他程序可以调用聚水潭的编码处理
第一步先计算本次聚水潭出库中对应不上的大库的编码
第二步将差异处理后，再运行形成差异较少的出库明细表
'''

import numpy as np
import pandas as pd
import re
import easygui
import openpyxl
import os
def getCunhuoConcent(fname):
    pattern = r'(?P<num>\d+)本/件'
    regex = re.compile(pattern)
    df = pd.read_excel(fname,dtype = {'存货编码':'str','其它属性4':'str'})
    def addBen(string):
        if string == '本':
            string =0
            return string
        else :
            mat = regex.search(string)
            
            if mat:
                string = int(mat.group('num'))
                return string
            else :
                string = 0
                return string
    df['计量单位1'] = df['计量单位'].map(addBen)
    content_dic = dict(zip(df['存货编码'],df['计量单位1'] ))
    return content_dic

def getCunhuoConcentFile(path,content_dic):
    riqi = easygui.enterbox('请输入日期')
    dicfile = f'存货档案字典{riqi}.xlsx'
    sheet_name = '存货档案'
    df_dic = pd.DataFrame.from_dict(content_dic,orient='index')
    df_dic = df_dic.reset_index()
    df_dic.columns = ['存货编码','含量']
    df_dic.to_excel(dicfile,sheet_name = sheet_name,index = False)
    return os.path.join(path,dicfile)

def chuliColor(string,pattern):
    mat = re.search(pattern,string)
    if mat :
        str1 = mat.group(1)
    else :
        str1 = string
    return str1

def chuluJusuitanBianma(fname_jusuitan,jusuitan_bianma,dic_jusuitanDiff,content_dic):
    df_jusuitanXiaoshou0 = pd.read_excel(fname_jusuitan,dtype = {f'{jusuitan_bianma}':'str'})
    df_jusuitanXiaoshou0 = df_jusuitanXiaoshou0[~df_jusuitanXiaoshou0[f'{jusuitan_bianma}'].isna()]
    df_jusuitanXiaoshou = pd.pivot_table(df_jusuitanXiaoshou0,index = jusuitan_bianma,values = '数量',aggfunc = 'sum',fill_value = 0)
    df_jusuitanXiaoshou = df_jusuitanXiaoshou.reset_index()
    df_jusuitanXiaoshou = df_jusuitanXiaoshou.assign(content = df_jusuitanXiaoshou[f'{jusuitan_bianma}'].map(content_dic))
    #规则1，结尾为A商品编码改为a
    jusuitan_bianma1 = f'{jusuitan_bianma}1'
    df_jusuitanXiaoshou[f'{jusuitan_bianma1}'] = np.where(df_jusuitanXiaoshou[f'{jusuitan_bianma}'].str.endswith('A'), df_jusuitanXiaoshou[f'{jusuitan_bianma}'].str[:-1]+'a',df_jusuitanXiaoshou[f'{jusuitan_bianma}'])
    # df_jusuitan_diff1['商品编码1'] = np.where(df_jusuitan_diff1['商品编码'].str.endswith('B'), df_jusuitan_diff1['商品编码'].str[:-1]+'b',df_jusuitan_diff1['商品编码'])
    #规则2，带颜色的编码弄短
    regax = r'(.+)-\D+色$'
    pattern = re.compile(regax)
    df_jusuitanXiaoshou[f'{jusuitan_bianma1}'] = df_jusuitanXiaoshou[f'{jusuitan_bianma1}'].apply(lambda x:chuliColor(x,pattern))
    df_jusuitanXiaoshou.content = np.where(df_jusuitanXiaoshou.content.isna(),df_jusuitanXiaoshou[f'{jusuitan_bianma}1'].map(content_dic),df_jusuitanXiaoshou.content)
    df_jusuitanXiaoshou[f'{jusuitan_bianma1}'] = np.where(df_jusuitanXiaoshou.content.isna(),df_jusuitanXiaoshou[f'{jusuitan_bianma1}'].map(dic_jusuitanDiff),df_jusuitanXiaoshou[f'{jusuitan_bianma1}'])
    df_jusuitanXiaoshou = df_jusuitanXiaoshou.rename(columns = {'编码2':'jusuitan','编码21':'xiaoshoubu'})
    df_jusuitanXiaoshou = df_jusuitanXiaoshou[['jusuitan','数量','content','xiaoshoubu']]
    return df_jusuitanXiaoshou0,df_jusuitanXiaoshou

def GetDicJusuitanDiff(fname_jusuitanDiff):
    df_jusuitanDiff = pd.read_excel(fname_jusuitanDiff ,dtype = {'jusuitan':'str','xiaoshoubu':'str'})
    dic_jusuitanDiff = dict(zip(df_jusuitanDiff['jusuitan'],df_jusuitanDiff['xiaoshoubu']))
    return dic_jusuitanDiff

def chukliAddDiff(fname_addDiff):
    add_diff = pd.read_excel(fname_addDiff ,dtype = {'jusuitan':'str','xiaoshoubu':'str'})
    add_diff = add_diff[['jusuitan','xiaoshoubu']]
    return add_diff
    
def main():
    lst = ['商品编码',
                 '实发数量',
                 '净销量',
                 '净销售额',
                 '净销售成本',
                 '净销售毛利',
                 '基本金额',
                 '已付金额',
                 '优惠金额',
                 '运费收入',
                 '运费支出',
                 '净毛利率',
                 '基本售价',
                 ]
    fname  = easygui.fileopenbox('请点选存货档案')
    path,_ = os.path.split(fname)
    os.chdir(path)
    #生成存货档案字典
    content_dic = getCunhuoConcent(fname)
    getCunhuoConcentFile(path,content_dic)
    
     #聚水潭编码与大库差异
    fname_jusuitanDiff = r"F:\a00nutstore\008\zww08\002电商\聚水潭\聚水潭编码与大库差异.xlsx"
    wb = openpyxl.load_workbook(fname_jusuitanDiff)
    ws = wb.active
    max_row = ws.max_row
    dic_jusuitanDiff = GetDicJusuitanDiff(fname_jusuitanDiff)
    fname_jusuitan = r"F:\a00nutstore\008\zww08\002电商\聚水潭\聚水潭各平台销售\聚水潭202503\编码处理后的聚水潭销售（已处理成本价）2025-03.xlsx"
    # fname_jusuitan =easygui.fileopenbox('请打开编码处理后的聚水潭销售主题文件')
    jusuitan_bianma = '编码2'
    df_jusuitanXiaoshou0,df_jusuitanXiaoshou = chuluJusuitanBianma(fname_jusuitan,jusuitan_bianma,dic_jusuitanDiff,content_dic)
    #形成本次聚水潭编码差异
    add_diff = df_jusuitanXiaoshou[df_jusuitanXiaoshou.content.isna()]
    add_diff = add_diff[['jusuitan','xiaoshoubu']]
    add_diff.to_excel('本次聚水潭编码异常add_diff.xlsx',index = False)
    #打开本次需要添加的聚水潭编码与大库的差异文件
    msg = '是否将本次聚水潭差异添加到总差异文件中？'
    ISNO = easygui.boolbox(msg= msg)
    if ISNO:
        fname_addDiff = easygui.fileopenbox('本次需要添加的聚水潭编码与大库的差异文件"本次聚水潭编码异常addDiff.xlsx"')
        add_diff = chukliAddDiff(fname_addDiff)
        with pd.ExcelWriter(fname_jusuitanDiff, engine='openpyxl',mode='a', if_sheet_exists='overlay')  as writer:
            add_diff.to_excel(writer, sheet_name = 'Sheet1',startrow=max_row, header = False,index = False)
        
    else:
        qijian = easygui.enterbox(msg = '请输入期间"2025-01"')
        #求和
        df_jusuitanXiaoshou_chuku = df_jusuitanXiaoshou.copy()   
        # df_jusuitanXiaoshou_chuku = df_jusuitanXiaoshou_chuku[[jusuitanbianma1,'实发数量','content','商品编码']] 
        df_jusuitanXiaoshou_chuku = df_jusuitanXiaoshou_chuku.rename(columns = {'jusuitan': '存货编码', '数量': '数量', 'content': '含量', 'xiaoshoubu': '聚水潭编码'}) 
        df_jusuitanXiaoshou_chuku = df_jusuitanXiaoshou_chuku.set_index('存货编码')
        #加合计数
        total = df_jusuitanXiaoshou_chuku['数量'].sum()
        df_jusuitanXiaoshou_chuku.loc['合计'] = [total,'','']
       
        df_jusuitanXiaoshou18 = df_jusuitanXiaoshou0[lst]
    
        #df_jusuitanXiaoshou8求和
        df_jusuitanXiaoshou18 = df_jusuitanXiaoshou18[~df_jusuitanXiaoshou18['商品编码'].isna()]
        total8 = []
        lst1 = lst[1:]
        for i in lst1:
            if i != '净毛利率':
                total = df_jusuitanXiaoshou18[i].sum()
                total8.append(total)
            else :
                total = ''
                total8.append(total)
      
        df_jusuitanXiaoshou18 = df_jusuitanXiaoshou18.set_index('商品编码')
        df_jusuitanXiaoshou18.loc['合计'] = total8
        df_jusuitanXiaoshou18 = df_jusuitanXiaoshou18.reset_index()
        # df_jusuitanXiaoshou18['商品编码1'] = df_jusuitanXiaoshou18['商品编码'].map(content_dic)
                 
        path = r"F:\a00nutstore\008\zww08\002电商\聚水潭"
        file = f'电商产品销售出库明细-{qijian}.xlsx'
        fname_jst = os.path.join(path,file)
        wb1 = openpyxl.Workbook()
        ws1 = wb1.active
        ws1.title = '原表'
        wb1.save(fname_jst)
        wb1 = openpyxl.load_workbook(fname_jst)
        ws1 = wb.active
        with pd.ExcelWriter(fname_jst, engine='openpyxl',mode='a', if_sheet_exists='overlay')  as writer:
            df_jusuitanXiaoshou0.to_excel(writer,sheet_name = '原表',header = True,index = False)
            df_jusuitanXiaoshou_chuku.to_excel(writer, sheet_name = '电商出库', header = True)
            df_jusuitanXiaoshou18.to_excel(writer, sheet_name = '销售出库', header = True,index = False)
        easygui.msgbox(msg = '程序结束')
        os.startfile(fname_jst)

if __name__=='__main__':
    main()               
       

