'''
仓库供应商，运行后形成的列表保表到py文件
'''
import openpyxl
import pprint

def cangkuGongyingshang(fname):

    wb = openpyxl.load_workbook(fname)
    ws1 = wb.active

    if  '仓库供应商' not in wb.sheetnames :
        wb.create_sheet(title = '仓库供应商')
    else :
        pass

    ws2 = wb['仓库供应商']

    cangkuGongyingshangs = []

    for row in range(1,ws1.max_row):
        if row == 1 :
            ws2.cell(row, 1).value = '仓库供应商'
        else :
            ws2.cell(row, 1).value = ws1.cell(row, 3).value
            cangkuGongyingshangs.append(ws2.cell(row, 1).value)

    wb.save(fname)
    resultFile = open('cangkuGongyingshangs.py', 'w', encoding='utf-8')
    resultFile.write('cangku = ' + pprint.pformat(cangkuGongyingshangs))
    return cangkuGongyingshangs

def main():
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\仓库供应商.XLSx'
    cangkuGongyingshang(fname)

if __name__=='__main__':
    main()




