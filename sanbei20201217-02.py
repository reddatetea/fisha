# 导入必要的库
import openpyxl
from openpyxl.styles import Font,Border,Side

# 新建一个工作簿
wb = openpyxl.Workbook()
ws = wb.active


# 设定单元格内容
ws['c1'].value = '杰尼！'

# 先设定框线的样式和颜色
medium_blue = Side(border_style='medium', color='0000ff')
# 组装边框
medium_blue_border = Border(
  top=medium_blue,
  bottom=medium_blue,
  left = medium_blue,
  right = medium_blue)
ws['C1'].border = medium_blue_border

# 请不要修改最后一行的文件保存代码哦
wb.save('杰尼龟.xlsx')