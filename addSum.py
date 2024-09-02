# 纸箱求和公式，指定工作簿、工作表和指定求和字段
import os
import excelmessage
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import easygui

#装饰器
def addsums(fun):
    def wrapper(fname1,sheetnames,letters,*args):  # 传递的参数分别是：excel文件名、工作表名、需要合计的列名所对应的字母
        ret_val = fun(fname1,sheetnames,letters,*args)

        wb = openpyxl.load_workbook(fname1)
        for sheetname in sheetnames:
            ws = wb[sheetname]
            max_row = ws.max_row
            for letter in letters:
                ws['A{}'.format(max_row + 1)].value = '合计'
                ws['{}{}'.format(letter, max_row + 1)].value = '=sum({}2:{}{})'.format(letter, letter, max_row)
            wb.save(fname1)
        return ret_val

    return wrapper

def sumheji(fname1,sheetnames,letters):
    wb = openpyxl.load_workbook(fname1)
    for sheetname in sheetnames:
        ws = wb[sheetname]
        max_row = ws.max_row
        for letter in letters:
            ws['A{}'.format(max_row + 1)].value = '合计'
            ws['{}{}'.format(letter, max_row + 1)].value = '=sum({}2:{}{})'.format(letter, letter, max_row)
        wb.save(fname1)

def main():
    msg = '请点选要加计合计数的excel文件'
    print(msg)
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    wb = openpyxl.load_workbook(fname)
    strs = wb.sheetnames
    msg = '请点选要求和的工作表'
    sheetnames = easygui.multchoicebox(msg, choices=strs)
    wb.close()
    letters = 'KPRT'
    sumheji(fname,sheetnames,letters)

if __name__ == '__main__':
    main()
