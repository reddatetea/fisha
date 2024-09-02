'''
核对计算苏州三鑫背条的的价格和金额
'''
import beitiaojisuan
import openpyxl
from openpyxl.utils import column_index_from_string
import tiqushuju
import qijiansplit
import os
import easygui
import addSum

@addSum.addsums
def beitiaoHedui(fname1,sheetnames,letters,ws_name,im_gongyingshang):
    wb = openpyxl.load_workbook(fname1)
    print('ws_name',ws_name)
    ws = wb[ws_name]
    sheetnames = wb.sheetnames
    maxrow = ws.max_row
    print(maxrow)
    pinming_num = column_index_from_string('D')
    shuliang_num = column_index_from_string('F')
    hetong_danjia_num = column_index_from_string('S')
    hetong_jiner_num = column_index_from_string('T')
    ruku_jiner_num = column_index_from_string('H')
    danjia_duoji_num = column_index_from_string('U')
    jiner_duoji_num = column_index_from_string('V')

    for row in range(1,maxrow+1):
        if row == 1 :
            ws.cell(row,hetong_danjia_num,value='合同单价')
            ws.cell(row, hetong_jiner_num, value='合同金额')
            ws.cell(row, danjia_duoji_num, value='单价多计')
            ws.cell(row, jiner_duoji_num, value='金额多计')

        else :
            pinming = ws.cell(row,pinming_num).value
            kuan,beitiao_price= beitiaojisuan.beitiaoJisuan(pinming,im_gongyingshang)
            hetong_danjia = beitiao_price*kuan
            ws.cell(row,hetong_danjia_num,value = hetong_danjia)
            hetong_jiner = round(ws.cell(row,shuliang_num).value*hetong_danjia,2)
            ws.cell(row, hetong_jiner_num, value=hetong_jiner)
            danjia_duoji =   ws.cell(row,shuliang_num+1).value - hetong_danjia
            ws.cell(row,danjia_duoji_num,value = danjia_duoji)
            jiner_duoji  =  ws.cell(row,shuliang_num+2).value - hetong_jiner
            ws.cell(row,jiner_duoji_num,value = jiner_duoji)

    wb.save(fname1)
    return fname1,sheetnames


def main():
    # fname = easygui.fileopenbox(msg = '请选择原材料实时流水账',title = '实时流水账')
    # path,filename = os.path.split(fname)
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    im_gongyingshangs = ['苏州市三鑫','昆山楚宏']
    msg = '请点选供应商'
    im_gongyingshangs = easygui.multchoicebox(msg, title='', choices=im_gongyingshangs)
    for im_gongyingshang in im_gongyingshangs:
        im_qijian0 = easygui.enterbox(msg='请输入要查询{}的期间,如果查询多个期间，中间用and连接'.format(im_gongyingshang), title='期间')
        im_qijians = qijiansplit.shijiqijian(im_qijian0)
        fname1, ws_name = tiqushuju.tiquShuju(fname, im_gongyingshang, im_qijians)
        print(ws_name)
        # fname1, sheetnames = beitiaoHedui(fname1, ws_name)
        letters = 'FHTV'
        # zhixiangheji.sumheji(fname1, sheetnames, letters)
        sheetnames = [ws_name]
        fname1,sheetnames = beitiaoHedui(fname1,sheetnames,letters,ws_name,im_gongyingshang)
        path, filename = os.path.split(fname1)
        choice = easygui.ccbox(msg='是否打开要查询的文件"%s"' % filename, title='{}背条价格核对'.format(im_gongyingshang), choices=('yes', 'no'))
        if choice == True:
            os.startfile(fname1)
        else:
            pass


if __name__ == '__main__' :
    main()









