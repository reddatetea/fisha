import os
import openpyxl
from openpyxl.styles import Font,Border,Side,Fill,Alignment
import excelmessage
import  easygui
import datetime

def addnewTable(fname,riqi):
    path,filename = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    if 'AAA' in sheetnames:
        ws1 = wb['AAA']
    else :

        msg = '请点选"当日"工作表'
        print(msg)
        choice = easygui.choicebox(msg,choices=sheetnames)
        ws1 = wb[choice]

    max_row1 = ws1.max_row
    leibie_dic = {ws1['A{}'.format(row)].value: (ws1[row][3].value, ws1[row][4].value, ws1[row][5].value) for row in  range(2, max_row1 + 1)}
    pinming_dic = {ws1['B{}'.format(row)].value:(ws1[row][3].value,ws1[row][4].value,ws1[row][5].value) for row in range(2,max_row1+1)}
    #print(pinming_dic)                             #生成品名的当日字典

    if 'BBB' in sheetnames:
        ws2 = wb['BBB']
    else:

        msg = '请点选"累计"工作表'
        print(msg)
        choice = easygui.choicebox(msg, choices=sheetnames)
        ws2 = wb[choice]

    #新建原材料日报表工作簿
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = '原材料日报表'
    for  index,row in enumerate(ws2.values):
        ws3.append(row)

    ws3['D1'].value = '月初'
    ws3['E1'].value = '入库累计'
    ws3['F1'].value = '出库累计'
    ws3['G1'].value = '结余'

    ws3.insert_cols(5,2)                      #列为实数，第几列就是第几列
    ws3.insert_cols(8,1)
    ws3['E1'].value = '上日'
    ws3['F1'].value = '本日入库'
    ws3['H1'].value = '本日出库'

    fname1 = '原材料日报表{}.xlsx'.format(riqi)
    wb3.save(fname1)
    wb3 = openpyxl.load_workbook(fname1)
    ws3 = wb3['原材料日报表']
    max_row3 = ws3.max_row

    for row in range(2,max_row3+1):
        leibie = ws3.cell(row,1).value
        pinming = ws3.cell(row, 2).value
        if  pinming not in [None,'','合   计']:
            shangri = pinming_dic.get(pinming, (0, 0, 0))[0]
            benriruku = pinming_dic.get(pinming, (0, 0, 0))[1]
            benrichuku = pinming_dic.get(pinming, (0, 0, 0))[2]
            ws3.cell(row, 5).value = shangri
            ws3.cell(row, 6).value = benriruku
            ws3.cell(row, 8).value = benrichuku

        else :
            shangri = leibie_dic.get(leibie, (0, 0, 0))[0]
            benriruku = leibie_dic.get(leibie, (0, 0, 0))[1]
            benrichuku = leibie_dic.get(leibie, (0, 0, 0))[2]
            ws3.cell(row, 5).value = shangri
            ws3.cell(row, 6).value = benriruku
            ws3.cell(row, 8).value = benrichuku



    wb3.create_sheet(title = '原材料零结存累计表')
    ws4 = wb3['原材料零结存累计表']
    del_rows = []
    for index,row in enumerate(ws3.values):
        if index ==0:
            ws4.append(row)
        else :
            if row[-1] ==0:
                del_rows.append(index+1)
                ws4.append(row)

    del_rows.reverse()
    #print(del_rows)
    for j in del_rows:
        ws3.delete_rows(j)
    wb.close()
    wb3.save(fname1)
    return fname1


def printseting(fname1,riqi):
    #设置字体、边框、对齐等常量
    font = Font(name='宋体', size=10)
    font_xiaoji = Font(name='宋体', bold=True, size=10)
    font_firstrow = Font(name='宋体', bold=True, size=20)
    thin_danbian = Side(border_style='thin')
    double_danbian = Side(border_style='double')
    thin_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=thin_danbian)
    double_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=double_danbian)

    wb = openpyxl.load_workbook(fname1)
    for ws in wb.worksheets :
        #单元格宽度
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 4
        for j in 'DFGHI':
            ws.column_dimensions['{}'.format(j)].width = 7.44
        ws.column_dimensions['J'].width = 9.44

        #插入第一行
        ws.insert_rows(1)
        ws['A1'].value = '双佳' + ws.title +' ' +riqi
        ws['A1'].font = font_firstrow
        ws['A1'].alignment = Alignment(horizontal='center',vertical='center')
        ws.merge_cells('A1:J1')

        max_row = ws.max_row
        max_column = ws.max_column

        for row in range(2,max_row+1):                  #从第二行开始，设置单元格格式
            if row == 2:
                for col in range(1, max_column + 1):
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
            else :
                for col in range(1, max_column + 1):
                    if col <= 3:
                        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        ws.cell(row, col).alignment = Alignment(horizontal='right', vertical='center')


        for cell in ws['A']:
            row = cell.row
            if row == 1:
                continue
            else:
                if '小计' in cell.value:
                    for col in range(1, max_column + 1):
                        ws.cell(row, col).font = font_xiaoji
                        ws.cell(row, col).border = double_bian
                        ws.merge_cells('A{}:B{}'.format(row, row))
                else:
                    for col in range(1, max_column + 1):
                        ws.cell(row, col).font = font
                        ws.cell(row, col).border = thin_bian

        ws.freeze_panes = ws['D3']
        ws.print_title_rows = '1:2'  # the first row
        # 页脚设置
        ws.oddFooter.center.text = " &[Page] / &N"  # 1/n
        ws.oddFooter.center.size = 12  # 页脚中字体大小
        ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
        ws.oddFooter.center.color = "000000"  # 页脚中字体颜色

        ws.oddFooter.right.text = "张文伟 &[Date]"  # 页脚右 文字
        ws.oddFooter.right.size = 12  # 页脚右 字体大小
        ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
        ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
        ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.1, left=0.51, right=0.1, footer=0.5,
                                                              bottom=1)
    wb.save(fname1)

def  main():
    msg = '请点选"原材料当日和累计"工作表'
    easygui.msgbox(msg=msg)
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    today_date = datetime.date.today()
    yesterday_date = (today_date + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")
    msg = '报表日期是{}?'.format(yesterday_date)
    choice = easygui.ccbox(msg, title='请选择"是"或"否"', choices=('是', '否'))
    if choice:
        riqi = yesterday_date
    else:
        msg = '请输入报表日期'
        riqi = easygui.enterbox(msg, title=" 昨天日期")
    fname1 = addnewTable(fname, riqi)
    printseting(fname1, riqi)
    easygui.msgbox(msg = '程序结束运行')
    os.system(fname1)

if __name__ == '__main__':
    main()
