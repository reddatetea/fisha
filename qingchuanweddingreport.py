import  easygui
import   os
import shutil
import  datetime
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font,Border,Side,Alignment,Fill,Protection

monthdic = {'1': 'January', '2': 'February', '3': 'March', '4': 'April', '5': 'May', '6': 'June', '7': 'July', '8': 'August', '9': 'September', '10': 'October', '11': 'November', '12': 'December'}
report_title ='Monthly Wedding Report'
sjqy2_title = [['No.', 'Book date', 'Wedding Date', 'Time', 'Name', 'Pax', 'BQT Lunch', None, 'BQT Dinner', None, '场租', '进场费', 'LED屏', '服务费+VAT     （1.166）', 'Total Revenue', '房费'], [None, None, None, None, None, 'BQT', 'Food', 'Beverage', 'Food', 'Beverage', None, None, None, None, None,None]]

#选取文件
msg = '请点选monthly weding report excel文件'
print(msg)
fname = easygui.fileopenbox(msg)
#fname = r'D:\a00nutstore\fishc\qingchuan\yanhui.xlsx'
path,filename = os.path.split(fname)
qian,hou = os.path.splitext(filename)
wb = openpyxl.load_workbook(fname,data_only=True)

sheetnames = wb.sheetnames
if len(sheetnames)==1:
    ws = wb.active
else :
    msg = '请点选源工作表：'
    print(msg)
    title = '源工作表'
    choice = easygui.buttonbox(msg,title,choices = sheetnames)
    ws = wb[choice]
ws_name  = ws.title
ws2 = wb.create_sheet(title = 'msr')
ws2.title = 'mwr'
ws2.append((report_title,))
ws2.append(('',))
ws2.append(('Month:',))
for row in sjqy2_title :
    ws2.append(row)

#合并单元格
for i in 'abcdeklmnop':
    merge_range = '{}4:{}5'.format(i,i)
    ws2.merge_cells(merge_range)

ws2.merge_cells('G4:H4')
ws2.merge_cells('I4:J4')
ws2.merge_cells('A1:R1')
#字体边框和对齐方式设置
ft0 = Font(name='Times New Roman',size=12)        #字体
ft = Font(name='Times New Roman',size=11)        #字体
bd = Border(left=Side(border_style="thin"),

        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
                    )
alignment=Alignment(vertical='center',horizontal='center',shrink_to_fit=True)    #,horizontal='center'
for row in ws2['A4:P5']:
    for cell in row :
        cell.font = ft0
        cell.border = bd
        cell.alignment = alignment

ws2['A1'].font = Font(name='Times New Roman',size=14)
ws2['A1'].alignment = alignment

ws2['A3'].font = Font(name='Times New Roman',size=12)
ws2['A3'].alignment = alignment

max_row1 = ws.max_row
max_column1  = ws.max_column
#print(max_row1,max_column1)
max_row2 = ws2.max_row

#当前日期转字符串
dqrq = datetime.date.today().strftime('%Y%m%d')

#备份以防万一
backup_path = r'd:\qingchuangbackup'
backup_fname =  os.path.join(backup_path,qian+dqrq+hou)
if not os.path.exists(backup_path):
      os.makedirs(backup_path)
if not os.path.exists(backup_fname):
    shutil.copyfile(fname,backup_fname)

#确定Sheet1数据区域
for  j in ws['A1:A{}'.format(max_row1)]:
    #print(j[0].row,j[0].value)
    if  j[0].value == 'Month:':
        ws2_min_row = j[0].row
        sjqy_min_row = j[0].row +3
        break

for j in ws['A{}:A{}'.format(sjqy_min_row+1,max_row1)]:
    if j[0].value == None:
        sjqy_max_row = j[0].row -1
        break

max_alp = openpyxl.utils.get_column_letter(max_column1)
sjqy = ws['A{}:{}{}'.format(sjqy_min_row,max_alp,sjqy_max_row)]
ws2.delete_rows(6,max_row2-6+1)

data = []
for row in sjqy:
    if row[6].value not in [None,0,''] :
        meihang = [cell.value for cell in row]
        data.append(meihang)
    else :
        continue

newdata =[]
for i in data :
    hang = []
    for j in i:
        if j in [None,'']:
            j = 0
            hang.append(j)
        else :
            j = j
            hang.append(j)
    newdata.append(hang)

#循环取数
pax_bqt_lunch ,pax_bqt_dinner= 0,0
bqt_lunch_food_t,bqt_lunch_bev_t = 0,0
bqt_dinner_food_t,bqt_dinner_bev_t = 0,0
btq_rev = 0
fangfei_t = 0

for i in newdata:
    pax = i[6]       #人数
    bqt_lunch_food = i[9]
    #bqt_lunch_food = format(bqt_lunch_food,'.2f')

    bqt_lunch_bev = i[10]
    bqt_dinner_food = i[11]
    bqt_dinner_bev = i[12]


    cp_lunch_food = i[13]
    cp_lunch_bev = i[14]
    cp_dinner_food = i[15]
    cp_dinner_bev = i[16]

    te_lunch_food = i[17]
    te_lunch_bev = i[18]
    te_dinner_food = i[19]
    te_dinner_bev = i[20]

    changzhu = i[21]
    jinchang = i[22]
    led = i[23]

    #total_rev = int(i[25]-(changzhu+jinchang+led)*1.06-(cp_lunch_food+cp_lunch_bev+cp_dinner_food+cp_dinner_bev)*1.166)
    total_rev = round(round(i[25],2) - (changzhu + jinchang + led) * 1.06 - (cp_lunch_food + cp_lunch_bev + cp_dinner_food + cp_dinner_bev) * 1.166,0)
    bqt_lunch_food = round(bqt_lunch_food,2)
    bqt_lunch_bev = round(bqt_lunch_bev, 2)
    bqt_dinner_food = round(bqt_dinner_food, 2)
    bqt_dinner_bev = round(bqt_dinner_bev, 2)

    fuwu = total_rev - (bqt_lunch_food+bqt_lunch_bev+bqt_dinner_food+bqt_dinner_bev)
    fangfei = round( i[26]/1.166,2)

    changzhu = round(changzhu,2)
    jinchang = round(jinchang,2)
    led = round(led,2)



    newhang = [i[0],i[1],i[2],i[3],i[4],pax,bqt_lunch_food,bqt_lunch_bev,bqt_dinner_food,bqt_dinner_bev,changzhu,jinchang,led,fuwu,int(total_rev),fangfei]

    ws2.append(newhang)
    #求各项合计数
    bqt_lunch_food_t+=round(bqt_lunch_food,2)
    bqt_lunch_bev_t += round(bqt_lunch_bev,2)
    bqt_dinner_food_t+=round(bqt_dinner_food,2)
    bqt_dinner_bev_t += round(bqt_dinner_bev,2)
    fangfei_t+=fangfei
    if bqt_lunch_food !=0 or bqt_lunch_food !=0 :
        pax_bqt_lunch+=pax
    else :
        pax_bqt_dinner+=pax

bqt_lunch_food_t = bqt_lunch_food_t -fangfei_t
bqt_rev = bqt_lunch_food_t + bqt_lunch_bev_t + bqt_dinner_food_t + bqt_dinner_bev_t


#添加公式
heji_row = int(5 + len(newdata) + 1)

col =  column_index_from_string('F')
alp = get_column_letter(col)
# ws['F16'].value  =  "=SUM('{}6:{}{}')".format(alp, alp, heji_row - 1)
for  row in ws2['F{}:P{}'.format(heji_row,heji_row)]:
    for j in row:
        alp = get_column_letter(col)
        j.value  =  "=SUM({}6:{}{})".format(alp, alp, heji_row - 1)
        col = col + 1

#修改序号
k = 1
for row in ws2['A6:A{}'.format(heji_row-1)]:
    row[0].value = k
    k = k + 1

rq0 = ws2['C6'].value
rq = format(rq0,'%Y年%m月')
month_num = str(int(format(rq0,'%m')))
month = monthdic[month_num]

#第二列日期格式处理,处理前先取出年月

for row in ws2['B6:C{}'.format(heji_row-1)]:
    for cell in row :
        a = cell.value
        a = format(a,'%m/%d/%Y')
        cell.value = a

#设置时间为24进制
for row in ws2['D6:D{}'.format(heji_row-1)]:
    for cell in row :
        a = cell.value
        a = format(a,'%H:%M')
        cell.value = a

#字体边框和对齐方式设置
ft = Font(name='Times New Roman',size=11)        #字体

bd = Border(left=Side(border_style="thin"),

        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
                    )
alignment=Alignment(vertical='center',horizontal='center',shrink_to_fit=True)    #,horizontal='center'
alignment2 = Alignment(vertical='center',horizontal='center',wrap_text=True)

for  row in ws2['A6:P{}'.format(heji_row)]:
    for cell in row:
        cell.font = ft
        cell.border = bd
        cell.alignment = alignment
        #cell.number_format = number_format

#合计区域添加数据
heji_item=[rq, '宴会午餐人数', '宴会午餐食品', '宴会午餐酒水', '宴会晚餐人数', '宴会晚餐食品', '宴会晚餐酒水', '', '', '总合计数']
hejishus = ['',pax_bqt_lunch,bqt_lunch_food_t,bqt_lunch_bev_t,pax_bqt_dinner,bqt_dinner_food_t,bqt_dinner_bev_t,'','',bqt_rev]
heji_quyu_start_row = len(newdata) + 5 +4
heji_quyu_end_row = heji_quyu_start_row+9
heji_quyu = ws2['E{}:F{}'.format(heji_quyu_start_row,heji_quyu_end_row)]

index = 0
for row in heji_quyu:
    row[0].value = heji_item[index]
    row[1].value = hejishus[index]
    index = index + 1

for row in ws2['E{}:F{}'.format(heji_quyu_start_row,heji_quyu_end_row-3)]:
    for cell in row :
        cell.font = ft
        cell.border = bd
        cell.alignment = alignment

for row in ws2['E{}:F{}'.format(heji_quyu_end_row,heji_quyu_end_row)]:
    for cell in row :
        cell.font = ft
        cell.border = bd
        cell.alignment = alignment

ws2['B3'].value = month
ws2['B3'].font = Font(name='Times New Roman',size=12)        #字体
ws2['B3'].alignment = alignment

ws2.column_dimensions['N'].width = 15               #列宽
ws2['N4'].alignment = alignment2                    #自动换行
ws2.column_dimensions['E'].width = 25             #列宽
wb.save(fname)
os.system(fname)
