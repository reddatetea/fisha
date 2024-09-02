'''
打开彩皇线圈价格表，形成彩皇线圈规格和价格字典
'''
import openpyxl

def caihuang_price():
    fname = r'D:\a00nutstore\006\zw\xianquan(xianhuan)\彩皇线环线圈价格表03.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb['线圈价格表']
    maxrows = ws.max_row

    caihuang_xianquan_dic = {}

    for row in range(2,maxrows+1):
        guige = ws.cell(row,1).value
        price = ws.cell(row,2).value
        caihuang_xianquan_dic[guige]= price
        wb.close()

    print(caihuang_xianquan_dic)
    return caihuang_xianquan_dic

def main():
    caihuang_price()

if __name__ == '__main__' :
    main()


