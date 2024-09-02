import easygui
import openpyxl
import excelmessage
import os
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill

msg = '请点选"销售模板文件sale_template.xlsx"文件'
print(msg)
moban = easygui.fileopenbox(msg)
path0,filename = os.path.split(moban)
os.chdir(path0)
wb0 = openpyxl.load_workbook(moban)
ws0 = wb0.active
mingchens = [j.value for j in ws0['B']][2:]
print(mingchens)
max_row = ws0.max_row
max_column = ws0.max_column
meitian = r'leiji_xiaoshou'
sale_huzhong_path = os.path.join(path0, meitian)
if not os.path.exists(sale_huzhong_path):
    os.makedirs(sale_huzhong_path)

def saleHuzhong(fname):
    path, file = os.path.split(fname)
    os.chdir(path)
    # fname = excelmessage.excelMessage(fname)
    wb1 = openpyxl.load_workbook(fname)
    ws1 = wb1.active
    sale_mingchens = [j.value for j in ws1['C']]
    print(sale_mingchens)
    a = ws1['A4'].value
    b = a.replace('(', '')
    c = b.replace(')', '')
    d = c.split('-')
    riqi = str(d[1]) + str(d[2])
    print(riqi)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = riqi
    file_riqi = '{}.xlsx'.format(riqi)
    fname_riqi = os.path.join(sale_huzhong_path,file_riqi)
    for index, row in enumerate(ws0.values):
        ws2.append(row)
    wb2.save(fname_riqi)
    wb2 = openpyxl.load_workbook(fname_riqi)
    ws2 = wb2[riqi]
    ws2.cell(1, 4).value = riqi
    ws2.cell(2, 4).value = '数量'
    ws2.cell(2, 5).value = '金额'
    for i in range(3, len(mingchens) + 3+1):
        sale_mingchen = ws2.cell(i,2).value
        if sale_mingchen in sale_mingchens:
            sale_row = sale_mingchens.index(sale_mingchen)+1
            shuliang = ws1['AH{}'.format(sale_row)].value
            jiner = ws1['AM{}'.format(sale_row)].value
        else:
            shuliang = ''
            jiner =''
        ws2.cell(i, max_column+1).value = shuliang
        ws2.cell(i, max_column+2).value = jiner
    wb2.save(fname_riqi)
    wb1.close()

msg = '请点选itemize_sales_analysis每日销售所在文件夹'
print(msg)
path1 = easygui.diropenbox(msg)
filenames = os.listdir(path1)
for filename in filenames:
    filename = os.path.join(path1,filename)
    newfname =excelmessage.excelMessage(filename)
    saleHuzhong(newfname)
#点选需要累计计算的sale文件，后缀为.xlsx
salefiles =[j for j in os.listdir(sale_huzhong_path) if j.split('.')[-1]=='xlsx']
salefiles.sort()
msg = '请点选需要累计计算的sale文件'
salefiles = easygui.multchoicebox(msg,choices=salefiles)
total_name = '模板累计销售.xlsx'
sale_grand_total_name = os.path.join(sale_huzhong_path,total_name)
wb_t = openpyxl.Workbook()
ws_t = wb_t.active
ws_t.title = '累计'
for index, row in enumerate(ws0.values):
    ws_t.append(row)
for j in range(len(salefiles)):
    column1 = max_column + 2*j +1
    column2 = column1 +1
    file = os.path.join(sale_huzhong_path,salefiles[j])
    wb_j = openpyxl.load_workbook(file)
    ws_j = wb_j.active
    str1 = get_column_letter(max_column+1)
    str2 = get_column_letter(max_column +2)
    shuliangs = [k.value for k in ws_j['{}'.format(str1)]][2:]
    jiners = [k.value for k in ws_j['{}'.format(str2)]][2:]
    ws_t.cell(1,column1).value = ws_j.cell(1,max_column+1).value
    ws_t.cell(2, column1).value = ws_j.cell(2, max_column+1).value
    ws_t.cell(2, column2).value = ws_j.cell(2, max_column+2).value
    for i in range(3,len( shuliangs)+3):
        ws_t.cell(i,column1).value = shuliangs[i-3]
        ws_t.cell(i, column2).value = jiners[i - 3]
    column_letter1 = get_column_letter(column1)
    column_letter2 = get_column_letter(column2)
    wb_j.close()
#横向求和
max_row = ws_t.max_row
for k in range(3,max_row+1):
    shuliang_t = ''
    for n in range(max_column+1,column1+1,2):
        str = get_column_letter(n)
        shu_liang = '+{}{}'.format(str,k)
        shuliang_t ='{}{}'.format(shuliang_t,shu_liang)
    ws_t.cell(k,column1+2).value = '={}'.format(shuliang_t)
for k in range(3,max_row +1):
    jiner_t = ''
    for n in range(max_column+2,column2+1,2):
        str = get_column_letter(n)
        jin_er = '+{}{}'.format(str,k)
        jiner_t ='{}{}'.format(jiner_t,jin_er)
    ws_t.cell(k,column2+2).value = '={}'.format(jiner_t)
# 竖向求和
for m in range(max_column+1,column2+2+1):
    str = get_column_letter(m)
    ws_t['{}{}'.format(str,i+1)].value = '=sum({}3:{}{})'.format(str,str,i)
ws_t['A{}'.format(i+1)].value = '合计'
ws_t.cell(1,column1+2).value = '合计'
ws_t.cell(2, column1 + 2).value = '数量'
ws_t.cell(2, column2 + 2).value = '金额'
wb_t.save(sale_grand_total_name)
wb0.close()


def setStyles(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb['累计']
    max_row = ws.max_row
    max_column = ws.max_column
    font = Font(name = '等线',size = '11')
    bd = Border(left=Side(border_style="thin"),

                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
                )
    alignment1 = Alignment(vertical='center', horizontal='center')                           #垂直方向和水平方向居中
    alignment2 = Alignment(vertical='center', horizontal='right')
    pattern_fill = PatternFill(fill_type='solid', fgColor='fdeff2')  # 薄樱
    for i in range(3,max_row+1):
        for j in range(1,3+1):
            if j==1 or j ==3:
                ws.cell(i,j).alignment = alignment1
            else :
                continue
    for  i in range(1,max_column+1+1):                                  #设置列宽
        str = get_column_letter(i)
        if  i ==1 :
            ws.column_dimensions['{}'.format(str)].width = '6'
        elif  i==2 :
            ws.column_dimensions['{}'.format(str)].width = '40'
        elif  i==3 :
            ws.column_dimensions['{}'.format(str)].width = '7'
        else :
            ws.column_dimensions['{}'.format(str)].width = '6'
    ws.merge_cells('A1:A2')
    ws.cell(1,1).alignment = alignment1
    ws.merge_cells('B1:B2')
    ws.cell(1,2).alignment = alignment1
    ws.merge_cells('C1:C2')
    ws.cell(1,3).alignment = alignment1

    for  j in range(4,max_column+1,2):
        str1 = get_column_letter(j)
        str2 = get_column_letter(j+1)
        ws.merge_cells('{}1:{}1'.format(str1,str2))
        ws.cell(1,j).alignment = alignment1
        ws.cell(2,j).alignment = alignment1
        ws.cell(2, j+1).alignment = alignment1

    #设置第一行、第二行和最后一行填充
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            ws.cell(i,j).border = bd
            if i == 1 or i == 2 or i == max_row:
                ws.cell(i,j).fill =pattern_fill
            else :
                continue
    ws.freeze_panes = 'D3'                     #冻结单元格
    title = '晴川特卖数量金额明细账'
    name = '贾莉霞'
    # 打印标题
    ws.print_title_cols = 'A:C'  # the first two cols   #左端标题列（即从左侧重复的列数）
    ws.print_title_rows = '1:2'  # the first row     #顶端标题行（即从上端重复的行数）
    # 页眉设置
    ws.oddHeader.center.text = '%s' % title  # 1/n      页眉中文字
    ws.oddHeader.center.size = 18  # 页眉中字体大小
    ws.oddHeader.center.font = '宋体,Bold Italic'  # 页眉中字体,宋体、加粗、倾斜
    ws.oddHeader.center.color = "000000"  # 页眉中字体颜色
    # 页脚设置
    ws.oddFooter.center.text = " &[Page] / &N"  # 1/n      页脚中文字
    ws.oddFooter.center.size = 12  # 页脚中字体大小
    ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
    ws.oddFooter.center.color = "000000"  # 页脚中字体颜色
    ws.oddFooter.right.text = "{} &[Date]".format(name)  # 页脚右 文字
    ws.oddFooter.right.size = 12  # 页脚右 字体大小
    ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
    ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
    wb.save(fname)

setStyles(sale_grand_total_name)
os.startfile(sale_grand_total_name)