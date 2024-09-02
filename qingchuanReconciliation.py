# -*-coding = utf-8 -*-
import pdfplumber
import re
import openpyxl
import easygui
import os
from  xlsxlsx import xlsXlsx


def shengji(fname):
    path, filename = os.path.split(fname)
    if os.path.splitext(filename)[1] == '.xls':
        fname = xlsXlsx(fname)
    else:
        fname = fname
    return fname,path

def writeDangtian(fname):
    pdf = pdfplumber.open(fname)
    # print('开始读取数据')
    filename = 'BALANCE SHEET.txt'
    with open(filename, 'w') as f:
        f.write('')
    with open(filename, 'a') as f:
        for page in pdf.pages:
            content = page.extract_text()
            f.write(content)
    pdf.close()
    return filename

def txtToData(filename):
    data0 = []
    with open(filename) as f:
       for readline in f.readlines():
            data0.append(readline.strip())
    return data0


def getSubjectCredit(str, qita, qita_subject_dic):
    shuzhi = r'.*\s{1}(?P<credit>\(?\d{1,3}(,\d{3})*\.\d{2}\)?)\s{1}(?:\(?\d{1,3}(,\d{3})*\.\d{2}\)?)\s{1}(?:\(?\d{1,3}(,\d{3})*\.\d{2}\)?)$'
    pattern = r'^(?P<subject>\d{5}).*\s{1}(?P<credit>\(?\d{1,3}(,\d{3})*\.\d{2}\)?)\s{1}(?:\(?\d{1,3}(,\d{3})*\.\d{2}\)?)\s{1}(?:\(?\d{1,3}(,\d{3})*\.\d{2}\)?)$'
    regex_subject_credit = re.compile(pattern)
    mat_subject_credit = regex_subject_credit.search(str)
    if mat_subject_credit:
        sub = mat_subject_credit.group('subject')
        cre = mat_subject_credit.group('credit')
    else:
        for i in qita:
            pattern_qita = r'^{}{}'.format(i, shuzhi)
            regex_qita = re.compile(pattern_qita)
            mat = regex_qita.search(str)
            if mat:
                sub = qita_subject_dic[i]
                cre = mat.group('credit')
                break
            else:
                sub = ''
                cre = ''
    return sub, cre

def main():
    qita = ['LESS: PROV. FOR BAD DEBT', 'ACCOUNT PAYABLE - TRADE', 'AMOUNT DUE TO IHG',
            'PROVISION FOR CAPITAL REPLACEMENT', 'WITHDRAWAL - OWNING COMPANY', 'BALANCE BROUGHT FORWARD',
            'ACCUMULATED EARNING YEAR-TO-DATE']
    qita_subject = ['12340', '21010', '21050', '25100', '29010', '29020', '29030']
    qita_subject_dic = dict(zip(qita, qita_subject))
    msg = '请点选BALANCE SHEET .pdf文件'
    fname = easygui.fileopenbox(msg)
    path, file = os.path.split(fname)
    os.chdir(path)
    filename = writeDangtian(fname)
    data0 = txtToData(filename)
    dic = {}
    for i in data0:
        sub, cre = getSubjectCredit(i, qita, qita_subject_dic)
        cre = cre.strip()
        cre = cre.replace(',', '')
        if sub not in [None, '']:
            if cre.startswith('('):
                cre = cre.replace('(', '')
                cre = cre.replace(')', '')
                cre = float(cre) * (-1)
                dic[sub] = cre
            else:
                cre = float(cre)
                dic[sub] = cre
        else:
            continue
    fname_rec = easygui.fileopenbox('请点选Reconciliation  .xls文件')
    fname_rec, path = shengji(fname_rec)
    wb = openpyxl.load_workbook(fname_rec)
    sheetnames = wb.sheetnames
    for key, value in dic.items():
        if key == '21050':
            ws = wb['2105&21100']
            ws['k19'].value = value
        else :
            if int(key) >= 12010 and key in sheetnames:
                ws = wb[key]
                ws['k19'].value = value
            else:
                continue
    wb.save(fname_rec)
    os.startfile(fname_rec)

if __name__ == '__main__':
    main()


