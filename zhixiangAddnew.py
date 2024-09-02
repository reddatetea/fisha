'''
将当月以前没有的纸箱添加到2020纸箱‘纸箱尺寸’中
'''
# _*_ conding:utf-8 _*_

import os
import openpyxl
from openpyxl.utils import column_index_from_string
import easygui

def addnew():
    fname = r'F:\a00nutstore\006\zw\ZHIXIANG\2020纸箱入库.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb['纸箱尺寸']
    maxrows = ws.max_row
    msg = '请打开新纸箱所在excel文件'
    title='当月纸箱'
    fname1 = easygui.fileopenbox(msg=msg,title=title)
    wb1 = openpyxl.load_workbook(fname1)
    sheetnames = wb1.sheetnames
    msg1 = '请选择当月正工作表'
    title1 = '新纸箱工作表'
    choice = easygui.buttonbox(msg=msg1,title=title1,choices=sheetnames)
    ws1 = wb1[choice]
    maxrow1s = ws1.max_row
    pinming_num = column_index_from_string('E')
    chang_num = column_index_from_string('G')
    kuan_num = column_index_from_string('H')
    gao_num = column_index_from_string('I')
    beizhu_num = column_index_from_string('U')

    i = 0
    for row in range(2,maxrow1s+1):
        beizhu = ws1.cell(row, beizhu_num).value
        if beizhu not in ['',None]:
            i = i+1
            gongyingshang = ws1.cell(row, 3).value
            riqi = ws1.cell(row, 1).value
            danhao = ws1.cell(row, 2).value
            pinming = ws1.cell(row,pinming_num).value
            chang = ws1.cell(row, chang_num).value
            kuan = ws1.cell(row, kuan_num).value
            gao = ws1.cell(row, gao_num).value
            #添加到纸箱尺寸中
            ws.cell(maxrows+i,2,value = gongyingshang)
            ws.cell(maxrows + i, 3, value= riqi)
            ws.cell(maxrows + i, 3).number_format = 'yyyy-m-d'
            ws.cell(maxrows + i, 4, value= danhao)
            ws.cell(maxrows + i, 5, value= pinming)
            ws.cell(maxrows +i, 6, value= chang)
            ws.cell(maxrows + i, 7, value= kuan)
            ws.cell(maxrows + i, 8, value= gao)
        else :
            continue
    wb.save(fname)
    wb1.close()
    easygui.msgbox('程序结束')


def main():
    os.chdir(r'F:\a00nutstore\006\zw\duizhang')
    addnew()


if __name__ == '__main__':
    main()

