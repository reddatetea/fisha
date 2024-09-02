'''
bypricedic.py
根据白云入库制作白云指定期间的价格字典
本版增加传入期间参数处理
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl

def bypriceDic(qijian):
    path = r'd:\a00nutstore\006\zw\baiyun'
    os.chdir(path)
    filename = r'2020白云入库.xlsx'
    fname = os.path.join(path,filename)
    wb = openpyxl.load_workbook(fname,data_only=True,read_only=True)
    ws = wb['2020']
    maxrow = ws.max_row
    byprices = []
    gongying_price = []
    ling_dun_jiner = []
    gongyingshang = '驻马店白云纸业有限公司'
    for row in range(2, maxrow + 1):
        dangqian_qijian = ws['E{}'.format(row)].value
        jizhang = ws['N{}'.format(row)].value
        if (dangqian_qijian == qijian) and (jizhang not in ['', None]):
            pricename = ws['Q{}'.format(row)].value
            price = ws['K{}'.format(row)].value
            lingshu = ws['H{}'.format(row)].value
            dunshu = ws['J{}'.format(row)].value
            jiner = ws['L{}'.format(row)].value
            if lingshu in ['', None]:
                lingshu = 0
            if dunshu in ['', None]:
                dunshu = 0
            if jiner in ['', None]:
                jiner = 0
            if [gongyingshang, pricename, price] not in gongying_price:
                gongying_price.append([gongyingshang, pricename, price])
                ling_dun_jiner.append([lingshu, dunshu, jiner])
                c = [gongyingshang, pricename, price], [lingshu, dunshu, jiner]
                byprices.append(c)
            else:
                k = [byprice[0] for byprice in byprices]
                print(k)
                j = k.index([gongyingshang, pricename, price])
                byprices[j][1][0] = byprices[j][1][0] + lingshu
                byprices[j][1][1] = byprices[j][1][1] + dunshu
                byprices[j][1][2] = byprices[j][1][2] + jiner
    print(byprices)
    wb.close()
    return byprices

def main():
    qijian = input('请输入期间，如2020-04\n')
    byprices = bypriceDic(qijian)

if __name__ == '__main__':
    main()







