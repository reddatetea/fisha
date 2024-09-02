import openpyxl
from openpyxl.styles import Font,PatternFill,GradientFill,Side,Border,Alignment
import easygui
import os

def excelcellFill(fname,ws_name,max_row,max_col,min_row=1,min_col=1):
    font = Font(name='微软雅黑', size=12, bold=True, italic=True, color='000000')
    font1 = Font(name='微软雅黑', size=10, bold=False, italic=False, color='000000')
    pattern_fill = PatternFill(fill_type='solid', fgColor='40E0D0')  # 绿宝石
    pattern_fill1 = PatternFill(fill_type='solid', fgColor='AFEEEE')  # 苍白的绿宝石
    side = Side(style=None, color='000000')  # 边框 线型为thin,黑色
    border = Border(left=side, right=side, top=side, bottom=side)  # 定义边框的左右上下

    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False)
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    # ws.print_title_cols = 'A:B'  # the first two cols
    ws.print_title_rows = '1:1'  # the first row
    ws.freeze_panes = 'B2'
    for i in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in i:
            # cell.font = font1
            if cell.row == min_row or  cell.row == max_row :
                cell.font = font
                cell.border = border
                cell.fill = pattern_fill
                cell.alignment = alignment

            elif cell.row % 2 != 0 :
                cell.font = font1
                cell.fill = pattern_fill1
            else :
                cell.font = font1
                continue
    wb.save(fname)
    return fname

def main():
    msg = '请点选要设置单元格填充的excel文件'
    fname = easygui.fileopenbox(msg,title=msg)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选要设置单元格填充的excel工作表'
    ws_name = easygui.buttonbox(msg, title='工作表', choices=sheetnames)
    ws = wb[ws_name]
    max_row = ws.max_row
    max_col = ws.max_column
    wb.close()
    excelcellFill(fname,ws_name,max_row, max_col)

if __name__== '__main__':
    main()








