'''
本模块是从流水账中将纸箱入库数据引入
'''
# _*_ coding:utf-8 _*_
import openpyxl
import os
import datetime
import xianquanchicun
import caihuangpricedic
import addSum


def Dangyue(qijian):
    # 当天日期
    dtrq = datetime.date.today().strftime('%Y%m%d')
    path = r'F:\a00nutstore\006\zw\xianquan(xianhuan)'
    os.chdir(path)

    filename = '线圈当月入库%s.xlsx' % dtrq
    fname = os.path.join(path, filename)

    wb1 = openpyxl.Workbook()
    ws1 = wb1.create_sheet(title='当月')

    fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    # fname2 = r'D:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['流水账']

    # qijian = '2020-04'
    print(ws2.max_row)
    jianchen = {'彩皇': '武汉彩皇伟业包装材料有限公司',
                '东莞市双吉装订': '东莞市双吉装订器材有限公司',
                '博源装订': ''}
    for i in range(1, ws2.max_row + 1):
        if i == 1:
            taitou = (
                '日期', '单据号', '线圈类型', '规格', '齿数', '入库条数', '单价（元/齿）', '单价（元/条)', '金额', '对账单价', '对账金额', '差额', '供应商')
            ws1.append(taitou)
        else:
            if (ws2.cell(i, 3).value in jianchen) and (ws2.cell(i, 11).value == qijian):

                ws1.cell(i, 1, value=ws2.cell(i, 1).value)
                ws1.cell(i, 2, value=ws2.cell(i, 2).value)
                ws1.cell(i, 3, value=ws2.cell(i, 4).value)
                ws1.cell(i, 6, value=ws2.cell(i, 6).value)
                ws1.cell(i, 10, value=ws2.cell(i, 7).value)
                ws1.cell(i, 11, value=ws2.cell(i, 8).value)
                ws1.cell(i, 13, value=ws2.cell(i, 3).value)



            else:
                continue
    wb1.save(fname)
    wb2.close()

    # 当月正
    fname3 = r'F:\a00nutstore\006\zw\xianquan(xianhuan)\线圈当月入库%s.xlsx' % dtrq
    # fname3 = r'D:\a00nutstore\006\zw\xianquan(xianhuan)\线圈当月入库%s.xlsx' % dtrq
    wb3 = openpyxl.load_workbook(fname3)
    ws3 = wb3['当月']

    ws4 = wb3.create_sheet(title='当月正')

    for index, row in enumerate(ws3.values):
        if index == 0:
            ws4.append(row)
        else:
            if row[0] != None:
                ws4.append(row)
            else:
                continue

    wb3.save(fname3)

    # 添加合同价格，计算合同金额，多计单计，多计金额
    wb3 = openpyxl.load_workbook(fname3)
    ws3 = wb3['当月正']
    max_rows = ws3.max_row
    caihuang_xianquan_dic = caihuangpricedic.caihuang_price()

    for row in range(2, max_rows + 1):
        gongyingshang = ws3.cell(row, 13).value
        xianquanleixing = ws3.cell(row, 3).value
        shuliang = ws3.cell(row, 6).value
        songhuo_jiner = ws3.cell(row, 11).value
        guige, cishu = xianquanchicun.xianquan_chicun(xianquanleixing)

        if '彩皇' in ws3.cell(row, 13).value:
            yuan_ci = caihuang_xianquan_dic[guige]

        elif '双吉' in ws3.cell(row, 1).value:
            yuan_ci = caihuang_xianquan_dic[guige]

        else:
            continue

        # 写入规格
        ws3.cell(row, 4, value=guige)
        # 写入齿数
        ws3.cell(row, 5, value=cishu)
        # 写入元/齿
        ws3.cell(row, 7, value=yuan_ci)

        # 合同单价
        hetong_danjia = round(yuan_ci * cishu, 4)

        # 颜色加成 黑色和白色加10%，金和银加20%
        if ('黑' in ws3.cell(row, 3).value) or ('白' in ws3.cell(row, 3).value):
            hetong_danjia = hetong_danjia
        elif ('金' in ws3.cell(row, 3).value) or ('银' in ws3.cell(row, 3).value):
            hetong_danjia = round(hetong_danjia * 1.2, 2)
        else:
            hetong_danjia = round(hetong_danjia*1.1,2)

        # 写入元/条
        ws3.cell(row, 8, value=hetong_danjia)

        # 合同金额
        hetong_jiner = round(hetong_danjia * shuliang, 2)

        # 写入合同金额
        ws3.cell(row, 9, value=round(hetong_danjia * shuliang, 2))

        # 写入多计金额
        ws3.cell(row, 12, value=songhuo_jiner - hetong_jiner)

    wb3.save(fname3)

    fname3 = r'F:\a00nutstore\006\zw\xianquan(xianhuan)\线圈当月入库%s.xlsx' % dtrq
    # fname3 = r'D:\a00nutstore\006\zw\xianquan(xianhuan)\线圈当月入库%s.xlsx' % dtrq

    wb3 = openpyxl.load_workbook(fname3)
    ws3 = wb3['当月正']

    qishu = qijian[2:4] + qijian[-2:]

    ws4 = wb3.create_sheet(title='彩皇%s' % qishu)
    ws5 = wb3.create_sheet(title='双吉%s' % qishu)
    # ws6 = wb3.create_sheet(title='孝感鑫荣%s' % qishu)
    sheetnames = [ws4.title,ws5.title]

    for index, row in enumerate(ws3.values):
        if index == 0:
            ws4.append(row)
            ws5.append(row)
            # ws6.append(row)
        else:
            if row[12] == '彩皇':
                ws4.append(row)

            elif row[12] == '东莞市双吉装订':
                ws5.append(row)

            else:
                continue

    wb3.save(fname3)
    return fname,sheetnames


def main():
    qijian = input('请输入期间：格式为2020-04：')
    fname,sheetnames = xianquanDangyue(qijian)
    letters = 'FIKL'
    addSum.sumheji(fname,sheetnames,letters)
    os.startfile(fname)


if __name__ == '__main__':
    main()



















