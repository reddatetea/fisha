'''
从实时流水账中引入当月的纸入库
'''
import os
import pandas as pd
import numpy as np
import re
import openpyxl
import xlwings as xw
import baiyunleibie
from beifenFile import beifen
import easygui

#实时吨价
def dunjiaDic():
    fname2 = r'F:\a00nutstore\006\zw\baiyun\2020白云入库.xlsx'
    df = pd.read_excel(fname2, sheet_name='2020',dtype = {'入库单号':str})
    df = df.loc[~df['记账'].isnull()]  # 删除记账为空的记录
    df1 = df[df['pricename'].str.contains('返利|价差|冲减|多计|折扣') == False]  # 品名中包含返利、价差、冲减等不计入字典
    df1 = df1.assign(gongsi='驻马店白云纸业有限公司')  # 增加一列‘'驻马店白云纸业有限公司'
    gys_pinming = [(x, y) for x, y in zip(df1['gongsi'].tolist(), df1['pricename'].tolist())]
    dunjia_dic = dict(zip(gys_pinming, df1['单价'].tolist()))
    return dunjia_dic

def liushuizhang():
    df = pd.read_excel(r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx', sheet_name = '流水账',dtype = {'单据号':str})
    df = df.loc[df['供货单位'] == '驻马店白云纸业有限公司']
    df = df.loc[:,
         [ '日期','送货日期','期间', '白云期间', '单据号', 'cwName','令数', '吨数',  '吨价','入库金额','记账','品名', 'priceName']]
    df.insert(1, '公司', '白云')
    df.insert(9, '数量(吨)', 0)
    df.insert(12, '不含税金额', 0)
    df.insert(15, '备注', 0)
    for j in ['批次', '类别','车间','备注2']:
        df[j] = None
    return df

def query():
    start_riqi = pd.Timestamp(easygui.enterbox('请输入入库起始日期期间：格式为2021-11-4：'))
    end_riqi = pd.Timestamp(easygui.enterbox('请输入入库结束日期期间：格式为2021-11-4：'))
    piaojuhao = easygui.enterbox('请输入票据号90031065：')
    return start_riqi,end_riqi,piaojuhao

def qushu(df,start_riqi,end_riqi,piaojuhao):
    df.set_index('日期', inplace=True)  # 索引
    df.sort_index(inplace=True)  # 对索引排序
    df = df.truncate(before=start_riqi, after=end_riqi)
    df = df.loc[df['单据号'] >= piaojuhao]
    return df

def getruku():
    fname_ruku = r'F:\a00nutstore\006\zw\baiyun\2020白云入库.xlsx'
    ws_name_ruku = '2020'
    beifen(fname_ruku)        #重要文件备份
    wb = openpyxl.load_workbook(fname_ruku,data_only=True)  #将有公式的全部转为值，避免以后excel中有公式出现空值
    wb.save(fname_ruku)
    df2= pd.read_excel(fname_ruku,ws_name_ruku,dtype = {'入库单号':str})
    max_row = df2.shape[0] + 1
    leibie_dic = dict(zip(df2['pricename'].tolist(), df2['类别'].tolist()))
    return fname_ruku,ws_name_ruku,max_row,leibie_dic

def chuli(df1):
    df1.reset_index(inplace=True)
    df1['吨数'] = round(df1['吨数'],3)
    df1['数量(吨)'] = df1['吨数']
    dunjia_dic = dunjiaDic()
    # print(dunjia_dic)
    # print(len(dunjia_dic))
    df1 = df1.assign(吨价=df1.apply(lambda x: dunjia_dic.get(('驻马店白云纸业有限公司', x['priceName']),0), axis=1))
    fname_ruku, ws_name_ruku, max_row,leibie_dic = getruku()
    df1 = df1.assign(类别=df1.apply(lambda x:leibie_dic.get(x.priceName,0),axis=1))
    # wb = openpyxl.load_workbook(fname_ruku)
    # with pd.ExcelWriter(fname_ruku, engine='openpyxl')  as writer:
    #     writer.book = wb
    #     writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    #     df1.to_excel(writer, ws_name_ruku, header=None, index=False, startrow=max_row + 1)
    with pd.ExcelWriter(fname_ruku, engine='openpyxl',date_format='yyyy-mm-dd', mode='a', if_sheet_exists='overlay') as writer:
        df1.to_excel(writer, ws_name_ruku, header=None, index=False, startrow=max_row  )
    #writer.save()

    return fname_ruku

def delchongfu(fname,sheetname,in_subject,sort_cols):
    data = pd.DataFrame(pd.read_excel(fname, sheetname,dtype = {'入库单号':str}))
    data.drop_duplicates(subset=in_subject, keep='first', inplace=True)
    data = data.sort_values(by=sort_cols)
    # data = data.set_index(in_subject[0])
    wb = openpyxl.load_workbook(fname)
    ws = wb[sheetname]
    max_row = ws.max_row
    ws.delete_rows(2, max_row)
    wb.save(fname)
    # wb = openpyxl.load_workbook(fname)
    # with pd.ExcelWriter(fname, engine='openpyxl')  as writer:
    #     writer.book = wb
    #     writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    #     data.to_excel(writer, sheetname, header=None, index=False, startrow=1)
    with pd.ExcelWriter(fname, engine='openpyxl',date_format='yyyy-mm-dd', mode='a', if_sheet_exists='overlay') as writer:
        data.to_excel(writer, sheetname, header=None, index=False, startrow = 1) #excel中实际行+1才是
    wb = openpyxl.load_workbook(fname)
    ws = wb[sheetname]
    max_row = ws.max_row
    for i in range(2, max_row + 1):
        ws[f'A{i}'].number_format = 'yyyy-m-d'
        ws[f'C{i}'].number_format = 'yyyy-m-d'
        ws[f'N{i}'].number_format = 'yyyy-m-d'
    wb.save(fname)
    return fname

def jiagongsi(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb['2020']
    max_row = ws.max_row
    #未记账的加公式
    for i in range(2,max_row+1):
        if ws.cell(i,14).value in ['',None]:
            ws.cell(i, 12, value='=round(J' + str(i) + ' * ' + 'K' + str(i) + ', 2)')
            # 不含税金额用公式表达
            ws.cell(i, 13, value='=round(L' + str(i) + '/1.13 , 2)')
        else:
            continue
    wb.save(fname)

def main():
    start_riqi, end_riqi, piaojuhao = query()
    df = liushuizhang()    #流水账df
    df1 = qushu(df,start_riqi,end_riqi,piaojuhao)
    fname_ruku = chuli(df1)
    in_subject = ['开票日期', '入库单号','品名', '数量(令)', '计算重量','仓库材料']
    sort_cols = ['开票日期', '入库单号','品名']
    delchongfu(fname_ruku,'2020', in_subject,sort_cols)
    jiagongsi(fname_ruku)
    os.startfile(fname_ruku)

if __name__ == '__main__':
    main()


