'''
根据原材料库盘存表和流水账，统计出库存商品分别是从哪些供应商采购回的，以及采购数量、金额
2024-4-5加一列'计量单位’
2024-4-16按品名分组后，吨价先向下填充，然后向上填充，最后以5000填充，处理后财合并数据
'''
import pandas as pd
import numpy as np
import easygui
import os
import re
import openpyxl


def getKucun(caigous,pinmings,early_riqi):
    datas = []
    fname1 = easygui.fileopenbox('请选择原材料日报表文件')
    path, filename = os.path.split(fname1)
    regax = r'\.*(?P<riqi>\d+\-\d+\-\d+)\.xlsx'
    pattern = re.compile(regax)
    mat = pattern.search(filename)
    if mat != None:
        riqi = mat.group('riqi')

    else:
        riqi = easygui.enterbox('请输入"库存原材料分布"日期')

    dfs = pd.read_excel(fname1, sheet_name=None)
    sheet_names = [i for i in dfs]
    for sheet_name in sheet_names:
        if sheet_name == '原材料日报表':
            break

        elif '盘存表' in sheet_name:
            break
        else:
            sheet_name = easygui.choicebox('"日报表"or"盘存表"', '请点选工作表', sheet_names)
            break
    kucun = pd.read_excel(fname1,sheet_name ,usecols=[0, 1,2, 9],skiprows = 1)
    kucun.columns = ['存货大类名称','品名','计量单位','期末结存数量']
    kucun_cailiao = kucun[kucun['期末结存数量']!=0]
    kucun_cailiao = kucun_cailiao.dropna(subset=['品名'])
    # kucun_cailiao = kucun_cailiao.dropna(subset=['计量单位'])
    cailiaos = kucun_cailiao['品名'].to_list()
    for cailiao in cailiaos:
        kucun = kucun_cailiao[kucun_cailiao['品名']==cailiao]
        kucun_shuliang = kucun.iloc[-1][-1]  # 提取库存数
        caigou = getCaigou(caigous,pinmings,cailiao,kucun_shuliang)
        df11 = lianjie(kucun,kucun_shuliang,caigou,early_riqi)
        datas.append(df11)
    return datas,riqi


def caigouLiushuizhang():
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    caigous = pd.read_excel(fname, '流水账',usecols=[0, 1, 2, 3, 5, 6, 7, 8, 9], index_col=0,dtype = {'单据号':str})
    caigous = caigous.sort_index(ascending=False)
    early_riqi = caigous.index.to_list()[-1]
    caigous = caigous.reset_index()
    pinmings = list(caigous['品名'].unique())
    return caigous,pinmings,early_riqi

def getCaigou(caigous,pinmings,cailiao,kucun_shuliang):
    if cailiao in pinmings:
        caigou = caigous[caigous['品名'] == cailiao]
    else:
        caigou = caigous.iloc[-1:]
        caigou.iloc[-1, 2] = '待查'
        caigou.iloc[-1, 3] = cailiao
        caigou.iloc[-1, -5] = kucun_shuliang
        caigou.iloc[-1, -4] = 0
        caigou.iloc[-1, -3] = 0
        caigou.iloc[-1, -2] = None
        caigou.iloc[-1, -1] = None
    return caigou

def lianjie(kucun,kucun_shuliang,caigou,early_riqi):
    df8 = pd.merge(kucun, caigou, on='品名', how='left')
    df9 = df8
    df9['累加数量'] = df9['入库数量'].cumsum()  # 累加求和
    df9['剩余数量'] = df9['期末结存数量'] - df9['累加数量']  # 累加求和
    leiji_ruku = df9['入库数量'].sum()          #对应于期末结存数量的累计入库数量
    shuliangs = df9['剩余数量'].to_list()

    if leiji_ruku >= kucun_shuliang:
        shuliangs = df9['剩余数量'].to_list()
        for j in range(len(shuliangs)):
            if shuliangs[j] < 0:
                row = j
                break
            else:
                row = len(shuliangs)
                continue
        row = row +1
        df10 = df9.iloc[:row,:]
        df11 = df10
        shengyu = df11.iloc[-1, -1]
        danjia = df11.iloc[-1, -6]
        last_ruku = df11.iloc[-1, -7]
        leiji = df11.iloc[-1, -2]
        df11.iloc[-1, -7] = last_ruku + shengyu
        df11.iloc[-1, -5] = round((last_ruku + shengyu) * danjia, 2)  # 调整金额，按调整后的数量和单价计算
        df11.iloc[-1, -2] = leiji + shengyu
        df11.iloc[-1, -1] = 0
    else:
        df11 = df9
        last_col_ser = df9.iloc[-1]
        df11 = df11._append(last_col_ser)
        df11.iloc[-1,4] = early_riqi    #加‘计量单位’这一列后，df11.iloc[-1,3]变为df11.iloc[-1,4]
        df11.iloc[-1, 6] = '待查'       #加‘计量单位’这一列后，df11.iloc[-1,5]变为df11.iloc[-1,6]
        df11.iloc[-1, -7] = kucun_shuliang - leiji_ruku
        danjia = df9.iloc[-1,-6]
        df11.iloc[-1, -5] = round((kucun_shuliang - leiji_ruku) * danjia, 2)  # 调整金额，按调整后的数量和单价计算
        df11.iloc[-1, -2] = kucun_shuliang
        df11.iloc[-1, -1] = 0
    # grouped = df11.groupby(['供货单位','品名','本日数量'])
    return df11


def dunshu(x):
    pattern = r'(?P<ke>\d{2,3})g(?:.*?)(?P<chang>\d{3,4})\*(?P<kuan>\d{3,4})'
    regexp = re.compile(pattern)
    dunshu = x['吨数']
    if x['计量单位'] == '令':
        string = x['品名']
        mat = regexp.search(string)

        if mat == None:
            ke, chang, kuan = 0, 0, 0
            dunshu = x['吨数']

        else:
            # 克重
            ke = int(mat.group('ke'))
            # 长和宽
            chang = int(mat.group('chang'))
            kuan = int(mat.group('kuan'))
            dunshu = round(ke / 1000 / 1000 * chang / 1000 * kuan / 1000 * 500 * x['入库数量'], 3)
    else:
        pass
    return dunshu

def mergeKucunZhi(left,right):
    columns = ['存货大类名称',
 '品名',
 '计量单位',
 '期末结存数量',
 '日期',
 '单据号',
 '供货单位',
 '入库数量',
 '入库单价',
 '入库金额',
 'cwName',
 'priceName',
 '累加数量',
 '剩余数量',
 '吨数',
 '吨价',
 '金额',
]
    right.columns = ['单据号', '品名', 'dunjia']
    merge_left_right = pd.merge(left,right,on = ['单据号','品名'],how = 'left')
    merge_left_right['吨价'] = np.where(merge_left_right['吨价'] == 0,merge_left_right.dunjia,merge_left_right['吨价'])
    merge_left_right = merge_left_right.loc[:,columns]
    return merge_left_right

def chuliDf(df,fname,sheet_name):
    wb = openpyxl.load_workbook(fname)
    ws = wb[sheet_name]
    max_row = ws.max_row
    ws.delete_rows(2, max_row)
    wb.save(fname)
    # wb = openpyxl.load_workbook(fname)
    # with pd.ExcelWriter(fname, engine='openpyxl')  as writer:
    #     writer.book = wb
    #     writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    with pd.ExcelWriter(fname, engine='openpyxl', date_format='yyyy-m-d', mode='a',
                        if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name, header=None, index=False, startrow=1)
    wb = openpyxl.load_workbook(fname)
    ws = wb[sheet_name]
    max_row = ws.max_row
    for i in range(2, max_row + 1):
        ws[f'E{i}'].number_format = 'yyyy-m-d'

    wb.save(fname)
    return fname

def fillPrice(df0):
    df = df0.sort_values('日期')
    df['吨价'] = df['吨价'].fillna(method='ffill').fillna(method='bfill').fillna(5000)
    return df


def main():
    data = []
    caigous,pinmings,early_riqi =caigouLiushuizhang()
    path = r'F:\a00nutstore\006\zw\原材料实时流水账'
    os.chdir(path)
    datas,riqi = getKucun(caigous,pinmings,early_riqi)
    filename0 = f'库存原材料分布{riqi}.xlsx'
    filename = f'库存原材料分布.xlsx'
    fname = os.path.join(path,filename)
    sheet_name = '库存分配'
    df = pd.concat(datas)
    df = df[df['入库数量']!=0]
    grouped = df.groupby(['供货单位', '品名', '期末结存数量'])
    # for key, value in grouped:
    #     print(value[['日期', '供货单位', '入库数量', '入库金额']])
    df['单据号'] = df['单据号'].astype('str')
    df['吨数'] = 0
    df['吨价'] = pd.NA
    df['金额'] = 0
    df['吨数'] = np.where(df['计量单位'].isin(['kg', '公斤']), df['入库数量'] / 1000,
                          0)  # 如果计量单位是'kg'or'公斤",吨数是公斤数/1000
    df['吨价'] = np.where(df['计量单位'].isin(['kg', '公斤']), df['入库单价'] * 1000, 0)
    df['金额'] = df['入库金额']
    df = df.assign(吨数=df.apply(lambda x: dunshu(x), axis=1))
    df_zhi = pd.read_excel(r"F:\a00nutstore\006\zw\else\2020入库.xlsx", '入库', dtype={'单号': str})
    df_zhi = df_zhi.loc[:, ['单号', '材料', '吨价']]
    mergedf = mergeKucunZhi(df, df_zhi)
    df_by = pd.read_excel(r"F:\a00nutstore\006\zw\baiyun\2020白云入库.xlsx", '2020', dtype={'入库单号': str})
    df_baiyun = df_by.loc[:, ['入库单号', '仓库材料', '单价']]
    merge_kucun_zhi_baiyun = mergeKucunZhi(mergedf, df_baiyun)
    zhis = ['AA有光纸', 'AB双胶纸', 'AD白卡纸', 'AE涂布纸', 'AF铜版纸', 'AG牛皮纸', 'AH特种纸',
            'AI复写纸', 'AJ不干胶', 'AK灰板', 'AP卷筒纸']
    df8 = merge_kucun_zhi_baiyun.copy()
    group = df8.groupby('品名')
    for key, gp in group:
        gp['吨价']  = np.where(gp['存货大类名称'].isin(zhis), fillPrice(gp)['吨价'], gp['吨价'])
        data.append(gp)
    df9 = pd.concat(data)


    df9['金额'] = np.where(df9['存货大类名称'].isin(zhis), df9['吨数'] * df9['吨价'], df9['金额'])

    fname = chuliDf(df9,fname,sheet_name)
    # df8.to_excel(filename0, sheet_name, index=False)
    # df8.to_excel(filename,sheet_name,index=False)
    os.startfile(filename)

if __name__ == '__main__':
    main()

