'''
本模块是从流水账中将白云入库数据引入
本版本模块化
本版本增加输入期间处理，和当月入库工作后添加当前日期
'''
# _*_ coding:utf-8 _*_
import openpyxl
import os
import datetime
import easygui
import excelmessage
import time


def lszTobaiyundangyue(baiyunQijian,piaojuhao):
    #当天日期
    dtrq = datetime.date.today().strftime('%Y%m%d')
    print('选择白云当月入库路径')
    #path = easygui.diropenbox(msg='请选择白云当月入库路径', title='')
    path = r'F:\a00nutstore\006\zw\baiyun'
    os.chdir(path)

    filename = '白云当月入库%s.xlsx'%dtrq
    fname = os.path.join(path, filename)

    wb1 = openpyxl.Workbook()
    ws1 = wb1.create_sheet(title='当月')

    print('选择原材料实时流水账')
    #fname2 = excelmessage06.excelMessage()
    fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['流水账']

    #qijian = '2020-04'
    print(ws2.max_row)
    for i in range(1, ws2.max_row + 1):
        if i == 1:
            taitou = ('开票日期',
            '公司',  '送货单日期', '期间', '白云期间', '单号', '品名', '数量', '计算数量', '数量(吨)', '单价', '金额', '不含税金额', '记账', '仓库名称','备注','pricename')
            ws1.append(taitou)
        else:
           if (ws2.cell(i, 3).value == '驻马店白云纸业有限公司') and  (ws2.cell(i,13).value >= baiyunQijian) and (ws2.cell(i,2).value >= int(piaojuhao)):

                #ws1.cell(i, 1, value='双佳')
                #ws1.cell(i, 2, value=ws2.cell(i, 1).value)
                ws1.cell(i, 1, value=ws2.cell(i, 1).value)
                ws1.cell(i, 2, value='双佳')
                ws1.cell(i, 3, value=ws2.cell(i, 12).value)
                ws1.cell(i, 4, value=ws2.cell(i, 11).value)
                ws1.cell(i, 5, value=ws2.cell(i, 13).value)
                ws1.cell(i, 6, value=ws2.cell(i, 2).value)
                ws1.cell(i, 7, value=ws2.cell(i, 9).value)


                # 计算数量
                ws1.cell(i, 9, value=round(ws2.cell(i, 15).value,3))

                # 重量手工10
                ws1.cell(i,10, value=round(ws2.cell(i, 15).value, 3))

                # 吨价
                ws1.cell(i, 11, value=ws2.cell(i, 17).value)

                # 数量令数
                ws1.cell(i, 8, value=ws2.cell(i, 14).value)

                # 送货单金额
                #ws1.cell(i, 18, value=round(ws2.cell(i, 8).value, 2))

                #计算金额
                ws1.cell(i,12, value=round(round(ws2.cell(i, 15).value, 3)*ws2.cell(i, 17).value,2))

                # 计算不含税金额
                ws1.cell(i, 13, value=round(round(ws2.cell(i, 15).value, 3) * ws2.cell(i, 17).value/1.13, 2))

                ws1.cell(i, 15, value=ws2.cell(i, 4).value)

                #pricename
                ws1.cell(i,17,value=ws2.cell(i,10).value)



           else :
               continue
    wb1.save(fname)
    wb2.close()

    #删除行
    fname3 = fname
    wb3 = openpyxl.load_workbook(fname3)
    ws3 = wb3['当月']


    ws4 = wb3.create_sheet(title= '当月正')

    for index,row in enumerate(ws3.values):
        if index == 0 :
            ws4.append(row)
        else :
            if row[0]!= None:
                ws4.append(row)
            else:
                continue

    wb3.save(fname3)
    time.sleep(3)

def main():
    baiyunQijian = input('请输入白云期间：格式为2020-04：\n')
    piaojuhao = input('请输入票据号(包含该号)：\n')
    lszTobaiyundangyue(baiyunQijian,piaojuhao)

if __name__ == '__main__':
    main()

















