#复制贴粘板的内容到'zhangdan当前日期.txt'.并自动打开，修改后再次复制，关闭文件。出来两选项，如继续则按后来复制的内容，否则还是原来内容

import  re
import  os
import  pyperclip
import  datetime
import easygui
import openpyxl
import pytesseract
from PIL import Image
import baiduOcr

def ocrImg(pic_file):
    img = Image.open(pic_file)
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    string = pytesseract.image_to_string(img, lang='chi_sim')  #不加lang参数的话，默认进行英文识别
    #string = pytesseract.image_to_string(img, lang='chi_tra')  # 繁体字
    print(string)
    return string

def  getzhangdantxt(string):
    dqrq = datetime.date.today().strftime('%Y%m%d')
    print(dqrq)
    program_dir = os.getcwd()
    fname_txt = program_dir + os.sep + 'zhangdan{}.txt'.format(dqrq)
    fname_excel = program_dir + os.sep + 'zhangdan{}.xlsx'.format(dqrq)

    with open(fname_txt, 'w', encoding='utf-8') as f:
        f.write(string)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws_name = dqrq
    ws.title = ws_name
    wb.save(fname_excel)

    os.system(fname_txt)
    input_text = string
    print(input_text)
    return input_text,fname_excel,ws_name
#
def  countzhangdans(input_text):
    input_text.replace('\n', '')
    zhangdans = list(input_text.split('+'))
    zhangdanlist = [j.replace('\r\n', '').strip() for j in zhangdans]
    regax = r'(?P<xiangmu>.*?)(?P<jiner>([1-9]\d*\.\d*$)|([1-9]\d*$))'
    pattern = re.compile(regax)

    xiangmus = []
    jiners =[]
    for j in zhangdanlist:
        mat = pattern.search(j)
        print(mat)
        xiangmu = mat.group('xiangmu')
        jiner = mat.group('jiner')
        print(jiner)
        xiangmus.append(xiangmu)
        jiners.append(float(jiner))
    print(xiangmus)
    print(jiners)
    jiner_total = sum(jiners)
    print(jiner_total)
    return xiangmus,jiners,jiner_total

def  zhangdanToexcel(fname_excel,ws_name,xiangmus,jiners):
    wb = openpyxl.load_workbook(fname_excel)
    ws  =wb[ws_name]
    title = ['name','xiangmu','jiner']
    ws.append(title)

    for  j in range( len(xiangmus)):
        hang = ['xixi',xiangmus[j],jiners[j]]
        ws.append(hang)

    max_row = len(xiangmus)+1
    ws.cell( max_row+1,1).value = '合计'
    ws.cell( max_row+1,3).value = '=SUM(C2:C{})'.format(max_row)
    wb.save(fname_excel)

def  main():
    # APP_ID, API_KEY, SECRET_KEY, client = baiduOcr.ocrSeting()
    # msg = '请点选需要ocr的图片文件'
    # pic_file = easygui.fileopenbox(msg)
    # result =  baiduOcr.img_to_str(pic_file,client)
    # wenben = []
    # for i in result.get('words_result'):
    #     wenben.append(i.get('words'))
    # string =''.join(wenben)
    # print(wenben)

    input_text,fname_excel,ws_name = getzhangdantxt(string)
    print('请复制调整后的文本')
    msg = '现在是否可以继续'
    print(msg)

    titleYN = easygui.ccbox(msg, title='请选择Yes or No', choices=('yes', 'no'), image=None)
    if titleYN == True:
        input_text = pyperclip.paste()
        xiangmus,jiners,jiner_total = countzhangdans(input_text)
    else:
        xiangmus,jiners,jiner_total = countzhangdans(input_text)

    zhangdanToexcel(fname_excel, ws_name, xiangmus, jiners)
    myjiner = float(input('请输入这段时期你的付款金额\n'))
    myyingfu = jiner_total*0.6 - myjiner*0.4
    print('本次应付：',round(myyingfu,2))

if __name__=='__main__':
    main()





