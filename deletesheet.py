import openpyxl
import os
import easygui

def deleteSheet(fname):
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请选择要删除的工作表'
    print(msg)
    deletesheet_names = easygui.multchoicebox(msg,title = 'sheet',choices =sheetnames)
    print(deletesheet_names)
    for deletesheet_name in deletesheet_names:
        del wb[deletesheet_name]
    wb.save(fname)

def main():
    msg = '请点选要删除工作表的excel文件'
    print(msg)
    fname = easygui.fileopenbox(msg,title='excel文件')
    deleteSheet(fname)

if __name__ == '__main__':
    main()

