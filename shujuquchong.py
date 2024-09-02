'''
#本模块删除excel中的重复行,增加不变字段subject列表传入，增加程序的通用性
'''
# 导入pandas包并重命名为pd
import pandas as pd
import os
import xlsxlsx
import excelmessage
import openpyxl


def delchongfu(fname, excelname, sheetname,in_subject):
    # dirname, excelname = os.path.split(fname)
    # wb = openpyxl.load_workbook(fname)
    # ws = wb.active
    # sheetname = ws.title
    # wb.close()
    # excelname = fname

    # 读取Excel中Sheet1中的数据
    data = pd.DataFrame(pd.read_excel(excelname, sheetname))

    # 查看读取数据内容
    #print(data)

    # 查看是否有重复行
    re_row = data.duplicated()
    #print(re_row)

    # 查看去除重复行的数据,
    no_re_row = data.drop_duplicates(subset=in_subject, keep='first',inplace=True)
    #no_re_row = data.drop_duplicates(keep='first')
    #print(no_re_row)


    # 查看基于[科目编码]列去除重复行的数据
    #wp = data.drop_duplicates()
    # wp = data.drop_duplicates(['科目编码'])
    #print(wp)

    # 将去除重复行的数据输出到excel表中
    duplicated_name = fname[:-5] + '过滤重复行' + '.xlsx'
    #inplace=True ***************************************
    data.to_excel("%s" % duplicated_name)
    #inplace=False,默认，那就是下面的代码！**************
    #no_re_row.to_excel("%s" % duplicated_name)
    return duplicated_name


def main():
    print('请选择要删除重复数据的excel文件:')
    fname = excelmessage.excelMessage()
    dirname, excelname = os.path.split(fname)
    os.chdir(dirname)

    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    sheetname = sheetnames[0]
    ws = wb[sheetname]
    print(sheetname)
    wb.close()

    #in_subject = ['公司', '开票日期', '入库单号','品名', '数量(令)', '计算重量','仓库材料']

    in_subject = ['日期', '单据号', '供货单位', '品名', '单位', '入库数量', '入库单价', '入库金额', 'cwName', 'priceName', '期间', '送货日期', '白云期间',
                  '令数', '吨数', '令价', '吨价']

    delchongfu(fname, excelname, sheetname,in_subject)


if __name__ == '__main__':
    main()










