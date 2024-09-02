import openpyxl
import os
import easygui

def chaxunTianjia(gongyingshang_dic,dic):
    new_gongyingshang = easygui.enterbox('请输入新供应商名字:')
    if new_gongyingshang not in gongyingshang_dic:
        easygui.msgbox('该供应商不在原供应商列表中')
        msg = '请输入"{}"的简称'.format(new_gongyingshang)
        new_jianchen = easygui.enterbox(msg)
        dic[new_gongyingshang] = new_jianchen
    else:
        easygui.msgbox('该供应商已在原供应商列表中,不需要添加')
    return dic

def zhiGongyingshang():
    fname = r'F:\a00nutstore\006\zw\else\纸供应商.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    maxrows = ws.max_row
    zhi_gongyingshang_dic = {}
    for row in range(2, maxrows + 1):
        quchen = ws.cell(row, 1).value
        jianchen = ws.cell(row, 2).value
        zhi_gongyingshang_dic[quchen] = jianchen
    wb.save(fname)
    return zhi_gongyingshang_dic

def main():
    fname = r'F:\a00nutstore\006\zw\else\纸供应商.xlsx'
    gongyingshang_dic = zhiGongyingshang()
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    jieshuchaxun = True
    dic = {}
    while jieshuchaxun:
        choice = easygui.ccbox('查询和新增纸供应商', title='继续查询和添加or退出', choices=('yes', 'no'))
        if choice == True:
            dic = chaxunTianjia(gongyingshang_dic,dic)
        else:
            jieshuchaxun = False
    for key,value in dic.items():
        ws.append((key,value))
    wb.save(fname)
    os.startfile(fname)

if __name__=='__main__':
    main()




