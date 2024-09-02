'''
本模块将盘存表中的材料全部更名为财务标准名称
'''
# -*-coding = utf-8 -*-
from openpyxl import Workbook,load_workbook
#from openpyxl.utils import get_column_letter
import re
import csv
import openpyxl
import win32com.client as win32
import time
import os

def jiacaiwuname(fname):
    # 打开删除空行后的盘存表
    dirname,filename = os.path.split(fname)
    wb=load_workbook(fname)
    sheet = wb.active

    #print(sheet.max_row)
    #print(sheet.max_column)
    jishu = 0
    first_list = []
    for index,row in enumerate(sheet.values):
        print(row[1])
        first_list.append(row[1])

        if row[0] == None:
            jishu = jishu +1
            #print(jishu)

    mrows = sheet.max_row -jishu


    #file1 = open('F:\\a00nutstore\\fishc\\材料标准命名.txt','w',encoding = 'utf-8')
    #file2 = open('F:\\a00nutstore\\fishc\\异常命名明细.txt','w',encoding = 'utf-8')
    file1 = open('材料标准命名.txt','w',encoding = 'utf-8')
    file2 = open('异常命名明细.txt','w',encoding = 'utf-8')

    cw_names = []
    #不能匹配的列表abnormal_list
    abnormal_lists = []

    for in_string in first_list:

        if in_string =='' :
            continue
            #cw_name = ''
            #cw_names.append(cw_name)

        elif in_string ==None :
            continue

        elif  '小计' in in_string  :
            continue

        elif  '合计' in in_string  :
            continue

        else :
            if '卷筒' in in_string:
                if "上白" in in_string or "中白" in in_string or "下白" in in_string or "上红" in in_string or "中红" in in_string or "上红" in in_string or "上蓝" in in_string or "中蓝" in in_string or "下蓝" in in_string or "下红" in in_string:
                    leixing = '无碳复写卷筒纸'

                else:
                    leixing = '卷筒纸'

                pattern = r'(?P<ke>\d{2,3})g?(?:.*)(?P<leixing>卷筒).*'
                regexp = re.compile(pattern)
                mat = regexp.search(in_string)

                if mat == None :
                    cw_name = in_string
                    cw_names.append(cw_name)
                    abnormal_lists.append(cw_name)

                else :
                    # 克重
                    ke = mat.group('ke')

                    zddd = ''
                    # 该材料的财务软件标准命名
                    # cw_name = ke + 'g' + leixing + zddd
                    cw_name = leixing + zddd
                    # print(cw_name)

                    # 该材料的财务软件标准命名
                    # cw_name = ke + 'g' + leixing + zddd

                    # print(cw_name)

                    cw_names.append(cw_name)

            elif '不干胶' in in_string:
                cw_name = '不干胶'
                cw_names.append(cw_name)

            elif '箱' in in_string:
                if '封箱'  in in_string:
                    cw_name = '封箱胶'
                    cw_names.append(cw_name)
                else :
                    cw_name = '纸箱'
                    cw_names.append(cw_name)


            elif '卷膜POF' in in_string:
                cw_name = '卷膜pof'
                cw_names.append(cw_name)

            elif '哑膜' in in_string:
                cw_name = '哑膜'
                cw_names.append(cw_name)

            elif '光膜' in in_string:
                cw_name = '光膜'
                cw_names.append(cw_name)

            elif '缠绕卷膜' in in_string:
                cw_name = '缠绕膜'
                cw_names.append(cw_name)

            elif '打包带' in in_string:
                cw_name = '打包带'
                cw_names.append(cw_name)

            elif '收缩袋' in in_string:
                cw_name = '收缩袋'
                cw_names.append(cw_name)

            elif '空白袋' in in_string or '吊袋' in in_string:
                cw_name = '空白袋'
                cw_names.append(cw_name)

            elif '手提' in in_string or '提手' in in_string:
                cw_name = '手提袋'
                cw_names.append(cw_name)

            elif '书套' in in_string:
                cw_name = '书套'
                cw_names.append(cw_name)

            elif '小文具胶' in in_string:
                cw_name = '小透明胶'
                cw_names.append(cw_name)

            elif  ('封面' in in_string or '衬膜' in in_string) and  ('PP片' not in in_string):
                cw_name = '封面'
                cw_names.append(cw_name)

            elif '胶套' in in_string:
                if '水胶套' in in_string :
                    cw_name = in_string
                    cw_names.append(cw_name)
                else :
                    cw_name = '胶套'
                    cw_names.append(cw_name)


            elif '热熔胶' in in_string or '波士胶' in in_string:
                cw_name = '热熔胶'
                cw_names.append(cw_name)

            elif '816胶' in in_string or '8475胶' in in_string:
                cw_name = '816粘合剂'
                cw_names.append(cw_name)

            elif '封箱胶' in in_string:
                cw_name = '封箱胶'
                cw_names.append(cw_name)

            elif '啫喱胶' in in_string:
                cw_name = '啫喱胶'
                cw_names.append(cw_name)

            elif '大双面胶德莎' in in_string or '3mmE1060H双面胶' in in_string:
                cw_name = '进口双面胶'
                cw_names.append(cw_name)

            elif '聚乙烯醇' in in_string:
                cw_name = '聚乙烯醇'
                cw_names.append(cw_name)

            elif '线圈' in in_string or '线环' in in_string:
                cw_name = '线圈'
                cw_names.append(cw_name)

            elif '书边带' in in_string:
                cw_name = '堵头布'
                cw_names.append(cw_name)

            elif '无纺布' in in_string:
                cw_name = '无纺布'
                cw_names.append(cw_name)

            elif '电化铝' in in_string:
                cw_name = '电化铝'
                cw_names.append(cw_name)

            elif '票据夹' in in_string:
                cw_name = '票据夹'
                cw_names.append(cw_name)

            elif '票据夹' in in_string:
                cw_name = '票据夹'
                cw_names.append(cw_name)

            elif '16KPP账夹' in in_string:
                cw_name = '16kpp帐夹'
                cw_names.append(cw_name)

            elif '25KPP账夹' in in_string:
                cw_name = '25kpp帐夹'
                cw_names.append(cw_name)

            elif 'PP铁夹' in in_string:
                cw_name = 'PP铁夹'
                cw_names.append(cw_name)

            elif '袖珍账夹' in in_string:
                cw_name = '袖珍帐夹'
                cw_names.append(cw_name)

            elif '账钉' in in_string or '活页钉' in in_string:
                cw_name = '账钉'
                cw_names.append(cw_name)

            elif '挂锁' in in_string:
                cw_name = '挂锁'
                cw_names.append(cw_name)

            elif '圆珠笔' in in_string:
                cw_name = '圆珠笔'
                cw_names.append(cw_name)

            elif '背条' in in_string:
                cw_name = '背条（米）'
                cw_names.append(cw_name)

            elif 'PP片' in in_string:
                cw_name = 'PP片'
                cw_names.append(cw_name)

            elif '铁丝' in in_string:
                cw_name = '铁丝'
                cw_names.append(cw_name)

            elif '白海绵' in in_string:
                cw_name = '海绵（张）'
                cw_names.append(cw_name)

            elif '3MM海绵' in in_string:
                cw_name = '海绵'
                cw_names.append(cw_name)

            elif '橡皮筋' in in_string:
                cw_name = '橡皮筋'
                cw_names.append(cw_name)

            elif '7MM油双面胶' in in_string:
                cw_name = '7MM双面胶'
                cw_names.append(cw_name)

            elif '13MM双面胶' in in_string:
                cw_name = '13mm双面胶'
                cw_names.append(cw_name)

            elif '白乳胶' in in_string:
                cw_name = '白乳胶'
                cw_names.append(cw_name)

            elif '丝线' in in_string:
                cw_name = '丝线'
                cw_names.append(cw_name)

            elif '丝带' in in_string :
                cw_name = '1分丝带'
                cw_names.append(cw_name)

            elif '灰板' in in_string or '双灰' in in_string:
                #cw_name = '灰板'
                #cw_names.append(cw_name)

                pattern = r'(?P<ke>\d\.?\d?\d?)(MM)?(?:.*?)(?P<leixing>灰板|双灰)(?:.*?)'
                regexp = re.compile(pattern)
                mat = regexp.search(in_string)

                if mat == None :
                    cw_name = in_string
                    cw_names.append(cw_name)
                    abnormal_lists.append(cw_name)

                else :
                    # 克重
                    ke = mat.group('ke')

                    # 类型
                    leixing = mat.group('leixing')

                    if 'MM' in in_string:
                        ke = str(int(10 * float(ke)))
                        cw_name = ke + 'mm' + leixing

                    else:
                        cw_name = ke + 'g' + leixing

                    # print(cw_name)
                    cw_names.append(cw_name)

            else:
                pattern = r'(?P<ke>\d{2,3})g?(?:.*?)(?P<leixing>双胶|白卡|牛|铜|涂布|无碳|有光纸|黑卡|莱妮纹|镭射金|浅绿皮纹|特种纸)(?:.*?)(?P<changKuan>(\d{3}\*\d{4})|(\d{4}\*\d{3})|(\d{3}\*\d{3}))'
                regexp = re.compile(pattern)
                mat = regexp.search(in_string)

                if mat == None :
                    cw_name = in_string
                    cw_names.append(cw_name)
                    abnormal_lists.append(cw_name)

                else:
                    # 克重
                    ke = mat.group('ke')

                    # 类型
                    leixing = mat.group('leixing')

                    # 长和宽
                    changKuan = mat.group('changKuan')

                    if '牛' in leixing or '牛卡' in leixing or '牛皮' in leixing:
                        leixing = '牛卡'

                    elif '铜' in leixing or '铜版' in leixing:
                        leixing = '铜版'

                    elif '莱妮纹' in leixing or '特种纸' in leixing or '镭射金' in leixing or '浅绿皮纹' in leixing:
                        leixing = '特种纸'

                    elif '有光' in leixing:
                        leixing = '有光纸'

                    elif '黑卡' in leixing:
                        leixing = '黑卡'

                    elif '白卡' in leixing:
                        leixing = '白卡'

                    elif '无碳' in leixing or '无碳复写' in leixing or '无碳复写纸' in leixing:
                        leixing = '无碳复写纸'

                    else:
                        leixing = leixing

                    # 标准命名
                    pinming = ke + 'g' + leixing + changKuan
                    # print(pinming)

                    # 长宽
                    chang = re.compile('(?P<chang>\d{3,4})\*(?P<kuan>\d{3,4})').search(changKuan).group('chang')
                    kuan = re.compile('(?P<chang>\d{3,4})\*(?P<kuan>\d{3,4})').search(changKuan).group('kuan')

                    # 判断该材料是正度还是大度
                    mianji = int(chang) * int(kuan)
                    # print(mianji)
                    if mianji / 787 / 1092 <= 1.05:
                        zddd = '正度'
                    else:
                        zddd = '大度'
                    # print(zddd)

                    # 该材料的财务软件标准命名
                    cw_name = ke + 'g' + leixing + zddd
                    cw_names.append(cw_name)

    cw_Newnames = []
    for cw_name in cw_names:
        i = 0
        if cw_name == '17mm灰板':
            cw_name = '18mm灰板'
            cw_Newnames.append(cw_name)
        elif cw_name == '7mm灰板':
            cw_name = '75dm灰板'
            cw_Newnames.append(cw_name)

        elif cw_name == '18mm双灰':
            cw_name = '18mm灰板'
            cw_Newnames.append(cw_name)
        elif cw_name == '15mm双灰':
            cw_name = '15mm灰板'
            cw_Newnames.append(cw_name)
        elif cw_name == '450g双灰':
            cw_name = '450g灰板'
            cw_Newnames.append(cw_name)
        elif cw_name == '550g双灰':
            cw_name = '550g灰板'
            cw_Newnames.append(cw_name)
        elif '无碳复写纸' in cw_name:
            cw_name = '无碳复写纸'
            cw_Newnames.append(cw_name)

        else:
            cw_name = cw_name
            cw_Newnames.append(cw_name)
        file1.write('%s\n'%cw_name.strip())

    print(cw_names)
    print(len(cw_names))
    print(cw_Newnames)
    print(len(cw_Newnames))


    #打印没有匹配的材料名称
    #print('**********************')
    #print('下面是没有匹配的材料：')
    for abnormal_list in abnormal_lists:
        file2.write('%s\n'%abnormal_list)
        #print(abnormal_list)

    file1.close()
    file2.close()

    #数据去向工作簿 Workbook
    ss_wb = Workbook()
    ss_sheet = ss_wb.active
    ss_sheet.title = '流水账'


    for index,row in enumerate(sheet.values):
        #print(index,cw_Newnames[index])
        if row[1] in ['',None]:
            continue
        else:
            ss_sheet.append(row+(cw_Newnames[index],))

    filename = '盘存表材料更名.xlsx'
    fname = os.path.join(dirname,filename)


    ss_wb.save(fname)
    return cw_names, cw_Newnames,fname

def main():
    fname= r'F:\a00nutstore\006\zw\2020\202010\2020年10月盘存表\双佳2020年10月份原材料盘存表.xlsx'
    jiacaiwuname(fname)

if __name__=='__main__':
    main()

