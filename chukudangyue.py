import os
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import easygui
import chukuprintseting
import re

def chukuDangyue(fname):
    path,filename = os.path.split(fname)
    #qian,hou = os.path.splitext(filename)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname,data_only=True)
    sheetnames = wb.sheetnames
    msg = '请点选出库参考copy，注意：原材料202106出库参考'
    print(msg)
    ws_name = easygui.buttonbox(msg,title=msg,choices=sheetnames)
    ws = wb[ws_name]

    regax = r'(?:原材料)(\d{4}-\d{2})(?:出库参考)'
    pattern = re.compile(regax)
    mat = pattern.search(filename)
    if mat :
        qijian_string = mat.group(1)
        chuku_name = qijian_string[-2:]
    else :
        qijian_string = input('请输入期间"202202"\n')
        chuku_name = qijian_string[-2:]

    yemei_title = '原材料出库明细' +qijian_string[:-2] +'-' + qijian_string[-2:]

    new_name = '原材料出库' + chuku_name +'.xlsx'
    fname1 = os.path.join(path,new_name)
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = chuku_name
    # wb.create_sheet(chuku_name)
    # ws1 = wb[ws_name]
    # ws = wb[chuku_name]
    title = ['科目编码','科目名称','出库数量','单价','金额']
    kemuminchen_number = column_index_from_string('I')-1
    kemubianma_number = column_index_from_string('H') - 1
    shuliang_number = column_index_from_string('U')-1
    danjia_number = column_index_from_string('V')-1
    jiner_number = column_index_from_string('W')-1
    for index,row in enumerate(ws.values):
        if index == 0 :
            ws1.append(title)
        else :

            jiner = row[jiner_number]

            if jiner != 0:
                kemubianma = row[kemubianma_number]
                kemumingchen = row[kemuminchen_number]
                shuliang = row[shuliang_number]
                danjia = row[danjia_number]
                newrow = (kemubianma,kemumingchen,shuliang,danjia,jiner)
                ws1.append(newrow)
            else :
                continue

    wb1.save(new_name)
    wb.close()

    wb1 = openpyxl.load_workbook(new_name)
    ws1 = wb1[chuku_name]
    for index,row in enumerate(ws1.values):
        if index == 0:
            continue
        else :
            if row[4]==0 :
                dele

    wb1.save(new_name)
    return fname1,yemei_title

def main():
    msg = '请点选原材料出库参考“原材料202012出库参考"'
    print(msg)
    fname = easygui.fileopenbox(msg,title=msg)
    fname1,yemei_title = chukuDangyue(fname)                                 #title页眉‘原材料出库明细2020112‘
    fname,ws_name = chukuprintseting.setPrintArea(fname1)
    fname = chukuprintseting.firstRowSeting(fname,ws_name)
    chukuprintseting.yemei(fname,ws_name,yemei_title)
    os.startfile(fname)


if __name__ == '__main__':
    main()
