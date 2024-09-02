'''
本模块是从流水账中将白云入库数据引入

'''
# _*_ coding:utf-8 _*_
import openpyxl
import os




path = r'F:\a00nutstore\006\zw\baiyun'
os.chdir(path)

filename = '白云当月入库.xlsx'
fname = os.path.join(path,filename)

wb1 = openpyxl.Workbook()
ws1 = wb1.create_sheet(title = '当月')

fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
wb2 = openpyxl.load_workbook(fname2)
ws2 = wb2['流水账']

qijian = '2020-04'
print(ws2.max_row)
for i in range(1,ws2.max_row+1):
    if i==1 :
        taitou = ('公司','开票日期','送货单日期','期间','白云期间','单号','品名','数量','计算数量','数量(吨)','单价','金额','不含税金额','记账','备注')
        ws1.append(taitou)
    else :
        if  ws2.cell(i,3).value == '驻马店白云纸业有限公司':
            ws1.cell(i, 1,value = '双佳')
            ws1.cell(i, 2,value = ws2.cell(i, 1).value)
            ws1.cell(i, 3,value = ws2.cell(i, 12).value)
            ws1.cell(i, 4,value = ws2.cell(i, 11).value)
            ws1.cell(i, 5,value = ws2.cell(i, 13).value)
            ws1.cell(i, 6,value = ws2.cell(i, 2).value)
            ws1.cell(i, 7,value = ws2.cell(i, 9).value)
            

            # 计算数量
            ws1.cell(i, 9,value = ws2.cell(i, 15).value)

            # 重量手工10

            # 吨价
            ws1.cell(i, 11,value = ws2.cell(i, 18).value)

            # 数量令数
            ws1.cell(i, 8,value = ws2.cell(i,14).value)




        else :
            continue

wb1.save(fname)
wb2.close()














