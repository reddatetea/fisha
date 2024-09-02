'''
本模块将白云当月入库添加到白云2020入库
本版本模块化
'''
# _*_ conding:utf-8 _*_
import openpyxl
import os
import datetime
import easygui


def addBaiyundangyue():
    dtrq = datetime.date.today().strftime('%Y%m%d')
    print('请选择白云当月入库路径')
    path = easygui.diropenbox(msg='请选择白云当月入库路径', title='')
    os.chdir(path)
    filename = '白云当月入库%s.xlsx' % dtrq
    fname = os.path.join(path,filename)
    wb1 = openpyxl.load_workbook(fname)
    ws1 = wb1['当月正']

    print('请选择白云年入库文件')
    fname2 = easygui.fileopenbox(msg = '请选择白云年入库文件',title = '2020白云入库')
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['2020']
    maxrow2 =ws2.max_row

    for index, row in enumerate(ws1.values):
        if index == 0:
            pass
        else:
            ws2.append(row)

    wb1.close()
    wb2.save(fname2)
    return fname2,maxrow2

def jiagongsi(fname2,maxrow2):
    wb = openpyxl.load_workbook(fname2)
    ws = wb['2020']
    maxrow = ws.max_row
    for i in range(maxrow2+1,maxrow+1):
        #金额用公式表达
        ws.cell(i,12, value='=round(J' + str(i) + ' * ' + 'K' + str(i) + ', 2)')
        #不含税金额用公式表达
        ws.cell(i,13, value='=round(L' + str(i) + '/1.13 , 2)')


    wb.save(fname2)

def main():
    fname2,maxrow2 = addBaiyundangyue()
    jiagongsi(fname2,maxrow2)

if __name__ == '__main__':
    main()









