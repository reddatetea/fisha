'''
根据聚水潭出库进行编码关联后计算电商002出库明细表
2025-1-27程序的通用性，其他程序可以调用聚水潭的编码处理
第一步先计算本次聚水潭出库中对应不上的大库的编码
第二步将差异处理后，再运行形成差异较少的出库明细表
'''
import numpy as np
import pandas as pd
import re
import easygui
import openpyxl
import os
import TcloudCunhuoDic
import bianmabiaozhunhua06
from bianmabiaozhunhua06 import chuliNumAfterHengGang,quShu,chuliFirstHenggang,bianmaBiaozhun 
from bianmabiaozhunhua06 import huizhongXiaoshoubuDiaobojia,chuliXiaoshoubuDiaobojia


def getCunhuoConcentFile(path,content_dic):
    riqi = easygui.enterbox('请输入日期')
    dicfile = f'存货档案字典{riqi}.xlsx'
    sheet_name = '存货档案'
    df_dic = pd.DataFrame.from_dict(content_dic,orient='index')
    df_dic = df_dic.reset_index()
    df_dic.columns = ['存货编码','含量']
    df_dic.to_excel(dicfile,sheet_name = sheet_name,index = False)
    return os.path.join(path,dicfile)

def isnoInCunhuoLst(s,lst):
    if s in lst:
        return True
    else :
        return False

#如果字符在列表的某个元素中，则返回该元素
def strIsinLststr(str,lst):
    lst.sort()
    for i in lst:
        if str in i or str == i:
            return i
        else :
            continue
    return None
  

def chuliColor(string,pattern):
    mat = re.search(pattern,string)
    if mat :
        str1 = mat.group(1)
    else :
        str1 = string
    return str1

def chuluJusuitanBianma(fname_jusuitan,jusuitan_bianma,dic_jusuitanDiff,content_dic):
    df_jusuitanXiaoshou0 = pd.read_excel(fname_jusuitan,dtype = {f'{jusuitan_bianma}':'str'})
    df_jusuitanXiaoshou0 = df_jusuitanXiaoshou0[~df_jusuitanXiaoshou0[f'{jusuitan_bianma}'].isna()]
    df_jusuitanXiaoshou = pd.pivot_table(df_jusuitanXiaoshou0,index = jusuitan_bianma,values = '数量',aggfunc = 'sum',fill_value = 0)
    df_jusuitanXiaoshou = df_jusuitanXiaoshou.reset_index()
    df_jusuitanXiaoshou = df_jusuitanXiaoshou.assign(content = df_jusuitanXiaoshou[f'{jusuitan_bianma}'].map(content_dic))
    #规则1，结尾为A商品编码改为a
    jusuitan_bianma1 = f'{jusuitan_bianma}1'
    df_jusuitanXiaoshou[f'{jusuitan_bianma1}'] = np.where(df_jusuitanXiaoshou[f'{jusuitan_bianma}'].str.endswith('A'), df_jusuitanXiaoshou[f'{jusuitan_bianma}'].str[:-1]+'a',df_jusuitanXiaoshou[f'{jusuitan_bianma}'])
    # df_jusuitan_diff1['商品编码1'] = np.where(df_jusuitan_diff1['商品编码'].str.endswith('B'), df_jusuitan_diff1['商品编码'].str[:-1]+'b',df_jusuitan_diff1['商品编码'])
    #规则2，带颜色的编码弄短
    regax = r'(.+)-\D+色$'
    pattern = re.compile(regax)
    df_jusuitanXiaoshou[f'{jusuitan_bianma1}'] = df_jusuitanXiaoshou[f'{jusuitan_bianma1}'].apply(lambda x:chuliColor(x,pattern))
    df_jusuitanXiaoshou.content = np.where(df_jusuitanXiaoshou.content.isna(),df_jusuitanXiaoshou[f'{jusuitan_bianma}1'].map(content_dic),df_jusuitanXiaoshou.content)
    df_jusuitanXiaoshou[f'{jusuitan_bianma1}'] = np.where(df_jusuitanXiaoshou.content.isna(),df_jusuitanXiaoshou[f'{jusuitan_bianma1}'].map(dic_jusuitanDiff),df_jusuitanXiaoshou[f'{jusuitan_bianma1}'])
    df_jusuitanXiaoshou = df_jusuitanXiaoshou.rename(columns = {'编码2':'jusuitan','编码21':'xiaoshoubu'})
    df_jusuitanXiaoshou = df_jusuitanXiaoshou[['jusuitan','数量','content','xiaoshoubu']]
    return df_jusuitanXiaoshou0,df_jusuitanXiaoshou

def GetDicJusuitanDiff(fname_jusuitanDiff):
    df_jusuitanDiff = pd.read_excel(fname_jusuitanDiff ,dtype = {'jusuitan':'str','xiaoshoubu':'str'})
    dic_jusuitanDiff = dict(zip(df_jusuitanDiff['jusuitan'],df_jusuitanDiff['xiaoshoubu']))
    return dic_jusuitanDiff

def chukliAddDiff(fname_addDiff):
    add_diff = pd.read_excel(fname_addDiff ,dtype = {'商品编码':'str','bianma4':'str'})
    add_diff.columns =['jusuitan','xiaoshoubu']
    return add_diff
    
def main():
    lst = ['存货编码',
             '实发数量1',
             '净销量1',
             '净销售额',
             '净销售成本',
             '净销售毛利',
             '基本金额',
             '已付金额',
             '优惠金额',
             '运费收入',
             '运费支出',
             '净毛利率',
             '基本售价',
                   ]
    lst1 = ['存货编码',
                 '实发数量1',
                 '净销量1',
                 '净销售额',
                 '净销售成本',
                 '净销售毛利',
                 '基本金额',
                 '已付金额',
                 '优惠金额',
                 '运费收入',
                 '运费支出',
                 '净毛利率',
                 '基本售价',
                        ]
    fname  = easygui.fileopenbox('请点选存货档案')
    path,_ = os.path.split(fname)
    os.chdir(path)
    #生成存货档案字典
    # content_dic = getCunhuoConcent(fname)
    # getCunhuoConcentFile(path,content_dic)
    content_dic = TcloudCunhuoDic.getCunhuoConcent(fname)
    getCunhuoConcentFile(path,content_dic)
    cunhuo_lst = list(content_dic.keys())
    
     #聚水潭编码与大库差异
    fname_jusuitanDiff = r"F:\a00nutstore\008\zww08\002电商\聚水潭\聚水潭编码与大库差异.xlsx"
    wb = openpyxl.load_workbook(fname_jusuitanDiff)
    ws = wb.active
    max_row = ws.max_row
    dic_jusuitanDiff = GetDicJusuitanDiff(fname_jusuitanDiff)
    # fname_jusuitan = r"F:\a00nutstore\008\zww08\002电商\聚水潭\聚水潭各平台销售\聚水潭202503\各平台总销售2025-03.xlsx"
    fname_jusuitan =easygui.fileopenbox('请打开编码处理后的聚水潭销售主题文件“各平台总销售2025-03”')
    path,_ = os.path.split(fname_jusuitan)
    os.chdir(path)
    df0 = pd.read_excel(fname_jusuitan,dtype = {'商品编码':str,'其他属性4':str})
    df = df0[~df0['商品编码'].isna()]
    df = df[df['pingtai'] != '全平台']
    pattern = r'(.*)-(\d{1,2})$'
    df1 = df.copy()
    #处理最右边的-，形成bianma1
    df1 = df1.assign(bianma1 = df1.商品编码.apply(lambda x:chuliNumAfterHengGang(x)[0]))
    df1 = df1.assign(shuliang = df1.商品编码.apply(lambda x:chuliNumAfterHengGang(x)[1]))
    #处理+,形成bianma2
    df1['bianma2'] = df1['bianma1'].str.split('+')
    df1['len'] = df1['bianma2'].str.len()
    df1 = df1.assign(shuliang1 = np.where(df1.len == 1,df1.shuliang ,df1['shuliang']/df1['len']))
    df1 = df1.explode('bianma2')
    df1['净销量1'] =  df1['净销量']*df1['shuliang1']
    df1['实发数量1'] =  df1['实发数量']*df1['shuliang1']
    for i in ['实发金额','销售金额','销售毛利','净销售额','净销售成本','净销售毛利','基本金额','已付金额','优惠金额','运费收入',
                 '运费支出']:
        df1[f'{i}'] = np.where(df1.len == 1,df1[f'{i}'] ,df1[f'{i}']/df1['len'])
        df1['bianma3'] = df1['bianma2']   #为下一步处理未匹配编码做准备
    df1['bianma4'] = df1['bianma2']
    df11 = df1[df1.商品编码.map(lambda x:isnoInCunhuoLst(x,cunhuo_lst))]
    df12 = df1[~df1.商品编码.map(lambda x:isnoInCunhuoLst(x,cunhuo_lst))]
    #直接匹配639个匹配成功，另外124个没有匹配，需要另外进一步处理
    #对bianma2处理成标准编码，形成bianma3
    df12['bianma3'] = df12.bianma2.map(chuliFirstHenggang)
    bianma3 =  df12.bianma3.map(bianmaBiaozhun)
    df12['bianma3'] = bianma3
    df12['bianma4'] = df12['bianma3']
        # df12.bianma3.map(content_dic)
    df121 = df12[df12.bianma3.map(lambda x:isnoInCunhuoLst(x,cunhuo_lst))]
    df122 = df12[~df12.bianma3.map(lambda x:isnoInCunhuoLst(x,cunhuo_lst))]
    df121 = df121.assign(bianma4 = df121.bianma3)
    df122 = df122.assign(bianma4 = df122.bianma3.apply(lambda x:strIsinLststr(x,cunhuo_lst)))
    df122 = df122.assign(bianma4 = np.where(df122.bianma4.isna(),df122['商品编码'].map(dic_jusuitanDiff),df122.bianma4))
    #未匹配
    df_weipipei = df122[df122.bianma4.isna()]
    df122 = df122.assign(bianma4 = np.where(df122.bianma4.isna(),df122.bianma3,df122.bianma4))
    df_jst0 = pd.concat([df11,df121,df122])
    df_jst0.to_excel('df_jst0.xlsx',index = False)
     
    df_weipipei.to_excel('df_weipipei.xlsx',index = False)
    add_diff = df_weipipei.copy()
    add_diff = add_diff[['商品编码','bianma3']]
    add_diff = add_diff.drop_duplicates(subset = '商品编码')
    add_diff.to_excel('本次聚水潭编码异常add_diff.xlsx',index = False)

    #打开本次需要添加的聚水潭编码与大库的差异文件
    msg = '是否将本次聚水潭差异添加到总差异文件中？'
    ISNO = easygui.boolbox(msg= msg)
    def chukliAddDiff(fname_addDiff):
        add_diff = pd.read_excel(fname_addDiff ,dtype = {'商品编码':'str','bianma4':'str'})
        add_diff.columns =['jusuitan','xiaoshoubu']
        return add_diff
    if ISNO:
        
        fname_addDiff = easygui.fileopenbox('本次需要添加的聚水潭编码与大库的差异文件"本次聚水潭编码异常addDiff.xlsx"')
        add_diff = chukliAddDiff(fname_addDiff)
        with pd.ExcelWriter(fname_jusuitanDiff, engine='openpyxl',mode='a', if_sheet_exists='overlay')  as writer:
            add_diff.to_excel(writer, sheet_name = 'Sheet1',startrow=max_row, header = False,index = False)
    
    else:
        
        qijian = easygui.enterbox(msg = '请输入期间"2025-01"')
        #求和
        # df_jusuitanXiaoshou0 = pd.read_excel(fname_jusuitan,dtype = {f'{jusuitan_bianma}':'str'})
        # df_jusuitanXiaoshou0 = df_jusuitanXiaoshou0[~df_jusuitanXiaoshou0[f'{jusuitan_bianma}'].isna()]
        chuku = df_jst0.copy()
        chuku = chuku[['bianma4','实发数量1','商品编码','净销量1']]
        bianma4_bianma = dict(zip(chuku.bianma4,chuku.商品编码))
        df_chuku = pd.pivot_table(chuku,index = 'bianma4',values = ['实发数量1','净销量1'],fill_value=0,aggfunc = np.sum)
        df_chuku = df_chuku.reset_index()
        df_chuku = df_chuku.assign(content = df_chuku.bianma4.map(content_dic))
        df_chuku = df_chuku.assign(商品编码 = df_chuku.bianma4.map(bianma4_bianma))
        df_chuku = df_chuku.rename(columns = {'bianma4':'电商库002存货编码','content':'含量','商品编码':'聚水潭003商品编码'})
        df_chuku = df_chuku[['电商库002存货编码','实发数量1','含量','聚水潭003商品编码','净销量1']]
        total = df_chuku.sum()
        total[0] = '合计'
        total[2] = ''
        total[3] = ''
        df_chuku = df_chuku.set_index('电商库002存货编码')
        df_chuku.loc['合计'] = total
    
        lst1 = ['bianma4',
                 '实发数量1',
                 '净销量1',
                 '净销售额',
                 '净销售成本',
                 '净销售毛利',
                 '基本金额',
                 '已付金额',
                 '优惠金额',
                 '运费收入',
                 '运费支出',
                  '基本售价',
                 ]
        xiaoshou = df_jst0[lst1]
        xiaoshou = xiaoshou.rename(columns = {'bianma4':'存货编码'})
        df_jusuitanXiaoshou18 = pd.pivot_table(xiaoshou,index = '存货编码',fill_value=0,aggfunc = np.sum)
        df_jusuitanXiaoshou18 = df_jusuitanXiaoshou18.reset_index()
        total18 = df_jusuitanXiaoshou18.sum()
        total18[0] = '合计'
        df_jusuitanXiaoshou18 = df_jusuitanXiaoshou18.set_index('存货编码')
        df_jusuitanXiaoshou18.loc['合计'] = total18
        # df_jusuitanXiaoshou18 = df_jusuitanXiaoshou18.reset_index()
        df_jusuitanXiaoshou18
           
                 
        path = r"F:\a00nutstore\008\zww08\002电商\聚水潭"
        file = f'电商产品销售出库明细-{qijian}.xlsx'
        fname_jst = os.path.join(path,file)
        wb1 = openpyxl.Workbook()
        ws1 = wb1.active
        ws1.title = '原表'
        wb1.save(fname_jst)
        wb1 = openpyxl.load_workbook(fname_jst)
        ws1 = wb.active
        with pd.ExcelWriter(fname_jusuitan, engine='openpyxl',mode='a', if_sheet_exists='overlay')  as writer:
            df_jst0.to_excel(writer,sheet_name = '原表',header = True,index = False)
            df_chuku.to_excel(writer, sheet_name = '电商出库', header = True)
            df_jusuitanXiaoshou18.to_excel(writer, sheet_name = '销售出库', header = True)
        easygui.msgbox(msg = '程序结束')
        os.startfile(fname_jusuitan)

if __name__=='__main__':
    main()             
        




