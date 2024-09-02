'''
选定目录,对目录下的工作簿xls,下面的工作表进行汇总合并,指定AB列名
'''
import os
import openpyxl
import easygui
import excelmessage


nwb = openpyxl.Workbook()
nws = nwb.active
nws.title = '合并结果'# 新建工作簿和工作表。
msg = '请点选要汇总的文件夹'
path = easygui.diropenbox(msg)
files = os.listdir(path)  # 获取销售表文件夹下的所有工作簿名称。
lst = []
fname = os.path.join(path,os.listdir(path)[0])
fname = excelmessage.excelMessage(fname)
wb = openpyxl.load_workbook(fname)  # 根据工作簿名称读取工作簿对象。
ws = wb.worksheets[0]
old_title = [j.value  for j in ws[1] ]
print(old_title)
wb.close()
daidin1 = input('请输入合并结果报表中表头第一项，类似每个工作簿的名字')
daidin2 = input('请输入合并结果报表中表头第一项，类似每个工作表的名字')
new_title = [daidin1, daidin2] + old_title
print(new_title)
nws.append(new_title)
hz_filename = r'合并结果openpyxl.xlsx'
hz_fname = os.path.join(path, hz_filename)
for file in files:  # 循环工作簿名称。
    fname = os.path.join(path,file)
    fname = excelmessage.excelMessage(fname)
    wb = openpyxl.load_workbook(fname)  # 根据工作簿名称读取工作簿对象。
    for ws in wb.worksheets:  # 循环工作簿下所有工作表。
        n = 0
        for index,row in enumerate(ws.values): # 循环工作表下所有行记录
            if index ==0 :
                continue
            else :
                newhang = (file.split('.')[0], ws.title)+row
                nws.append(newhang)

    wb.close()

nwb.save(hz_fname)  # 保存新建的工作簿。








