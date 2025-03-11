'''
每月月底固定凭证，根据工次表生成
'''
import re
import os
import easygui
import pandas as pd
import openpyxl

sheet_name = r'工资'
def getDf(msg):
    fname = easygui.fileopenbox(msg=f'请点选{msg}工资表.xlsx')
    df = pd.read_excel(fname, sheet_name=sheet_name)
    return fname

#生产工资
fname_shengcan = getDf('请点选当月“生产人员工资.xlsx”')
# fname_shengcan = r"F:\a00nutstore\008\zw08\gongzi\生产人员工资.xlsx"
path,filename = os.path.split(fname_shengcan)
os.chdir(path)
df_shengcan = pd.read_excel(fname_shengcan,sheet_name = sheet_name)
yingfa = df_shengcan['实发数'].sum() + df_shengcan['代扣社保'].sum() + df_shengcan['扣税'].sum() \
- df_shengcan['补扣税'].sum()

#双佳行管工资
fname = getDf('请点选行管工资')
# fname = r"F:\a00nutstore\008\zw08\gongzi\行管工资.xlsx"
df = pd.read_excel(fname,sheet_name = sheet_name)
gs = '双佳'
df =  df[df['公司'] == gs]
gp = df.groupby('部门')
data = {}
for k,v in gp:
    shebao = gp['社保'].sum()[k]
    yingfashu = gp['应发数'].sum()[k]
    data[k] = shebao+yingfashu
renshu_yanfa = gp['姓名'].get_group('设计研发部').count()
yanglao = round(renshu_yanfa * 588,2)      #养老588元/人
shiye = round(renshu_yanfa * 25.73,2)        #失业25.73元/人
gongshang = round(renshu_yanfa * 60.64,2)        #工伤60.64元/人
yiliao = round(renshu_yanfa * 316.38,2)        #医疗316.38元/人
expense_zhizhao = data['仓库搬运'] + data['生产部']
expense_xingguan = data['行政部'] + data['财务部']
expense_design = data['设计研发部']
expense_sale = data['营销部']

#每月固定凭证
fname_fix = r"F:\a00nutstore\008\zww08\2024\每月固定凭证.xlsx"
wb = openpyxl.load_workbook(fname_fix)
ws = wb['pingzheng']
#写入数据
ws['D5'].value = yingfa
ws['E6'].value = yingfa
ws['D7'].value = expense_zhizhao
ws['D8'].value = expense_xingguan
ws['D9'].value = expense_design
ws['D10'].value = expense_sale
ws['E11'].value = expense_zhizhao + expense_xingguan + expense_design + expense_sale
#社保调整写入
ws['D24'].value = yanglao * -1
ws['D25'].value = yanglao
ws['D26'].value = shiye * -1
ws['D27'].value = shiye
ws['D28'].value = gongshang * -1
ws['D29'].value = gongshang
ws['D30'].value = yiliao * -1
ws['D31'].value = yiliao

newfname = os.path.join(path,'每月固定凭证.xlsx')
wb.save(newfname)
os.startfile(newfname)