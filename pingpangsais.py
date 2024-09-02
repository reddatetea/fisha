import openpyxl
import easygui
import  pingpangaoyunhui
import  pingpangshijinsai
import  pingpangshijiebei
import  pingpangxunhuisai
import  os

def  jinyinton(fname,yundongyuans,guanjun_dic,yajun_dic,jijun_dic):
    wb = openpyxl.load_workbook(fname)
    ws2 = wb.create_sheet(title='金银铜')
    title = ('运动员', '金牌', '银牌', '铜牌', '总奖牌数')
    ws2.append(title)
    yundongyuans = list(yundongyuans)

    for j in range(2, len(yundongyuans) + 2):
        ws2.cell(j, 1).value = yundongyuans[j - 2]

    for index, row in enumerate(ws2.values):
        if index == 0:
            continue
        else:
            yundongyuan = row[0]
            ws2.cell(index + 1, 2).value = guanjun_dic[yundongyuan]
            ws2.cell(index + 1, 3).value = yajun_dic[yundongyuan]
            ws2.cell(index + 1, 4).value = jijun_dic[yundongyuan]
            ws2.cell(index + 1, 5).value = guanjun_dic[yundongyuan] + yajun_dic[yundongyuan] + jijun_dic[yundongyuan]

    wb.save(fname)

def  main():
    msg = '请选择要查询项目的性别：'
    print(msg)
    title = '男子or女子'
    sexs = ['男子', '女子']
    sex = easygui.buttonbox(msg, title, choices=sexs)
    msg = '请选择要查询项目：'
    print(msg)
    title = '单打or双打等'
    items = ['单打', '双打']
    item = easygui.buttonbox(msg, title, choices=items)
    choice = sex[0]+item[0]
    
    msg = '请选择要查询的赛事：'
    print(msg)
    title = msg
    saishis = ['奥运会', '世锦赛','世界杯','巡回赛总决赛']
    saishi = easygui.buttonbox(msg, title, choices=saishis)

    msg = '请选择要查询的项目：'
    fname = r'乒乓球{}历届{}战绩.xlsx'.format(choice,saishi)
    ws_name = choice
    if  saishi == '奥运会':
        yundongyuans, guanjun_dic, yajun_dic, jijun_dic = pingpangaoyunhui.aoyunhui(fname,ws_name)
    elif saishi == '世锦赛':
        yundongyuans, guanjun_dic, yajun_dic, jijun_dic = pingpangshijinsai.shijinsai(fname, ws_name)
    elif saishi == '世界杯':
        yundongyuans, guanjun_dic, yajun_dic, jijun_dic = pingpangshijiebei.shijiebei(fname, ws_name)
    else :
        yundongyuans, guanjun_dic, yajun_dic, jijun_dic = pingpangxunhuisai.xunhuisai(fname, ws_name)
    jinyinton(fname, yundongyuans, guanjun_dic, yajun_dic, jijun_dic)
    os.system(fname)


if  __name__ =='__main__':
    main()



