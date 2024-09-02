import openpyxl
import  pingpangaoyunhui
import  pingpangshijinsai
import  pingpangshijiebei
import  easygui
import os

def  sandaisaijiangpaishu(choice):

    ws_name = choice

    fname = r'乒乓球{}历届奥运会战绩.xlsx'.format(choice)
    ay_yundongyuans,ay_guanjun_dic,ay_yajun_dic,ay_jijun_dic = pingpangaoyunhui.aoyunhui(fname,ws_name)

    fname = r'乒乓球{}历届世锦赛战绩.xlsx'.format(choice)
    sjs_yundongyuans,sjs_guanjun_dic,sjs_yajun_dic,sjs_jijun_dic = pingpangshijinsai.shijinsai(fname,ws_name)

    fname = r'乒乓球{}历届世界杯战绩.xlsx'.format(choice)
    sjb_yundongyuans,sjb_guanjun_dic,sjb_yajun_dic,sjb_jijun_dic = pingpangshijiebei.shijiebei(fname,ws_name)

    yundongyuans = set()
    for j in list(ay_yundongyuans):
        yundongyuans.add(j)
    for j in list(sjs_yundongyuans):
        yundongyuans.add(j)
    for j in list(sjb_yundongyuans):
        yundongyuans.add(j)

    yundongyuans = list(yundongyuans)
    print(yundongyuans)

    fname =r'乒乓球{}三大赛战绩.xlsx'.format(choice)
    ws_name = r'三大赛'
    ws_name2 = r'排行榜'
    xishus = [11,8,4,5.5,4,2,2.75,2,1]                             #奖牌系数
    xishus = [11, 8, 4, 5.5, 4, 2, 0, 0, 0]  # 奖牌系数
    wb = openpyxl.load_workbook(fname)
    ws  = wb.create_sheet(title =ws_name)
    ws_name = ws.title
    ws2 = wb.create_sheet(title =ws_name2)
    ws_name2=ws2.title
    title =[ '运动员','奥运会金牌','奥运会银牌','奥运会铜牌','世锦赛金牌','世锦赛银牌','世锦赛铜牌','世界杯金牌','世界杯银牌','世界杯铜牌','奖牌总数']
    title2 = ['运动员', '奥运会金牌', '奥运会银牌', '奥运会铜牌', '世锦赛金牌', '世锦赛银牌', '世锦赛铜牌', '世界杯金牌', '世界杯银牌', '世界杯铜牌', '奖牌分值']
    ws.append(title)
    ws2.append(title2)
    for  j in range(2,len(yundongyuans)+2):
        ws.cell(j,1).value = yundongyuans[j-2]
        ws2.cell(j, 1).value = yundongyuans[j - 2]
    wb.save(fname)
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    ws2 = wb[ws_name2]
    for index,row in enumerate(ws.values):
        if  index==0:
            continue
        else :
            yundongyuan = row[0]
            ay_guanjun_dic.setdefault(yundongyuan, 0)
            ay_yajun_dic.setdefault(yundongyuan, 0)
            ay_jijun_dic.setdefault(yundongyuan, 0)
            sjs_guanjun_dic.setdefault(yundongyuan, 0)
            sjs_yajun_dic.setdefault(yundongyuan, 0)
            sjs_jijun_dic.setdefault(yundongyuan, 0)
            sjb_guanjun_dic.setdefault(yundongyuan, 0)
            sjb_yajun_dic.setdefault(yundongyuan, 0)
            sjb_jijun_dic.setdefault(yundongyuan, 0)

            ayjp = ay_guanjun_dic[yundongyuan]
            ayyp = ay_yajun_dic[yundongyuan]
            aytp = ay_jijun_dic[yundongyuan]

            sjsjp = sjs_guanjun_dic[yundongyuan]
            sjsyp = sjs_yajun_dic[yundongyuan]
            sjstp = sjs_jijun_dic[yundongyuan]

            sjbjp = sjb_guanjun_dic[yundongyuan]
            sjbyp = sjb_yajun_dic[yundongyuan]
            sjbtp = sjb_jijun_dic[yundongyuan]

            #下面计算积分
            ayjp_fenzhi = ayjp*xishus[0]
            ayyp_fenzhi = ayyp*xishus[1]
            aytp_fenzhi = aytp*xishus[2]

            sjsjp_fenzhi = sjsjp * xishus[3]
            sjsyp_fenzhi = sjsyp * xishus[4]
            sjstp_fenzhi = sjstp * xishus[5]

            sjbjp_fenzhi = sjbjp * xishus[6]
            sjbyp_fenzhi = sjbyp * xishus[7]
            sjbtp_fenzhi = sjbtp * xishus[8]

            heji = ayjp+ayyp+aytp+sjsjp+sjsyp+sjstp+sjbjp+sjbyp+sjbtp
            heji_fenzhi= ayjp_fenzhi + ayyp_fenzhi + aytp_fenzhi + sjsjp_fenzhi + sjsyp_fenzhi + sjstp_fenzhi + sjbjp_fenzhi + sjbyp_fenzhi + sjbtp_fenzhi
            # hang = [yundongyuan,ayjp,ayyp,aytp,sjsjp,sjsyp,sjstp,sjbjp,sjbyp,sjbtp,heji]
            # ws.append(hang)
            #ws.cell(index, 1).value = yundongyuan
            ws.cell(index+1,2).value = ayjp
            ws.cell(index+1, 3).value = ayyp
            ws.cell(index+1, 4).value = aytp
            ws.cell(index+1, 5).value = sjsjp
            ws.cell(index+1, 6).value = sjsyp
            ws.cell(index+1, 7).value = sjstp
            ws.cell(index+1, 8).value = sjbjp
            ws.cell(index+1, 9).value = sjbyp
            ws.cell(index+1, 10).value = sjbtp
            ws.cell(index+1, 11).value = heji

            ws2.cell(index + 1, 2).value = ayjp_fenzhi
            ws2.cell(index + 1, 3).value = ayyp_fenzhi
            ws2.cell(index + 1, 4).value = aytp_fenzhi
            ws2.cell(index + 1, 5).value = sjsjp_fenzhi
            ws2.cell(index + 1, 6).value = sjsyp_fenzhi
            ws2.cell(index + 1, 7).value = sjstp_fenzhi
            ws2.cell(index + 1, 8).value = sjbjp_fenzhi
            ws2.cell(index + 1, 9).value = sjbyp_fenzhi
            ws2.cell(index + 1, 10).value = sjbtp_fenzhi
            ws2.cell(index + 1, 11).value = heji_fenzhi


    wb.save(fname)

def main():
    path = r'F:\a00nutstore\娱乐\乒乓球比赛成绩'
    os.chdir(path)
    msg = '请选择要查询项目的性别：'
    print(msg)
    title = '男子or女子'
    sexs = ['男子', '女子']
    sex = easygui.buttonbox(msg, title, choices=sexs)
    msg = '请选择要查询项目：'
    print(msg)
    title = '单打or双打等'
    items = ['单打', '双打']
    item = easygui.buttonbox(msg, title, choices=items)
    choice = sex[0]+item[0]
    sandaisaijiangpaishu(choice)

if __name__ == '__main__':
    main()










