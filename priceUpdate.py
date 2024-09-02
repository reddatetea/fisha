import pandas as pd
import numpy as np
import easygui
import os
import xlwings as xw
from beifenFile import beifen
import openpyxl

# df1 = pd.read_clipboard(sep='\\n+',usecols = [0,1,2,3,19])
keys = ['日期','单据号','供货单位','品名']
df1 = pd.read_clipboard(sep='\\n+',dtype = {'单据号':str})
df1['日期']=pd.to_datetime(df1['日期'],format='%Y-%m-%d')
df1 =df1[keys+['合同金额']]
df1 = df1.astype({'日期':'datetime64[ns]'})
df1.dropna(subset=['合同金额'],inplace=True)
fname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
ws_name = '流水账'
df0 = pd.read_excel(fname,ws_name,index_col=0,dtype = {'单据号':str})
df8 = pd.merge(df0,df1,on=keys,how='left')
df8['入库金额'] = np.where(df8['合同金额'].notna(), df8['合同金额'],
                                df8['入库金额'])
df8['入库单价'] = np.where(df8['合同金额'].notna(),df8.apply(lambda  x :round(x['入库金额'] /x['入库数量'],2) if  x['入库数量']!=0 else x['入库单价'],axis=1),df8['入库单价'])
df8['吨价'] = np.where((df8['入库单价'] .notna())*(df8['单位']=='kg')!=1, df8['吨价'],df8['入库单价']*1000)
df8['吨价'] = np.where((df8['入库单价'].notna())*(df8['单位']=='公斤')!=1, df8['吨价'],df8['入库单价']*1000)
riqi = easygui.enterbox('请输入记账日期"2021-12-20"')
df8['记账'] = np.where(df8['合同金额'].notna(), [riqi]*len(df8),
                                df8['记账'])
df9 = df8.iloc[:,:-1]
df9 = df9.sort_values(by=['日期','单据号'])
fname1 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
sheet_name = '流水账'
beifen(fname1)   #价格更新前进行备份
app = xw.App(visible = False,add_book = False)
wb = app.books.open(fname1)
ws = wb.sheets[sheet_name]
ws.clear_contents()
wb.save()
wb.close()
app.quit()

with pd.ExcelWriter(fname1,engine = 'openpyxl',
date_format="YYYY-MM-DD",
datetime_format="YYYY-MM-DD HH:MM:SS",mode='a',
                    if_sheet_exists='overlay') as writer:
    df9.to_excel(writer, sheet_name, index=False)
wb = openpyxl.load_workbook(fname1)
ws = wb[sheet_name]
max_row = ws.max_row
for i in range(2, max_row + 1):
    ws[f'A{i}'].number_format = 'yyyy-m-d'
    ws[f'L{i}'].number_format = 'yyyy-m-d'
    ws[f'S{i}'].number_format = 'yyyy-m-d'
wb.save(fname1)

easygui.msgbox('程序结束')
# os.startfile(fname1)