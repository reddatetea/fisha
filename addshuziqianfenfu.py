'''
对指定工作簿的工作表，数字加千分符
'''
import openpyxl


def addShuziStyle(fname,ws_name,min_row=2,min_col =2):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    taitou = [j.value for j in ws[1]]
    max_row = ws.max_row
    max_col = ws.max_column
    dun_col = []
    for j in range(1,len(taitou)+1):
        for i in ['重量','吨']:
            if i in taitou[j-1]:
                dun_col.append(j)
    print(dun_col)

    for i in range(2,max_col+1):
        for j in range(2,max_row+1):
            if i not in dun_col:
                if ws.cell(j,i).value == 0:
                    continue
                else :
                    ws.cell(j, i).number_format = '#,##0.00'
            else :
                if ws.cell(j, i).value == 0:
                    continue
                else:
                    ws.cell(j, i).number_format = '#,##0.000'
    wb.save(fname)

def main():
    fname =r'材料入库单.xlsx'
    ws_name ='入库'
    addShuziStyle(fname, ws_name)

if __name__ == '__main__':
    main()