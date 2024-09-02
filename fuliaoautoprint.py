'''制作数据透视表的通用程序！但筛选项是单选 ，日期选项 里datetime.datetime(2020,09,30,0,0)，输入 9/30/2020
#2021-09-29更新为toushibiaoAutoPrint.py,此程序留着以后参考
'''
import os
import xlwings as xw
import pandas as pd
import openpyxl
import easygui
from openpyxl.utils import  get_column_letter,column_index_from_string
import excelcellfillruku
# import excelfastsetingruku
import excelprint

def choiceSheet():
    fname =r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    path, filename = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname,data_only=True)
    ws_name = '流水账'
    ws = wb[ws_name]
    zhiziduan =  [ '入库数量', '入库金额']
    hangziduan ='cwName'
    jisuanfangsi = 'sum'
    saixuanliezhi = input('请输入期间（如2020-10)\n')
    wb.close()
    return fname, ws_name, zhiziduan,hangziduan, jisuanfangsi,saixuanliezhi

def toushiBiao(fname, ws_name, gongyingshang,zhiziduan,hangziduan,jisuanfangsi,saixuanliezhi):
    app = xw.App(visible = False,add_book = False)
    wb = app.books.open(fname)
    ws = wb.sheets[ws_name]
    values = ws.range('A1').expand('table').options(pd.DataFrame).value
    max_row = values.shape[0] + 1

    values = ws.range('A1').expand('table').options(pd.DataFrame).value
    filtered =values[(values[u'期间']=='{}'.format(saixuanliezhi)) & (values[u'供货单位']=='{}'.format(gongyingshang))]

    # print(fname, ws_name, zhiziduan,hangziduan, jisuanfangsi,saixuanliezhi)
    pivottable = pd.pivot_table(filtered, values=zhiziduan, index=hangziduan, columns=None, aggfunc=jisuanfangsi, fill_value =0,margins=True, margins_name='合计')
    order = zhiziduan
    pivottable = pivottable[order]
    nws = worksheets['{}'.gongyingshang]  # 新表
    nws.clear()
    nws.range('A1').value = pivottable
    nws.range('A1').value = '品名'
    nws.range('D1').value = '不含税金额'
    value1s = nws.range('A1').expand('table').options(pd.DataFrame).value

    max_row = value1s.shape[0] + 1
    for j in range(2,max_row):
        nws.range('D{}'.format(j)).formula = '=ROUND(C{}/1.13,2)'.format(j)
    last_row = max_row - 1
    nws.range('D{}'.format(max_row)).formula = '=SUM(D2:D{})'.format(last_row)
        #nws.range('D{}'.format(j)).formula ='=round(C{}/1.13,2)'.format(j)
    nws.range('B1').expand('down').number_format = '#,##0.00'
    nws.range('D1').expand('down').number_format = '#,##0.00'
    nws.range('C1').expand('down').number_format = '#,##0.00'

    workbook.save()
    workbook.close()
    app.quit()
    return gongyingshang

def  choiceJPrintGongyingshangs(fname, ws_name, zhiziduan,hangziduan,jisuanfangsi,saixuanliezhi):
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open(fname)
    worksheets = workbook.sheets
    ws = worksheets[ws_name]
    values = ws.range('A1').expand('table').options(pd.DataFrame).value
    max_row = values.shape[0] + 1

    gongyingshangs = set()
    for j in range(2, max_row):

        if ws.range('k{}'.format(j)).value == saixuanliezhi:
            gongyingshang = ws.range('C{}'.format(j)).value
            gongyingshangs.add(gongyingshang)
        else:
            continue
    gongyingshangs = list(gongyingshangs)
    print(gongyingshangs)
    msg = '请选择供应商'
    print(msg)
    gongyingshangs = easygui.multchoicebox(msg, title=msg, choices=gongyingshangs)
    workbook.close()
    app.quit()
    return gongyingshangs


def main():
    fname, ws_name, zhiziduan, hangziduan, jisuanfangsi,saixuanliezhi= choiceSheet()
    gongyingshangs  = choiceJPrintGongyingshangs(fname, ws_name, zhiziduan,hangziduan,jisuanfangsi,saixuanliezhi)
    for gongyingshang in gongyingshangs :
        gongyingshang =toushiBiao(fname, ws_name, gongyingshang,zhiziduan,hangziduan,jisuanfangsi,saixuanliezhi)
        wb = openpyxl.load_workbook(fname)
        ws_name1 = 'toushibiao'
        ws1 = wb[ws_name1]
        max_row = ws1.max_row
        max_col = ws1.max_column
        min_row = 1
        min_col = 1
        excelcellfillruku.excelcellFill(fname,ws_name1,min_row,min_col,max_row,max_col)
        fname, ws_name1 = excelfastsetingruku.setPrintArea(fname)
        excelfastsetingruku.firstRowSeting(fname, ws_name1)
        excelfastsetingruku.yemei(fname, ws_name1,gongyingshang)
        #os.system(fname)
        excelprint.wsPrint(fname,ws_name1)

if __name__ == '__main__':
    main()