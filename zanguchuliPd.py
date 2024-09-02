'''
20211014用pandas进行暂估处理
'''
import os
import openpyxl
import easygui
import pandas as pd
import numpy as np
import re

def tojianchen(zaiyao,jianchens):
    for j in jianchens:
        if j in zaiyao:
            return j
            break

def zanguChuli(jianchens):
    easygui.msgbox('请复制当月暂估数据区域')
    df = pd.read_clipboard()
    year = df.columns[0][:4]
    df.columns = [re.sub(r'方向(\S*)$', '方向', i) for i in df.columns]
    df.columns = [re.sub(r'\d{4}年月$', 'month', i) for i in df.columns]
    df.columns = [re.sub(r'\d{4}年日$', 'day', i) for i in df.columns]
    df = df[[i for i in df.columns if i not in ['供应商编号', '供应商名称', '科目编号', '科目名称']]]
    for i in ['贷方本币','借方本币','余额本币']:
        df[f'{i}'] = df[f'{i}'].str.replace('¥', '').str.replace(',', '')
        df[f'{i}'] = df[f'{i}'].astype('float')
    df.month = df.month.fillna(99)
    df.day = df.day.fillna(99)
    df['month'] = df['month'].astype('int').astype('str')
    df['day'] = df['day'].astype('int').astype('str')
    df = df.assign(month=[i.zfill(2) for i in df.month])
    df = df.assign(day=[i.zfill(2) for i in df.day])
    df['month'] = df['month'].replace('99','')
    df['day'] = df['day'].replace('99','')
    df1 = df.copy()
    df1 = df.loc[df.isnull().sum(1) < 1]
    df1['余额本币'] = ''
    df1['借方本币'] = df1['摘要'].apply(lambda x: tojianchen(x, jianchens))  #供应商简称
    df1['方向'] = round(df1['贷方本币'] * 1.13, 2)   #含税应付账款
    df1 = df1.assign(余额本币=lambda s: np.where(df1.摘要.str.contains('冲|调整|&'), '有', '无'))

    df1 = df1.assign(shijian=year + '/' + df1.month + '/' + df1.day)
    # df1.shijian = df1.shijian.transform(lambda x: pd.Timestamp(x))
    # df1.shijian = df1.shijian.map(lambda x: x.strftime('%Y/%m/%d'))
    df1.shijian = df1.shijian.astype('datetime64[ns]')

    fname1 = easygui.fileopenbox('请选择当月暂估文件：')
    path,_ = os.path.split(fname1)
    os.chdir(path)
    wb1 = openpyxl.load_workbook(fname1,data_only=True)
    sheet0 = wb1.sheetnames[1]  # 默认暂估工作表
    regax = r'\d{4}'
    pattern = re.compile(regax)
    for sheetname in wb1.sheetnames:
        mat = pattern.search(sheetname)
        if mat != None:
            sheet0 = mat.group()
            break
        else:
            continue
    ws1 = wb1['双佳']
    ws0= wb1[sheet0]

    max_row = ws1.max_row
    max_row2 = ws0.max_row
    with pd.ExcelWriter(fname1,engine='openpyxl') as writer:
        writer.book = wb1
        writer.sheets = dict((ws1.title,ws1) for ws1 in wb1.worksheets)
        df.to_excel(writer,sheet_name = '双佳',index = False,header=None,startrow=max_row)
        df1.to_excel(writer, sheet_name=sheet0, index=False, header=None, startrow=max_row2)
        for i in range(max_row2 + 1, len(df1) + max_row2 + 1):
            ws0[f'I{i}'].number_format = 'yyyy-mm-dd'

    return fname1

def main():
    path = r'F:\a00nutstore\006\zw\yingfu'
    os.chdir(path)
    fname_jianchen = r'F:\a00nutstore\006\zw\yingfu\暂估供应商简称.xlsx'
    sheet_name0 = r'简称'
    df_jianchen = pd.read_excel(fname_jianchen, sheet_name0)
    jianchens = df_jianchen['简称'].to_list()  # 简称列表
    fname1=zanguChuli(jianchens)
    os.startfile(fname1)

if __name__ == '__main__':
    main()



