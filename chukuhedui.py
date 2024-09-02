'''
出库录入后核对，将数量金额明细账与材料出库EXCEL核对
学习openpyxl有写入 excel公式
字母转列号from openpyxl.utils import get_column_letter, column_index_from_string的用法
'''
# _*_ coding:utf-8 _*_
import chaiwutotal
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import dataonly
import excelmessage
import easygui


def daiyu():
    kemuList, kemuname_dic, fname2 = chaiwutotal.chaiwuTotal()
    # wb2 = openpyxl.load_workbook(fname2, read_only=True)  #取公式
    wb2 = openpyxl.load_workbook(fname2, data_only=True)  # 直接取值
    # 数量金额账--数量金额汇总
    ws2 = wb2['数量金额汇总']
    maxrows2 = ws2.max_row
    # 货方数量和期末数量字典daiyu_dic
    daiyu_dic = {}
    # 科目名称所在列号
    kemunum = column_index_from_string('r')
    # 货方数量所在列号
    dainum = column_index_from_string('k')
    # 期末数量所在列号
    qimonum = column_index_from_string('v')

    for row in range(2, maxrows2 - 2 + 1):
        kemu = ws2.cell(row, kemunum).value
        daifangshuliang = ws2.cell(row, dainum).value
        qimoshuliang = ws2.cell(row, qimonum).value
        daiyu_dic[kemu] = [daifangshuliang, qimoshuliang]

    print(daiyu_dic)

    wb2.close()
    return daiyu_dic


def chukuHedui(daiyu_dic):
    msg ='请点选当月出库参考'
    fname = easygui.fileopenbox(msg,title='出库excel')
    path,filename = os.path.split(fname)
    os.chdir(path)
    fname = os.path.join(path, filename)
    wb1 = openpyxl.load_workbook(fname)
    sheetnames = wb1.sheetnames
    msg = '请点选出库参考copy'
    sheetname1 = easygui.buttonbox(msg,title ='出库参考copy',choices=sheetnames)
    ws1 = wb1[sheetname1]
    maxrows1 = ws1.max_row

    # 科目名称所在列号
    kemu_num1 = column_index_from_string('R')

    # 实际出库数量所在列号
    dai_num1 = column_index_from_string('AK')
    # 差异所在列号
    chayi_num1 = column_index_from_string('AL')
    # 实际期末数量所在列号
    qimo_num1 = column_index_from_string('AM')
    # 实际差异所在列号
    shijichayi_num1 = column_index_from_string('AN')

    for row1 in range(1, maxrows1 - 2 + 1):
        if row1 == 1:
            ws1.cell(row1, dai_num1, value='实际出库数量')
            ws1.cell(row1, chayi_num1, value='出库差异')
            ws1.cell(row1, qimo_num1, value='实际期末数量')
            ws1.cell(row1, shijichayi_num1, value='期末差异')

        else:
            kemu1 = ws1.cell(row1, kemu_num1).value
            daiyu_dic.setdefault(kemu1, [0, 0])
            dai = daiyu_dic[kemu1][0]
            yu = daiyu_dic[kemu1][1]

            # 写入贷方数量
            ws1.cell(row1, dai_num1, value=dai)
            # 写入期末数量
            ws1.cell(row1, qimo_num1, value=yu)
            # 写入出库差异数量，用公式
            ws1.cell(row1, chayi_num1, value='=AH' + str(row1) + '-AK' + str(row1))
            # 写入期末数量实际差异，用公式
            # 铁丝特例
            if kemu1 == '铁丝':
                ws1.cell(row1, shijichayi_num1, value='=AD' + str(row1) + '/2-AM' + str(row1) + '+AF' + str(row1))
            else:
                ws1.cell(row1, shijichayi_num1, value='=AD' + str(row1) + '-AM' + str(row1) + '+AF' + str(row1))

    wb1.save(fname)
    return (fname, sheetname1)


def chukujinjian(newname,sheetname1):
    path,filename5=os.path.split(newname)
    os.chdir(path)
    wb3 = openpyxl.load_workbook(newname, data_only=True)
    sheetnames = wb3.sheetnames
    ws3 = wb3[sheetname1]
    maxrow3 = ws3.max_row

    # 科目编码所在列号
    kemu_daima_num = column_index_from_string('C')
    # 科目名称所在列号
    kemu_mingchen_num = column_index_from_string('R')
    # 出库数量列号
    chuku_shuliang_num = column_index_from_string('AH')
    # 单价所在列号
    chuku_danjia_num = column_index_from_string('AI')
    # 出库金额所在列号
    chuku_jiner_num = column_index_from_string('AJ')

    # 建新工作簿‘出库202004'
    chuku_month = input('请输入出库月份:')
    filename4 = '%s月出库.xlsx' % chuku_month
    fname4 = os.path.join(path, filename4)
    dangyue = '%s月出库' % chuku_month


    wb4 = openpyxl.Workbook(fname4)
    ws4 = wb4.create_sheet(title=dangyue)
    wb4.save(fname4)

    wb5 = openpyxl.load_workbook(fname4, data_only=True)
    ws5 = wb5.active

    taitou = ('科目代码', '科目名称', '出库数量', '单价', '金额')

    for row in range(1, maxrow3 + 1):
        if row == 1:
            ws5.cell(row, 1, value='科目代码')
            ws5.cell(row, 2, value='科目名称')
            ws5.cell(row, 3, value='出库数量')
            ws5.cell(row, 4, value='单价')
            ws5.cell(row, 5, value='金额')
            # ws5.append(taitou)

        else:
            if  ws3.cell(row, chuku_jiner_num).value in [0,'',None]:
                continue
            else :

                kemu_daima = ws3.cell(row, kemu_daima_num).value
                kemu_mingchen = ws3.cell(row, kemu_mingchen_num).value
                chuku_shuliang = ws3.cell(row, chuku_shuliang_num).value
                chuku_danjia = ws3.cell(row, chuku_danjia_num).value
                chuku_jiner = ws3.cell(row, chuku_jiner_num).value
                # hang = (kemu_daima,kemu_mingchen,chuku_shuliang,chuku_danjia,chuku_jiner)
                # ws5.append(hang)
                ws5.cell(row, 1, value=kemu_daima)
                ws5.cell(row, 2, value=kemu_mingchen)
                ws5.cell(row, 3, value=chuku_shuliang)
                ws5.cell(row, 4, value=chuku_danjia)
                ws5.cell(row, 5, value=chuku_jiner)

    wb5.save(fname4)
    wb3.close()
    return fname4,dangyue

def delekonghang(fname4,dangyue):
    wb = openpyxl.load_workbook(fname4)
    ws_name = dangyue
    ws = wb[ws_name]
    max_row = ws.max_row

    deleterows = []

    for index, rows in enumerate(ws.values):
        if index == 0:
            continue
        else:
            if ws.cell(index, 5).value in [0, '', None]:
                deleterows.append(index)

    deleterows.sort(reverse=True)
    for i in deleterows:
        ws.delete_rows(idx=i)
    wb.save(fname4)


def main():
    print('找开已输完出库的双佳数量金额xls')
    daiyu_dic = daiyu()
    fname, sheetname1 = chukuHedui(daiyu_dic)
    #dataonly
    newname = dataonly.dataOnly(fname)

    #纯出库数据

    fname4,dangyue = chukujinjian(newname,sheetname1)

    delekonghang(fname4,dangyue)


if __name__ == '__main__':
    main()











