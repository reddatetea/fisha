import openpyxl
import re
import easygui

def shijiebei(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    max_row = ws.max_row
    guanjuns = []
    guanjuns_range = ws['D2:' + 'D{}'.format(max_row)]
    for row in guanjuns_range:
        for j in row:
            guanjuns.append(j.value.strip())
    print(guanjuns)

    yajuns = []
    yajuns_range = ws['E2:' + 'E{}'.format(max_row)]
    for row in yajuns_range:
        for j in row:
            yajuns.append(j.value.strip())
    print(yajuns)

    jijuns = []
    jijuns_range = ws['F2:' + 'F{}'.format(max_row)]
    for row in jijuns_range:
        for j in row:
            jijuns.append(j.value.strip())
    print(jijuns)
    yundongyuans = set(guanjuns + yajuns + jijuns)
    print(yundongyuans)

    # guanjun_dic 冠军字典
    guanjun_dic = {}
    yajun_dic = {}
    jijun_dic = {}

    for yundongyuan in yundongyuans:
        guanjun_dic.setdefault(yundongyuan, 0)
        for row in guanjuns_range:
            for j in row:
                guanjun_dic[yundongyuan] = guanjuns.count(yundongyuan)

    for yundongyuan in yundongyuans:
        yajun_dic.setdefault(yundongyuan, 0)
        for row in yajuns_range:
            for j in row:
                yajun_dic[yundongyuan] = yajuns.count(yundongyuan)

    for yundongyuan in yundongyuans:
        jijun_dic.setdefault(yundongyuan, 0)
        for row in jijuns_range:
            for j in row:
                jijun_dic[yundongyuan] = jijuns.count(yundongyuan)

    print(guanjun_dic)
    print(yajun_dic)
    print(jijun_dic)
    wb.close()

    return  yundongyuans,guanjun_dic,yajun_dic,jijun_dic

def  jinyinton(fname,yundongyuans,guanjun_dic,yajun_dic,jijun_dic):
    wb = openpyxl.load_workbook(fname)
    ws2 = wb.create_sheet(title='金银铜')
    title = ('运动员', '金牌', '银牌', '铜牌', '总奖牌数')
    ws2.append(title)
    yundongyuans = list(yundongyuans)

    for j in range(2, len(yundongyuans) + 2):
        ws2.cell(j, 1).value = yundongyuans[j - 2]

    for index, row in enumerate(ws2.values):
        if index == 0:
            continue
        else:
            yundongyuan = row[0]
            ws2.cell(index + 1, 2).value = guanjun_dic[yundongyuan]
            ws2.cell(index + 1, 3).value = yajun_dic[yundongyuan]
            ws2.cell(index + 1, 4).value = jijun_dic[yundongyuan]
            ws2.cell(index + 1, 5).value = guanjun_dic[yundongyuan] + yajun_dic[yundongyuan] + jijun_dic[yundongyuan]

    wb.save(fname)

def  main():
    msg = '请选择要查询项目的性别：'
    print(msg)
    title = '男子or女子'
    sexs = ['男子', '女子']
    sex = easygui.buttonbox(msg, title, choices=sexs)
    msg = '请选择要查询项目：'
    print(msg)
    title = '单打or双打等'
    items = ['单打', '双打']
    item = easygui.buttonbox(msg, title, choices=items)
    choice = sex[0]+item[0]
    fname = r'乒乓球{}历届世界杯战绩.xlsx'.format(choice)
    ws_name = choice
    yundongyuans, guanjun_dic, yajun_dic, jijun_dic = shijiebei(fname,ws_name)
    jinyinton(fname, yundongyuans, guanjun_dic, yajun_dic, jijun_dic)

if  __name__ =='__main__':
    main()



