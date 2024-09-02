# _*_ conding = utf-8 _*_
'''
查看excel文件基本信息，如果是低版的xls文件，则自动将其转化为高版本的xlsx
需输入完整路路径和带后缀的文件名
本版解决后缀xls大小写混写
'''
from xlsxlsx import xlsXlsx
from openpyxl import Workbook,load_workbook
import os
import re
import easygui


def wenjian(msg='请点选要处理的excel文件'):
    fname = easygui.fileopenbox(msg,title='file')
    return fname

def excelMessage(fname):
    if os.path.splitext(fname)[-1]=='.xls':
        xlsXlsx(fname)
        fname = fname + 'x'

    else:
        fname = fname

    # print('下面将显示excel基本信息')
    # print('文件名为:%s' % fname)

    wb = load_workbook(fname)
    sheetnames = wb.sheetnames

    # 第一个工作表的名字
    sheetname = wb.sheetnames[0]
    sheet = wb[sheetname]

    jishu = 0
    first_list = []
    #print('上面打印的是excel表的第1列所有单元格')
    for index, row in enumerate(sheet.values):
        #print(row[0])
        first_list.append(row[1])

        if row[0] == None:
            jishu = jishu + 1
            # print(jishu)
    #print('上面打印的是excel表的第1列所有单元格' )
    mrows = sheet.max_row - jishu
    # print('工作表名称：', sheetnames)
    # print('实际操作的工作表名称：', sheetname)
    # print('工作表的最大行数是：', sheet.max_row)
    # print('工作表的最大列数是：', sheet.max_column)
    # print('工作表计数是：', jishu)
    # print('工作表可操作的最大行数：', mrows)
    wb.close()
    return fname

def main():
    msg = '"请点选产成品当日和累计"excel文件'
    fname =wenjian(msg)
    print(fname)
    excelMessage(fname)

if __name__=='__main__':
    main()






