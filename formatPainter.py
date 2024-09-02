'''
用openpyxl库对excel进行格式刷操作
'''

import openpyxl
import easygui
import os
import excelmessage


def painter(ws,cell):
    j = cell
    #字休
    font_name = j.font.name
    font_size = j.font.size
    font_italic = j.font.italic
    font_bold = j.font.bold
    font_color = j.font.color
    font = openpyxl.styles.Font(name=font_name, size=font_size , bold=font_bold, italic=font_italic, color=font_color)
    #边框
    border_left = j.border.left
    border = openpyxl.styles.Border(left=border_left, right=border_left , top=border_left, bottom=border_left)
    #填充
    fill_type = j.fill.fill_type
    fgColor = j.fill.fgColor
    pattern_fill = openpyxl.styles.PatternFill(fill_type=fill_type, fgColor=fgColor)
    #倾斜
    horizontal = j.alignment.horizontal
    vertical = j.alignment.vertical
    text_rotation = j.alignment.text_rotation
    wrap_text = j.alignment.wrap_text
    alignment = openpyxl.styles.Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text)
    #数字格式
    number_format = j.number_format
    #protection
    locked = j.protection.locked
    hidden = j.protection.locked
    protection =  openpyxl.styles.protection.Protection(locked=locked,hidden=hidden)
    #行高
    height = ws.row_dimensions[j.row].height
    # 列宽
    col_letter = openpyxl.utils.get_column_letter(j.column)
    width = ws.column_dimensions[col_letter].width
    return font,border,pattern_fill,alignment,number_format,protection,height,width

def getStyles(ws,area0):
    styles = []
    for row in area0:
        hang = []
        for j in row:
            font, border, pattern_fill, alignment, number_format, protection, height, width = painter(ws, j)
            cell_sytle = [font,border,pattern_fill,alignment,number_format,protection,height,width]
            hang.append(cell_sytle)
        styles.append(hang)
    return styles

def mAres(ws):    #获取工作表所有合并单元格地址
    m_list = ws.merged_cells
    cr = []
    for m_area in m_list:
        # print(m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col)
        # 合并单元格的起始行坐标、终止行坐标。。。。，
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # 纵向合并单元格的位置信息提取出
        if c2 - c1 > 0:
            cr.append((r1, r2, c1, c2))
        else:
            continue

    cr.sort()
    return cr

def m_Area0(ws,area0):   #获取某区域所有合并单元格地址
    marea = mAres(ws)
    cr = set()
    for i in marea:
        row_min = i[0]
        col_min = i[1]
        row_max = i[2]
        col_max = i[3]

        for row in area0:
            for j in row:
                row = j.row
                col = j.column
                if row >= row_min and row <= row_max and col >= col_min and col <= col_max:
                    cr.add(i)
                else:
                    continue

    cr1 = list(cr)
    cr1.sort()
    return cr1

def unmergeAres(nws,area):
    nws.unmerge_cells(start_row=area[0],start_column=area[2],end_row=area[1],end_column=area[3])

def mergeAres(nws,area):
    nws.merge_cells(start_row=area[0],start_column=area[2],end_row=area[1],end_column=area[3])

def writeStyles(nws,styles,len_result,max_column,row_start,col_start):
    for row in range(row_start,len_result+row_start):
        styles_row = styles[row-row_start]
        for col in range(col_start,max_column+col_start):
            nws.cell(row,col).font  = styles_row[col-col_start][0]
            nws.cell(row,col).border  = styles_row[col-col_start][1]
            nws.cell(row,col).fill  = styles_row[col-col_start][2]
            nws.cell(row,col).alignment  = styles_row[col-col_start][3]
            nws.cell(row,col).number_format  = styles_row[col-col_start][4]
            nws.cell(row,col).protection  = styles_row[col-col_start][5]
            height  = styles_row[col-col_start][6]
            width = styles_row[col - col_start][7]
            col_letter = openpyxl.utils.get_column_letter(col)
            nws.column_dimensions[col_letter].width = width
        nws.row_dimensions[row].height = height
    return nws

def stylesFormat(nws,area0,cell_start1):
    merger_area0 = m_Area0(nws, area0)
    styles = getStyles(nws, area0)
    row_start0 = area0[0][0].row
    col_start0 = area0[0][0].column
    row_end0 = area0[-1][-1].row
    col_end0 = area0[-1][-1].column
    len_result0 = len(styles)
    max_col = col_end0 - col_start0 + 1
    row_start1 = cell_start1.row
    col_start1 = cell_start1.column
    row_offset = row_start1 - row_start0
    col_offset = col_start1 - col_start0
    merge_area1 = [(area[0] + row_offset, area[1] + row_offset, area[2] + col_offset, area[3] + col_offset) for area in
                   merger_area0]
    ws = writeStyles(nws, styles, len_result0, max_col, row_start1, col_start1)
    for area in merge_area1:
        mergeAres(ws, area)

def choiceFname0(fname0):
    wb0 = openpyxl.load_workbook(fname0)
    sheetnames = wb0.sheetnames
    ws_name = easygui.choicebox('请点选源工作表',choices=sheetnames)
    ws0 = wb0[ws_name]
    # area0 = ws['J4:N4']
    area0_input = easygui.enterbox('请输入源工作表区域“J4:N4”')
    area0 = ws0[area0_input]
    return fname0,wb0,ws0,area0

def choiceFname1(fname1):
    wb1 = openpyxl.load_workbook(fname1)
    sheetnames1 = wb1.sheetnames
    ws_name1 = easygui.choicebox('请点选目标工作表',choices=sheetnames1)
    ws1 = wb1[ws_name1]
    # area0 = ws['J4:N4']
    cell_start1_input = easygui.enterbox('请输入目标工作表区域初始单元格“E7”')
    return fname1,wb1,ws1,cell_start1_input


def main():
    fname0 = excelmessage.wenjian('请点选源工作薄')
    fname0,wb0,ws0,area0 = choiceFname0(fname0)
    fname1= excelmessage.wenjian('请点选源工作薄')
    fname1,wb1,ws1,cell_start1_input = choiceFname1(fname1)
    cell_start1 = ws1[cell_start1_input]
    stylesFormat(ws1,area0,cell_start1)
    # for cell_start1 in ws['E7:E23']:                   #一行刷多行
    #     stylesFormat(ws, area0, cell_start1[0])
    wb1.save(fname1)
    wb0.close()


if __name__ == '__main__':
    main()