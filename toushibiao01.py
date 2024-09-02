import os
import xlwings as xw
import pandas as pd
import openpyxl
import easygui

def choiceSheet():
    msg = '请点选要进行数据透视表的 excel文件'
    fname = easygui.fileopenbox(msg, title='excel')
    path, filename = os.path.split(fname)
    os.chdir(path)

    wb = openpyxl.load_workbook(fname,data_only=True)
    sheetnames = wb.sheetnames

    if len(sheetnames) == 1:
        ws_name = sheetnames[0]
    else:
        msg = '请选择要排序的工作表'
        print(msg)
        ws_name = easygui.buttonbox(msg, title='sheet', choices=sheetnames)

    ws = wb[ws_name]

    max_row = ws.max_row
    max_column = ws.max_column

    firstrow = ws[1]
    print(firstrow)
    ziduans = []
    for i in firstrow:
        ziduans.append(i.value)

    ziduans.append(None)

    msg = '请选择值字段'
    print(msg)
    zhiziduan = easygui.multchoicebox (msg, title=msg, choices=ziduans)


    # hungziduans = set()
    # for i in ws['A']:
    #     hungziduans.add(i.value)
    msg = '请选择行字段'
    print(msg)
    hangziduan = easygui.multchoicebox(msg, title=msg, choices=ziduans)

    msg = '请选择列字段'
    print(msg)
    lieziduan = easygui.multchoicebox(msg, title=msg, choices=ziduans)


    msg = '请选择值字段的计算方式'
    jisuanfangsis = ['sum','mean']
    print(msg)
    jisuanfangsi = easygui.buttonbox(msg, title=msg, choices=jisuanfangsis)


    wb.close()
    return fname, ws_name, zhiziduan,hangziduan, lieziduan,jisuanfangsi

def toushiBiao(fname, ws_name, zhiziduan,hangziduan, lieziduan,jisuanfangsi):

    app = xw.App(visible = False,add_book = False)

    workbook = app.books.open(fname)
    worksheets = workbook.sheets
    ws = worksheets[ws_name]
    values = ws.range('A1').expand('table').options(pd.DataFrame).value

    newsheetname = input('请输入创建的新的透视表名字\n')
    worksheets.add(newsheetname)
    nws = worksheets[newsheetname]    #新表

    #values = '销售金额'对应的是值字段
    #index = '销售地区‘对应的是行字段
    #cloumns = '销售分部' 对应的是列字段
    #aggfunc = 'sum'，是值字段的计算方式
    #margins = True 是否需要汇总
    #margins_name = '堂堂' ，汇总命名
    print(fname, ws_name, zhiziduan,hangziduan, lieziduan,jisuanfangsi)

    pivottable = pd.pivot_table(values,values = zhiziduan,index = hangziduan,columns = None,aggfunc = jisuanfangsi,fill_value =0,margins = True,margins_name = '合计')
    nws.range('A1').value = pivottable
    workbook.save()
    workbook.close()

    app.quit()


def main():
    fname, ws_name, zhiziduan, hangziduan, lieziduan, jisuanfangsi = choiceSheet()
    toushiBiao(fname, ws_name, zhiziduan,hangziduan, lieziduan,jisuanfangsi)

if __name__ == '__main__':
    main()