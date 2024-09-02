#通用加求和公式，指定工作簿、工作表和指定求和字段
import  os
import excelmessage
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import easygui

def  sumheji(fname):
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    msg = '请点选要求和的工作表'
    sts  = easygui.multchoicebox(msg,choices = sheetnames)
    for st in sts :
        ws = wb[st]
        max_row = ws.max_row
        fields  = [j.value for j in ws[1]]
        msg = '请点选工作表{}要求和的列名'.format(st)
        lienames = easygui.multchoicebox(msg,choices = fields)
        letters = [get_column_letter(fields.index(liename)+1)  for  liename in lienames]
        for letter in letters :
            ws['A{}'.format(max_row+1)].value = '合计'
            ws['{}{}'.format(letter,max_row+1)].value = '=sum({}2:{}{})'.format(letter,letter,max_row)
    wb.save(fname)

def main():
    msg = '请点选要加计合计数的excel文件'
    print(msg)
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    sumheji(fname)

if __name__ == '__main__':
    main()
