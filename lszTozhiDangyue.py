'''
本模块是从流水账中将纸类材料入库数据引入，原材料实时流水账----纸当月入库20200524
jianchen字典的键是全称，值是简称
'''
# _*_ coding:utf-8 _*_
import openpyxl
import os
import datetime
import easygui
import excelmessage
import time
import zhigongyingshang

def lszTozhidangyue(qijian,piaojuhao):
    # 当天日期
    dtrq = datetime.date.today().strftime('%Y%m%d')
    print('选择纸当月入库路径')
    #path = easygui.diropenbox(msg = '请选择纸当月入库路径',title = '')
    path = r'F:\a00nutstore\006\zw\else'
    os.chdir(path)
    filename = '纸当月入库%s.xlsx' % dtrq
    fname = os.path.join(path,filename)
    wb1 = openpyxl.Workbook()
    ws1 = wb1.create_sheet(title='当月')

    print('选择原材料实时流水账')
    #fname2 = excelmessage06.excelMessage()
    fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'

    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['流水账']

    print(ws2.max_row)
    jianchen = zhigongyingshang.jianchen
    for i in range(1, ws2.max_row + 1):
        if i == 1:
            taitou = (
                '公司', '供应商', '期间', '开票日期', '单号', '仓库品名', '令数', '财务品名', '单位', '批次', '计算重量', '入库kg', '入库吨', '吨价0', '吨价',
                '令价', '金额', '送货单金额', '不含税金额', '多计', '令价0', '记账', '备注','pricename')
            ws1.append(taitou)
        else:
            if (ws2.cell(i, 3).value in jianchen) and (ws2.cell(i, 11).value == qijian) and (ws2.cell(i,2).value >= int(piaojuhao)):

                ws1.cell(i, 1, value='双佳')
                ws1.cell(i, 2, value=ws2.cell(i, 3).value)
                ws1.cell(i, 3, value=ws2.cell(i, 11).value)
                ws1.cell(i, 4, value=ws2.cell(i, 1).value)
                ws1.cell(i, 5, value=ws2.cell(i, 2).value)
                ws1.cell(i, 6, value=ws2.cell(i, 4).value)
                ws1.cell(i, 7, value=ws2.cell(i, 14).value)
                ws1.cell(i, 8, value=ws2.cell(i, 9).value)
                ws1.cell(i, 9, value=ws2.cell(i, 5).value)
                ws1.cell(i, 11, value=round(ws2.cell(i, 15).value, 3))
                ws1.cell(i, 12, value=round(ws2.cell(i, 15).value, 3))
                ws1.cell(i, 13, value=round(ws2.cell(i, 15).value, 3))

                ws1.cell(i, 14, value=round(ws2.cell(i, 17).value, 2))
                ws1.cell(i, 15, value=round(ws2.cell(i, 17).value, 2))
                ws1.cell(i, 16, value=round(ws2.cell(i, 16).value, 2))

                ws1.cell(i, 17, value=round(round(ws2.cell(i, 15).value, 3) * ws2.cell(i, 17).value, 2))
                #ws1.cell(i, 17, value='=round(round(M'+str(i)+', 3) * '+'N'+str(i)+', 2)')

                #送货单金额
                ws1.cell(i, 18, value=round(ws2.cell(i,8).value,2))
                ws1.cell(i, 19, value=round(round(ws2.cell(i, 15).value, 3) * ws2.cell(i, 17).value / 1.13, 2))
                ws1.cell(i,24, value=ws2.cell(i,10).value)



            else:
                continue
    wb1.save(fname)
    wb2.close()

    # 删除行
    #fname3 = r'F:\a00nutstore\006\zw\else\纸当月入库%s.xlsx'%dtrq
    fname3 = fname
    wb3 = openpyxl.load_workbook(fname3)
    ws3 = wb3['当月']

    nodelelist = []
    for i in range(1, ws3.max_row + 1):
        if ws3.cell(i, 1).value != None:
            nodelelist.append(i)

    ws4 = wb3.create_sheet(title='当月正')

    for index, row in enumerate(ws3.values):
        if index == 0:
            ws4.append(row)
        else:
            if row[0] != None:
                ws4.append(row)
            else:
                continue

    wb3.save(fname3)
    time.sleep(3)
    return fname3

def main():
    qijian = input('请输入期间：格式为2020-04：\n')
    piaojuhao = input('请输入票据号(包含该号)：\n')

    lszTozhidangyue(qijian,piaojuhao)


if __name__ == '__main__':
    main()



















