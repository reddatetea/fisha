'''
根据all月原材料价格表.xlsx制作价格字典gongying_dic，并形成文件yuancailiaoPriceDic.py
'''
import openpyxl
import pprint
import excelmessage
import easygui

def yuancailiaoprice():
    print('现在更新价格字典，请选allyue原材料价格表')
    #fname = excelmessage06.excelMessage()
    fname = r'F:\a00nutstore\006\zw\price\all月原材料价格表.xlsx'
    wb = openpyxl.load_workbook(fname)
    #sheetnames = wb.sheetnames
    #choice = easygui.buttonbox(msg = '请点选原材料价格表',title = '价格表',choices=sheetnames)
    choice = 'all月'
    ws = wb[choice]
    mrows = ws.max_row
    rcols = ws.max_column
    gongying_dic = {}
    qijian_dic = {}
    for i in range(2,mrows+1):
        gongyingshang = ws.cell(i,1).value
        qijian = ws.cell(i,2).value
        pinming = ws.cell(i,3).value
        dunjia = ws.cell(i,4).value
        gongying_dic.setdefault(gongyingshang,{})
        gongying_dic[gongyingshang].setdefault(qijian,{})
        gongying_dic[gongyingshang][qijian][pinming] = dunjia


    print(qijian_dic)
    print(gongying_dic)
    resultFile = open('yuancailiaoPriceDic.py', 'w',encoding='utf-8')
    resultFile.write('allDate = ' + pprint.pformat(gongying_dic))
    resultFile.close()

    wb.close()
    print(gongying_dic)
    return gongying_dic

def main():
    yuancailiaoprice()


if __name__=="__main__":
    main()




