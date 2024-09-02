import openpyxl
import glob

# 汇总表的路径设定好了
path = '/Users/wenwen/Desktop/部门预算统计/10月预算汇总表.xlsx'

# 请写下你的代码
# 打开汇总表并打开找到数据表
wb_total = openpyxl.load_workbook(path)
ws = wb_total.active

# 获取所有部门的数据表，在同目录下，以“10月预算.xlsx”结尾
deps = glob.glob('/Users/wenwen/Desktop/部门预算统计/*10月预算.xlsx')

# 将部门按名称排序
deps.sort()

# 对 deps 中的文件依次读取，写入汇总表
for dep in deps:
  wb1 = openpyxl.load_workbook(dep)
  ws1 = wb1.active
  dep = dep.replace('/Users/wenwen/Desktop/部门预算统计/','')
  dep_name =dep.replace('10月预算.xlsx','')
  values = [row[0].value for row in ws1['c4:c12']]
  values.insert(0,dep_name)
  ws.append(values)
  wb1.close



# 依次填入 K 列的汇总公式
rowtotal = ws['k3:k12']
for row in rowtotal:
  for cell in row :
    #cell.vlaue = '=sum(B{}:J{})'.format(cell.row,cell.row)
    cell.value = '=SUM(B{}:J{})'.format(cell.row,cell.row)


# 依次填入 13 行的汇总公式
coltotal = ws['b13:k13']
for row in coltotal:
  for cell in row:
    cell.value = '=sum({}3:{}12)'.format(cell.column_letter,cell.column_letter)


# 别忘了填入 A13 的文字
ws['a13'].value = '合计'



# 为了方便云服务识别文件，不要改动最后一行的代码哦
wb_total.save('数据汇总.xlsx')