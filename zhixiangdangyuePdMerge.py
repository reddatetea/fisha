'''
本模块是从流水账中将纸箱入库数据引入
'''
# _*_ coding:utf-8 _*_
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
import os
import re
import datetime
import easygui
import NewGongyinshang

def Dangyue(start_riqi,end_riqi):
    dtrq = datetime.date.today().strftime('%Y%m%d')    #当天日期
    path = r'F:\a00nutstore\006\zw\duizhang'
    os.chdir(path)
    filename = '纸箱当月入库%s.xlsx'%dtrq
    fname = os.path.join(path, filename)
    fname_gy = r'F:\a00nutstore\006\zw\price\纸箱供应商.xlsx'
    jianchen = NewGongyinshang.Gongyingshang(fname_gy)
    # price_dic = zhixiangdanjiajisuan.priceDic()
    fname2 = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    df = pd.read_excel(fname2, sheet_name='流水账',usecols=[0,1,2,3,5,6,7,9,10],index_col=0,dtype = {'单据号':str})    #usecols直接取所取的行
    df.sort_index(inplace=True)  # 对索引排序
    df = df.truncate(before=start_riqi, after=end_riqi)
    df = df.loc[df['供货单位'].isin (jianchen)] #isin非常实用
    df.insert(3,'箱型',df['品名'])
    for j in ['长', '宽', '高', '长*宽*高', '适装产品'][::-1] :
        df.insert(5,j,None)
    df = df.iloc[:,:-2]     #去掉最后两列，pricename，期间
    for j in ['单个面积', '面积','制版费', '合同单价', '合同金额'][::-1]:
        df.insert(11, j, 0.0)
    for j in [ '多计', '多计金额', '备注']:
        df[j] = 0.0                             #0.0是flaost64,0是ind64
    df [ '备注'] = None
    zhixiang_dic = zhixiangdic()                   #生成纸箱字典
    df['长'] = df['箱型'].map(lambda x:zhixiang_dic.get(x,(0,0,0))[0])                 #分别根据字典取长宽高的值
    df['宽'] = df['箱型'].map(lambda x:zhixiang_dic.get(x,(0,0,0))[1])
    df['高'] =df['箱型'].map(lambda x:zhixiang_dic.get(x,(0,0,0))[2])
    df['长*宽*高'] =df['长'].map(str)+'*'+df['宽'].map(str)+'*'+df['高'].map(str)                     #str在前面不行！
    df['类别'] = df['箱型'].map(lambda x:zhixiangLeibian(x))
    df['单个面积'] = (df['长'] + df['宽'] + 80) / 1000 * (df['宽'] + df['高'] + 50) / 1000 * 2
    df['面积'] =df['单个面积'] * df['入库数量']
    df = df.reset_index()
    return fname,jianchen,df

def zhixiangdic():
    #fname = r'D:\a00nutstore\006\zw\ZHIXIANG\2020纸箱入库.xlsx'
    fname = r'F:\a00nutstore\006\zw\ZHIXIANG\2020纸箱入库.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb['纸箱尺寸']
    maxrows = ws.max_row
    pinming_number = column_index_from_string('E')
    chang_number = column_index_from_string('F')
    kuan_number = column_index_from_string('G')
    gao_number = column_index_from_string('H')
    zhixiang_dic = {}
    for row in range(2,maxrows+1):
        pinming = ws.cell(row,pinming_number).value
        chang = ws.cell(row, chang_number).value
        kuan = ws.cell(row, kuan_number).value
        gao = ws.cell(row, gao_number).value
        zhixiang_dic[pinming]=(chang,kuan,gao)                  #(长，宽，高)
        wb.close()
    return zhixiang_dic

def zhixiangLeibian(pinming):
    leixiangs = ['内箱']
    # 质量好一点的纸箱
    haoxiangs = ['C7箱', 'C8箱', 'C9箱', 'F10箱', 'F11箱', 'F12箱', 'F13箱', 'F14箱', 'F15箱', 'F16箱']
    ruiyixiangs = ['锐意箱', '锐意外箱']
    xinruixiangs = ['新锐']
    # 外贸箱
    waimaoxiangs = ['外贸箱', '外贸外箱']
    zhixiangs = leixiangs + haoxiangs + ruiyixiangs + xinruixiangs + waimaoxiangs
    for zhixiang in zhixiangs:
        r = r'.*(?P<xiaolei>%s).*' % zhixiang
        pattern = re.compile(r)
        pipei = re.search(pattern, pinming)

        if pipei is not None:
            zhixiang_xiaolei = pipei.group('xiaolei')
            if zhixiang_xiaolei in leixiangs :
                zhixiang_leibie = 'leixiang'
                return zhixiang_leibie
            elif zhixiang_xiaolei in haoxiangs :
                zhixiang_leibie = 'haoxiang'
                return zhixiang_leibie
            elif zhixiang_xiaolei in xinruixiangs :
                zhixiang_leibie = 'xinruixiang'
                return zhixiang_leibie
            elif zhixiang_xiaolei in ruiyixiangs :
                zhixiang_leibie = 'ruiyixiang'
                return zhixiang_leibie
            elif zhixiang_xiaolei in waimaoxiangs :
                zhixiang_leibie = 'waimaoxiang'
                return zhixiang_leibie

            else :
                zhixiang_leibie = 'leimaoxiang'
                return zhixiang_leibie

        else:
            continue
    zhixiang_leibie = 'leimaoxiang'
    return zhixiang_leibie

'''
def jiuanyuan(zhixiang_leibie,chang,kuan,gao):
    dange_mianji = (chang+kuan+80)/1000*(kuan+gao+50)/1000*2
    if zhixiang_leibie == 'leixiang':
        zx_singlePrice = 2.16
    elif zhixiang_leibie == 'haoxiang':
        zx_singlePrice = 3.82
    elif zhixiang_leibie == 'ruiyixiang':
        zx_singlePrice = 3.82
    elif zhixiang_leibie == 'xinruixiang':
        zx_singlePrice = 3.26
    elif zhixiang_leibie == 'waimaoxiang':
        zx_singlePrice = 3.82
    else :
        zx_singlePrice = 3.42
    hetong_danjia = round(dange_mianji*zx_singlePrice,2)
    print(hetong_danjia)
    return dange_mianji, hetong_danjia
'''

def Hetong(end_riqi):
    fname = r'F:\a00nutstore\006\zw\price\纸箱合同价格.xlsx'
    sheet_name = '合同单价'
    df = pd.read_excel(fname, sheet_name, index_col=0)
    grouped = df.groupby('供货单位')
    all_values = []
    for gys, values in grouped:
        print('gys', gys)
        print('values', values)
        last_day = pd.Timestamp(end_riqi)
        values.loc[last_day, '供货单位'] = gys
        values = values.resample('D').ffill()
        values.fillna(method='ffill', inplace=True)
        all_values.append(values)
    df2 = pd.concat(all_values, axis=0)
    return df2

def dfMerge(df1,df2):
    df2 = df2.reset_index()
    df8 = pd.merge(df1, df2, on=['日期', '供货单位'], how='left')  #连接！从python的传统字典到pandas的连接！量变到质变的飞跃！
    return df8

def huizhong(jianchen,fname,df8):
    df8 = df8.assign(singlePrice=df8.apply(lambda x: x[x.类别], axis=1))    #根据类别，确定sinlePrice的值,横竖交叉定位！非常实用！
    df8['合同单价'] = round(df8['单个面积'] * df8['singlePrice'], 2)
    df8['合同金额'] = round(df8['合同单价'] * df8['入库数量'], 2)
    df8['多计'] = df8['入库单价'] - df8['合同单价']
    df8['多计金额'] = df8['入库金额'] - df8['合同金额']
    df9  = df8.iloc[:,:-8]
    grouped = df9.groupby('供货单位')
    with pd.ExcelWriter(fname,datetime_format='yyyy-m-d')  as writer:
        df8 = df8.append(df8.sum(numeric_only=True), ignore_index=True)
        df8.to_excel(writer, '当月正', index=False)
        for gys, values in grouped:
            gys = jianchen[gys]
            values1 = values.copy()
            values1.loc['合计', '日期'] = '合计'
            values1.at['合计', '入库数量'] = sum(values['入库数量'])
            values1.at['合计', '入库金额'] = sum(values['入库金额'])
            values1.at['合计', '合同金额'] = sum(values['合同金额'])
            values1.at['合计', '多计金额'] = sum(values['多计金额'])
            values1.to_excel(writer, sheet_name='{}'.format(gys), index=False)

def main():
    start_riqi = pd.Timestamp(easygui.enterbox('请输入入库起始日期期间：格式为2021-11-4:'))
    end_riqi = pd.Timestamp(easygui.enterbox('请输入入库结束日期期间：格式为2021-11-4：'))
    fname,jianchen,df1  = Dangyue(start_riqi,end_riqi)
    df2 = Hetong(end_riqi)
    df8 = dfMerge(df1, df2)
    huizhong(jianchen, fname, df8)
    os.startfile(fname)

if __name__ == '__main__':
    main()



















