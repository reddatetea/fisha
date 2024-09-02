'''
本模块将纸当月入库添加到纸2020入库,纸当月入库20200524----2020入库
'''
# _*_ conding:utf-8 _*_
import openpyxl
import os
import datetime
import easygui

def addzhiDangyue():
    dtrq = datetime.date.today().strftime('%Y%m%d')
    print('请选择纸当月入库路径')
    path = easygui.diropenbox(msg = '请选择纸当月入库路径',title = '')
    os.chdir(path)

    filename = '纸当月入库%s.xlsx' % dtrq
    fname = os.path.join(path, filename)

    wb1 = openpyxl.load_workbook(fname)
    ws1 = wb1['当月正']

    filename2 = '2020入库.xlsx'
    fname2 = os.path.join(path, filename2)
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['入库']
    maxrow2 = ws2.max_row

    for index,row in enumerate(ws1.values):
        if index == 0:
            pass
        else :
            ws2.append(row)

    wb1.close()
    wb2.save(fname2)
    return fname2,maxrow2

def jiagongsi(fname2,maxrow2):
    wb = openpyxl.load_workbook(fname2)
    ws = wb['入库']
    maxrow = ws.max_row
    for i in range(maxrow2+1,maxrow+1):
        #金额用公式表达
        ws.cell(i, 17, value='=round(round(M' + str(i) + ', 3) * ' + 'O' + str(i) + ', 2)')
        #不含税金额用公式表达
        ws.cell(i, 19, value='=round(R' + str(i) + '/1.13 , 2)')
        # 多计用公式表达R2-Q2
        ws.cell(i, 20, value='=R' + str(i) + '-Q'+str(i))

    wb.save(fname2)

def main():
    fname2,maxrow2 = addzhiDangyue()
    jiagongsi(fname2,maxrow2)

if __name__ == '__main__':
    main()
