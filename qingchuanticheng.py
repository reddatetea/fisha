import easygui
import os
from openpyxl import load_workbook,Workbook

def tichengDic(fname1):              #按项目、折扣区间生成提成字典
    wb1 = load_workbook(fname1)
    sheetnames = wb1.sheetnames
    if len(sheetnames) == 1:
        ws1 = wb1.active
    else:
        msg = '请选择提成比例工作表sheet'
        sheetname = easygui.buttonbox(msg, choices=sheetnames)
        ws1 = wb1[sheetname]
    interval = [i.value for i in ws1[1]][1:]          #折扣区间
    ticheng_propotion = {}
    for index,row in enumerate(ws1.rows):
        if index == 0:
            continue
        else :
            zekou_dic = {}
            for i in range(len(interval)):
                tichengbili = [j.value for j in row[1:]]
                zekou_dic[interval[i]] = tichengbili[i]
                ticheng_propotion[row[0].value]= zekou_dic

    # 折扣区间，折扣区间字典
    qujian = [k.split('<=x<') for k in interval]
    qujians = [[float(k[0]), float(k[-1])] for k in qujian]
    qujian_dic = {}
    for i in range(len(qujians)):
        qujian_dic[str(qujians[i])] = interval[i]
    wb1.close()
    return ticheng_propotion,zekou_dic,qujians,qujian_dic

def judgePorpotions(qujians,qujian_dic,zekou):            #判断区间
    porpotion = '0'
    for k in range(len(qujians)):
        xiao = qujians[k][0]
        da = qujians[k][-1]
        if zekou == 1:
            porpotion = '1'
        elif zekou == 0:
            porpotion = '0'
        else :
            if zekou>=xiao and zekou<da:
                porpotion = qujian_dic[str(qujians[k])]
            else :
                continue
    return(porpotion)

def tichengCount(fname2,fname1,ticheng_propotion,zekou_dic,qujians, qujian_dic):
    wb2 = load_workbook(fname2,data_only=True)
    sheetnames = wb2.sheetnames
    if len(sheetnames) == 1:
        ws2 = wb2.active
    else:
        msg = '请选择新年礼盒销售记录工作表sheet'
        sheetname = easygui.buttonbox(msg, choices=sheetnames)
        ws2 = wb2[sheetname]
    fname = r'新年礼盒销售提成.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = '销售提成'
    first_row = [k.value for k in ws2[1]]
    first_row.extend(['折扣','提成金额'])
    for index,row in enumerate(ws2.values):
        if index == 0:
            ws.append(first_row)
        else :
            ws.append(row)
    max_row = ws.max_row
    for i in range(2,max_row+1):
        leibie = ws.cell(i,4).value
        shuliang = ws.cell(i,5).value
        if ws.cell(i,7).value != 0:
            zekou = round(ws.cell(i,6).value/ws.cell(i,7).value,2)
        else:
            zekou = 0
        porpotion = judgePorpotions(qujians, qujian_dic, zekou)
        try :
            ticheng = ticheng_propotion[leibie][porpotion]
        except:
            ticheng = 0
        ws.cell(i,8).value = zekou
        try :
            ws.cell(i,9).value = shuliang*ticheng            #提成金额
        except :
            ws.cell(i, 9).value = 0
    wb.save(fname)
    wb2.close()
    os.startfile(fname)

def main():
    msg = '请点选"提成比例"excel文件'
    print(msg)
    fname1 = easygui.fileopenbox(msg)
    path,filename = os.path.split(fname1)
    os.chdir(path)
    ticheng_propotion, zekou_dic, qujians, qujian_dic = tichengDic(fname1)
    msg = '请点选"新年礼盒销售记录"excel文件'
    print(msg)
    fname2 = easygui.fileopenbox(msg)
    tichengCount(fname2, fname1, ticheng_propotion,zekou_dic, qujians, qujian_dic)

if __name__ == '__main__':
    main()
