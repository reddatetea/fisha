'''
本模块将纸当月入库添加到纸2020入库,纸当月入库20200524----2020入库
'''
# _*_ conding:utf-8 _*_
import openpyxl
import os
import datetime
import easygui
import pandas as pd
import re

#价格字典，20210923更新

def tichuNum(temp):
    pattern = r'卷筒(\d+)$'
    regexp = re.compile(pattern)
    pipei = regexp.search(temp)
    if pipei == None:
        string = temp
    else:
        string = re.sub(pattern, '卷筒', temp)
    return string

def dunjiaDic():
    fname2 = r'F:\a00nutstore\006\zw\else\2020入库.xlsx'
    df = pd.read_excel(fname2, sheet_name='入库')
    df1 = df[df['pricename'].str.contains('返利|价差|冲减|多计') == False]  # 品名中包含返利、价差、冲减等不计入字典
    gongyingshangs = list(df1['供应商'])
    pinmings = list(df1['pricename'])
    cailiaos = list(df1['材料'])                                       #list(map(tichuNum,df1['材料']))
    dunjias = list(df1['吨价'])
    jizhangs = list(df1['记账'])
    dunjia_dic = {}
    for j in range(len(dunjias)):
        gongyingshang = gongyingshangs[j]
        if jizhangs[j] not in ["", None]:
            if gongyingshang != '河南省江河纸业有限公司':
                dunjia_dic[(gongyingshangs[j], pinmings[j])] = dunjias[j]
            else :
                cailiao = tichuNum(cailiaos[j])
                dunjia_dic[(gongyingshangs[j], cailiao)] = dunjias[j]
        else :
            continue

    print(dunjia_dic)
    return dunjia_dic

def zhidangyueTozhiruku():
    dtrq = datetime.date.today().strftime('%Y%m%d')
    print('选择纸当月入库路径')
    #path = easygui.diropenbox(msg = '请选择纸当月入库路径',title = '')
    path = r'F:\a00nutstore\006\zw\else'
    os.chdir(path)
    filename = '纸当月入库%s.xlsx' % dtrq
    fname = os.path.join(path, filename)
    wb1 = openpyxl.load_workbook(fname)
    ws1 = wb1['当月正']
    filename2 = '2020入库.xlsx'
    fname2 = os.path.join(path, filename2)
    wb2 = openpyxl.load_workbook(fname2)
    ws2 = wb2['入库']
    maxrow2 = ws2.max_row
    dic = {}
    danjuhaos = [j.value for j in ws2['E'][1:]]
    jizhangriqis = [j.value for j in ws2['V'][1:]]
    for j in range(len(danjuhaos)):
        if jizhangriqis[j] not in ['',None]:
            dic[danjuhaos[j]] = jizhangriqis[j]          #单据号-记账日期字典
    for index,row in enumerate(ws1.values):
        if index == 0:
            pass
        else :
            if row[4] not in dic:           #如果未记账（单据号不在单据号-记账日期字典中），则添加！
                ws2.append(row)
            else :
                continue
    wb1.close()
    wb2.save(fname2)
    return  fname2,maxrow2

def jiagongsi(fname2,maxrow2,dunjia_dic):
    wb = openpyxl.load_workbook(fname2)
    ws = wb['入库']
    maxrow = ws.max_row
    for i in range(maxrow2+1,maxrow+1):
        #吨价
        gongyingshang = ws.cell(i,2).value
        pinming = ws.cell(i, 24).value
        if gongyingshang == '河南省江河纸业有限公司':
            pinming = tichuNum(pinming)
        else :
            continue
        ws.cell(i, 15).value = dunjia_dic.get((gongyingshang, pinming), 0)     #取最新价格字典，没有则为0
        #金额用公式表达
        ws.cell(i, 17, value='=round(round(M' + str(i) + ', 3) * ' + 'O' + str(i) + ', 2)')
        #不含税金额用公式表达
        ws.cell(i, 19, value='=round(R' + str(i) + '/1.13 , 2)')
        # 多计用公式表达R2-Q2
        ws.cell(i, 20, value='=R' + str(i) + '-Q'+str(i))
    wb.save(fname2)

def main():
    dunjia_dic = dunjiaDic()
    print(dunjia_dic)
    fname2,maxrow2 = zhidangyueTozhiruku()
    jiagongsi(fname2,maxrow2,dunjia_dic)

if __name__ == '__main__':
    main()
