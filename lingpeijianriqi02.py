import openpyxl
import datetime
fname = r'F:\a00nutstore\006\zw\lingpeijian\零配件实时入库2020.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb['ssrk']
# riqi_string0 = ws['b2'].value
for row in ws['B2:B2347']:
    riqi_string0 = row[0].value
    riqi_string= datetime.datetime.strftime(riqi_string0, '%Y-%m-%d')         #标准日期变字符串
    row[0].value = riqi_string

wb.save(fname)
