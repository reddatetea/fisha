'''
09可以运行，不过第143行IF语句，还没有完全搞明白！见文末注释
'''
import os
import sys
import openpyxl
from openpyxl.styles import Font,Border,Side,Fill,Alignment
import excelmessage
import  easygui
import datetime
import formatPainter


def fourshus(ws,row):
    yuechu = ws.cell(row, 5).value  # 月初
    ruku = ws.cell(row, 8).value  # 入库累计
    chuku = ws.cell(row, 10).value  # 出库累计
    jieyu = ws.cell(row, 11).value  # 结余
    return yuechu,ruku,chuku,jieyu

def listsum(data):
    yuechus = []
    rukus = []
    chukus = []
    jieyus = []
    if data != None:
        for j in data:
            if j != None and j[3]!=0:
                yuechus.append(j[0])
                rukus.append(j[1])
                chukus.append(j[2])
                jieyus.append(j[3])
            else :
                yuechus.append(0)
                rukus.append(0)
                chukus.append(0)
                jieyus.append(0)
        lstsum = [sum(yuechus), sum(rukus), sum(chukus), sum(jieyus)]
    else :
        lstsum = [0,0,0,0]
    return lstsum

def addnewTable(fname,riqi):
    dic = {}
    path,filename = os.path.split(fname)
    os.chdir(path)
    wb = openpyxl.load_workbook(fname)
    sheetnames = wb.sheetnames
    if 'AAA' in sheetnames:
        ws1 = wb['AAA']
    else :
        msg = '请点选"当日"工作表'
        print(msg)
        choice = easygui.choicebox(msg,choices=sheetnames)
        ws1 = wb[choice]
    max_row1 = ws1.max_row
    leijis =  {ws1['A{}'.format(row)].value+ws1['A{}'.format(row-1)].value:(ws1[row][5].value, ws1[row][6].value, ws1[row][7].value) \
               for row in  range(2, max_row1 + 1) \
               if ('累计' in ws1['A{}'.format(row)].value) or ('合   计' in ws1['A{}'.format(row)].value)}
    leibie_dic = {ws1['A{}'.format(row)].value: (ws1[row][5].value, ws1[row][6].value, ws1[row][7].value) \
                  for row in  range(2, max_row1 + 1) \
                  if '小计' in ws1['A{}'.format(row)].value}
    pinming_dic = {ws1['B{}'.format(row)].value+ws1['C{}'.format(row)].value:(ws1[row][5].value,ws1[row][6].value,ws1[row][7].value) \
                   for row in range(2,max_row1+1) \
                   if  (ws1['B{}'.format(row)].value!=None) or (ws1['C{}'.format(row)].value!=None)}
    if 'BBB' in sheetnames:
        ws2 = wb['BBB']
    else:
        msg = '请点选"累计"工作表'
        print(msg)
        choice = easygui.choicebox(msg, choices=sheetnames)
        ws2 = wb[choice]
    #新建原材料日报表工作簿
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = '抄本作业及订制'
    for  index,row in enumerate(ws2.values):
        ws3.append(row)
    ws3.delete_cols(13)
    ws3.delete_cols(12)
    ws3.delete_cols(11)
    ws3.delete_cols(10)
    ws3.delete_cols(5)
    ws3['E1'].value = '月初'
    ws3['F1'].value = '入库累计'
    ws3['G1'].value = '出库累计'
    ws3.insert_cols(6,2)                      #列为实数，第几列就是第几列
    ws3.insert_cols(9,1)
    ws3['F1'].value = '上日'
    ws3['G1'].value = '本日入库'
    ws3['I1'].value = '本日出库'
    fname1 = '产成品日报表{}.xlsx'.format(riqi)
    wb3.save(fname1)
    wb3 = openpyxl.load_workbook(fname1)
    ws3 = wb3.active
    max_row3 = ws3.max_row
    for row in range(2,max_row3+1):
        leibie = ws3.cell(row,1).value
        if ws3.cell(row, 2).value!=None and  ws3.cell(row, 3).value!=None :
            pinming = ws3.cell(row, 2).value + ws3.cell(row, 3).value
            shangri = pinming_dic.get(pinming, (0, 0, 0))[0]
            benriruku = pinming_dic.get(pinming, (0, 0, 0))[1]
            benrichuku = pinming_dic.get(pinming, (0, 0, 0))[2]
        else :
            if  '小计' in  leibie:
                shangri = leibie_dic.get(leibie, (0, 0, 0))[0]
                benriruku = leibie_dic.get(leibie, (0, 0, 0))[1]
                benrichuku = leibie_dic.get(leibie, (0, 0, 0))[2]
            else :                                #累计行、合 计行的处理
                leiji = ws3.cell(row,1).value + ws3.cell(row-1,1).value
                shangri = leijis.get(leiji,(0,0,0))[0]
                benriruku =  leijis.get(leiji,(0,0,0))[1]
                benrichuku =leijis.get(leiji,(0,0,0))[2]
        ws3.cell(row, 6).value = shangri
        ws3.cell(row, 7).value = benriruku
        ws3.cell(row, 9).value = benrichuku

    #加抄本合计和账本合计
    for cell in ws3['A']:
        if cell.value == '(I订制缝线本)小计:':
            chaoben_row = cell.row + 1
        else :
            if cell.value == '(H作业本)小计:':
                chaoben_row = cell.row + 1
        if cell.value == '(X装订配件)小计:':
            zhangben_row = cell.row + 1          #2022-7-26需求，此行删除
        elif cell.value == '(Y订制产品)小计:':
            dingzhicanpin_row = cell.row
        elif cell.value == '(Y订制账簿产品)小计:':
            dingzhizhangben_row = cell.row
        elif cell.value == '(Y订制账簿产品)小计:':
            dingzhizhangben_row = cell.row
        elif cell.value == '(Y出口订制产品)小计:':
            chukou_row = cell.row
        elif cell.value == '(Z废次品)小计:':
            feicipin_row = cell.row
        elif cell.value == '合   计':
            zhong_row = cell.row
        else:
            continue
            #print(zhangben_row)
    ws3.cell(chaoben_row, 1).value = '抄本作业本合计'
    yuechu, ruku, chuku, jieyu = fourshus(ws3,chaoben_row)
    chaoben_heji = yuechu, ruku, chuku, jieyu
    dic['抄本作业本合计'] = chaoben_heji    #2022-7-26需求，此行要减掉电商数据
    # ws3.cell(dingzhicanpin_row, 1).value = '订制本合计'
    yuechu, ruku, chuku, jieyu = fourshus(ws3,dingzhicanpin_row)
    dic['订制本合计'] = [yuechu, ruku, chuku, jieyu]
    # ws3.cell(dingzhizhangben_row, 1).value = '订制账本合计'
    yuechu, ruku, chuku, jieyu = fourshus(ws3,dingzhizhangben_row)
    dic['订制账本合计'] = [yuechu, ruku, chuku, jieyu]
    # ws3.cell(feicipin_row, 1).value = '废次品合计'
    yuechu, ruku, chuku, jieyu = fourshus(ws3,feicipin_row)
    dic['废次品合计'] = [yuechu, ruku, chuku, jieyu]
    ws3.cell(chukou_row, 1).value = '出口订制产品合计'
    yuechu, ruku, chuku, jieyu = fourshus(ws3,chukou_row)
    dic['外贸合计'] = [yuechu, ruku, chuku, jieyu]
    # dic['废次品合计'] = [yuechu, ruku, chuku, jieyu]
    # ws3.cell(zhong_row, 1).value = '总计'
    yuechu, ruku, chuku, jieyu = fourshus(ws3, zhong_row)
    dic['总计'] = [yuechu, ruku, chuku, jieyu]
    ws3.cell(zhangben_row, 1).value = '账本合计'      #2022-07-26需求，此行删掉
    wb3.create_sheet(title = '零结存')
    ws4 = wb3['零结存']
    wb3.create_sheet(title='新锐')
    ws5 = wb3['新锐']
    wb3.create_sheet(title='Royal')
    ws6 = wb3['Royal']
    wb3.create_sheet(title='账本')
    ws7 = wb3['账本']
    wb3.create_sheet(title='出口订制')
    ws8 = wb3['出口订制']
    wb3.create_sheet(title='电商')
    ws9 = wb3['电商']
    del_rows = set()
    for index,row in enumerate(ws3.values):
        if index ==0:
            ws4.append(row)
        else :
            if (row[-1] ==0) and ('小计' not in  row[0] and   '累计'  not in  row[0]):
                del_rows.add(index+1)
                ws4.append(row)
            elif   (row[-1] ==0) and ('小计'  in  row[0]) :
                del_rows.add(index+1)
            elif   '累计'  in  row[0]:
                del_rows.add(index + 1)
            else :
                continue
    xinrui =[]
    ruiyi = []
    zhangben = []
    chukoudinzhi =[]
    dianshang = []
    #新锐
    for index, row in enumerate(ws3.values):
        if index == 0:
            ws5.append(row)
            ws6.append(row)
            ws7.append(row)
            ws8.append(row)
            ws9.append(row)
        else:
            if row[-1]==0 and '小计' not in row[0]:      #20220127修改跳过小计为0:
                # 计算零结存各产品，上日数、入库累计、、出库累计
                newhang = [row[4], row[7],  row[9], row[10]]
                if row[0][0]== 'A' or row[0][:2] == '(A':
                    dianshang.append(newhang)
                elif row[0][0] in ['J','K','L'] or row[0][:2] in ['(J','(K','(L']:    #新锐
                    xinrui.append(newhang)
                elif (row[0].startswith('M')) or (row[0].startswith('(M')):  # M
                    ruiyi.append(newhang)
                elif row[0][0] in ['N', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X'] or row[0][:2] in ['(N', '(P', '(Q',
                                                                                                       '(R', '(S', '(T',
                                                                                                    '(U', '(V', '(W',
                                                                                                      '(X']:
                    zhangben.append(newhang)
                elif 'Y出口订制产品' in row[0]:
                    chukoudinzhi.append(newhang)
                else:
                    continue
            else :
                if row[0][0] == 'A' or row[0][:2] == '(A':
                    ws9.append(row)
                elif row[0][0] in ['J', 'K', 'L'] or row[0][:2] in ['(J', '(K', '(L']:  # 新锐
                    ws5.append(row)
                elif (row[0].startswith('M')) or  (row[0].startswith('(M')) :                    #M
                    ws6.append(row)
                elif   row[0][0] in ['N','P','Q','R','S','T','U','V','W','X'] or row[0][:2] in ['(N','(P','(Q','(R','(S','(T','(U','(V','(W','(X'] :
                    ws7.append(row)
                elif '出口订制' in row[0]:
                    ws8.append(row)
                    chukoudinzhi.append([row[4],row[7],row[9],row[10]])
                else :
                    continue
                del_rows.add(index + 1)
    del_rows = list(del_rows)
    del_rows.sort()
    del_rows.reverse()
    #删除行
    for j in del_rows:
        ws3.delete_rows(j)
    #给新锐日报表和产成品日报表加合计公式
    st_names = ['新锐','Royal','账本','电商']
    for st_name in st_names:
        ws_st = wb3[st_name]
        xiaoji_row = [cell.row for cell in ws_st['A'] if '小计' in cell.value]
        max_row_st = ws_st.max_row
        for j in 'EFGHIJK':
            newhang = []
            xiaoji = sum([ws_st['{}{}'.format(j,i)].value for i in xiaoji_row])
            ws_st['{}{}'.format(j, max_row_st + 1)].value = xiaoji
    wb3.save(fname1)
    for st_name in st_names:
        ws_st = wb3[st_name]
        max_row_st = ws_st.max_row
        hang = []
        newhang = []
        for j in 'EFGHIJK':
            i = ws_st['{}{}'.format(j,max_row_st)].value
            hang.append(i)
        newhang.append(hang[0])
        newhang.append(hang[3])
        newhang.append(hang[5])
        newhang.append(hang[6])
        if st_name == '新锐' :
            ws_st['A{}'.format(max_row_st)].value = '新锐合计'
            xinrui.append(newhang)
        elif st_name == '账本' :
            ws_st['A{}'.format(max_row_st)].value = '账本合计'
            zhangben.append(newhang)
        elif st_name == '电商' :
            ws_st['A{}'.format(max_row_st)].value = '电商合计'
            dianshang.append(newhang)
        else :
            ws_st['A{}'.format(max_row_st)].value = 'Royal合计'
            ruiyi.append(newhang)
    xinrui_heji = listsum(xinrui)
    ruiyi_heji = listsum(ruiyi)
    zhangben_heji = listsum(zhangben)
    dianshang_heji = listsum(dianshang)
    # chukoudinzhi_heji = listsum(chukoudinzhi)
    dic['新锐合计'] = xinrui_heji
    dic['锐意合计'] = ruiyi_heji
    dic['账本合计'] = zhangben_heji
    dic['电商合计'] = dianshang_heji
    new_chaoben_heji = [i-j for i,j in zip(chaoben_heji,dianshang_heji)]
    dic['抄本作业本合计'] = new_chaoben_heji   #2022-7-26需求，
    wb3.create_sheet(title='累计')
    ws_t = wb3['累计']
    first_row = ['类别','月初','入库累计','出库累计','结存']
    lsts = ['账本合计','锐意合计','新锐合计','抄本作业本合计','电商合计','订制本合计','订制账本合计','废次品合计','外贸合计','总计']
    ws_t.append(first_row)
    for j in lsts:
        newhang = [j]+dic[j]
        ws_t.append(newhang)
    chaoben_row = [i.value for i in ws3['A']].index('抄本作业本合计') + 1
    zhangben_row = chaoben_row + 1
    # print(chaoben_row, zhangben_row)
    ws3.cell(chaoben_row, 5).value = new_chaoben_heji[0]
    ws3.cell(chaoben_row, 8).value = new_chaoben_heji[1]
    ws3.cell(chaoben_row, 10).value = new_chaoben_heji[2]
    ws3.cell(chaoben_row, 11).value = new_chaoben_heji[3]
    ws3.delete_rows(idx=zhangben_row)
    wb3.save(fname1)
    return fname1

def printseting(fname1,riqi):
    #设置字体、边框、对齐等常量
    font = Font(name='宋体', size=10)
    font_xiaoji = Font(name='宋体', bold=True, size=10)
    font_firstrow = Font(name='宋体', bold=True, size=20)
    thin_danbian = Side(border_style='thin')
    double_danbian = Side(border_style='double')
    thin_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=thin_danbian)
    double_bian = Border(
        left=thin_danbian,
        right=thin_danbian,
        top=thin_danbian,
        bottom=double_danbian)
    wb = openpyxl.load_workbook(fname1)
    for ws in wb.worksheets :
        # 插入第一行
        ws.insert_rows(1)
        ws['A1'].value = '双佳' + ws.title + ' ' + riqi
        ws['A1'].font = font_firstrow
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        max_row = ws.max_row
        max_column = ws.max_column
        if ws.title !='累计':
            # 单元格宽度
            ws.column_dimensions['A'].width = 8.56
            ws.column_dimensions['B'].width = 17.33
            ws.column_dimensions['C'].width = 13.56
            ws.column_dimensions['D'].width = 3.67
            for j in 'EFGHIJK':
                ws.column_dimensions['{}'.format(j)].width = 7.44
                ws.merge_cells('A1:K1')
            for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
                if row == 2:
                    for col in range(1, max_column + 1):
                        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                else:
                    for col in range(1, max_column + 1):
                        if col <= 3:
                            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        else :
            # 单元格宽度
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 18.8
            ws.column_dimensions['C'].width = 18.8
            ws.column_dimensions['D'].width = 18.8
            ws.column_dimensions['D'].width = 18.8
            ws.merge_cells('A1:E1')
            for row in range(2, max_row + 1):  # 从第二行开始，设置单元格格式
                if row == 2:
                    for col in range(1, max_column + 1):
                        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                else:
                    for col in range(1, max_column + 1):
                        if col <= 1:
                            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            ws.cell(row, col).alignment = Alignment(horizontal='right', vertical='center')
        for cell in ws['A']:
            row = cell.row
            if row == 1:
                continue
            else:
                if '小计' in cell.value:
                    for col in range(1, max_column + 1):
                        ws.cell(row, col).font = font_xiaoji
                        ws.cell(row, col).border = double_bian
                        ws.merge_cells('A{}:C{}'.format(row, row))
                else:
                    for col in range(1, max_column + 1):
                        ws.cell(row, col).font = font
                        ws.cell(row, col).border = thin_bian
        ws.freeze_panes = ws['E3']
        ws.print_title_rows = '1:2'  # the first row
        # 页脚设置
        ws.oddFooter.center.text = " &[Page] / &N"  # 1/n
        ws.oddFooter.center.size = 12  # 页脚中字体大小
        ws.oddFooter.center.font = "Tahoma"  # 页脚中字体
        ws.oddFooter.center.color = "000000"  # 页脚中字体颜色
        ws.oddFooter.right.text = "张文伟 &[Date]"  # 页脚右 文字
        ws.oddFooter.right.size = 12  # 页脚右 字体大小
        ws.oddFooter.right.font = "书体坊米芾体"  # 页脚右 字体
        ws.oddFooter.right.color = "000000"  # 页脚右 字体颜色
        ws.page_margins = openpyxl.worksheet.page.PageMargins(top=0.48, header=0.5, left=0.51, right=0.34, footer=0.5,
                                                              bottom=1)
    wb.save(fname1)

def res_path(relative_path):
    """获取资源绝对路径"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def  main():
    msg = '请点选"产成品当日和累计"工作表'
    fname = excelmessage.wenjian(msg)
    fname = excelmessage.excelMessage(fname)
    today_date = datetime.date.today()
    yesterday_date = (today_date + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")
    msg = '报表日期是{}?'.format(yesterday_date)
    choice = easygui.ccbox(msg,title = '请选择"是"或"否"',choices=('是','否'))
    if choice :
        riqi = yesterday_date
    else :
        msg = '请输入报表日期'
        riqi = easygui.enterbox(msg,title=" 昨天日期")
    fname1 = addnewTable(fname,riqi)
    #2022-7-27需求，增加抄本小计行的上日数、本日入库
    wb = openpyxl.load_workbook(fname1)
    ws = wb['抄本作业及订制']
    col_A = [i.value for i in ws['A']]
    caoben_row = col_A.index('抄本作业本合计')
    pinmings = [i.value for i in ws['A']][2:caoben_row]
    shangris = [i.value for i in ws['F']][2:caoben_row]
    benris = [i.value for i in ws['G']][2:caoben_row]
    chukus = [i.value for i in ws['I']][2:caoben_row]
    shangris_xiaoji = [j for i, j in zip(pinmings, shangris) if '小计' in i]
    shangri = sum(shangris_xiaoji)
    benris_xiaoji = [j for i, j in zip(pinmings, benris) if '小计' in i]
    benri = sum(benris_xiaoji)
    chukus_xiaoji = [j for i, j in zip(pinmings, chukus) if '小计' in i]
    chuku = sum(chukus_xiaoji)
    ws.cell(caoben_row + 1, 6).value = shangri
    ws.cell(caoben_row + 1, 7).value = benri
    ws.cell(caoben_row + 1, 9).value = chuku
    wb.save(fname1)
    printseting(fname1, riqi)
    # easygui.msgbox(msg = '程序结束运行')
    fname0 = res_path('img/leiji.xlsx')
    wb0 = openpyxl.load_workbook(res_path('img/leiji.xlsx'))
    ws0 = wb0.active
    area0 = ws0['A1:E12']
    wb = openpyxl.load_workbook(fname1)
    ws = wb['累计']
    cell_start1 = ws['A1']
    formatPainter.stylesFormat(ws, area0, cell_start1)
    wb0.close()
    wb.save(fname1)
    os.startfile(fname1)

if __name__ == '__main__':
    main()

