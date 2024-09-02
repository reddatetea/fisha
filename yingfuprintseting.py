import openpyxl
import os
import easygui
import excelseting
import time
# import excelcellfill
import xlwings as xw
import excelautofit
import excelsort
import excelcellfill
from openpyxl.utils import get_column_letter,column_index_from_string

def yingfudanyue(fname1,jiner_wu_total):
    wb1 = openpyxl.load_workbook(fname1)
    ws1 = wb1['应付账款参考']
    max_row1 = ws1.max_row
    path,filename = os.path.split(fname1)
    os.chdir(path)
    ws_name = input('请输入本月应付账款名字"0730"\n')
    newfilname = '应付账款{}.xlsx'.format(ws_name)
    fname = os.path.join(path,newfilname)
    wb = openpyxl.Workbook()
    wb.save(fname)
    time.sleep(3)
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    ws.title = ws_name

    for index,row in enumerate(ws1.values):
        kemubianma = row[0]
        gongyingshang = row[1]
        qichu = row[2]
        fukuan = row[3]
        goujin = row[4]
        qimo = row[5]
        shouxin = row[13]
        print(gongyingshang,shouxin)
        # qimo = '=C{}-D{}+E{}'.format(index+1,index+1,index+1)
        # firsthang = (row[0],row[1],row[2],row[3],row[4],row[5])
        newhang = (kemubianma,gongyingshang,qichu,fukuan,goujin,qimo,shouxin)
        # if index == 0 :
        #     ws.append(firsthang)
        if row[2]==0 and row[3]==0 and row[4]==0 and row[5]==0 :
            continue

        ws.append(newhang)

    wb.save(fname)

    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    max_row = ws.max_row
    max_column = ws.max_column

    ziduan_col_letter = 'B'   #供应商的列号为‘B’
    max_column_letter = 'G'    #最后一列的列号‘F’
    fname = excelsort.excelSort(fname, ws_name, ziduan_col_letter, max_column_letter, max_row, max_column,sort_choice=1)
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    hejihang = ('','合计')
    ws.append(hejihang)
    max_row = ws.max_row
    max_col = ws.max_column
    shuju_row = ws.max_row
    shuju_col = ws.max_column
    print(max_row)
    # ws['C{}'.format(max_row)]='=SUM(C2:C{})'.format(max_row-1)
    # ws['D{}'.format(max_row)] = '=SUM(D2:D{})'.format(max_row-1)
    # ws['E{}'.format(max_row)] = '=SUM(E2:E{})'.format(max_row-1)
    # ws['F{}'.format(max_row)]= '=SUM(F2:F{})'.format(max_row-1)
    ws['C{}'.format(max_row)] = sum([j.value for j in ws['c'][1:] if j.value not in ['',None]])
    ws['D{}'.format(max_row)] = sum([j.value for j in ws['D'][1:]if j.value not in ['',None]])
    ws['E{}'.format(max_row)] = sum([j.value for j in ws['E'][1:] if j.value not in ['',None]])
    ws['F{}'.format(max_row)] = sum([j.value for j in ws['F'][1:] if j.value not in ['',None]])
    konghang = ('',)
    ws.append(konghang)
    wb.save(fname)
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    for index,row in enumerate(ws.values):
        if index == 0:
            continue
        elif index == max_row-1:
            break
        else :
            if row[1] == '暂估' :
                ws['C{}'.format(index + 1)] =0
                ws['D{}'.format(index + 1)] = 0
                ws['E{}'.format(index + 1)] = jiner_wu_total
                ws['F{}'.format(index + 1)] = row[2]-row[3]+row[4]
            else :
                ws['F{}'.format(index+1)] = row[2]-row[3]+row[4]
    wb.save(fname)
    return fname, ws_name, shuju_row, shuju_col

def addShuziStyle(fname,ws_name,shuju_row, shuju_col):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    for row in ws.iter_rows(min_row = 2,min_col =3,max_row = shuju_row,max_col = shuju_col):
        for cell in row:
            cell.number_format = '#,##0.00'
    wb.save(fname)

def addBeizhu(fname,ws_name):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    msg = '请点选上次应付账款excel文件'
    print(msg)
    fname2 = easygui.fileopenbox(msg,title = 'excel')
    wb2 = openpyxl.load_workbook(fname2)
    sheetnames = wb2.sheetnames
    if len(sheetnames) == 1:
        ws2 = wb2.active
    else :
        msg = '请点选工作表'
        print(msg)
        choice = easygui.buttonbox(msg,title='sheet',choices=sheetnames)
        ws2 = wb2['%s'%choice]

    max_row2 = ws2.max_row
    print(max_row2)
    for i in ws2.iter_cols(min_row = 2,min_col =1,max_row = max_row2,max_col = 1):
        for cell in i :
            print(cell.value)
            if  cell.value=='备注' :
                start_row = cell.row
            elif  cell.value=='注意':
                end_row = cell.row
            else:
                continue
    print(start_row,end_row)
    konghang = ('',)
    ws.append(konghang)
    for index,row in enumerate(ws2.values):
        if index < start_row -1:
            continue
        elif index >= end_row -1:
            break
        else :
            newhang = (row[0],)
            ws.append(newhang)
    wb.save(fname)
    return fname

def copyTwoBiao(fname,ws_name,shuju_row):
    wb = openpyxl.load_workbook(fname,data_only=True)
    ws = wb[ws_name]
    caigou = ws_name +'采购'
    chaiwu = ws_name +'财务'
    wb.copy_worksheet(ws).title = caigou
    ws_caigou = wb[caigou]
    ws_caigou.move_range('a{}:a{}'.format(shuju_row+2,ws_caigou.max_row),0,1)
    ws_caigou.delete_cols(1, 1)
    ws_caigou.column_dimensions['A'].width = 30.88  # 列宽
    ws_caigou.column_dimensions['B'].width = 16.63
    ws_caigou.column_dimensions['C'].width = 15.5
    ws_caigou.column_dimensions['D'].width = 15.5
    ws_caigou.column_dimensions['E'].width = 16.63
    ws_caigou.column_dimensions['F'].hidden = True
    ws_caigou.page_margins = openpyxl.worksheet.page.PageMargins(top=1, header=0.5, left=0.38, right=0.38, footer=0.5,
                                                          bottom=1)

    wb.copy_worksheet(ws_caigou).title = chaiwu
    ws_chaiwu = wb[chaiwu]
    ws_chaiwu.column_dimensions['A'].width = 35  # 列宽
    ws_chaiwu.column_dimensions['B'].width = 19.25
    ws_chaiwu.column_dimensions['C'].width = 17.5
    ws_chaiwu.column_dimensions['D'].width = 17.5
    ws_chaiwu.column_dimensions['E'].width = 19.25
    ws_chaiwu.column_dimensions['F'].width = 16
    ws_chaiwu.page_margins = openpyxl.worksheet.page.PageMargins(top=1, header=0.5, left=0.38, right=0.38, footer=0.5,
                                                                 bottom=1)
    wb.save(fname)
    return fname,ws_name,caigou,chaiwu

def main(jiner_wu_total):
    msg = '请点选应付账款参考excel文件'
    print(msg)
    fname1 = easygui.fileopenbox(msg,title='excel')
    # fname1 =r'F:\a00nutstore\006\zw\yingfu\2020yingfu\应付账款参考.xlsx'
    #jiner_wu_total = 1000000
    fname,ws_name,shuju_row,shuju_col = yingfudanyue(fname1,jiner_wu_total)
    # print(fname)
    # #fname = r'D:\a00nutstore\006\zw\yingfu\2020yingfu\应付账款0930.xlsx'
    # excelcellfill.excelcellFill(fname,ws_name,max_row=shuju_row,max_col=shuju_col,min_row=1,min_col=1)
    excelautofit.excelAutofit(fname,ws_name)
    addShuziStyle(fname,ws_name,shuju_row,shuju_col)   #数字格式为千分符
    addBeizhu(fname,ws_name)
    excelseting.setPrintArea(fname,ws_name)
    excelcellfill.excelcellFill(fname,ws_name,shuju_row,shuju_col,min_row=1,min_col=1)
    # excelseting.firstRowSeting(fname, ws_name)
    # excelseting.yemei(fname,ws_name)

    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    ws.append(('注意',))
    wb.save(fname)
    fname, *ws_names = copyTwoBiao(fname, ws_name, shuju_row)

    yingfu_name = '应付账款' + ws_name
    for name in ws_names:
        # wb = openpyxl.load_workbook(fname)
        # ws = wb[name]
        # for i in range(2,shuju_row):
        #     ws.cell(i,5).value = '=C{}+D{}-E{}'.format(i,i,i)
        # for i in 'BCDE':
        #     ws['{}{}'.format(i,shuju_row)].value = '=sum({}2:{}{}'.format(i,i,shuju_row-1)
        # wb.save(fname)

        excelseting.yemei(fname,name,yingfu_name)



if __name__ == '__main__':
    jiner_wu_total =533198.49
    main(jiner_wu_total)

# fname = r'D:\a00nutstore\006\zw\yingfu\2021yingfu\应付账款0711.xlsx'
# ws_name = '0711'
# shuju_row = 114
# copyTwoBiao(fname,ws_name,shuju_row)