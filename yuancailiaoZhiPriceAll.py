'''
根据纸入库制作纸价格字典，制作纸所有期间价格字典
注意金额不能为空！
'''
# _*_ conding:utf-8 _*_
import os
import openpyxl
from openpyxl.utils import column_index_from_string

def zhipriceDic():
    #print('请输入纸入库文件所在路径：')
    #path = input('')

    path = r'F:\a00nutstore\006\zw\else'
    os.chdir(path)
    filename = r'2020入库.xlsx'
    fname = os.path.join(path,filename)
    # wb = openpyxl.load_workbook(fname)
    # nws = wb.copy_worksheet(wb['入库'])
    # nws.title = '入库copy'
    # wb.save(fname)

    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb['入库']
    zhiprices = {}

    maxrow = ws.max_row
    # for i in 'GKLMNOPQ':
    #     for j in range(2,maxrow+1):
    #         if ws['{}{}'.format(i,j)].value in ['',None]:
    #             ws['{}{}'.format(i,j)].value = 0
    #         else :
    #             continue
    # 获取列号
    gongyingshang_num = column_index_from_string('B')
    qijian_num = column_index_from_string('C')
    cailiao_num = qijian_num = column_index_from_string('F')
    pricename_num = column_index_from_string('X')
    price_num = column_index_from_string('O')
    lingshu_num = column_index_from_string('G')
    dunshu_num = column_index_from_string('M')
    jiner_num = column_index_from_string('R')
    jizhang_num = column_index_from_string('V')

    for row in range(2,maxrow+1):
        dangqian_qijian = ws.cell(row,qijian_num).value
        gongyingshang = ws.cell(row,gongyingshang_num).value
        pricename = ws.cell(row,pricename_num).value
        cailiao = ws.cell(row, cailiao_num).value
        price = ws.cell(row, price_num).value
        lingshu = ws.cell(row, lingshu_num).value
        dunshu = ws.cell(row, dunshu_num).value
        jiner = ws.cell(row, jiner_num).value
        jizhang = ws.cell(row, jizhang_num).value

        if gongyingshang != '河南省江河纸业有限公司':
            pricename = pricename              #
        else :
            pricename = cailiao


        #期间、令数不为None、吨数不为None、pricename不为None
        if jizhang not in  ['',None]:          #有吨数和有记账标识的才做字典！
            key =  (gongyingshang,dangqian_qijian,pricename,price)
            if key not in zhiprices.keys():
                zhiprices[key] = (lingshu,dunshu,jiner)
            else :
                lingshu += zhiprices[key][0]
                dunshu +=zhiprices[key][1]
                jiner +=zhiprices[key][2]
                zhiprices[key] = (lingshu, dunshu, jiner)
        else :
            continue

    wb.close()
    return zhiprices

def main():
    #qijian = input('请输入期间，如2020-04\n')
    zhiprices = zhipriceDic()
    print(zhiprices)


if __name__ == '__main__':
    main()







