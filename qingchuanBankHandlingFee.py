# -*-coding = utf-8 -*-
import pdfplumber
import re
import openpyxl
import easygui
import os
from itertools import groupby

def writeDangtian(fname):
    pdf = pdfplumber.open(fname)
    # print('开始读取数据')
    filename = '晴川发票RECEIPT pdf.txt'
    with open(filename, 'w') as f:
        f.write('')
    with open(filename, 'a') as f:
        for page in pdf.pages:
            content = page.extract_text()
            f.write(content)
    pdf.close()
    return filename

def chuliData(data0):
    kehuhao_rows = []
    for j in range(len(data0)):
        if data0[j].startswith('客户号：'):
            kehuhao_rows.append(j)
    data1 = data0.copy()
    for j in kehuhao_rows:
        data1[j] = 'kehuhao'
    df0 = [['kehuhao']+ list(g) for k, g in groupby(data1, lambda x: x == 'kehuhao') if not k]
    df0 = df0[1:]
    df = df0.copy()
    for j in range(len(df)):
        kehuhao = data0[kehuhao_rows[j]]
        df[j][0] = kehuhao
    return df

def txtToData(filename):
    data0 = []
    with open(filename) as f:
       for readline in f.readlines():
            data0.append(readline.strip())
    return data0

def matchNumber(df8):
    regex_riqi = re.compile(r'日期：(?P<year>\d{4})年(?P<month>\d{2})月(?P<day>\d{2})日')
    regex_jiner = re.compile(r'金额：CNY(\d{1,3}(,\d{3})*\.\d{1,2})')
    regex_liushuihao = re.compile(r'交易流水号：(\d{6,10}-\d{1,4})')
    regex_name = re.compile(r'(?P<fee_name>.*\s*.*手续费)\s*\d{1,3}(,\d{3})*\.\d{2}')
    # regex_name = re.compile(r'费用名称：(\w+手续费)')
    riqi = '2010 年1月12日'
    jiner = 0
    name = '对公跨行柜台转账汇款手续费'
    liushuihao = '88888888-888'
    result = []
    for hang in df8:
        tmp_hang = []
        for j in hang:
            mat_riqi = regex_riqi.search(j)
            mat_jiner = regex_jiner.search(j)
            mat_liushuihao = regex_liushuihao.search(j)
            mat_name = regex_name.search(j)
            if mat_riqi:
                year = mat_riqi.group('year')
                month = mat_riqi.group('month')
                day = mat_riqi.group('day')
                riqi = year + '/' + str(int(month)) + '/' + str(int(day))
            elif mat_jiner:
                jiner = mat_jiner.group(1)
                jiner = jiner.replace(',','')
                jiner = float(jiner)
            elif mat_name:
                name = mat_name.group('fee_name')
            elif mat_liushuihao:
                liushuihao = mat_liushuihao.group(1)
            else:
                continue
        year_month = year + month
        for j in [riqi,jiner,name,liushuihao]:
            tmp_hang.append(j)
        result.append(tmp_hang)
    return result,year_month

def mAres(ws):    #获取合并单元格地址
    m_list = ws.merged_cells
    cr = []
    for m_area in m_list:
        # print(m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col)
        # 合并单元格的起始行坐标、终止行坐标。。。。，
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # 纵向合并单元格的位置信息提取出
        if c2 - c1 > 0:
            cr.append((r1, r2, c1, c2))
        else:
            continue

    cr.sort()
    return cr

def unmergeAres(nws,area):
    nws.unmerge_cells(start_row=area[0],start_column=area[2],end_row=area[1],end_column=area[3])

def mergeAres(nws,area):
    nws.merge_cells(start_row=area[0],start_column=area[2],end_row=area[1],end_column=area[3])

def painter(ws,cell):
    j = cell
    #字休
    font_name = j.font.name
    font_size = j.font.size
    font_italic = j.font.italic
    font_bold = j.font.bold
    font_color = j.font.color
    font = openpyxl.styles.Font(name=font_name, size=font_size , bold=font_bold, italic=font_italic, color=font_color)
    #边框
    border_left = j.border.left
    border = openpyxl.styles.Border(left=border_left, right=border_left , top=border_left, bottom=border_left)
    #填充
    fill_type = j.fill.fill_type
    fgColor = j.fill.fgColor
    pattern_fill = openpyxl.styles.PatternFill(fill_type=fill_type, fgColor=fgColor)
    #倾斜
    horizontal = j.alignment.horizontal
    vertical = j.alignment.vertical
    text_rotation = j.alignment.text_rotation
    wrap_text = j.alignment.wrap_text
    alignment = openpyxl.styles.Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text)
    #数字格式
    number_format = j.number_format
    #protection
    locked = j.protection.locked
    hidden = j.protection.locked
    protection =  openpyxl.styles.protection.Protection(locked=locked,hidden=hidden)
    #行高
    height = ws.row_dimensions[j.row].height
    return font,border,pattern_fill,alignment,number_format,protection,height

def writeToExcel(ws,nws,result,kehuming,shibihao,kemuhao,fapiao,fenshu,suilu,row_start=5):
    for j in range(len(result)):
        riqi = result[j][0]
        jiner = float(result[j][1])
        name = result[j][2]
        liushuihao = result[j][3]
        row = j + row_start
        nws.cell(row, 1).value = j + 1
        nws.cell(row, 2).value = kehuming
        nws.cell(row, 3).value = shibihao
        nws.cell(row, 4).value = kemuhao
        nws.cell(row, 5).value = riqi
        nws.cell(row, 6).value = liushuihao
        nws.cell(row, 7).value = name
        nws.cell(row, 8).value = fapiao
        nws.cell(row, 9).value = fenshu
        nws.cell(row, 10).value = jiner
        nws.cell(row, 11).value = round(jiner / 1.06, 2)
        nws.cell(row, 12).value = suilu
        nws.cell(row, 13).value = jiner - round(jiner / 1.06, 2)

def getStyles(ws,area0):
    styles = []
    for j in area0:
        font,border,pattern_fill,alignment,number_format,protection,height = painter(ws,j)
        style = [font,border,pattern_fill,alignment,number_format,protection,height]
        styles.append(style)
    return styles

def writeStyles(nws,styles,result,max_column,row_start=5):
    len_result = len(result)
    for i in range(row_start,len_result+row_start):
        for j in range(1,max_column+1):
            nws.cell(i,j).font  = styles[j-1][0]
            nws.cell(i,j).border  = styles[j-1][1]
            nws.cell(i,j).fill  = styles[j-1][2]
            nws.cell(i,j).alignment  = styles[j-1][3]
            nws.cell(i,j).number_format  = styles[j-1][4]
            nws.cell(i,j).protection  = styles[j-1][5]
            height  = styles[j-1][6]
        nws.row_dimensions[i].height = height
    return nws

def toEexcel(result,year_month):
    row_start = 5
    msg = '请点选"增值税发票开具申请表.xlsx"'
    fname = easygui.fileopenbox(msg)
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    ws_name = ws.title
    max_column = ws.max_column
    max_row = ws.max_row
    tail_row_start = max_row - 7
    tail_areas0 = list(ws)[tail_row_start-1:]
    nws = wb.copy_worksheet(ws)
    nws.title = year_month
    kehuming = nws.cell(row_start,2).value
    shibihao = nws.cell(row_start,3).value
    kemuhao = nws.cell(row_start, 4).value
    fapiao = nws.cell(row_start, 8).value
    fenshu = nws.cell(row_start, 9).value
    suilu = nws.cell(row_start, 12).value
    xuhao_col = [j.value for j in nws['A']]
    total_row = max_row - 7
    len_old = total_row  - 5
    m_area = mAres(ws)                  #获取合并单元格地址
    merge_area = [j for j in m_area if j[0] >= total_row]
    for area in merge_area:
        unmergeAres(nws,area)
    if len(result) > len_old:
        nws.insert_rows(total_row, len(result) - len_old)
    elif  len(result) < len_old:
        for j in range(total_row-1,total_row-1-(len_old-len(result)),-1):
            nws.delete_rows(j)
    else :
        pass
    wb.save(fname)
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    nws = wb[year_month]
    writeToExcel(ws,nws, result, kehuming, shibihao, kemuhao, fapiao, fenshu, suilu, row_start)
    nws.cell(len(result)+row_start,10).value = '=sum(j5:j{})'.format(len(result)+row_start-1)
    chayi = len(result) - len_old
    new_merge_arer = [(j[0]+chayi,j[1]+chayi,j[2],j[3]) for j in merge_area]
    area0 = ws[row_start]
    styles = getStyles(ws,area0)
    nws = writeStyles(nws,styles,result,max_column,row_start)
    for j in new_merge_arer:
        mergeAres(nws,j)
    nws = wb[year_month]
    wb.save(fname)
    os.startfile(fname)

def main():
    row_start = 5
    msg = '请点选银行RECEIPT.pdf文件'
    fname = easygui.fileopenbox(msg)
    # fname = r'd:\a00nutstore\fishc\qingchuan\晴川发票RECEIPT-20210701-20211228-91871730.pdf'
    path, file = os.path.split(fname)
    os.chdir(path)
    filename = writeDangtian(fname)
    data0 = txtToData(filename)
    df = chuliData(data0)
    df8 = [j for j in df for i in j if ('手续费' in i) and (i.startswith('附言')==False)]
    df9 = df8.copy()
    nums = set()
    for j in range(len(df9)):
        for i in df9[j]:
            if ('支票手续费' in i) or ('退货手续费' in i) or ('退还手续费' in i) :
                nums.add(j)
            else:
                continue
    ys = [df9[num] for num in list(nums)]
    for j in list(ys):
        df9.remove(j)
    result, year_month = matchNumber(df9)
    toEexcel(result, year_month)


if __name__ == '__main__':
    main()


